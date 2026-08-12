"""sbepv.model

Our own physics-based PV prediction model for the SBE Innovation Center PV
(STAC1) East array. Consumes the historian CSV (stac1.csv) and predicts AC power
and cumulative energy for the two systems, then compares against measured:

  - SolarEdge  -> module-level optimization (sum of module Pmp)
  - Solectria  -> string-level mismatch (pvmismatch)

This is a fresh, automated reimplementation of the physics in pvmismatch_Ho_v8.py
(the reference blueprint, which is NOT modified). Differences from the reference:
  - Runs with no prompts; all settings are the constants below.
  - Reads OUR column names (see COLUMN_RENAME) from stac1.csv.
  - Uses MEASURED DHI from the CSV when present (only derives it from GHI/DNI when
    DHI is missing).

INPUT  (stac1.csv, produced by sbepv.ingest.bazefield):
  timestamp, solaredge_measured_power, solectria_measured_power,
  dni, ghi, dhi, temp_air, wind_speed          (power in W, irradiance W/m2, degC, m/s)

OUTPUT:
  <OUTPUT_BASE>_ac_power.png, <OUTPUT_BASE>_cumulative_energy.png
  <OUTPUT_BASE>.xlsx  (tabs: time_series, tilts_and_strings, run_info)
"""

from __future__ import annotations

from collections.abc import Mapping
from copy import copy, deepcopy
import hashlib
import json
import secrets

import numpy as np
import pandas as pd
import matplotlib

matplotlib.use("Agg")  # non-interactive: save PNGs, never block on show()
import matplotlib.dates as mdates
import matplotlib.pyplot as plt

import pvlib as pvl
import pvmismatch as pvm
from pvmismatch.contrib import gen_coeffs

from sbepv.calibration import (
    _bounded_interval_hours,
    apply_frozen_seasonal_calibration,
    apply_seasonal_calibration,
)

__version__ = "5"

# -----------------------------------------------------------------------------
# RUN SETTINGS (edit here -- no command-line prompts)
# -----------------------------------------------------------------------------
INPUT_CSV = "stac1.csv"
OUTPUT_BASE = "stac1_model"
INPUT_IS_UTC = True  # historian writes UTC timestamps
TIMEZONE = "America/Denver"  # local tz for display/indexing
MAX_TEMPERATURE_INTERPOLATION_GAP_ROWS = 2

# Solectria XGI 1500-250 nameplate constraints. The CEC efficiency is used as
# the documented constant-efficiency approximation until a validated part-load
# curve is available for this site.
SOLECTRIA_INVERTER_MODEL = "Solectria XGI 1500-250"
SOLECTRIA_INVERTER_MPPT_MIN_V = 860.0
SOLECTRIA_INVERTER_MPPT_MAX_V = 1_250.0
SOLECTRIA_INVERTER_AC_RATING_W = 250_000.0
SOLECTRIA_INVERTER_CEC_EFFICIENCY = 0.985

# AC conversion efficiencies (predicted DC * eff). 1.0 = no derate.
SE_EFF = 1.0
SOL_EFF = SOLECTRIA_INVERTER_CEC_EFFICIENCY
DEFAULT_CURTAILMENT_LIMIT_KW = 125.0

# Incidence-angle modifier. Physical is the canonical default; INCLUDE_IAM is
# retained for callers that still use the deployed dashboard's legacy flag.
IAM_MODEL_PHYSICAL = "physical"
IAM_MODEL_MARTIN_RUIZ = "martin_ruiz"
IAM_MODELS = frozenset({IAM_MODEL_PHYSICAL, IAM_MODEL_MARTIN_RUIZ})
INCLUDE_IAM = None
A_R = 0.2


def resolve_iam_model(
    iam_model: str | None = None,
    include_iam: bool | None = None,
) -> str:
    """Return a canonical IAM model, honoring an explicit selection first.

    ``include_iam`` is the deployed dashboard's legacy customization flag.
    Either boolean value implies Martin-Ruiz; an omitted legacy flag and an
    omitted explicit model select Physical.
    """
    if iam_model is None:
        return (
            IAM_MODEL_PHYSICAL
            if include_iam is None
            else IAM_MODEL_MARTIN_RUIZ
        )

    selected = str(iam_model).strip().lower()
    if selected not in IAM_MODELS:
        choices = ", ".join(sorted(IAM_MODELS))
        raise ValueError(f"IAM model must be one of: {choices}.")
    return selected


def resolve_iam_settings(
    iam_model: str | None = None,
    iam_a_r: float | None = A_R,
    include_iam: bool | None = None,
) -> tuple[str, float | None]:
    """Resolve model selection and validate ``a_r`` only for Martin-Ruiz."""
    selected = resolve_iam_model(iam_model, include_iam)
    if selected == IAM_MODEL_PHYSICAL:
        return selected, None

    # Legacy ``include_iam=False`` meant "use Martin-Ruiz's default a_r" and
    # ignored the submitted coefficient. Canonical explicit selection always
    # uses the supplied coefficient.
    candidate = A_R if iam_model is None and include_iam is False else iam_a_r
    try:
        value = float(candidate)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            "Martin-Ruiz a_r must be a positive finite value."
        ) from exc
    if not np.isfinite(value) or value <= 0:
        raise ValueError("Martin-Ruiz a_r must be a positive finite value.")
    return selected, value


def resolve_iam_a_r(include_iam: bool, iam_a_r: float | None = A_R) -> float:
    """Return the legacy selector's effective coefficient.

    Kept for API compatibility. A false customization flag returns the default
    without validating the unused submitted value.
    """
    _, value = resolve_iam_settings(
        iam_model=None,
        iam_a_r=iam_a_r,
        include_iam=include_iam,
    )
    return A_R if value is None else value


# -----------------------------------------------------------------------------
# SITE / GEOMETRY / LAYOUT (copied verbatim from pvmismatch_Ho_v8.py)
# -----------------------------------------------------------------------------
LAT, LON = 39.7552, -104.6184  # SolarTAC SBE coordinates

AXIS_AZIMUTH = 180
MAX_ANGLE = 60
GCR = 0.4
BACKTRACK = True

INVERTER_PARAMETERS = {"pdc0": 9e9, "eta_inv_nom": 1}

MODULES_PER_BAY = 6
SOLECTRIA_STRINGS = 10
SOLECTRIA_BAYS_PER_STRING = 4
SOLAREDGE_STRINGS = 5
SOLAREDGE_BAYS_PER_STRING = 8

# Sandia module temperature model parameters
TEMPERATURE_MODEL_PARAMETERS = {"a": -3.47, "b": -0.0594, "deltaT": 0}

# As-built torque-tube slopes (axis_tilt, degrees) -- verbatim from reference.
SOLAREDGE_TILT_ASBUILT = [
    [3.05, 0.96, -1.35, -7.27, 7.31, -0.58, -0.05, -3.46],
    [-4.95, 7.62, -7.19, -0.26, -3.52, -5.27, 4.22, 2.67],
    [3.16, 0.71, -1.26, -7.46, 7.56, -0.57, 0.56, 2.87],
    [-3.4, -5.29, 7.06, -0.53, -0.12, -7.27, 2.86, 2.4],
    [2.78, -3.38, -5, 7.31, -0.57, -0.12, -7.28, 2.59],
]

SOLECTRIA_TILT_ASBUILT = [
    [3.14, 0.93, -1.41, -7.44],
    [7.52, -0.84, -0.03, -3.47],
    [-4.9, 7.63, -7.32, 4.27],
    [-0.5, -0.11, -3.58, -5.1],
    [0.73, -1.3, -7.45, 7.2],
    [2.78, 2.76, -3.26, -4.99],
    [7.28, -0.82, -0.08, -7.21],
    [2.59, 2.78, -3.9, 2.55],
    [0.62, -4.96, 7.61, -0.86],
    [-0.23, -7.16, 2.57, 2.6],
]

# -----------------------------------------------------------------------------
# MODULE PARAMETERS -- WAAREE BiN-08-580 (bifacial), verbatim from reference
# -----------------------------------------------------------------------------
_modules = pvl.pvsystem.retrieve_sam("CECMod")
_modules["WAAREE_BIN_08_580"] = _modules.index.map(
    {
        "Technology": "Mono-c-Si",
        "Bifacial": 1,
        "STC": 579.92,
        "PTC": 550.7,
        "A_c": 2.56,
        "Length": 0,
        "Width": 0,
        "N_s": 72,
        "I_sc_ref": 13.93,
        "V_oc_ref": 52.5,
        "I_mp_ref": 13.18,
        "V_mp_ref": 44,
        "alpha_sc": 0.0045969,
        "beta_oc": -0.12548,
        "T_NOCT": 43.2,
        "a_ref": 1.82068,
        "I_L_ref": 13.9415,
        "I_o_ref": 4.12e-12,
        "R_s": 0.206355,
        "R_sh_ref": 249.122,
        "Adjust": 3.05545,
        "gamma_r": -0.302,
        "BIPV": "N",
        "Version": "SAM 2023.12.17",
        "Date": "11/14/2024",
    }
)
MODULE_PARAMETERS = _modules["WAAREE_BIN_08_580"]
MODULE_NAME = "WAAREE_BIN_08_580"

# The installed module nameplate photographed at the site identifies BIN-08-580
# and matches the configured STC points below. These cell-level coefficients were
# frozen after a physically seeded, converged PVMismatch 4.1 two-diode fit. The
# runtime builder never invokes the nonlinear solver, so a dependency update
# cannot silently move the production electrical baseline.
SOLECTRIA_PHYSICS_VERSION = "solectria-bin08-580-pvmismatch-v3"
CALIBRATION_PHYSICS_VERSION = "sbe-stac1-calibration-physics-v1"
SOLECTRIA_TWO_DIODE_COEFFICIENTS = (
    6.051338067069484e-12,
    7.246006548395709e-7,
    0.002562727221945137,
    9.52605775638587,
)
SOLECTRIA_BAND_GAP_EV = 1.16
SOLECTRIA_FIT_RESIDUAL_TOLERANCE = 1e-6
SOLECTRIA_DATASHEET_RELATIVE_TOLERANCE = 0.005


def _module_parameter_float(name: str) -> float:
    return float(MODULE_PARAMETERS[name])


def solectria_relative_alpha_isc() -> float:
    """Return the relative 1/K coefficient expected by PVMismatch."""

    return _module_parameter_float("alpha_sc") / _module_parameter_float(
        "I_sc_ref"
    )


def solectria_equivalent_cell_area_cm2() -> float:
    """Return module area divided over the configured equivalent series cells."""

    return (
        _module_parameter_float("A_c")
        * 10_000.0
        / _module_parameter_float("N_s")
    )


def solectria_physics_manifest() -> dict[str, object]:
    """Return the immutable electrical/topology contract used for calibration."""

    parameter_names = (
        "STC",
        "PTC",
        "A_c",
        "N_s",
        "I_sc_ref",
        "V_oc_ref",
        "I_mp_ref",
        "V_mp_ref",
        "alpha_sc",
        "beta_oc",
        "T_NOCT",
        "a_ref",
        "I_L_ref",
        "I_o_ref",
        "R_s",
        "R_sh_ref",
        "Adjust",
        "gamma_r",
    )
    return {
        "physics_version": SOLECTRIA_PHYSICS_VERSION,
        "module_name": MODULE_NAME,
        "module_parameters": {
            name: _module_parameter_float(name) for name in parameter_names
        },
        "two_diode_coefficients": list(SOLECTRIA_TWO_DIODE_COEFFICIENTS),
        "relative_alpha_isc_per_k": solectria_relative_alpha_isc(),
        "effective_band_gap_ev": SOLECTRIA_BAND_GAP_EV,
        "equivalent_cell_area_cm2": solectria_equivalent_cell_area_cm2(),
        "equivalent_series_cells": 72,
        "pvmismatch": {
            "curve_grid_points": int(pvm.pvconstants.NPTS),
            "module_topology": "STD72",
            "substring_count": 3,
            "cells_per_substring": 24,
            "substring_bypass_voltage_v": float(pvm.pvmodule.VBYPASS),
        },
        "module_count": (
            SOLECTRIA_STRINGS * SOLECTRIA_BAYS_PER_STRING * MODULES_PER_BAY
        ),
        "string_layout": "10 strings x 4 bays x 6 modules",
        "string_tilts_deg": [list(row) for row in SOLECTRIA_TILT_ASBUILT],
        "mppt_assumption": (
            "ten strings combined in parallel at one common inverter voltage"
        ),
        "inverter": {
            "model": SOLECTRIA_INVERTER_MODEL,
            "mppt_count": 1,
            "mppt_min_v": SOLECTRIA_INVERTER_MPPT_MIN_V,
            "mppt_max_v": SOLECTRIA_INVERTER_MPPT_MAX_V,
            "ac_rating_w": SOLECTRIA_INVERTER_AC_RATING_W,
            "cec_efficiency_default": SOLECTRIA_INVERTER_CEC_EFFICIENCY,
            "bos_efficiency_default": 1.0,
            "dc_aggregation_algorithm": "parallel_iv_common_mppt_v1",
            "conversion_model": "constant_efficiency_with_nameplate_clip",
        },
        "rear_side_irradiance": "not modeled",
    }


def solectria_physics_fingerprint() -> str:
    """Return a stable hash used to prevent cross-physics factor reuse."""

    payload = json.dumps(
        solectria_physics_manifest(),
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


SOLECTRIA_PHYSICS_FINGERPRINT = solectria_physics_fingerprint()


def calibration_physics_manifest() -> dict[str, object]:
    """Return the complete prediction contract bound to calibration factors.

    Calibration factors cover both arrays, so compatibility must include every
    shared and system-specific constant or algorithm that can change either
    prediction. Runtime scenario settings are preserved separately with the
    baseline; this manifest records their implementation and defaults.
    """

    shared_module_parameters = {
        name: _module_parameter_float(name)
        for name in (
            "STC",
            "PTC",
            "A_c",
            "N_s",
            "I_sc_ref",
            "V_oc_ref",
            "I_mp_ref",
            "V_mp_ref",
            "alpha_sc",
            "beta_oc",
            "T_NOCT",
            "a_ref",
            "I_L_ref",
            "I_o_ref",
            "R_s",
            "R_sh_ref",
            "Adjust",
            "gamma_r",
        )
    }
    return {
        "physics_version": CALIBRATION_PHYSICS_VERSION,
        "model_version": __version__,
        "dependencies": {
            "pvlib": str(getattr(pvl, "__version__", "unknown")),
            "pvmismatch": str(getattr(pvm, "__version__", "unknown")),
        },
        "input_weather": {
            "input_is_utc": INPUT_IS_UTC,
            "local_timezone": TIMEZONE,
            "maximum_temperature_interpolation_gap_rows": (
                MAX_TEMPERATURE_INTERPOLATION_GAP_ROWS
            ),
            "dhi_contract": "measured_when_finite_otherwise_ghi_dni_geometry",
        },
        "site": {"latitude": LAT, "longitude": LON},
        "module": {
            "name": MODULE_NAME,
            "parameters": shared_module_parameters,
            "rear_side_irradiance": "not modeled",
        },
        "tracker": {
            "axis_azimuth_deg": AXIS_AZIMUTH,
            "maximum_angle_deg": MAX_ANGLE,
            "gcr": GCR,
            "backtrack_default": BACKTRACK,
            "implementation": "pvlib.SingleAxisTrackerMount",
        },
        "modelchain": {
            "clearsky_model": "ineichen",
            "transposition_model": "haydavies",
            "solar_position_method": "nrel_numpy",
            "airmass_model": "kastenyoung1989",
            "dc_model": "cec",
            "spectral_model": "no_loss",
            "dc_ohmic_model": "no_loss",
            "losses_model": "no_loss",
            "placeholder_inverter_parameters": dict(INVERTER_PARAMETERS),
            "modelchain_ac_output_consumed": False,
        },
        "temperature": {
            "model": "sapm",
            "parameters": dict(TEMPERATURE_MODEL_PARAMETERS),
        },
        "iam": {
            "models": sorted(IAM_MODELS),
            "default_model": IAM_MODEL_PHYSICAL,
            "martin_ruiz_default_a_r": A_R,
            "application_count": 1,
            "implementation": "physical_in_modelchain_or_martin_ruiz_then_cec",
        },
        "solaredge": {
            "module_count": (
                SOLAREDGE_STRINGS
                * SOLAREDGE_BAYS_PER_STRING
                * MODULES_PER_BAY
            ),
            "string_count": SOLAREDGE_STRINGS,
            "bays_per_string": SOLAREDGE_BAYS_PER_STRING,
            "modules_per_bay": MODULES_PER_BAY,
            "string_tilts_deg": [
                list(row) for row in SOLAREDGE_TILT_ASBUILT
            ],
            "dc_aggregation_algorithm": "sum_individual_module_cec_pmp_v1",
            "ac_conversion": {
                "model": "request_efficiency_scalars",
                "inverter_efficiency_default": SE_EFF,
                "bos_efficiency_default": 1.0,
            },
            "hardware_clipping": "not modeled",
        },
        "solectria": solectria_physics_manifest(),
        "post_prediction": {
            "calibration": (
                "denver_meteorological_season_scalar_per_system_caps_v2"
            ),
            "solectria_hardware_cap_stage": "before_and_after_calibration",
            "factor_solver_caps": (
                "solaredge_optional_user_cap;solectria_min_hardware_user_cap"
            ),
            "curtailment": "optional_predicted_ac_cap_after_calibration",
            "default_curtailment_limit_kw": DEFAULT_CURTAILMENT_LIMIT_KW,
        },
    }


def calibration_physics_fingerprint() -> str:
    """Return the stable hash for both arrays and all shared physics."""

    payload = json.dumps(
        calibration_physics_manifest(),
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


CALIBRATION_PHYSICS_FINGERPRINT = calibration_physics_fingerprint()


def validate_calibration_profile_physics(
    profile: Mapping[str, object],
) -> None:
    """Reject factors not fitted with the complete current prediction core."""

    calibration_version = profile.get("calibration_physics_version")
    calibration_fingerprint = profile.get("calibration_physics_fingerprint")
    if not isinstance(calibration_version, str) or not calibration_version.strip():
        raise ValueError(
            "Calibration profile is missing its calibration physics version; "
            "run a fresh reviewed calibration."
        )
    if (
        not isinstance(calibration_fingerprint, str)
        or not calibration_fingerprint.strip()
    ):
        raise ValueError(
            "Calibration profile is missing its calibration physics fingerprint; "
            "run a fresh reviewed calibration."
        )
    if not secrets.compare_digest(
        calibration_version.strip(), CALIBRATION_PHYSICS_VERSION
    ):
        raise ValueError(
            "Calibration profile uses an incompatible calibration physics "
            "version; run a fresh reviewed calibration."
        )
    if not secrets.compare_digest(
        calibration_fingerprint.strip().lower(),
        CALIBRATION_PHYSICS_FINGERPRINT,
    ):
        raise ValueError(
            "Calibration profile uses an incompatible calibration physics "
            "fingerprint; run a fresh reviewed calibration."
        )

    version = profile.get("solectria_physics_version")
    fingerprint = profile.get("solectria_physics_fingerprint")
    if not isinstance(version, str) or not version.strip():
        raise ValueError(
            "Calibration profile is missing its Solectria physics version; "
            "run a fresh reviewed calibration."
        )
    if not isinstance(fingerprint, str) or not fingerprint.strip():
        raise ValueError(
            "Calibration profile is missing its Solectria physics fingerprint; "
            "run a fresh reviewed calibration."
        )
    if not secrets.compare_digest(version.strip(), SOLECTRIA_PHYSICS_VERSION):
        raise ValueError(
            "Calibration profile uses an incompatible Solectria physics version; "
            "run a fresh reviewed calibration."
        )
    if not secrets.compare_digest(
        fingerprint.strip().lower(), SOLECTRIA_PHYSICS_FINGERPRINT
    ):
        raise ValueError(
            "Calibration profile uses an incompatible Solectria physics "
            "fingerprint; run a fresh reviewed calibration."
        )

# Our CSV column names -> the canonical names used internally.
COLUMN_RENAME = {
    "solaredge_measured_power": "se_measured_power_w",
    "solectria_measured_power": "sol_measured_power_w",
    "dni": "dni_wm2",
    "ghi": "ghi_wm2",
    "dhi": "dhi_wm2",
    "temp_air": "temp_air_c",
    "wind_speed": "wind_speed_ms",
}

MIDC_COLUMNS = {
    "DATE (MM/DD/YYYY)": "date",
    "HOUR-MST": "hour_mst",
    "Avg Global Horizontal [W/m^2]": "ghi_wm2",
    "Avg Direct Normal [W/m^2]": "dni_wm2",
    "Avg Diffuse Horizontal [W/m^2]": "dhi_wm2",
    "Avg Air Temperature [deg C]": "temp_air_c",
    "Avg Avg Wind Speed @ 10m [m/s]": "wind_speed_ms",
}
MIDC_OPTIONAL_COLUMNS = {
    "MINUTE-MST": "minute_mst",
}


# -----------------------------------------------------------------------------
# I/O + PRE-PROCESSING
# -----------------------------------------------------------------------------
def _bounded_temperature_interpolation(values: pd.Series) -> pd.Series:
    """Interpolate only short consecutive temperature gaps."""

    numeric = pd.to_numeric(values, errors="coerce").replace(
        [np.inf, -np.inf], np.nan
    )
    missing = numeric.isna()
    if not missing.any() or missing.all():
        return numeric
    groups = missing.ne(missing.shift(fill_value=False)).cumsum()
    missing_run_lengths = missing.groupby(groups).transform("sum")
    interpolated = numeric.interpolate(limit_direction="both")
    interpolated.loc[
        missing
        & (
            missing_run_lengths
            > int(MAX_TEMPERATURE_INTERPOLATION_GAP_ROWS)
        )
    ] = np.nan
    return interpolated


def parse_input_csv(path: str) -> pd.DataFrame:
    """Read stac1.csv, rename to canonical columns, index by local (tz-aware) time."""
    df = pd.read_csv(path)
    df = df.rename(columns=COLUMN_RENAME)
    df = df.dropna(subset=["timestamp"]).copy()

    ts = pd.to_datetime(df["timestamp"], errors="coerce")
    df = df.loc[~ts.isna()].copy()
    ts = ts.loc[~ts.isna()]
    if df.empty:
        raise ValueError("Historian file contains no valid timestamp rows.")

    if getattr(ts.dt, "tz", None) is None:
        if INPUT_IS_UTC:
            ts_local = ts.dt.tz_localize(
                "UTC", ambiguous="infer", nonexistent="shift_forward"
            ).dt.tz_convert(TIMEZONE)
        else:
            ts_local = ts.dt.tz_localize(
                TIMEZONE, ambiguous="infer", nonexistent="shift_forward"
            )
    else:
        ts_local = ts.dt.tz_convert(TIMEZONE)

    df.insert(0, "timestamp_local", ts_local)
    df["timestamp_utc"] = df["timestamp_local"].dt.tz_convert("UTC")
    df = df.set_index("timestamp_local").sort_index()

    numeric = [
        "se_measured_power_w",
        "sol_measured_power_w",
        "dni_wm2",
        "ghi_wm2",
        "dhi_wm2",
        "temp_air_c",
        "wind_speed_ms",
    ]
    for c in numeric:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").replace(
                [np.inf, -np.inf], np.nan
            )

    required_weather = ("dni_wm2", "ghi_wm2", "temp_air_c")
    all_missing = [
        column for column in required_weather if not df[column].notna().any()
    ]
    if all_missing:
        labels = {
            "dni_wm2": "DNI",
            "ghi_wm2": "GHI",
            "temp_air_c": "air temperature",
        }
        raise ValueError(
            "Historian weather is missing for the entire date range: "
            + ", ".join(labels[column] for column in all_missing)
            + "."
        )

    df["dni_wm2"] = df["dni_wm2"].fillna(0.0)
    df["ghi_wm2"] = df["ghi_wm2"].fillna(0.0)
    df["temp_air_c"] = _bounded_temperature_interpolation(df["temp_air_c"])
    unfilled_temperature_rows = int(df["temp_air_c"].isna().sum())
    if unfilled_temperature_rows:
        raise ValueError(
            "Historian air-temperature gaps exceed the bounded fallback "
            f"of {MAX_TEMPERATURE_INTERPOLATION_GAP_ROWS} consecutive row(s); "
            "exclude the affected rows and retry."
        )
    df["wind_speed_ms"] = (
        df["wind_speed_ms"].interpolate(limit_direction="both").fillna(0.0)
    )
    return df


def parse_midc_csv(path: str) -> tuple[pd.DataFrame, list[str]]:
    """Read dashboard-generated MIDC interval data using fixed MST timestamps."""
    raw = pd.read_csv(path)
    missing_columns = sorted(set(MIDC_COLUMNS).difference(raw.columns))
    if missing_columns:
        raise ValueError(
            "MIDC file is missing required columns: " + ", ".join(missing_columns)
        )
    raw = raw.rename(columns={**MIDC_COLUMNS, **MIDC_OPTIONAL_COLUMNS})
    parsed_dates = pd.to_datetime(raw["date"], format="%m/%d/%Y", errors="coerce")
    hours = pd.to_numeric(raw["hour_mst"], errors="coerce")
    minutes = (
        pd.to_numeric(raw["minute_mst"], errors="coerce")
        if "minute_mst" in raw
        else pd.Series(0, index=raw.index, dtype="int64")
    )
    valid_time = (
        parsed_dates.notna()
        & hours.between(0, 23)
        & hours.mod(1).eq(0)
        & minutes.between(0, 59)
        & minutes.mod(1).eq(0)
    )
    warnings: list[str] = []
    invalid_time_count = int((~valid_time).sum())
    if invalid_time_count:
        warnings.append(
            f"Dropped {invalid_time_count} MIDC row(s) with invalid date/time values."
        )
    raw = raw.loc[valid_time].copy()
    parsed_dates = parsed_dates.loc[valid_time]
    hours = hours.loc[valid_time].astype(int)
    minutes = minutes.loc[valid_time].astype(int)
    if raw.empty:
        raise ValueError("MIDC file contains no valid date/time rows.")

    timestamp_mst = (
        parsed_dates
        + pd.to_timedelta(hours, unit="h")
        + pd.to_timedelta(minutes, unit="m")
    ).dt.tz_localize("Etc/GMT+7")
    timestamp_local = timestamp_mst.dt.tz_convert(TIMEZONE)

    data = pd.DataFrame(index=timestamp_local)
    for column in ("ghi_wm2", "dni_wm2", "dhi_wm2", "temp_air_c", "wind_speed_ms"):
        data[column] = pd.to_numeric(raw[column], errors="coerce").to_numpy()

    missing_counts = data[
        ["ghi_wm2", "dni_wm2", "dhi_wm2", "temp_air_c", "wind_speed_ms"]
    ].isna().sum()
    if int(missing_counts.sum()):
        detail = ", ".join(
            f"{column}={int(count)}"
            for column, count in missing_counts.items()
            if count
        )
        warnings.append(
            "MIDC interval gaps were filled for modeling "
            f"({detail}): GHI/DNI use zero, temperature/wind interpolate, "
            "and DHI uses a solar-position-derived fallback."
        )

    for column in ("ghi_wm2", "dni_wm2", "dhi_wm2"):
        data[column] = data[column].clip(lower=0.0)
    data["ghi_wm2"] = data["ghi_wm2"].fillna(0.0)
    data["dni_wm2"] = data["dni_wm2"].fillna(0.0)
    data["temp_air_c"] = data["temp_air_c"].interpolate(limit_direction="both")
    data["wind_speed_ms"] = (
        data["wind_speed_ms"].interpolate(limit_direction="both").fillna(0.0)
    )
    if data["temp_air_c"].isna().any():
        raise ValueError("MIDC air temperature is missing for the entire date range.")

    data["se_measured_power_w"] = 0.0
    data["sol_measured_power_w"] = 0.0
    data["timestamp_utc"] = data.index.tz_convert("UTC")
    data.index.name = "timestamp_local"
    if data.index.has_duplicates:
        raise ValueError("MIDC date/time keys contain duplicate timestamps.")
    return data.sort_index(), warnings


# -----------------------------------------------------------------------------
# PVLIB WEATHER + MODELCHAIN
# -----------------------------------------------------------------------------
def build_weather(
    df: pd.DataFrame, location: pvl.location.Location
) -> tuple[pd.DataFrame, str]:
    """Build the pvlib weather frame. Prefer MEASURED DHI; derive it only if absent."""
    times = df.index
    dni = df["dni_wm2"].to_numpy(dtype=float)
    ghi = df["ghi_wm2"].to_numpy(dtype=float)

    solpos = location.get_solarposition(times)
    cos_zen = np.cos(np.deg2rad(solpos["zenith"].clip(0, 90))).to_numpy()
    derived = np.maximum(ghi - dni * cos_zen, 0.0)

    if "dhi_wm2" in df.columns and df["dhi_wm2"].notna().any():
        dhi = df["dhi_wm2"].to_numpy(dtype=float)
        missing = np.isnan(dhi)
        dhi = np.where(missing, derived, dhi)
        dhi = np.clip(dhi, 0.0, None)
        dhi_source = (
            "measured with derived fallback"
            if missing.any()
            else "measured"
        )
    else:
        dhi = derived
        dhi_source = "derived (GHI-DNI*cos z)"

    weather = pd.DataFrame(
        index=times,
        data={
            "dni": dni,
            "ghi": ghi,
            "dhi": dhi,
            "temp_air": df["temp_air_c"].to_numpy(),
            "wind_speed": df["wind_speed_ms"].to_numpy(),
        },
    )
    return weather, dhi_source


def _cec_p_mp_from_effective_irradiance(
    effective_irradiance: pd.Series,
    cell_temperature: pd.Series,
    fallback_p_mp: pd.Series,
) -> pd.Series:
    """Recalculate CEC maximum power after a manually applied IAM.

    ModelChain's ``no_loss`` DC result was calculated before Martin-Ruiz was
    applied.  Re-running only the CEC electrical step makes SolarEdge consume
    the same adjusted irradiance that is passed to Solectria, without applying
    IAM a second time to either system.
    """

    effective = pd.Series(effective_irradiance, copy=False)
    temperature = pd.Series(cell_temperature, index=effective.index, copy=False)
    fallback = pd.Series(fallback_p_mp, index=effective.index, copy=False)
    finite_effective = pd.Series(
        np.isfinite(effective.to_numpy(dtype=float)),
        index=effective.index,
        dtype=bool,
    )
    positive_effective = finite_effective & effective.gt(0.0)
    recalculated = fallback.astype(float).copy()
    recalculated.loc[finite_effective & ~positive_effective] = 0.0

    # Zero irradiance is exactly zero power and need not enter pvlib's root
    # solver. Besides avoiding unnecessary annual-run work, this prevents the
    # solver's zero-current divide warnings while retaining the prior fallback
    # behavior for non-finite irradiance.
    if positive_effective.any():
        diode_parameters = pvl.pvsystem.calcparams_cec(
            effective.loc[positive_effective],
            temperature.loc[positive_effective],
            alpha_sc=MODULE_PARAMETERS["alpha_sc"],
            a_ref=MODULE_PARAMETERS["a_ref"],
            I_L_ref=MODULE_PARAMETERS["I_L_ref"],
            I_o_ref=MODULE_PARAMETERS["I_o_ref"],
            R_sh_ref=MODULE_PARAMETERS["R_sh_ref"],
            R_s=MODULE_PARAMETERS["R_s"],
            Adjust=MODULE_PARAMETERS["Adjust"],
        )
        recalculated.loc[positive_effective] = pd.Series(
            pvl.pvsystem.singlediode(*diode_parameters)["p_mp"],
            index=effective.index[positive_effective],
            dtype=float,
        )
    return recalculated


def run_modelchain_for_axis_tilt(
    axis_tilt: float,
    weather: pd.DataFrame,
    location: pvl.location.Location,
    backtrack: bool = BACKTRACK,
    include_iam: bool | None = INCLUDE_IAM,
    iam_a_r: float | None = A_R,
    iam_model: str | None = None,
) -> dict:
    """Run pvlib ModelChain for a single tracker axis tilt -> Ee (suns), Tk, p_mp (W)."""
    selected_iam_model, effective_iam_a_r = resolve_iam_settings(
        iam_model=iam_model,
        iam_a_r=iam_a_r,
        include_iam=include_iam,
    )
    array = pvl.pvsystem.Array(
        mount=pvl.pvsystem.SingleAxisTrackerMount(
            axis_tilt=axis_tilt,
            axis_azimuth=AXIS_AZIMUTH,
            max_angle=MAX_ANGLE,
            backtrack=bool(backtrack),
            gcr=GCR,
        ),
        module_parameters=MODULE_PARAMETERS,
        temperature_model_parameters=TEMPERATURE_MODEL_PARAMETERS,
        modules_per_string=1,
        strings=1,
    )
    system = pvl.pvsystem.PVSystem(
        arrays=[array], inverter_parameters=INVERTER_PARAMETERS
    )

    # Physical is handled inside ModelChain. Martin-Ruiz runs ModelChain with no
    # AOI loss, then applies IAM to effective irradiance and reruns the CEC DC
    # step so both modeled systems use the same single IAM application.
    aoi_model = (
        "no_loss"
        if selected_iam_model == IAM_MODEL_MARTIN_RUIZ
        else IAM_MODEL_PHYSICAL
    )
    mc = pvl.modelchain.ModelChain(
        system, location, aoi_model=aoi_model, spectral_model="no_loss"
    )
    mc.run_model(weather)

    effective = mc.results.effective_irradiance
    p_mp = mc.results.dc.p_mp
    if selected_iam_model == IAM_MODEL_MARTIN_RUIZ:
        aoi = getattr(mc.results, "aoi", None)
        if aoi is None:
            raise AttributeError(
                "ModelChain results did not include AOI needed for IAM."
            )
        iam = pvl.iam.martin_ruiz(aoi, a_r=effective_iam_a_r)
        effective = effective * iam
        p_mp = _cec_p_mp_from_effective_irradiance(
            effective,
            mc.results.cell_temperature,
            p_mp,
        )

    return {
        "Ee_suns": (effective / 1000.0).to_numpy(),
        "Tk": (mc.results.cell_temperature + 273.15).to_numpy(),
        "p_mp_w": p_mp.to_numpy(),
    }


def run_modelchain_for_axis_tilts(
    axis_tilts: list[float],
    weather: pd.DataFrame,
    location: pvl.location.Location,
    backtrack: bool = BACKTRACK,
    include_iam: bool | None = INCLUDE_IAM,
    iam_a_r: float | None = A_R,
    iam_model: str | None = None,
) -> dict[float, dict[str, np.ndarray]]:
    """Run all tracker tilts in one multi-array pvlib ModelChain.

    Solar position and other shared inputs are prepared once. pvlib returns one
    result object per array, in the same order as the supplied tilts.
    """
    tilts = tuple(dict.fromkeys(float(tilt) for tilt in axis_tilts))
    if not tilts:
        return {}

    selected_iam_model, effective_iam_a_r = resolve_iam_settings(
        iam_model=iam_model,
        iam_a_r=iam_a_r,
        include_iam=include_iam,
    )
    arrays = [
        pvl.pvsystem.Array(
            mount=pvl.pvsystem.SingleAxisTrackerMount(
                axis_tilt=tilt,
                axis_azimuth=AXIS_AZIMUTH,
                max_angle=MAX_ANGLE,
                backtrack=bool(backtrack),
                gcr=GCR,
            ),
            module_parameters=MODULE_PARAMETERS,
            temperature_model_parameters=TEMPERATURE_MODEL_PARAMETERS,
            modules_per_string=1,
            strings=1,
            name=f"axis_tilt_{index}",
        )
        for index, tilt in enumerate(tilts)
    ]
    system = pvl.pvsystem.PVSystem(
        arrays=arrays,
        inverter_parameters=INVERTER_PARAMETERS,
    )
    aoi_model = (
        "no_loss"
        if selected_iam_model == IAM_MODEL_MARTIN_RUIZ
        else IAM_MODEL_PHYSICAL
    )
    mc = pvl.modelchain.ModelChain(
        system,
        location,
        aoi_model=aoi_model,
        spectral_model="no_loss",
    )
    mc.run_model(weather)

    def per_array(value):
        return value if isinstance(value, tuple) else (value,)

    effective_results = per_array(mc.results.effective_irradiance)
    temperature_results = per_array(mc.results.cell_temperature)
    dc_results = per_array(mc.results.dc)
    aoi_results = per_array(mc.results.aoi)

    output: dict[float, dict[str, np.ndarray]] = {}
    for index, tilt in enumerate(tilts):
        effective = effective_results[index]
        p_mp = dc_results[index].p_mp
        if selected_iam_model == IAM_MODEL_MARTIN_RUIZ:
            iam = pvl.iam.martin_ruiz(
                aoi_results[index],
                a_r=effective_iam_a_r,
            )
            effective = effective * iam
            p_mp = _cec_p_mp_from_effective_irradiance(
                effective,
                temperature_results[index],
                p_mp,
            )

        output[tilt] = {
            "Ee_suns": (effective / 1000.0).to_numpy(),
            "Tk": (temperature_results[index] + 273.15).to_numpy(),
            "p_mp_w": p_mp.to_numpy(),
        }
    return output


# -----------------------------------------------------------------------------
# PVMISMATCH MODULE (validated, frozen two-diode fit)
# -----------------------------------------------------------------------------
def fit_solectria_two_diode_coefficients(
    *, residual_tolerance: float = SOLECTRIA_FIT_RESIDUAL_TOLERANCE
) -> tuple[float, float, float, float]:
    """Reproduce the offline fit and fail closed on solver or residual errors."""

    number_series_cells = int(_module_parameter_float("N_s"))
    initial = np.array(
        [
            _module_parameter_float("I_o_ref"),
            1e-6,
            _module_parameter_float("R_s") / number_series_cells,
            _module_parameter_float("R_sh_ref") / number_series_cells,
        ],
        dtype=float,
    )
    coefficients, solver = gen_coeffs.gen_two_diode(
        _module_parameter_float("I_sc_ref"),
        _module_parameter_float("V_oc_ref"),
        _module_parameter_float("I_mp_ref"),
        _module_parameter_float("V_mp_ref"),
        number_series_cells,
        1,
        25.0,
        x0=initial,
    )
    values = np.asarray(coefficients, dtype=float)
    residual_norm = float(np.linalg.norm(solver.fun))
    if not bool(solver.success):
        raise RuntimeError(
            "WAAREE BIN-08-580 two-diode fit failed: "
            f"status={solver.status}, message={solver.message!s}"
        )
    if not np.isfinite(residual_norm) or residual_norm > float(
        residual_tolerance
    ):
        raise RuntimeError(
            "WAAREE BIN-08-580 two-diode fit residual is too large: "
            f"{residual_norm:.6g} > {float(residual_tolerance):.6g}"
        )
    if values.shape != (4,) or not bool(np.isfinite(values).all()):
        raise RuntimeError(
            "WAAREE BIN-08-580 two-diode fit returned invalid coefficients."
        )
    if bool((values <= 0.0).any()):
        raise RuntimeError(
            "WAAREE BIN-08-580 two-diode fit returned non-positive coefficients."
        )
    return tuple(float(value) for value in values)


def _build_pvmismatch_module_from_coefficients(
    coefficients: tuple[float, float, float, float],
) -> pvm.pvmodule.PVmodule:
    pv_cell = pvm.pvcell.PVcell(
        Isat1_T0=coefficients[0],
        Isat2_T0=coefficients[1],
        Rs=coefficients[2],
        Rsh=coefficients[3],
        Isc0_T0=_module_parameter_float("I_sc_ref"),
        alpha_Isc=solectria_relative_alpha_isc(),
        Eg=SOLECTRIA_BAND_GAP_EV,
        pvconst=pvm.pvconstants.PVconstants(),
    )
    number_series_cells = int(_module_parameter_float("N_s"))
    if number_series_cells != 72:
        raise ValueError(
            "The Solectria PVMismatch model is validated only for the "
            "configured 72-equivalent-cell STD72 topology."
        )
    return pvm.pvmodule.PVmodule(
        cell_pos=pvm.pvmodule.STD72,
        cellArea=solectria_equivalent_cell_area_cm2(),
        pvcells=[pv_cell] * number_series_cells,
    )


def solectria_module_power_point(
    module: pvm.pvmodule.PVmodule,
    *,
    irradiance_suns: float = 1.0,
    temperature_c: float = 25.0,
) -> dict[str, float]:
    """Evaluate one uniform module operating point."""

    module.setSuns(float(irradiance_suns))
    module.setTemps(float(temperature_c) + 273.15)
    power = np.asarray(module.Pmod, dtype=float)
    voltage = np.asarray(module.Vmod, dtype=float)
    current = np.asarray(module.Imod, dtype=float)
    maximum_index = int(np.nanargmax(power))
    return {
        "pmp_w": float(power[maximum_index]),
        "vmp_v": float(voltage[maximum_index]),
        "imp_a": float(current[maximum_index]),
        "voc_v": float(np.asarray(module.Voc, dtype=float).sum()),
        "isc_a": float(np.asarray(module.Isc, dtype=float).mean()),
    }


def validate_solectria_module(
    module: pvm.pvmodule.PVmodule,
    *,
    relative_tolerance: float = SOLECTRIA_DATASHEET_RELATIVE_TOLERANCE,
) -> dict[str, object]:
    """Validate nameplate STC points and temperature response before use."""

    stc = solectria_module_power_point(module)
    targets = {
        "pmp_w": _module_parameter_float("STC"),
        "vmp_v": _module_parameter_float("V_mp_ref"),
        "imp_a": _module_parameter_float("I_mp_ref"),
        "voc_v": _module_parameter_float("V_oc_ref"),
        "isc_a": _module_parameter_float("I_sc_ref"),
    }
    errors = {
        name: abs(stc[name] / target - 1.0) for name, target in targets.items()
    }
    failed = {
        name: error
        for name, error in errors.items()
        if not np.isfinite(error) or error > float(relative_tolerance)
    }
    if failed:
        detail = ", ".join(
            f"{name}={100.0 * error:.3f}%" for name, error in failed.items()
        )
        raise ValueError(
            "WAAREE BIN-08-580 PVMismatch module misses datasheet tolerance: "
            + detail
        )

    cold = solectria_module_power_point(module, temperature_c=0.0)
    hot = solectria_module_power_point(module, temperature_c=65.0)
    if not cold["pmp_w"] > stc["pmp_w"] > hot["pmp_w"]:
        raise ValueError(
            "WAAREE BIN-08-580 temperature response is non-physical: expected "
            "Pmp(0 C) > Pmp(25 C) > Pmp(65 C)."
        )
    solectria_module_power_point(module)
    return {
        "physics_version": SOLECTRIA_PHYSICS_VERSION,
        "physics_fingerprint": SOLECTRIA_PHYSICS_FINGERPRINT,
        "stc": stc,
        "targets": targets,
        "relative_errors": errors,
        "pmp_0c_w": cold["pmp_w"],
        "pmp_65c_w": hot["pmp_w"],
        "pmp_temperature_coefficient_pct_per_c": (
            100.0 * (hot["pmp_w"] / stc["pmp_w"] - 1.0) / 40.0
        ),
        "voc_temperature_coefficient_v_per_c": (
            (hot["voc_v"] - stc["voc_v"]) / 40.0
        ),
        "relative_alpha_isc_per_k": solectria_relative_alpha_isc(),
        "cell_area_cm2": solectria_equivalent_cell_area_cm2(),
        "band_gap_ev": SOLECTRIA_BAND_GAP_EV,
    }


def build_pvmismatch_module() -> pvm.pvmodule.PVmodule:
    """Build and fail-closed validate the production Solectria module."""

    module = _build_pvmismatch_module_from_coefficients(
        SOLECTRIA_TWO_DIODE_COEFFICIENTS
    )
    validate_solectria_module(module)
    return module


def _uniform_series_topology(
    pvmod: pvm.pvmodule.PVmodule,
) -> tuple[int, int, float]:
    """Validate and summarize the uniform series topology used by this model.

    The fast mismatch path is exact for the configured STD72 module: every
    substring contains the same number of series-connected cells and uses the
    same substring bypass voltage. Keeping the checks here makes a future
    module-layout change fail loudly instead of silently changing the physics.
    """
    if pvmod.Vbypass_config != pvm.pvmodule.DEFAULT_BYPASS:
        raise ValueError(
            "Fast mismatch calculation requires substring bypass diodes."
        )

    cells_per_substring: list[int] = []
    for substring in pvmod.cell_pos:
        records = [record for column in substring for record in column]
        if any(bool(record["crosstie"]) for record in records):
            raise ValueError(
                "Fast mismatch calculation requires series-only substrings."
            )
        cells_per_substring.append(len(records))

    if not cells_per_substring or len(set(cells_per_substring)) != 1:
        raise ValueError(
            "Fast mismatch calculation requires equal-sized module substrings."
        )
    if any(cell is not pvmod.pvcells[0] for cell in pvmod.pvcells[1:]):
        raise ValueError(
            "Fast mismatch calculation requires uniform cells within a module."
        )
    if not np.isscalar(pvmod.Vbypass):
        raise ValueError("Fast mismatch calculation requires one bypass voltage.")

    return cells_per_substring[0], len(cells_per_substring), float(pvmod.Vbypass)


def _uniform_module_curve(
    template_cell: pvm.pvcell.PVcell,
    pvconst: pvm.pvconstants.PVconstants,
    cells_per_substring: int,
    number_substrings: int,
    bypass_voltage: float,
    effective_irradiance_suns: float,
    cell_temperature_k: float,
) -> tuple[np.ndarray, np.ndarray, float]:
    """Return one uniform module's current, voltage, and short-circuit current.

    ``PVmodule.calcMod`` interpolates every identical cell separately. For this
    model all cells within a module have the same irradiance and temperature,
    so one cell curve can be interpolated and its voltage multiplied by the
    number of series-identical cells/substrings. This is electrically identical
    to pvmismatch's generic calculation, aside from floating-point summation
    order, while avoiding hundreds of duplicate interpolations per module.
    """
    cell = copy(template_cell)
    cell.update(
        Ee=float(effective_irradiance_suns),
        Tcell=float(cell_temperature_k),
    )
    cell_current = cell.Icell.ravel()
    cell_voltage = cell.Vcell.ravel()

    current_at_breakdown = np.interp(
        cell.VRBD,
        cell_voltage,
        cell_current,
    )
    substring_current, resampled_cell_voltage = pvconst.calcSeries(
        cell_current[None, :],
        cell_voltage[None, :],
        cell.Isc,
        current_at_breakdown,
    )
    substring_voltage = resampled_cell_voltage * cells_per_substring
    substring_voltage[substring_voltage < bypass_voltage] = bypass_voltage

    substring_isc = np.interp(
        np.float64(0),
        substring_voltage,
        substring_current,
    )
    module_current, resampled_substring_voltage = pvconst.calcSeries(
        substring_current[None, :],
        substring_voltage[None, :],
        substring_isc,
        substring_current.max(),
    )
    module_voltage = resampled_substring_voltage * number_substrings
    return module_current, module_voltage, float(cell.Isc)


def _uniform_string_max_power(
    bay_curves: list[tuple[np.ndarray, np.ndarray, float]],
    pvconst: pvm.pvconstants.PVconstants,
) -> float:
    """Return max string power for four bay curves repeated six modules each."""
    string_current, string_voltage = _uniform_string_curve(
        bay_curves,
        pvconst,
    )
    return float(np.nanmax(string_current * string_voltage))


def _uniform_string_curve(
    bay_curves: list[tuple[np.ndarray, np.ndarray, float]],
    pvconst: pvm.pvconstants.PVconstants,
) -> tuple[np.ndarray, np.ndarray]:
    """Return one 24-module string I-V curve from its four bay curves."""

    module_current = np.asarray([curve[0] for curve in bay_curves])
    module_voltage = np.asarray([curve[1] for curve in bay_curves])
    mean_isc = float(np.mean([curve[2] for curve in bay_curves]))
    string_current, one_of_each_voltage = pvconst.calcSeries(
        module_current,
        module_voltage,
        mean_isc,
        module_current.max(),
    )

    # Each bay contributes MODULES_PER_BAY identical modules in series. The
    # generic pvmismatch path interpolates the same curve six times; multiplying
    # its voltage after one interpolation gives the same series I-V curve.
    string_voltage = one_of_each_voltage * MODULES_PER_BAY
    return string_current, string_voltage


def _common_mppt_power_point(
    string_curves: list[tuple[np.ndarray, np.ndarray]],
    pvconst: pvm.pvconstants.PVconstants,
    *,
    minimum_voltage: float = SOLECTRIA_INVERTER_MPPT_MIN_V,
    maximum_voltage: float = SOLECTRIA_INVERTER_MPPT_MAX_V,
) -> tuple[float, float, float]:
    """Return system DC power, voltage, and current at the single MPPT.

    The installed XGI inverter exposes one MPPT, so all ten strings operate at
    one voltage. Their currents are summed in parallel before maximizing power,
    and operating points outside the inverter's MPPT window are ineligible.
    """

    if not string_curves:
        raise ValueError("Common MPPT requires at least one string curve.")
    lower_limit = float(minimum_voltage)
    upper_limit = float(maximum_voltage)
    if (
        not np.isfinite(lower_limit)
        or not np.isfinite(upper_limit)
        or lower_limit < 0.0
        or upper_limit <= lower_limit
    ):
        raise ValueError("Common MPPT voltage limits must be finite and ordered.")

    string_currents: list[np.ndarray] = []
    string_voltages: list[np.ndarray] = []
    curve_size: int | None = None
    for current, voltage in string_curves:
        current_values = np.asarray(current, dtype=float).reshape(-1)
        voltage_values = np.asarray(voltage, dtype=float).reshape(-1)
        if (
            current_values.size < 2
            or current_values.shape != voltage_values.shape
            or not np.isfinite(current_values).all()
            or not np.isfinite(voltage_values).all()
        ):
            raise ValueError("Each common-MPPT string curve must be finite and aligned.")
        if curve_size is None:
            curve_size = current_values.size
        elif current_values.size != curve_size:
            raise ValueError("Common-MPPT string curves must use one shared grid size.")
        # PVMismatch legitimately repeats voltage samples where bypass clamps
        # the reverse-bias part of a string curve. Descending samples are not
        # valid input, but equal adjacent voltages are part of its native grid.
        if np.any(np.diff(voltage_values) < 0.0):
            raise ValueError(
                "Each common-MPPT string voltage curve must be nondecreasing."
            )
        string_currents.append(current_values)
        string_voltages.append(voltage_values)

    parallel_current, parallel_voltage = pvconst.calcParallel(
        np.stack(string_currents),
        np.stack(string_voltages),
        max(float(np.max(values)) for values in string_voltages),
        min(float(np.min(values)) for values in string_voltages),
    )
    parallel_current = np.asarray(parallel_current, dtype=float).reshape(-1)
    parallel_voltage = np.asarray(parallel_voltage, dtype=float).reshape(-1)
    finite = np.isfinite(parallel_current) & np.isfinite(parallel_voltage)
    if finite.sum() < 2:
        return 0.0, float("nan"), 0.0
    system_voltage = parallel_voltage[finite]
    system_current = parallel_current[finite]
    if np.any(np.diff(system_voltage) < 0.0):
        raise ValueError("Combined common-MPPT voltage curve must be nondecreasing.")
    system_voltage, unique_indices = np.unique(
        system_voltage,
        return_index=True,
    )
    system_current = system_current[unique_indices]
    if system_voltage.size < 2:
        return 0.0, float("nan"), 0.0

    lower = max(lower_limit, float(system_voltage[0]))
    upper = min(upper_limit, float(system_voltage[-1]))
    if upper < lower:
        return 0.0, float("nan"), 0.0

    in_window = (system_voltage >= lower) & (system_voltage <= upper)
    candidate_voltage = np.unique(
        np.concatenate(([lower], system_voltage[in_window], [upper]))
    )
    candidate_current = np.interp(
        candidate_voltage,
        system_voltage,
        system_current,
    )
    candidate_power = candidate_voltage * candidate_current
    eligible = (
        np.isfinite(candidate_power)
        & (candidate_voltage >= lower_limit)
        & (candidate_voltage <= upper_limit)
        & (candidate_current >= 0.0)
    )
    if not eligible.any():
        return 0.0, float("nan"), 0.0
    eligible_indices = np.flatnonzero(eligible)
    best_index = int(
        eligible_indices[np.argmax(candidate_power[eligible_indices])]
    )
    best_power = max(0.0, float(candidate_power[best_index]))
    return (
        best_power,
        float(candidate_voltage[best_index]),
        max(0.0, float(candidate_current[best_index])),
    )


# -----------------------------------------------------------------------------
# SYSTEM PREDICTION
# -----------------------------------------------------------------------------
def predict_ac_power(
    df: pd.DataFrame,
    progress_cb=None,
    backtrack: bool = BACKTRACK,
    se_eff: float = SE_EFF,
    sol_eff: float = SOL_EFF,
    include_iam: bool | None = INCLUDE_IAM,
    iam_a_r: float | None = A_R,
    iam_model: str | None = None,
) -> tuple[pd.DataFrame, str]:
    """Add predicted AC power columns for SolarEdge and Solectria.

    progress_cb(frac, msg): optional callback (frac in 0..1) for the Solectria
    time loop, so a UI can show a moving progress bar.
    """
    selected_iam_model, effective_iam_a_r = resolve_iam_settings(
        iam_model=iam_model,
        iam_a_r=iam_a_r,
        include_iam=include_iam,
    )
    location = pvl.location.Location(LAT, LON, tz=str(df.index.tz))
    weather, dhi_source = build_weather(df, location)

    all_tilts = (
        np.array(SOLECTRIA_TILT_ASBUILT).flatten().tolist()
        + np.array(SOLAREDGE_TILT_ASBUILT).flatten().tolist()
    )
    unique_tilts = sorted({float(t) for t in all_tilts})
    print(f"ModelChain over {len(unique_tilts)} unique tilts (DHI source: {dhi_source})...")

    tilt_out = run_modelchain_for_axis_tilts(
        unique_tilts,
        weather,
        location,
        backtrack=backtrack,
        include_iam=False,
        iam_a_r=effective_iam_a_r,
        iam_model=selected_iam_model,
    )

    # SolarEdge: module-level -> sum module p_mp over all bays x MODULES_PER_BAY
    se_dc = np.zeros(len(df), dtype=float)
    for t in np.array(SOLAREDGE_TILT_ASBUILT).flatten():
        se_dc += tilt_out[float(t)]["p_mp_w"] * MODULES_PER_BAY

    # Solectria: per-string mismatch via pvmismatch, per timestep. Modules in a
    # bay are identical, so calculate one curve per unique tilt and reuse the
    # series curve rather than asking the generic library to recalculate every
    # identical cell/module separately.
    pvm_mod = build_pvmismatch_module()
    cells_per_substring, number_substrings, bypass_voltage = (
        _uniform_series_topology(pvm_mod)
    )
    solectria_rows = [
        tuple(float(t) for t in row) for row in SOLECTRIA_TILT_ASBUILT
    ]
    solectria_tilts = sorted({tilt for row in solectria_rows for tilt in row})
    template_cell = pvm_mod.pvcells[0]

    n = len(df)
    print(f"Solectria string mismatch over {n} timesteps...")
    sol_dc = np.zeros(n, dtype=float)
    progress_interval = max(1, (n + 99) // 100)
    for j in range(n):
        if progress_cb is not None and (
            j % progress_interval == 0 or j == n - 1
        ):
            progress_cb(j / n if n else 1.0, f"Computing string mismatch… {j + 1}/{n}")
        # Skip night / no-irradiance steps (first string's first bay as proxy)
        proxy_irradiance = tilt_out[solectria_rows[0][0]]["Ee_suns"][j]
        if np.isnan(proxy_irradiance) or proxy_irradiance < 0.01:
            continue

        curves_by_tilt = {
            tilt: _uniform_module_curve(
                template_cell,
                pvm_mod.pvconst,
                cells_per_substring,
                number_substrings,
                bypass_voltage,
                tilt_out[tilt]["Ee_suns"][j],
                tilt_out[tilt]["Tk"][j],
            )
            for tilt in solectria_tilts
        }
        string_curves = [
            _uniform_string_curve(
                [curves_by_tilt[tilt] for tilt in row],
                pvm_mod.pvconst,
            )
            for row in solectria_rows
        ]
        sol_dc[j], _, _ = _common_mppt_power_point(
            string_curves,
            pvm_mod.pvconst,
        )

    out = df.copy()
    out["se_predicted_power_w"] = se_dc * float(se_eff)
    out["sol_predicted_power_w"] = np.minimum(
        sol_dc * float(sol_eff),
        SOLECTRIA_INVERTER_AC_RATING_W,
    )
    return out, dhi_source


def add_energy(
    df: pd.DataFrame,
    expected_interval_seconds: int | None = None,
    *,
    requested_start: str | None = None,
    requested_end: str | None = None,
) -> pd.DataFrame:
    """Add bounded interval durations and cumulative energies (kWh)."""
    out = df.copy()
    out["dt_hours"] = _bounded_interval_hours(
        pd.DatetimeIndex(out.index),
        expected_interval_seconds=expected_interval_seconds,
        requested_start=requested_start,
        requested_end=requested_end,
    )

    for sysname in ("se", "sol"):
        for kind in ("measured", "predicted", "uncalibrated"):
            power_column = f"{sysname}_{kind}_power_w"
            if power_column not in out:
                continue
            step = f"{sysname}_{kind}_energy_step_kwh"
            out[step] = (
                pd.to_numeric(out[power_column], errors="coerce")
                * out["dt_hours"]
                / 1000.0
            ).fillna(0.0)
            out[f"{sysname}_{kind}_energy_kwh"] = out[step].cumsum()
    return out


def calibrated_energy_balance_summary(
    df: pd.DataFrame,
    *,
    absolute_tolerance_kwh: float = 0.01,
    relative_tolerance: float = 1e-9,
) -> dict[str, object]:
    """Verify that a newly fitted baseline reconciles reported energy totals."""

    systems: dict[str, dict[str, float | str]] = {}
    for system, label in (("se", "solaredge"), ("sol", "solectria")):
        measured_column = f"{system}_measured_energy_kwh"
        predicted_column = f"{system}_predicted_energy_kwh"
        if measured_column not in df or predicted_column not in df or df.empty:
            raise ValueError(
                "Calibrated energy reconciliation requires measured and "
                "predicted cumulative energy for both systems."
            )
        measured_kwh = float(df[measured_column].iloc[-1])
        predicted_kwh = float(df[predicted_column].iloc[-1])
        if not np.isfinite(measured_kwh) or not np.isfinite(predicted_kwh):
            raise ValueError(
                "Calibrated measured and predicted energy must be finite."
            )
        residual_kwh = predicted_kwh - measured_kwh
        tolerance_kwh = max(
            float(absolute_tolerance_kwh),
            abs(measured_kwh) * float(relative_tolerance),
        )
        if abs(residual_kwh) > tolerance_kwh:
            raise ValueError(
                f"{label.title()} calibrated energy differs from measured "
                f"energy by {residual_kwh:.6f} kWh; expected no more than "
                f"{tolerance_kwh:.6f} kWh."
            )
        systems[label] = {
            "measured_kwh": measured_kwh,
            "predicted_kwh": predicted_kwh,
            "residual_kwh": residual_kwh,
            "tolerance_kwh": tolerance_kwh,
            "status": "balanced",
        }
    return {
        "scope": "all reviewed rows in the baseline calibration window",
        "status": "balanced",
        "systems": systems,
    }


def apply_curtailment(df: pd.DataFrame, limit_kw: float | None) -> pd.DataFrame:
    """Clip predicted AC power only to a per-system kW limit.

    Historian measurements are observations, not controllable modeled values.
    Keeping them immutable is required for trustworthy baseline/scenario
    comparisons and validation residuals.
    """
    if limit_kw is None:
        return df

    cap_kw = float(limit_kw)
    if not np.isfinite(cap_kw) or cap_kw <= 0:
        raise ValueError("Curtailment limit must be a positive kW value.")

    out = df.copy()
    cap_w = cap_kw * 1000.0
    for col in (
        "se_predicted_power_w",
        "sol_predicted_power_w",
        "se_uncalibrated_power_w",
        "sol_uncalibrated_power_w",
    ):
        if col in out:
            out[col] = out[col].clip(upper=cap_w)
    return out


# -----------------------------------------------------------------------------
# REPORTING (plots + Excel)
# -----------------------------------------------------------------------------
def tilt_summary() -> pd.DataFrame:
    """Bay tilt + string mapping for both systems."""
    recs = []
    for sysname, strings, bays, table in (
        ("Solectria", SOLECTRIA_STRINGS, SOLECTRIA_BAYS_PER_STRING, SOLECTRIA_TILT_ASBUILT),
        ("SolarEdge", SOLAREDGE_STRINGS, SOLAREDGE_BAYS_PER_STRING, SOLAREDGE_TILT_ASBUILT),
    ):
        bay_id = 1
        for s in range(strings):
            for b in range(bays):
                recs.append(
                    {
                        "system": sysname,
                        "string_id": s + 1,
                        "bay_in_string": b + 1,
                        "bay_id": bay_id,
                        "axis_tilt_deg": float(table[s][b]),
                        "modules_per_bay": MODULES_PER_BAY,
                    }
                )
                bay_id += 1
    return pd.DataFrame.from_records(recs)


def plot_results(
    df: pd.DataFrame,
    out_prefix: str,
    annual_mode: bool = False,
    calibrated: bool = False,
) -> None:
    """Save AC-power and cumulative-energy charts for the selected run mode."""
    prediction_label = "calibrated" if calibrated else "predicted"
    # AC power
    # Reserve a header band for the energy summary and series legend so neither
    # obscures the power curves. The extra figure height keeps the data axes at
    # approximately the same height as the previous 14 x 6 rendering.
    fig1, ax1 = plt.subplots(figsize=(14, 7.25))
    ax1.plot(
        df.index,
        df["se_predicted_power_w"] / 1000.0,
        "r-",
        label=f"SolarEdge {prediction_label}",
    )
    ax1.plot(
        df.index,
        df["sol_predicted_power_w"] / 1000.0,
        "b-",
        label=f"Solectria {prediction_label}",
    )
    if not annual_mode:
        ax1.plot(df.index, df["se_measured_power_w"] / 1000.0, "r--", label="SolarEdge measured")
        ax1.plot(df.index, df["sol_measured_power_w"] / 1000.0, "b--", label="Solectria measured")
    if annual_mode:
        ax1.set_title(
            "Calibration-Adjusted AC Power (kW) - Annual Simulation"
            if calibrated
            else "Predicted AC Power (kW) - Annual Simulation"
        )
    elif calibrated:
        ax1.set_title("Calibrated AC Power")
    else:
        ax1.set_title(
            f"Physics-Model AC Power (kW) — sbe_pv_model v{__version__}"
        )
    ax1.set_xlabel(f"Time ({df.index.tz})")
    ax1.set_ylabel("AC Power (kW)")
    ax1.grid(True, alpha=0.25)
    ax1.xaxis.set_major_formatter(
        mdates.DateFormatter("%m-%d-%Y %H:%M", tz=df.index.tz)
    )
    fig1.autofmt_xdate()
    fig1.subplots_adjust(top=0.78)

    se_meas = float(df["se_measured_energy_kwh"].iloc[-1])
    se_pred = float(df["se_predicted_energy_kwh"].iloc[-1])
    sol_meas = float(df["sol_measured_energy_kwh"].iloc[-1])
    sol_pred = float(df["sol_predicted_energy_kwh"].iloc[-1])

    def pct(pred, meas):
        return (pred - meas) / meas * 100.0 if meas != 0 else np.nan

    ac_txt = (
        f"SolarEdge: meas={se_meas:,.1f} kWh, pred={se_pred:,.1f} kWh, Δ={pct(se_pred, se_meas):+.2f}%\n"
        f"Solectria: meas={sol_meas:,.1f} kWh, pred={sol_pred:,.1f} kWh, Δ={pct(sol_pred, sol_meas):+.2f}%"
    )
    if annual_mode:
        pred_diff = se_pred - sol_pred
        pred_pct = pred_diff / sol_pred * 100.0 if sol_pred != 0 else np.nan
        ac_txt = (
            f"Period energy (kWh)\nSolarEdge: {se_pred:,.0f}\n"
            f"Solectria: {sol_pred:,.0f}\n"
            f"Difference: {pred_diff:,.0f} ({pred_pct:+.2f}%)"
        )
    fig1.text(
        0.08, 0.96, ac_txt, va="top", ha="left", fontsize=11,
        bbox=dict(boxstyle="round", facecolor="white", alpha=0.85, edgecolor="gray"),
    )
    handles, labels = ax1.get_legend_handles_labels()
    fig1.legend(
        handles,
        labels,
        loc="upper right",
        bbox_to_anchor=(0.97, 0.96),
        ncol=2,
        borderaxespad=0.0,
    )
    fig1.savefig(f"{out_prefix}_ac_power.png", dpi=200, bbox_inches="tight")

    # Cumulative energy
    fig2, ax2 = plt.subplots(figsize=(14, 6))
    ax2.plot(
        df.index,
        df["se_predicted_energy_kwh"],
        "r-",
        label=f"SolarEdge {prediction_label}",
    )
    ax2.plot(
        df.index,
        df["sol_predicted_energy_kwh"],
        "b-",
        label=f"Solectria {prediction_label}",
    )
    if not annual_mode:
        ax2.plot(df.index, df["se_measured_energy_kwh"], "r--", label="SolarEdge measured")
        ax2.plot(df.index, df["sol_measured_energy_kwh"], "b--", label="Solectria measured")
    if annual_mode:
        ax2.set_title(
            "Cumulative Calibration-Adjusted Energy (kWh) - Annual Simulation"
            if calibrated
            else "Cumulative Predicted Energy (kWh) - Annual Simulation"
        )
    elif calibrated:
        ax2.set_title("Calibrated Cumulative Energy")
    else:
        ax2.set_title(
            f"Physics-Model Cumulative Energy (kWh) — sbe_pv_model v{__version__}"
        )
    ax2.set_xlabel(f"Time ({df.index.tz})")
    ax2.set_ylabel("Cumulative Energy (kWh)")
    ax2.grid(True, alpha=0.25)
    ax2.legend(loc="best")
    ax2.xaxis.set_major_formatter(
        mdates.DateFormatter("%m-%d-%Y %H:%M", tz=df.index.tz)
    )
    fig2.autofmt_xdate()

    meas_sys_pct = (se_meas - sol_meas) / sol_meas * 100.0 if sol_meas != 0 else np.nan
    pred_sys_pct = (se_pred - sol_pred) / sol_pred * 100.0 if sol_pred != 0 else np.nan
    ce_txt = (
        f"SolarEdge\n measured : {se_meas:,.1f} kWh\n predicted: {se_pred:,.1f} kWh\n Δ: {pct(se_pred, se_meas):+.2f}%\n\n"
        f"Solectria\n measured : {sol_meas:,.1f} kWh\n predicted: {sol_pred:,.1f} kWh\n Δ: {pct(sol_pred, sol_meas):+.2f}%\n\n"
        f"SE vs Sol\n measured Δ : {meas_sys_pct:+.2f}%\n predicted Δ: {pred_sys_pct:+.2f}%"
    )
    if annual_mode:
        ce_txt = ac_txt
    ax2.text(
        0.01, 0.99, ce_txt, transform=ax2.transAxes, va="top", ha="left", fontsize=11,
        bbox=dict(boxstyle="round", facecolor="white", alpha=0.85, edgecolor="gray"),
    )
    fig2.savefig(f"{out_prefix}_cumulative_energy.png", dpi=200, bbox_inches="tight")
    plt.close("all")


def monthly_energy_table(
    df: pd.DataFrame,
    *,
    calibrated: bool = False,
) -> pd.DataFrame:
    """Return one annual energy summary row per fixed-MST source month."""
    monthly_frame = df.copy()
    monthly_index = pd.DatetimeIndex(monthly_frame.index)
    if monthly_index.tz is None:
        monthly_index = monthly_index.tz_localize(TIMEZONE)
    monthly_frame.index = monthly_index.tz_convert("Etc/GMT+7")
    # ``resample`` creates an empty bin for every month between the first and
    # last timestamp.  Selected MIDC years can be noncontiguous, so retain only
    # bins that contain source rows rather than reporting gap years as zero.
    occupied_months = (
        monthly_frame["se_predicted_energy_step_kwh"]
        .resample("MS")
        .size()
        .loc[lambda counts: counts > 0]
        .index
    )
    se_monthly = (
        monthly_frame["se_predicted_energy_step_kwh"]
        .resample("MS")
        .sum()
        .reindex(occupied_months)
    )
    sol_monthly = (
        monthly_frame["sol_predicted_energy_step_kwh"]
        .resample("MS")
        .sum()
        .reindex(occupied_months)
    )
    sol_monthly = sol_monthly.reindex(se_monthly.index)
    energy_label = "calibrated" if calibrated else "predicted"
    se_column = f"SolarEdge_{energy_label}_kWh"
    sol_column = f"Solectria_{energy_label}_kWh"
    result = pd.DataFrame(
        {
            "month": se_monthly.index.strftime("%b %Y"),
            "month_start": se_monthly.index.tz_localize(None),
            se_column: se_monthly.to_numpy(),
            sol_column: sol_monthly.to_numpy(),
        }
    )
    result["difference_kWh"] = result[se_column] - result[sol_column]
    result["difference_pct"] = np.where(
        result[sol_column] != 0,
        result["difference_kWh"] / result[sol_column] * 100.0,
        np.nan,
    )
    if calibrated and {
        "se_uncalibrated_energy_step_kwh",
        "sol_uncalibrated_energy_step_kwh",
    }.issubset(df.columns):
        result["SolarEdge_physics_only_kWh"] = (
            monthly_frame["se_uncalibrated_energy_step_kwh"]
            .resample("MS")
            .sum()
            .reindex(se_monthly.index)
            .to_numpy()
        )
        result["Solectria_physics_only_kWh"] = (
            monthly_frame["sol_uncalibrated_energy_step_kwh"]
            .resample("MS")
            .sum()
            .reindex(se_monthly.index)
            .to_numpy()
        )
    return result


def annual_energy_by_year(
    df: pd.DataFrame,
    periods: list[Mapping[str, object]] | None = None,
) -> tuple[list[dict[str, object]], dict[str, object]]:
    """Return per-period energy totals and an empirical complete-year CDF.

    MIDC labels are fixed MST even while dashboard plots use America/Denver.
    Convert the model index back to fixed MST before assigning rows to periods;
    otherwise the final 23:00 MST interval of a summer day is displayed on the
    following MDT date and can be attributed to the wrong requested period.
    """

    index = pd.DatetimeIndex(df.index)
    if index.tz is None:
        index = index.tz_localize(TIMEZONE)
    fixed_mst = index.tz_convert("Etc/GMT+7")
    fixed_dates = np.asarray(fixed_mst.date, dtype=object)

    normalized_periods: list[dict[str, object]] = []
    if periods:
        normalized_periods = [dict(period) for period in periods]
    elif len(fixed_mst):
        for year in sorted(set(int(value) for value in fixed_mst.year)):
            year_mask = fixed_mst.year == year
            period_start = min(fixed_dates[year_mask])
            period_end = max(fixed_dates[year_mask])
            complete = (
                period_start.isoformat() == f"{year:04d}-01-01"
                and period_end.isoformat() == f"{year:04d}-12-31"
            )
            normalized_periods.append(
                {
                    "year": year,
                    "period_start": period_start.isoformat(),
                    "period_end": period_end.isoformat(),
                    "coverage_status": "complete" if complete else "custom_range",
                    "complete_calendar_year": complete,
                    "cdf_eligible": complete,
                }
            )

    def energy_total(mask: np.ndarray, column: str) -> float:
        values = pd.to_numeric(df.loc[mask, column], errors="coerce")
        return round(float(values.fillna(0.0).sum()), 1)

    rows: list[dict[str, object]] = []
    physics_columns = {
        "se_uncalibrated_energy_step_kwh",
        "sol_uncalibrated_energy_step_kwh",
    }.issubset(df.columns)
    for period in normalized_periods:
        year = int(period["year"])
        try:
            period_start = pd.Timestamp(str(period["period_start"])).date()
            period_end = pd.Timestamp(str(period["period_end"])).date()
        except (KeyError, TypeError, ValueError) as exc:
            raise ValueError("Annual energy periods require valid start/end dates.") from exc
        mask = (fixed_dates >= period_start) & (fixed_dates <= period_end)
        se_predicted = energy_total(mask, "se_predicted_energy_step_kwh")
        sol_predicted = energy_total(mask, "sol_predicted_energy_step_kwh")
        row: dict[str, object] = {
            **period,
            "year": year,
            "period_start": period_start.isoformat(),
            "period_end": period_end.isoformat(),
            "row_count": int(mask.sum()),
            "se_predicted_kwh": se_predicted,
            "sol_predicted_kwh": sol_predicted,
            "combined_predicted_kwh": round(se_predicted + sol_predicted, 1),
        }
        if physics_columns:
            se_physics = energy_total(mask, "se_uncalibrated_energy_step_kwh")
            sol_physics = energy_total(mask, "sol_uncalibrated_energy_step_kwh")
            row.update(
                {
                    "se_physics_only_kwh": se_physics,
                    "sol_physics_only_kwh": sol_physics,
                    "combined_physics_only_kwh": round(
                        se_physics + sol_physics, 1
                    ),
                }
            )
        rows.append(row)

    eligible = [row for row in rows if row.get("cdf_eligible") is True]
    series: dict[str, list[dict[str, object]]] = {}
    for series_name, energy_field in (
        ("solaredge", "se_predicted_kwh"),
        ("solectria", "sol_predicted_kwh"),
        ("combined", "combined_predicted_kwh"),
    ):
        ordered = sorted(
            eligible,
            key=lambda row: (float(row[energy_field]), int(row["year"])),
        )
        count = len(ordered)
        cumulative_probability_by_energy = {
            float(row[energy_field]): round(rank / count, 6)
            for rank, row in enumerate(ordered, start=1)
        }
        series[series_name] = [
            {
                "year": int(row["year"]),
                "energy_kwh": float(row[energy_field]),
                # Equal observations share P(X <= x), as required by an ECDF.
                "cumulative_probability": cumulative_probability_by_energy[
                    float(row[energy_field])
                ],
            }
            for row in ordered
        ]

    cdf: dict[str, object] = {
        "population": "complete_calendar_years",
        "eligible_years": [int(row["year"]) for row in eligible],
        "excluded_years": [
            {
                "year": int(row["year"]),
                "reason": str(row.get("coverage_status") or "partial"),
            }
            for row in rows
            if row.get("cdf_eligible") is not True
        ],
        "series": series,
    }
    return rows, cdf


def plot_monthly_energy(
    df: pd.DataFrame,
    out_prefix: str,
    *,
    calibrated: bool = False,
) -> None:
    """Save the annual-result monthly energy comparison chart."""
    monthly = monthly_energy_table(df, calibrated=calibrated)
    x = np.arange(len(monthly))
    width = 0.38
    fig, ax = plt.subplots(figsize=(14, 6))
    energy_label = "calibrated" if calibrated else "predicted"
    se_column = f"SolarEdge_{energy_label}_kWh"
    sol_column = f"Solectria_{energy_label}_kWh"
    se_values = monthly[se_column].to_numpy(dtype=float)
    sol_values = monthly[sol_column].to_numpy(dtype=float)
    ax.bar(x - width / 2, se_values, width=width, color="red", label="SolarEdge (kWh)")
    ax.bar(x + width / 2, sol_values, width=width, color="blue", label="Solectria (kWh)")
    ax.set_xticks(x)
    ax.set_xticklabels(monthly["month"], rotation=35, ha="right")
    ax.set_xlabel("Month")
    ax.set_ylabel(
        "Calibration-Adjusted Energy (kWh)"
        if calibrated
        else "Predicted Energy (kWh)"
    )
    ax.set_title(
        "Monthly Calibration-Adjusted Energy - Annual Simulation"
        if calibrated
        else "Monthly Predicted Energy - Annual Simulation"
    )
    ax.grid(True, axis="y", alpha=0.25)
    ax.legend(loc="best")

    ymax = float(np.nanmax([se_values.max(), sol_values.max()])) if len(monthly) else 0.0
    pad = ymax * 0.02 if ymax > 0 else 0.0
    for index, row in monthly.iterrows():
        pct_value = float(row["difference_pct"])
        label = "n/a" if not np.isfinite(pct_value) else f"{pct_value:+.1f}%"
        ax.text(
            x[index],
            max(row[se_column], row[sol_column]) + pad,
            label,
            ha="center",
            va="bottom",
            fontsize=9,
            bbox=dict(boxstyle="round", facecolor="white", alpha=0.85, edgecolor="gray"),
        )

    se_total = float(df["se_predicted_energy_kwh"].iloc[-1])
    sol_total = float(df["sol_predicted_energy_kwh"].iloc[-1])
    total_diff = se_total - sol_total
    total_pct = total_diff / sol_total * 100.0 if sol_total else np.nan
    ax.text(
        0.01,
        0.99,
        f"Period energy (kWh)\nSolarEdge: {se_total:,.0f}\n"
        f"Solectria: {sol_total:,.0f}\nDifference: {total_diff:,.0f} ({total_pct:+.2f}%)",
        transform=ax.transAxes,
        va="top",
        ha="left",
        fontsize=10,
        bbox=dict(boxstyle="round", facecolor="white", alpha=0.85, edgecolor="gray"),
    )
    fig.tight_layout()
    fig.savefig(f"{out_prefix}_monthly_energy.png", dpi=200, bbox_inches="tight")
    plt.close(fig)


def _audit_value(value: object) -> object:
    """Return an Excel-safe scalar while retaining structured audit data."""

    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    try:
        return json.dumps(value, sort_keys=True, default=str, allow_nan=False)
    except (TypeError, ValueError):
        return str(value)


def _first_present(mapping: Mapping[str, object], names: tuple[str, ...]) -> object:
    for name in names:
        if name in mapping:
            return mapping[name]
    return None


def _settings_delta_rows(application: Mapping[str, object]) -> list[dict]:
    raw = _first_present(
        application,
        (
            "settings_delta",
            "settings_deltas",
            "setting_differences",
            "settings_differences",
        ),
    )
    rows: list[dict] = []
    if isinstance(raw, Mapping):
        items = raw.items()
    elif isinstance(raw, list):
        items = enumerate(raw)
    else:
        items = ()
    for key, raw_details in items:
        if isinstance(raw_details, Mapping):
            details = dict(raw_details)
            setting = _first_present(
                details,
                ("setting", "field", "name", "parameter"),
            ) or key
            inherited = _first_present(
                details,
                (
                    "calibrated_value",
                    "inherited_value",
                    "baseline_value",
                    "original_value",
                    "from",
                    "before",
                ),
            )
            annual = _first_present(
                details,
                (
                    "annual_value",
                    "requested_value",
                    "current_value",
                    "resolved_value",
                    "to",
                    "after",
                    "value",
                ),
            )
            changed = details.get("changed")
            extra = {
                name: _audit_value(value)
                for name, value in details.items()
                if name
                not in {
                    "setting",
                    "field",
                    "name",
                    "parameter",
                    "calibrated_value",
                    "inherited_value",
                    "baseline_value",
                    "original_value",
                    "from",
                    "before",
                    "annual_value",
                    "requested_value",
                    "current_value",
                    "resolved_value",
                    "to",
                    "after",
                    "value",
                    "changed",
                }
            }
            rows.append(
                {
                    "setting": setting,
                    "calibrated_value": _audit_value(inherited),
                    "annual_value": _audit_value(annual),
                    "changed": (
                        bool(changed)
                        if changed is not None
                        else inherited != annual
                    ),
                    **extra,
                }
            )
        else:
            rows.append(
                {
                    "setting": key,
                    "calibrated_value": None,
                    "annual_value": _audit_value(raw_details),
                    "changed": True,
                }
            )
    if not rows:
        rows.append(
            {
                "setting": "(none)",
                "calibrated_value": None,
                "annual_value": None,
                "changed": False,
            }
        )
    return rows


def _substitution_audit_rows(
    application: Mapping[str, object],
    calibration_factors: Mapping[str, object] | None,
) -> list[dict]:
    raw = _first_present(
        application,
        ("seasonal_substitution", "seasonal_fallback", "substitution"),
    )
    if raw is None and isinstance(calibration_factors, Mapping):
        raw = calibration_factors.get("seasonal_substitution")
    if isinstance(raw, Mapping):
        return [
            {
                "status": "used",
                **{
                    str(key): _audit_value(value)
                    for key, value in raw.items()
                },
            }
        ]
    if isinstance(raw, list):
        rows = []
        for item in raw:
            if isinstance(item, Mapping):
                rows.append(
                    {
                        "status": "used",
                        **{
                            str(key): _audit_value(value)
                            for key, value in item.items()
                        },
                    }
                )
        if rows:
            return rows
    return [{"status": "not_used"}]


def _calibration_lineage_rows(
    application: Mapping[str, object],
    calibration_factors: Mapping[str, object] | None,
) -> list[dict[str, object]]:
    values: dict[str, object] = {}
    if isinstance(calibration_factors, Mapping):
        for key in (
            "method",
            "application_mode",
            "profile_schema_version",
            "origin_job_id",
            "origin_review_id",
            "origin_source_sha256",
            "calibration_physics_version",
            "calibration_physics_fingerprint",
            "solectria_physics_version",
            "solectria_physics_fingerprint",
        ):
            if key in calibration_factors:
                values[key] = calibration_factors[key]
    excluded = {
        "settings_delta",
        "settings_deltas",
        "setting_differences",
        "settings_differences",
        "seasonal_substitution",
        "seasonal_fallback",
        "substitution",
        "origin_profile",
        "resolved_profile",
    }
    for key, value in application.items():
        if key not in excluded:
            values[str(key)] = value
    return [
        {"parameter": key, "value": _audit_value(value)}
        for key, value in values.items()
    ] or [{"parameter": "status", "value": "applied"}]


def write_excel(
    df: pd.DataFrame,
    excel_path: str,
    meta: dict,
    annual_mode: bool = False,
    calibration_factors: dict | None = None,
    factor_driver_diagnostics: dict | None = None,
    data_quality_context: dict | None = None,
    calibration_application_context: dict | None = None,
) -> None:
    """Write time series, calibration audit tables, layout, and metadata."""
    calibration_factors = (
        calibration_factors or meta.get("_calibration_factors")
    )
    factor_driver_diagnostics = (
        factor_driver_diagnostics or meta.get("_factor_driver_diagnostics")
    )
    data_quality_context = (
        data_quality_context or meta.get("_data_quality_context")
    )
    calibration_application_context = (
        calibration_application_context
        or meta.get("_calibration_application_context")
        or {}
    )
    if not isinstance(calibration_application_context, Mapping):
        calibration_application_context = {}
    out = df.copy()
    out["timestamp_local_naive"] = out.index.tz_localize(None)
    out["timestamp_utc_naive"] = out["timestamp_utc"].dt.tz_localize(None)

    cols = [
        "timestamp_local_naive",
        "timestamp_utc_naive",
        "se_measured_power_w",
        "se_uncalibrated_power_w",
        "se_predicted_power_w",
        "se_calibration_factor",
        "sol_measured_power_w",
        "sol_uncalibrated_power_w",
        "sol_predicted_power_w",
        "sol_calibration_factor",
        "se_measured_energy_kwh",
        "se_uncalibrated_energy_kwh",
        "se_predicted_energy_kwh",
        "sol_measured_energy_kwh",
        "sol_uncalibrated_energy_kwh",
        "sol_predicted_energy_kwh",
        "dni_wm2",
        "ghi_wm2",
        "dhi_wm2",
        "temp_air_c",
        "wind_speed_ms",
        "dt_hours",
    ]
    cols = [column for column in cols if column in out.columns]
    time_series = out[cols].copy()
    if meta.get("calibration_enabled") is True:
        time_series = time_series.rename(
            columns={
                "se_predicted_power_w": "se_calibrated_power_w",
                "sol_predicted_power_w": "sol_calibrated_power_w",
                "se_predicted_energy_kwh": "se_calibrated_energy_kwh",
                "sol_predicted_energy_kwh": "sol_calibrated_energy_kwh",
            }
        )

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        time_series.to_excel(writer, sheet_name="time_series", index=False)
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        time_series_sheet = writer.sheets["time_series"]
        time_series_sheet.freeze_panes = "A2"
        time_series_sheet.auto_filter.ref = time_series_sheet.dimensions
        header_fill = PatternFill(fill_type="solid", fgColor="DDEFEA")
        for cell in time_series_sheet[1]:
            cell.font = Font(bold=True, color="0B4F46")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for column_index, column_name in enumerate(time_series.columns, start=1):
            width = min(max(len(str(column_name)) + 2, 14), 34)
            time_series_sheet.column_dimensions[
                get_column_letter(column_index)
            ].width = width
        tilt_summary().to_excel(writer, sheet_name="tilts_and_strings", index=False)
        if calibration_factors:
            factor_rows: list[dict] = []
            for season in calibration_factors.get("seasons", []):
                for system, details in (season.get("systems") or {}).items():
                    factor_rows.append(
                        {
                            "season": season.get("season"),
                            "months": season.get("months"),
                            "first_timestamp": season.get("first_timestamp"),
                            "last_timestamp": season.get("last_timestamp"),
                            "system": system,
                            **dict(details or {}),
                        }
                    )
            pd.DataFrame(factor_rows).to_excel(
                writer, sheet_name="calibration_factors", index=False
            )
        if factor_driver_diagnostics:
            driver_rows: list[dict] = []
            for system, diagnostic in (
                factor_driver_diagnostics.get("systems") or {}
            ).items():
                drivers = diagnostic.get("drivers") or []
                if not drivers:
                    driver_rows.append(
                        {
                            "system": system,
                            "status": diagnostic.get("status"),
                            "sample_count": diagnostic.get("sample_count"),
                            "message": diagnostic.get("message"),
                        }
                    )
                for driver in drivers:
                    driver_rows.append(
                        {
                            "system": system,
                            "status": diagnostic.get("status"),
                            "sample_count": diagnostic.get("sample_count"),
                            "r_squared": diagnostic.get("r_squared"),
                            **dict(driver),
                        }
                    )
            pd.DataFrame(driver_rows).to_excel(
                writer, sheet_name="factor_drivers", index=False
            )
        if data_quality_context:
            decisions = {
                item.get("issue_id"): item.get("action")
                for item in (
                    data_quality_context.get("cleaning") or {}
                ).get("decisions", [])
            }
            quality_rows = [
                {
                    "issue_id": issue.get("id"),
                    "category": issue.get("category"),
                    "severity": issue.get("severity"),
                    "title": issue.get("title"),
                    "row_count": issue.get("row_count"),
                    "decision": decisions.get(
                        issue.get("id"), issue.get("recommended_action")
                    ),
                }
                for issue in (
                    (data_quality_context.get("report") or {}).get("issues")
                    or []
                )
            ]
            pd.DataFrame(quality_rows).to_excel(
                writer, sheet_name="data_quality_review", index=False
            )
        if annual_mode:
            monthly_energy_table(
                df,
                calibrated=meta.get("calibration_enabled") is True,
            ).to_excel(
                writer, sheet_name="monthly_energy", index=False
            )
            annual_rows = list(meta.get("_annual_energy_by_year") or [])
            pd.DataFrame(annual_rows).to_excel(
                writer, sheet_name="annual_energy_by_year", index=False
            )
            annual_cdf = meta.get("_annual_energy_cdf") or {}
            cdf_rows = [
                {
                    "series": series_name,
                    **dict(point),
                }
                for series_name, points in dict(
                    annual_cdf.get("series") or {}
                ).items()
                for point in points
            ]
            pd.DataFrame(
                cdf_rows,
                columns=[
                    "series",
                    "year",
                    "energy_kwh",
                    "cumulative_probability",
                ],
            ).to_excel(writer, sheet_name="annual_energy_cdf", index=False)
        if meta.get("calibration_enabled") is True:
            pd.DataFrame(
                _calibration_lineage_rows(
                    calibration_application_context,
                    calibration_factors,
                )
            ).to_excel(writer, sheet_name="calibration_lineage", index=False)
            pd.DataFrame(
                _settings_delta_rows(calibration_application_context)
            ).to_excel(writer, sheet_name="settings_delta", index=False)
            pd.DataFrame(
                _substitution_audit_rows(
                    calibration_application_context,
                    calibration_factors,
                )
            ).to_excel(writer, sheet_name="substitution_audit", index=False)
        public_meta = {
            key: value
            for key, value in meta.items()
            if not str(key).startswith("_")
        }
        pd.DataFrame(
            list(public_meta.items()), columns=["parameter", "value"]
        ).to_excel(
            writer, sheet_name="run_info", index=False
        )


# -----------------------------------------------------------------------------
# CALLABLE ENTRYPOINT + MAIN
# -----------------------------------------------------------------------------
def run_model(
    input_csv=INPUT_CSV,
    output_base=OUTPUT_BASE,
    progress_cb=None,
    backtrack: bool = BACKTRACK,
    solaredge_inverter_efficiency: float = 1.0,
    solaredge_bos_efficiency: float = 1.0,
    solectria_inverter_efficiency: float = SOL_EFF,
    solectria_bos_efficiency: float = 1.0,
    include_iam: bool | None = INCLUDE_IAM,
    iam_a_r: float | None = A_R,
    curtailment_enabled: bool = False,
    curtailment_limit_kw: float | None = None,
    input_kind: str = "historian",
    annual_mode: bool = False,
    iam_model: str | None = None,
    data_quality_context: dict | None = None,
    expected_interval_seconds: int | None = None,
    requested_start: str | None = None,
    requested_end: str | None = None,
    calibration_profile: dict | None = None,
    calibration_application_context: dict | None = None,
    calibrate_model: bool = True,
    annual_periods: list[Mapping[str, object]] | None = None,
) -> dict:
    """Run the full model on input_csv, write PNGs + Excel, return a stats dict.

    progress_cb(frac, msg): optional; forwarded to the Solectria time loop.
    Returns measured/predicted energy totals and pred-vs-meas % per system.
    """
    se_inv_eff = float(solaredge_inverter_efficiency)
    se_bos_eff = float(solaredge_bos_efficiency)
    sol_inv_eff = float(solectria_inverter_efficiency)
    sol_bos_eff = float(solectria_bos_efficiency)
    se_eff = se_inv_eff * se_bos_eff
    sol_eff = sol_inv_eff * sol_bos_eff
    selected_iam_model, effective_iam_a_r = resolve_iam_settings(
        iam_model=iam_model,
        iam_a_r=iam_a_r,
        include_iam=include_iam,
    )
    martin_ruiz_selected = selected_iam_model == IAM_MODEL_MARTIN_RUIZ
    iam_customized = bool(
        martin_ruiz_selected
        and effective_iam_a_r is not None
        and not np.isclose(effective_iam_a_r, A_R)
    )
    if calibration_profile is not None:
        validate_calibration_profile_physics(calibration_profile)
    application_context_supplied = calibration_application_context is not None
    if calibration_application_context is None:
        calibration_application: dict = {}
    elif isinstance(calibration_application_context, Mapping):
        calibration_application = deepcopy(
            dict(calibration_application_context)
        )
    else:
        raise ValueError("calibration_application_context must be an object.")

    quality_source = (
        ((data_quality_context or {}).get("report") or {}).get("source") or {}
    )
    if requested_start is None and requested_end is None:
        requested_start = quality_source.get("requested_start")
        requested_end = quality_source.get("requested_end")

    data_quality_warnings: list[str] = []
    if input_kind == "midc":
        df, data_quality_warnings = parse_midc_csv(input_csv)
    elif input_kind == "historian":
        df = parse_input_csv(input_csv)
    else:
        raise ValueError(f"Unsupported model input kind: {input_kind}")
    def prediction_progress(frac: float, msg: str) -> None:
        if progress_cb:
            progress_cb(max(0.0, min(1.0, frac)) * 0.84, msg)

    df, dhi_source = predict_ac_power(
        df,
        progress_cb=prediction_progress,
        backtrack=backtrack,
        se_eff=se_eff,
        sol_eff=sol_eff,
        include_iam=False,
        iam_a_r=effective_iam_a_r,
        iam_model=selected_iam_model,
    )
    if curtailment_enabled and curtailment_limit_kw is None:
        curtailment_limit_kw = DEFAULT_CURTAILMENT_LIMIT_KW
    calibration_factors: dict | None = None
    factor_driver_diagnostics: dict | None = None
    fitting_enabled = bool(
        calibrate_model
        and calibration_profile is None
        and input_kind == "historian"
        and not annual_mode
    )
    frozen_profile_enabled = calibration_profile is not None
    calibration_enabled = fitting_enabled or frozen_profile_enabled
    if calibration_enabled:
        if frozen_profile_enabled:
            if progress_cb:
                progress_cb(
                    0.85,
                    "Applying frozen baseline seasonal calibration factors",
                )
            profile_for_application = deepcopy(calibration_profile)
            context_substitution = calibration_application.get(
                "seasonal_substitution"
            )
            if (
                context_substitution is not None
                and "seasonal_substitution" not in profile_for_application
            ):
                # The backend keeps the resolved factor profile and consent
                # audit as separate immutable provenance objects. Join copies
                # only for model application so neither source is mutated.
                profile_for_application["seasonal_substitution"] = deepcopy(
                    context_substitution
                )
            (
                df,
                calibration_factors,
                factor_driver_diagnostics,
            ) = apply_frozen_seasonal_calibration(
                df,
                calibration_profile=profile_for_application,
            )
        else:
            if progress_cb:
                progress_cb(
                    0.85,
                    "Calculating and applying seasonal calibration factors",
                )
            (
                df,
                calibration_factors,
                factor_driver_diagnostics,
            ) = apply_seasonal_calibration(
                df,
                expected_interval_seconds=expected_interval_seconds,
                requested_start=requested_start,
                requested_end=requested_end,
                maximum_predicted_power_w_by_system={
                    "se": (
                        float(curtailment_limit_kw) * 1_000.0
                        if curtailment_enabled
                        and curtailment_limit_kw is not None
                        else None
                    ),
                    "sol": min(
                        SOLECTRIA_INVERTER_AC_RATING_W,
                        (
                            float(curtailment_limit_kw) * 1_000.0
                            if curtailment_enabled
                            and curtailment_limit_kw is not None
                            else SOLECTRIA_INVERTER_AC_RATING_W
                        ),
                    ),
                },
            )
        data_quality_warnings.extend(calibration_factors.get("warnings") or [])
        if data_quality_context:
            cleaning = data_quality_context.get("cleaning") or {}
            excluded_rows = int(cleaning.get("excluded_rows") or 0)
            retained_issues = cleaning.get("retained_issue_ids") or []
            if excluded_rows:
                data_quality_warnings.append(
                    f"Data-quality review excluded {excluded_rows} historian row(s)."
                )
            if retained_issues:
                data_quality_warnings.append(
                    "Data-quality review retained: "
                    + ", ".join(str(value) for value in retained_issues)
                )
    # Seasonal factors correct the physics estimate but cannot create an
    # operating point above the installed inverter's continuous AC nameplate.
    df["sol_predicted_power_w"] = pd.to_numeric(
        df["sol_predicted_power_w"],
        errors="coerce",
    ).clip(upper=SOLECTRIA_INVERTER_AC_RATING_W)
    if progress_cb:
        progress_cb(0.86, "Applying curtailment and integrating energy")
    active_curtailment_limit_kw = (
        float(curtailment_limit_kw) if curtailment_enabled else None
    )
    df = apply_curtailment(df, active_curtailment_limit_kw)
    df = add_energy(
        df,
        expected_interval_seconds=expected_interval_seconds,
        requested_start=requested_start,
        requested_end=requested_end,
    )
    if (
        fitting_enabled
        and calibration_factors is not None
    ):
        calibration_factors["energy_balance"] = (
            calibrated_energy_balance_summary(df)
        )

    if progress_cb:
        progress_cb(0.89, "Rendering power and energy charts")
    plot_results(
        df,
        out_prefix=output_base,
        annual_mode=annual_mode,
        calibrated=calibration_enabled,
    )
    if calibration_enabled:
        if progress_cb:
            progress_cb(0.92, "Rendering uncalibrated comparison charts")
        uncalibrated_plot_df = df.copy()
        for system in ("se", "sol"):
            uncalibrated_plot_df[f"{system}_predicted_power_w"] = df[
                f"{system}_uncalibrated_power_w"
            ]
            uncalibrated_plot_df[f"{system}_predicted_energy_kwh"] = df[
                f"{system}_uncalibrated_energy_kwh"
            ]
        plot_results(
            uncalibrated_plot_df,
            out_prefix=f"{output_base}_uncalibrated",
            annual_mode=annual_mode,
        )
    if annual_mode:
        if progress_cb:
            progress_cb(0.93, "Rendering monthly energy chart")
        plot_monthly_energy(
            df,
            out_prefix=output_base,
            calibrated=calibration_enabled,
        )

    annual_energy_rows: list[dict[str, object]] = []
    annual_energy_cdf: dict[str, object] = {
        "population": "complete_calendar_years",
        "eligible_years": [],
        "excluded_years": [],
        "series": {"solaredge": [], "solectria": [], "combined": []},
    }
    if annual_mode:
        annual_energy_rows, annual_energy_cdf = annual_energy_by_year(
            df,
            annual_periods,
        )

    seasonal_substitution = (
        deepcopy(calibration_factors.get("seasonal_substitution"))
        if calibration_factors
        and calibration_factors.get("seasonal_substitution") is not None
        else None
    )
    calibration_source = (
        {
            "origin_job_id": calibration_factors.get("origin_job_id"),
            "origin_review_id": calibration_factors.get("origin_review_id"),
            "origin_source_sha256": calibration_factors.get(
                "origin_source_sha256"
            ),
            "profile_fingerprint": _first_present(
                calibration_application,
                (
                    "resolved_profile_fingerprint",
                    "resolved_profile_sha256",
                    "profile_fingerprint",
                    "origin_profile_fingerprint",
                    "origin_profile_sha256",
                ),
            ),
        }
        if calibration_factors
        else None
    )

    meta = {
        "script": "sbe_pv_model.py",
        "version": __version__,
        "run_timestamp_utc": pd.Timestamp.now(tz="UTC").isoformat(),
        "input_csv": input_csv,
        "input_kind": input_kind,
        "annual_mode": bool(annual_mode),
        "input_is_utc": bool(INPUT_IS_UTC),
        "local_timezone": str(df.index.tz),
        "dhi_source": dhi_source,
        "SolarEdge_inverter_eff": float(se_inv_eff),
        "SolarEdge_BOS_eff": float(se_bos_eff),
        "SolarEdge_total_eff": float(se_eff),
        "Solectria_inverter_eff": float(sol_inv_eff),
        "Solectria_BOS_eff": float(sol_bos_eff),
        "Solectria_total_eff": float(sol_eff),
        "IAM_enabled": True,
        "IAM_customized": iam_customized,
        "IAM_model": selected_iam_model,
        "IAM_a_r": (
            float(effective_iam_a_r) if martin_ruiz_selected else "N/A"
        ),
        "IAM_reference_parity": True,
        "LAT": LAT,
        "LON": LON,
        "AXIS_AZIMUTH": AXIS_AZIMUTH,
        "MAX_ANGLE": MAX_ANGLE,
        "GCR": GCR,
        "BACKTRACK": bool(backtrack),
        "curtailment_enabled": bool(active_curtailment_limit_kw is not None),
        "curtailment_limit_kw": (
            float(active_curtailment_limit_kw)
            if active_curtailment_limit_kw is not None
            else "N/A"
        ),
        "curtailment_scope": "predicted_only",
        "MODULES_PER_BAY": MODULES_PER_BAY,
        "SOLECTRIA_STRINGS": SOLECTRIA_STRINGS,
        "SOLECTRIA_BAYS_PER_STRING": SOLECTRIA_BAYS_PER_STRING,
        "SOLAREDGE_STRINGS": SOLAREDGE_STRINGS,
        "SOLAREDGE_BAYS_PER_STRING": SOLAREDGE_BAYS_PER_STRING,
        "module_name": MODULE_NAME,
        "calibration_physics_version": CALIBRATION_PHYSICS_VERSION,
        "calibration_physics_fingerprint": CALIBRATION_PHYSICS_FINGERPRINT,
        "solectria_physics_version": SOLECTRIA_PHYSICS_VERSION,
        "solectria_physics_fingerprint": SOLECTRIA_PHYSICS_FINGERPRINT,
        "solectria_inverter_model": SOLECTRIA_INVERTER_MODEL,
        "solectria_mppt_count": 1,
        "solectria_mppt_min_v": SOLECTRIA_INVERTER_MPPT_MIN_V,
        "solectria_mppt_max_v": SOLECTRIA_INVERTER_MPPT_MAX_V,
        "solectria_inverter_ac_rating_w": SOLECTRIA_INVERTER_AC_RATING_W,
        "solectria_dc_aggregation": "parallel_iv_common_mppt_v1",
        "data_quality_warnings": " | ".join(data_quality_warnings) or "None",
        "calibration_method": (
            calibration_factors.get("method")
            if calibration_factors
            else "not_applied"
        ),
        "calibration_enabled": calibration_enabled,
        "calibration_kind": (
            "frozen_profile"
            if frozen_profile_enabled
            else "fitted_profile"
            if fitting_enabled
            else "not_applied"
        ),
        "calibration_factors": (
            json.dumps(calibration_factors, sort_keys=True)
            if calibration_factors
            else "None"
        ),
        "calibration_application": (
            json.dumps(
                {
                    key: value
                    for key, value in calibration_application.items()
                    if key not in {"origin_profile", "resolved_profile"}
                },
                sort_keys=True,
                default=str,
                allow_nan=False,
            )
            if application_context_supplied
            else "None"
        ),
        "calibration_source": (
            json.dumps(
                calibration_source,
                sort_keys=True,
                default=str,
                allow_nan=False,
            )
            if calibration_source
            else "None"
        ),
        "seasonal_substitution": (
            json.dumps(
                seasonal_substitution,
                sort_keys=True,
                default=str,
                allow_nan=False,
            )
            if seasonal_substitution
            else "None"
        ),
        "data_quality_reviewed": bool(data_quality_context),
        "_calibration_factors": calibration_factors,
        "_factor_driver_diagnostics": factor_driver_diagnostics,
        "_data_quality_context": data_quality_context,
        "_calibration_profile": deepcopy(calibration_profile),
        "_calibration_application_context": deepcopy(
            calibration_application
        ),
        "_annual_energy_by_year": deepcopy(annual_energy_rows),
        "_annual_energy_cdf": deepcopy(annual_energy_cdf),
    }
    if progress_cb:
        progress_cb(0.97, "Creating Excel workbook")
    write_excel(df, f"{output_base}.xlsx", meta, annual_mode=annual_mode)
    if progress_cb:
        progress_cb(
            1.0,
            "Calibrated annual model predictions ready"
            if annual_mode and calibration_enabled
            else "Annual model predictions ready"
            if annual_mode
            else (
                "Calibrated model predictions ready"
                if calibration_enabled
                else "Uncalibrated model predictions ready"
            ),
        )

    se_meas = float(df["se_measured_energy_kwh"].iloc[-1])
    se_pred = float(df["se_predicted_energy_kwh"].iloc[-1])
    sol_meas = float(df["sol_measured_energy_kwh"].iloc[-1])
    sol_pred = float(df["sol_predicted_energy_kwh"].iloc[-1])
    se_uncalibrated = (
        float(df["se_uncalibrated_energy_kwh"].iloc[-1])
        if "se_uncalibrated_energy_kwh" in df
        else se_pred
    )
    sol_uncalibrated = (
        float(df["sol_uncalibrated_energy_kwh"].iloc[-1])
        if "sol_uncalibrated_energy_kwh" in df
        else sol_pred
    )

    def _safe(x, ndigits=1):
        # JSON-safe: NaN/Inf -> None; else rounded float.
        try:
            xf = float(x)
        except (TypeError, ValueError):
            return None
        if not np.isfinite(xf):
            return None
        return round(xf, ndigits)

    def _pct(pred, meas):
        if not meas or not np.isfinite(meas):
            return None
        return _safe((pred - meas) / meas * 100.0, 2)

    predicted_difference = se_pred - sol_pred
    predicted_difference_pct = (
        predicted_difference / sol_pred * 100.0 if sol_pred else np.nan
    )

    return {
        "mode": "annual" if annual_mode else "validation",
        "model_version": __version__,
        "calibration_physics_version": CALIBRATION_PHYSICS_VERSION,
        "calibration_physics_fingerprint": CALIBRATION_PHYSICS_FINGERPRINT,
        "solectria_physics_version": SOLECTRIA_PHYSICS_VERSION,
        "solectria_physics_fingerprint": SOLECTRIA_PHYSICS_FINGERPRINT,
        "solectria_inverter_model": SOLECTRIA_INVERTER_MODEL,
        "solectria_mppt_count": 1,
        "solectria_mppt_min_v": SOLECTRIA_INVERTER_MPPT_MIN_V,
        "solectria_mppt_max_v": SOLECTRIA_INVERTER_MPPT_MAX_V,
        "solectria_inverter_ac_rating_w": SOLECTRIA_INVERTER_AC_RATING_W,
        "solectria_dc_aggregation": "parallel_iv_common_mppt_v1",
        "se_measured_kwh": None if annual_mode else _safe(se_meas),
        "se_predicted_kwh": _safe(se_pred),
        "sol_measured_kwh": None if annual_mode else _safe(sol_meas),
        "sol_predicted_kwh": _safe(sol_pred),
        "se_pct": None if annual_mode else _pct(se_pred, se_meas),
        "sol_pct": None if annual_mode else _pct(sol_pred, sol_meas),
        "uncalibrated": (
            {
                "se_predicted_kwh": _safe(se_uncalibrated),
                "sol_predicted_kwh": _safe(sol_uncalibrated),
                "se_pct": (
                    None
                    if annual_mode
                    else _pct(se_uncalibrated, se_meas)
                ),
                "sol_pct": (
                    None
                    if annual_mode
                    else _pct(sol_uncalibrated, sol_meas)
                ),
            }
            if calibration_enabled or not annual_mode
            else None
        ),
        "physics_only": (
            {
                "se_predicted_kwh": _safe(se_uncalibrated),
                "sol_predicted_kwh": _safe(sol_uncalibrated),
            }
            if calibration_enabled
            else None
        ),
        "calibration_adjusted": (
            {
                "se_predicted_kwh": _safe(se_pred),
                "sol_predicted_kwh": _safe(sol_pred),
            }
            if calibration_enabled
            else None
        ),
        "calibration_factors": calibration_factors,
        "calibration_enabled": calibration_enabled,
        "calibration_kind": (
            "frozen_profile"
            if frozen_profile_enabled
            else "fitted_profile"
            if fitting_enabled
            else "not_applied"
        ),
        "calibration_application": (
            deepcopy(calibration_application)
            if application_context_supplied
            else None
        ),
        "calibration_source": calibration_source,
        "calibration_profile_fingerprint": (
            calibration_source.get("profile_fingerprint")
            if calibration_source
            else None
        ),
        "seasonal_substitution": seasonal_substitution,
        "seasonal_substitution_warning": (
            "Fall used Spring substitute"
            if seasonal_substitution is not None
            else None
        ),
        "energy_result_label": (
            "calibration-adjusted"
            if calibration_enabled
            else "physics-only prediction"
        ),
        "factor_driver_diagnostics": factor_driver_diagnostics,
        "data_quality_review": data_quality_context,
        "predicted_difference_kwh": _safe(predicted_difference),
        "predicted_difference_pct": _safe(predicted_difference_pct, 2),
        "dhi_source": dhi_source,
        "data_quality_warnings": data_quality_warnings,
        "backtrack": bool(backtrack),
        "solaredge_inverter_efficiency": _safe(se_inv_eff, 4),
        "solaredge_bos_efficiency": _safe(se_bos_eff, 4),
        "solaredge_total_efficiency": _safe(se_eff, 4),
        "solectria_inverter_efficiency": _safe(sol_inv_eff, 4),
        "solectria_bos_efficiency": _safe(sol_bos_eff, 4),
        "solectria_total_efficiency": _safe(sol_eff, 4),
        "include_iam": martin_ruiz_selected,
        "iam_customized": iam_customized,
        "iam_model": selected_iam_model,
        "iam_a_r": (
            _safe(effective_iam_a_r, 4) if martin_ruiz_selected else None
        ),
        "iam_reference_parity": True,
        "curtailment_enabled": bool(active_curtailment_limit_kw is not None),
        "curtailment_limit_kw": (
            _safe(active_curtailment_limit_kw, 3)
            if active_curtailment_limit_kw is not None
            else None
        ),
        "curtailment_scope": "predicted_only",
        "n_rows": int(len(df)),
        "annual_energy_by_year": deepcopy(annual_energy_rows),
        "annual_energy_cdf": deepcopy(annual_energy_cdf),
        "ac_png": f"{output_base}_ac_power.png",
        "energy_png": f"{output_base}_cumulative_energy.png",
        "uncalibrated_ac_png": (
            f"{output_base}_uncalibrated_ac_power.png"
            if calibration_enabled
            else None
        ),
        "uncalibrated_energy_png": (
            f"{output_base}_uncalibrated_cumulative_energy.png"
            if calibration_enabled
            else None
        ),
        "monthly_png": (
            f"{output_base}_monthly_energy.png" if annual_mode else None
        ),
        "excel": f"{output_base}.xlsx",
    }


def main() -> None:
    print(f"sbe_pv_model.py (v{__version__}) — input: {INPUT_CSV}")
    stats = run_model(
        INPUT_CSV,
        OUTPUT_BASE,
        expected_interval_seconds=3_600,
    )
    print(f"Wrote: {stats['excel']}, {stats['ac_png']}, {stats['energy_png']}")
    print("End-of-period energy (kWh):")
    print(f"  SolarEdge  measured={stats['se_measured_kwh']:,.2f}  predicted={stats['se_predicted_kwh']:,.2f}")
    print(f"  Solectria  measured={stats['sol_measured_kwh']:,.2f}  predicted={stats['sol_predicted_kwh']:,.2f}")


if __name__ == "__main__":
    main()
