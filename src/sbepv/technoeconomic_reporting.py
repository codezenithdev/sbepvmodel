"""Deterministic Phase-4 exports for probabilistic technoeconomic results.

The worker calls :func:`generate_technoeconomic_exports` only after it has sealed
the Phase-3 calculation payload.  This module treats that no-pickle NPZ as the
authoritative realization table, re-verifies its durable identity, streams the
complete CSV/XLSX tables, renders static diagnostics on the non-GUI backend, and
returns an integrity manifest suitable for the terminal job record.
"""

from __future__ import annotations

from collections.abc import Callable, Iterable, Iterator, Mapping, Sequence
from datetime import datetime
import csv
import hashlib
import json
import math
import os
from pathlib import Path, PurePosixPath
import re
import secrets
import shutil
import stat
import struct
from typing import Any
import zipfile

import matplotlib

matplotlib.use("Agg")

import numpy as np
import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from sbepv import technoeconomic as technoeconomic_kernel
from sbepv.api import config
from sbepv.api import artifacts as api_artifacts
from sbepv.api import technoeconomic as technoeconomic_api
from sbepv.api import serializers


EXPORT_MANIFEST_SCHEMA_VERSION = "technoeconomic-exports-manifest-v1"
CSV_FORMAT_VERSION = "technoeconomic-csv-v1"
CSV_BUNDLE_SCHEMA_VERSION = "technoeconomic-csv-bundle-v1"
XLSX_SCHEMA_VERSION = "technoeconomic-xlsx-v1"
PNG_SCHEMA_VERSION = "technoeconomic-plot-v1"
XLSX_LOGICAL_HASH_VERSION = "technoeconomic-xlsx-logical-row-v1"

APPLIED_EXPORT_MANIFEST_SCHEMA_VERSION = "technoeconomic-exports-manifest-v2"
APPLIED_CSV_FORMAT_VERSION = "technoeconomic-csv-v2"
APPLIED_CSV_BUNDLE_SCHEMA_VERSION = "technoeconomic-csv-bundle-v2"
APPLIED_XLSX_SCHEMA_VERSION = "technoeconomic-xlsx-v2"

COMMERCIAL_SCALING_EXPORT_MANIFEST_SCHEMA_VERSION = (
    "technoeconomic-exports-manifest-v3"
)
COMMERCIAL_SCALING_CSV_FORMAT_VERSION = "technoeconomic-csv-v3"
COMMERCIAL_SCALING_CSV_BUNDLE_SCHEMA_VERSION = "technoeconomic-csv-bundle-v3"
COMMERCIAL_SCALING_XLSX_SCHEMA_VERSION = "technoeconomic-xlsx-v3"

STANDALONE_COMMERCIAL_EXPORT_MANIFEST_SCHEMA_VERSION = (
    "technoeconomic-exports-manifest-v4"
)
STANDALONE_COMMERCIAL_CSV_FORMAT_VERSION = "technoeconomic-csv-v4"
STANDALONE_COMMERCIAL_CSV_BUNDLE_SCHEMA_VERSION = (
    "technoeconomic-csv-bundle-v4"
)
STANDALONE_COMMERCIAL_XLSX_SCHEMA_VERSION = "technoeconomic-xlsx-v4"

PAIRED_COMMERCIAL_EXPORT_MANIFEST_SCHEMA_VERSION = (
    "technoeconomic-exports-manifest-v5"
)
PAIRED_COMMERCIAL_CSV_FORMAT_VERSION = "technoeconomic-csv-v5"
PAIRED_COMMERCIAL_CSV_BUNDLE_SCHEMA_VERSION = "technoeconomic-csv-bundle-v5"
PAIRED_COMMERCIAL_XLSX_SCHEMA_VERSION = "technoeconomic-xlsx-v5"

LIFECYCLE_EXPORT_MANIFEST_SCHEMA_VERSION = "technoeconomic-exports-manifest-v6"
LIFECYCLE_CSV_FORMAT_VERSION = "technoeconomic-csv-v6"
LIFECYCLE_CSV_BUNDLE_SCHEMA_VERSION = "technoeconomic-csv-bundle-v6"
LIFECYCLE_XLSX_SCHEMA_VERSION = "technoeconomic-xlsx-v6"
LIFECYCLE_XLSX_LOGICAL_HASH_VERSION = "technoeconomic-xlsx-logical-row-v2"
LIFECYCLE_FORMULA_TEMPLATE_HASH_VERSION = "tea-formula-template-hash-v1"
LIFECYCLE_DECISION_RULE_VERSION = "tea-upgrade-decision-v1"

CSV_BUNDLE_FILENAME = "technoeconomic-results-csv-v1.zip"
XLSX_FILENAME = "technoeconomic-results-v1.xlsx"
CDF_PLOT_FILENAME = "technoeconomic-cdf-v1.png"
SENSITIVITY_PLOT_FILENAME = "technoeconomic-sensitivity-v1.png"
CONVERGENCE_PLOT_FILENAME = "technoeconomic-convergence-v1.png"


def export_contract_versions(calculation_contract_version: str) -> dict[str, str]:
    """Return the export schema family pinned to one calculation contract."""

    if calculation_contract_version == technoeconomic_kernel.LEGACY_CALCULATION_CONTRACT_VERSION:
        return {
            "manifest": EXPORT_MANIFEST_SCHEMA_VERSION,
            "csv_format": CSV_FORMAT_VERSION,
            "csv_bundle": CSV_BUNDLE_SCHEMA_VERSION,
            "csv_bundle_manifest_filename": "csv-bundle-manifest-v1.json",
            "xlsx": XLSX_SCHEMA_VERSION,
            "png": PNG_SCHEMA_VERSION,
            "xlsx_logical_hash": XLSX_LOGICAL_HASH_VERSION,
        }
    if calculation_contract_version == technoeconomic_kernel.CALCULATION_CONTRACT_VERSION:
        return {
            "manifest": APPLIED_EXPORT_MANIFEST_SCHEMA_VERSION,
            "csv_format": APPLIED_CSV_FORMAT_VERSION,
            "csv_bundle": APPLIED_CSV_BUNDLE_SCHEMA_VERSION,
            "csv_bundle_manifest_filename": "csv-bundle-manifest-v2.json",
            "xlsx": APPLIED_XLSX_SCHEMA_VERSION,
            # The PNG rendering contract and logical-row hash algorithm are
            # unchanged.  The v2 manifest/calculation binding and XLSX schema
            # carry the applied-capacity semantics without mis-versioning either
            # byte-format algorithm.
            "png": PNG_SCHEMA_VERSION,
            "xlsx_logical_hash": XLSX_LOGICAL_HASH_VERSION,
        }
    if (
        calculation_contract_version
        == technoeconomic_kernel.COMMERCIAL_SCALING_CALCULATION_CONTRACT_VERSION
    ):
        return {
            "manifest": COMMERCIAL_SCALING_EXPORT_MANIFEST_SCHEMA_VERSION,
            "csv_format": COMMERCIAL_SCALING_CSV_FORMAT_VERSION,
            "csv_bundle": COMMERCIAL_SCALING_CSV_BUNDLE_SCHEMA_VERSION,
            "csv_bundle_manifest_filename": "csv-bundle-manifest-v3.json",
            "xlsx": COMMERCIAL_SCALING_XLSX_SCHEMA_VERSION,
            "png": PNG_SCHEMA_VERSION,
            "xlsx_logical_hash": XLSX_LOGICAL_HASH_VERSION,
        }
    if (
        calculation_contract_version
        == technoeconomic_kernel.PAIRED_COMMERCIAL_CALCULATION_CONTRACT_VERSION
    ):
        return {
            "manifest": PAIRED_COMMERCIAL_EXPORT_MANIFEST_SCHEMA_VERSION,
            "csv_format": PAIRED_COMMERCIAL_CSV_FORMAT_VERSION,
            "csv_bundle": PAIRED_COMMERCIAL_CSV_BUNDLE_SCHEMA_VERSION,
            "csv_bundle_manifest_filename": "csv-bundle-manifest-v5.json",
            "xlsx": PAIRED_COMMERCIAL_XLSX_SCHEMA_VERSION,
            "png": PNG_SCHEMA_VERSION,
            "xlsx_logical_hash": XLSX_LOGICAL_HASH_VERSION,
        }
    if calculation_contract_version == getattr(
        technoeconomic_kernel,
        "LIFECYCLE_CALCULATION_CONTRACT_VERSION",
        "tea-calculation-v6",
    ):
        return {
            "manifest": LIFECYCLE_EXPORT_MANIFEST_SCHEMA_VERSION,
            "csv_format": LIFECYCLE_CSV_FORMAT_VERSION,
            "csv_bundle": LIFECYCLE_CSV_BUNDLE_SCHEMA_VERSION,
            "csv_bundle_manifest_filename": "csv-bundle-manifest-v6.json",
            "xlsx": LIFECYCLE_XLSX_SCHEMA_VERSION,
            "png": PNG_SCHEMA_VERSION,
            "xlsx_logical_hash": LIFECYCLE_XLSX_LOGICAL_HASH_VERSION,
            "formula_template_hash": LIFECYCLE_FORMULA_TEMPLATE_HASH_VERSION,
        }
    if (
        calculation_contract_version
        == technoeconomic_kernel.STANDALONE_COMMERCIAL_CALCULATION_CONTRACT_VERSION
    ):
        return {
            "manifest": STANDALONE_COMMERCIAL_EXPORT_MANIFEST_SCHEMA_VERSION,
            "csv_format": STANDALONE_COMMERCIAL_CSV_FORMAT_VERSION,
            "csv_bundle": STANDALONE_COMMERCIAL_CSV_BUNDLE_SCHEMA_VERSION,
            "csv_bundle_manifest_filename": "csv-bundle-manifest-v4.json",
            "xlsx": STANDALONE_COMMERCIAL_XLSX_SCHEMA_VERSION,
            "png": PNG_SCHEMA_VERSION,
            "xlsx_logical_hash": XLSX_LOGICAL_HASH_VERSION,
        }
    raise TechnoeconomicExportError(
        f"Unsupported calculation contract for export: {calculation_contract_version!r}"
    )

_SAFE_FILENAME = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]*$")
_HASH_RE = re.compile(r"^[0-9a-f]{64}$")
_FIXED_ZIP_TIME = (1980, 1, 1, 0, 0, 0)
_CANCEL_INTERVAL = 1024

_BLUE = "#32679D"
_BLUE_LIGHT = "#B9CEE2"
_GOLD = "#C4932D"
_INK = "#26323D"
_GRID = "#D8DEE5"

LIFECYCLE_WORKBOOK_SHEET_ORDER = (
    "Summary",
    "Decision Charts",
    "Lifecycle Charts",
    "Reliability Charts",
    "Formula Catalog",
    "Calculation Audit",
    "Realizations",
    "Annual Lifecycle",
    "Reliability Summary",
    "Representative Event Traces",
    "Input Specifications",
    "Target Design",
    "Reliability Inputs",
    "Energy Snapshot",
    "Weather Summary",
    "Capacity and Basis",
    "Common-Cost Audit",
    "Cost-Coverage Audit",
    "Commercial Transfer",
    "Commercial LCOE",
    "Metric CDFs",
    "Sensitivity",
    "Convergence",
    "Provenance",
    "Checks",
)

LIFECYCLE_FORMULA_CATALOG_COLUMNS = (
    "formula_id",
    "name",
    "equation",
    "excel_template",
    "inputs",
    "units",
    "timing",
    "guards",
    "output",
    "contract_section",
)
LIFECYCLE_AUDIT_COLUMNS = (
    "audit_id",
    "formula_id",
    "selection_label",
    "realization_index",
    "system",
    "project_year",
    "component_id",
    "frozen_authority",
    "formula_replica",
    "difference",
    "binary64_tolerance",
    "status",
    "economic_decision_tolerance",
    "notes",
)
LIFECYCLE_TRACE_SELECTION_COLUMNS = (
    "selection_label",
    "quantile",
    "realization_index",
    "upgrade_npv_usd",
)
LIFECYCLE_TRACE_ANNUAL_COLUMNS = (
    "selection_label",
    "quantile",
    "realization_index",
    "system",
    "project_year",
    "weather_year",
    "source_energy_kwh",
    "target_source_energy_kwh",
    "degradation_factor",
    "base_availability",
    "component_availability",
    "common_cause_availability",
    "target_availability",
    "source_availability",
    "availability_adjustment",
    "delivered_energy_kwh",
    "discount_factor",
    "base_om_cost_usd",
    "scheduled_cost_usd",
    "preventive_cost_usd",
    "corrective_cost_usd",
    "common_cause_cost_usd",
    "terminal_cost_usd",
    "annual_cost_usd",
    "annual_cost_with_terminal_usd",
    "delta_energy_kwh",
    "delta_cost_usd",
    "electricity_value_usd_per_kwh",
    "incremental_cashflow_usd",
    "pv_incremental_cashflow_usd",
    "cumulative_upgrade_npv_usd",
)
LIFECYCLE_TRACE_COMPONENT_COLUMNS = (
    "selection_label",
    "realization_index",
    "system",
    "project_year",
    "component_id",
    "category",
    "cohort_age",
    "component_year_total_row",
    "start_count",
    "expected_start_count",
    "annual_failure_probability",
    "event_failures",
    "expected_failures",
    "preventive_replacements",
    "expected_preventive_replacements",
    "spares_start",
    "stocked_replacements",
    "emergency_replacements",
    "restock_quantity",
    "spares_end",
    "downtime_fraction",
    "expected_downtime_fraction",
    "hardware_cost_usd",
    "labor_cost_usd",
    "mobilization_cost_usd",
    "warranty_credit_usd",
    "corrective_cost_usd",
    "preventive_cost_usd",
)
LIFECYCLE_TRACE_COLUMNS = (
    "record_type",
    *LIFECYCLE_TRACE_SELECTION_COLUMNS,
    *tuple(
        column
        for column in LIFECYCLE_TRACE_ANNUAL_COLUMNS
        if column not in LIFECYCLE_TRACE_SELECTION_COLUMNS
    ),
    *tuple(
        column
        for column in LIFECYCLE_TRACE_COMPONENT_COLUMNS
        if column
        not in {
            *LIFECYCLE_TRACE_SELECTION_COLUMNS,
            *LIFECYCLE_TRACE_ANNUAL_COLUMNS,
        }
    ),
)


CHART_CONTRACTS: dict[str, dict[str, Any]] = {
    "cdf_v1": {
        "question": "What is the empirical distribution of each reportable output?",
        "family": "distribution",
        "variant": "right_continuous_empirical_cdf_small_multiples",
        "fields": ["metric_id", "value", "cumulative_probability"],
        "population": "finite metric-specific realization population",
        "denominator": "shown separately in every panel subtitle",
        "palette_policy": "single-root preferred",
        "palette": [_BLUE, _INK, _GRID],
        "non_color_cues": ["direct metric titles", "shared probability scale"],
        "render_point_cap_per_metric": 1200,
        "render_point_selection": "even-index grid plus endpoints and P5/P50/P95 neighbors",
        "filename": CDF_PLOT_FILENAME,
    },
    "sensitivity_v1": {
        "question": "Which entered predictors contribute most to each rank model?",
        "family": "comparison_and_ranking",
        "variant": "horizontal_incremental_r_squared_small_multiples",
        "fields": ["response_id", "predictor_id", "incremental_r_squared"],
        "population": "finite response-specific sensitivity population",
        "denominator": "sample count shown in each panel subtitle",
        "palette_policy": "hard two-root cap",
        "palette": [_BLUE, _GOLD, _INK, _GRID],
        "non_color_cues": ["signed beta glyph", "direct predictor labels"],
        "display_top_n_per_response": 15,
        "display_order": "incremental R-squared descending; entry order and stable ID ties",
        "filename": SENSITIVITY_PLOT_FILENAME,
    },
    "convergence_v1": {
        "question": "How do cumulative percentile estimates change with sample size?",
        "family": "uncertainty_and_benchmark",
        "variant": "p50_checkpoint_small_multiples_with_p5_p95_band",
        "fields": ["realization_count", "p5", "p50", "p95"],
        "population": "deterministic cumulative LHS prefixes",
        "denominator": "checkpoint realization count on the horizontal axis",
        "palette_policy": "single-root preferred",
        "palette": [_BLUE, _BLUE_LIGHT, _INK, _GRID],
        "non_color_cues": ["median line", "open percentile band"],
        "filename": CONVERGENCE_PLOT_FILENAME,
    },
}

STANDALONE_COMMERCIAL_CDF_CHART_CONTRACT_ID = (
    "standalone_commercial_solaredge_lcoe_cdf_v1"
)
STANDALONE_COMMERCIAL_CHART_CONTRACTS: dict[str, dict[str, Any]] = {
    STANDALONE_COMMERCIAL_CDF_CHART_CONTRACT_ID: {
        "question": "What is the probability distribution of commercial SolarEdge LCOE?",
        "family": "distribution",
        "variant": "single_right_continuous_empirical_cdf",
        "fields": ["value", "cumulative_probability"],
        "population": "finite standalone commercial SolarEdge LCOE realizations",
        "denominator": "shown in the panel subtitle",
        "palette_policy": "single-root preferred",
        "palette": [_BLUE, _INK, _GRID],
        "non_color_cues": ["direct metric title", "P10/P50/P90 neighbors"],
        "render_point_cap": 1200,
        "filename": CDF_PLOT_FILENAME,
    },
    "sensitivity_v1": CHART_CONTRACTS["sensitivity_v1"],
    "convergence_v1": CHART_CONTRACTS["convergence_v1"],
}

PAIRED_COMMERCIAL_CDF_CHART_CONTRACT_ID = "paired_commercial_lcoe_cdf_v3"
PAIRED_COMMERCIAL_CHART_CONTRACTS: dict[str, dict[str, Any]] = {
    PAIRED_COMMERCIAL_CDF_CHART_CONTRACT_ID: {
        "question": (
            "How do the commercial Solectria and SolarEdge lifecycle LCOE "
            "distributions compare?"
        ),
        "family": "distribution_comparison",
        "variant": "paired_right_continuous_empirical_cdf_decision_view",
        "fields": ["technology", "value", "cumulative_probability"],
        "source_value_unit": "constant USD/kWh_AC",
        "display_value_unit": "constant-dollar-year USD/MWh_AC",
        "display_value_transform": "source value multiplied by 1000",
        "probability_display": "percent",
        "population": "finite paired commercial lifecycle LCOE realizations",
        "denominator": "shown for each technology",
        "palette_policy": "two named systems with non-color line cues",
        "palette": [_GOLD, _BLUE, _INK, _GRID],
        "non_color_cues": [
            "Solectria dashed line",
            "SolarEdge solid line",
            "direct curve labels",
            "P10/P50/P90 value rows with P50 chart markers",
        ],
        "render_point_cap_per_metric": 1200,
        "filename": CDF_PLOT_FILENAME,
    },
    "sensitivity_v1": CHART_CONTRACTS["sensitivity_v1"],
    "convergence_v1": CHART_CONTRACTS["convergence_v1"],
}

LIFECYCLE_CDF_CHART_CONTRACT_ID = "lifecycle_upgrade_npv_and_lcoe_cdf_v1"
LIFECYCLE_CHART_CONTRACTS: dict[str, dict[str, Any]] = {
    LIFECYCLE_CDF_CHART_CONTRACT_ID: {
        **CHART_CONTRACTS["cdf_v1"],
        "question": (
            "What are the empirical Upgrade-NPV, standalone-LCOE, and "
            "incremental lifecycle distributions?"
        ),
        "primary_metric": "UpgradeNPV_se_minus_sol_USD",
    },
    "sensitivity_v1": CHART_CONTRACTS["sensitivity_v1"],
    "convergence_v1": CHART_CONTRACTS["convergence_v1"],
}


class TechnoeconomicExportError(ValueError):
    """A sealed calculation or generated export failed validation."""


class _SealedCalculation:
    def __init__(
        self,
        *,
        metadata: Mapping[str, Any],
        column_names: Sequence[str],
        columns: Sequence[np.ndarray],
        row_count: int,
    ) -> None:
        self.metadata = dict(metadata)
        self.column_names = tuple(column_names)
        self.columns = tuple(columns)
        self.row_count = row_count
        self.by_name = dict(zip(self.column_names, self.columns))

    def rows(self) -> Iterator[tuple[Any, ...]]:
        for row_index in range(self.row_count):
            yield tuple(_numpy_scalar(column[row_index]) for column in self.columns)


class _Table:
    def __init__(
        self,
        filename: str,
        sheet_name: str,
        columns: Sequence[str],
        rows_factory: Callable[[], Iterable[Sequence[Any]]],
    ) -> None:
        self.filename = filename
        self.sheet_name = sheet_name
        self.columns = tuple(columns)
        self.rows_factory = rows_factory


def _is_lifecycle_contract(value: Any) -> bool:
    return value == getattr(
        technoeconomic_kernel,
        "LIFECYCLE_CALCULATION_CONTRACT_VERSION",
        "tea-calculation-v6",
    )


def _mapping_sequence(value: Any, *, label: str) -> tuple[Mapping[str, Any], ...]:
    if value is None:
        return ()
    if not isinstance(value, (list, tuple)):
        raise TechnoeconomicExportError(f"{label} must be a row sequence")
    rows: list[Mapping[str, Any]] = []
    for row in value:
        if not isinstance(row, Mapping):
            raise TechnoeconomicExportError(f"{label} contains a non-object row")
        rows.append(dict(row))
    return tuple(rows)


def _records_to_table(
    records: Sequence[Mapping[str, Any]],
    *,
    preferred_columns: Sequence[str] = (),
    empty_columns: Sequence[str] = ("record_type", "notes"),
) -> tuple[tuple[str, ...], tuple[tuple[Any, ...], ...]]:
    keys = {str(key) for record in records for key in record}
    columns = tuple(column for column in preferred_columns if column in keys)
    columns += tuple(sorted(keys - set(columns)))
    if not columns:
        columns = tuple(empty_columns)
    rows = tuple(
        tuple(_safe_public_value(record.get(column)) for column in columns)
        for record in records
    )
    return columns, rows


def _formula_registry_records() -> tuple[Mapping[str, Any], ...]:
    registry_function = getattr(technoeconomic_kernel, "formula_registry", None)
    if not callable(registry_function):
        raise TechnoeconomicExportError("The v6 formula registry is unavailable")
    registry = _mapping_sequence(
        registry_function(),
        label="The v6 formula registry",
    )
    required = set(LIFECYCLE_FORMULA_CATALOG_COLUMNS)
    identifiers: set[str] = set()
    for record in registry:
        if set(record) != required:
            raise TechnoeconomicExportError(
                "A v6 formula-registry record has the wrong fields"
            )
        identifier = record.get("formula_id")
        if not isinstance(identifier, str) or not identifier or identifier in identifiers:
            raise TechnoeconomicExportError(
                "The v6 formula registry has an invalid formula ID"
            )
        identifiers.add(identifier)
        for text_field in (
            "name",
            "equation",
            "excel_template",
            "units",
            "timing",
            "guards",
            "output",
            "contract_section",
        ):
            if not isinstance(record.get(text_field), str):
                raise TechnoeconomicExportError(
                    f"Formula registry field {text_field!r} is not text"
                )
        inputs = record.get("inputs")
        if not isinstance(inputs, (list, tuple)) or not all(
            isinstance(item, str) for item in inputs
        ):
            raise TechnoeconomicExportError(
                "Formula registry inputs must be a sequence of strings"
            )
    if not registry:
        raise TechnoeconomicExportError("The v6 formula registry is empty")
    return registry


def _formula_registry_sha256(
    registry: Sequence[Mapping[str, Any]],
) -> str:
    registry_function = getattr(
        technoeconomic_kernel,
        "formula_registry_hash",
        None,
    )
    if callable(registry_function):
        digest = registry_function()
        return _require_digest(digest, "Formula-registry digest")
    return _canonical_json_sha256(list(registry))


def _formula_template_sha256(
    registry: Sequence[Mapping[str, Any]],
) -> str:
    """Hash canonical formula templates separately from exact sheet text."""

    normalized = [
        {
            "formula_id": record["formula_id"],
            "excel_template": str(record["excel_template"])
            .replace("\r\n", "\n")
            .replace("\r", "\n"),
        }
        for record in registry
    ]
    return _canonical_json_sha256(
        {
            "version": LIFECYCLE_FORMULA_TEMPLATE_HASH_VERSION,
            "templates": normalized,
        }
    )


def _lifecycle_summaries(metadata: Mapping[str, Any]) -> Mapping[str, Any]:
    summaries = metadata.get("summaries")
    if not isinstance(summaries, Mapping):
        raise TechnoeconomicExportError("Sealed v6 summaries are unavailable")
    return summaries


def _representative_trace_records(
    metadata: Mapping[str, Any],
) -> tuple[Mapping[str, Any], ...]:
    summaries = _lifecycle_summaries(metadata)
    raw = summaries.get("representative_event_traces")
    if raw is None:
        raw = metadata.get("representative_event_traces")
    bundles: list[Mapping[str, Any]] = []
    if isinstance(raw, Mapping) and (
        "annual" in raw or "components" in raw or "selection" in raw
    ):
        bundles.append(raw)
    elif isinstance(raw, Mapping):
        bundles.extend(
            value for _, value in sorted(raw.items()) if isinstance(value, Mapping)
        )
    elif isinstance(raw, (list, tuple)):
        bundles.extend(value for value in raw if isinstance(value, Mapping))
    else:
        raise TechnoeconomicExportError(
            "Representative v6 event traces are unavailable"
        )

    records: list[Mapping[str, Any]] = []
    selection_labels: set[str] = set()
    for bundle in bundles:
        raw_selection = bundle.get("selection")
        selection_defaults = (
            raw_selection if isinstance(raw_selection, Mapping) else {}
        )
        selection_label = selection_defaults.get("label")
        selection_rows = (
            (raw_selection,)
            if isinstance(raw_selection, Mapping)
            and "selection_label" in raw_selection
            else _mapping_sequence(
                raw_selection,
                label="Representative trace selection",
            )
            if raw_selection is not None and not isinstance(raw_selection, Mapping)
            else ()
        )
        for raw_row in selection_rows:
            row = dict(raw_row)
            row["record_type"] = "selection"
            label = row.get("selection_label")
            if isinstance(label, str):
                selection_labels.add(label)
            records.append(row)
        for record_type, field_name in (("annual", "annual"), ("component", "components")):
            for raw_row in _mapping_sequence(
                bundle.get(field_name),
                label=f"Representative trace {field_name}",
            ):
                row = dict(raw_row)
                if selection_label is not None:
                    row.setdefault("selection_label", selection_label)
                for key in ("quantile", "realization_index"):
                    if key in selection_defaults:
                        row.setdefault(key, selection_defaults[key])
                row["record_type"] = record_type
                label = row.get("selection_label")
                if isinstance(label, str):
                    selection_labels.add(label)
                records.append(row)
    expected_labels = {"NPV-P10", "NPV-P50", "NPV-P90"}
    if selection_labels != expected_labels:
        raise TechnoeconomicExportError(
            "Representative v6 traces must contain NPV-P10, NPV-P50, and NPV-P90"
        )
    return tuple(records)


def _canonical_json_text(value: Any) -> str:
    try:
        return json.dumps(
            value,
            allow_nan=False,
            ensure_ascii=True,
            separators=(",", ":"),
            sort_keys=True,
        )
    except (TypeError, ValueError, OverflowError) as exc:
        raise TechnoeconomicExportError(
            "Export metadata is not finite canonical JSON"
        ) from exc


def _canonical_json_sha256(value: Any) -> str:
    return hashlib.sha256(_canonical_json_text(value).encode("utf-8")).hexdigest()


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _numpy_scalar(value: Any) -> Any:
    if isinstance(value, np.generic):
        return value.item()
    return value


def _safe_filename(value: str) -> str:
    if not isinstance(value, str) or not _SAFE_FILENAME.fullmatch(value):
        raise TechnoeconomicExportError("Unsafe export filename")
    return value


def _require_digest(value: Any, label: str) -> str:
    if not isinstance(value, str) or not _HASH_RE.fullmatch(value):
        raise TechnoeconomicExportError(f"{label} is not a SHA-256 digest")
    return value


def _strict_regular_file(path: Path, *, label: str) -> os.stat_result:
    try:
        details = path.lstat()
    except OSError as exc:
        raise TechnoeconomicExportError(f"{label} cannot be inspected") from exc
    if stat.S_ISLNK(details.st_mode) or not stat.S_ISREG(details.st_mode):
        raise TechnoeconomicExportError(f"{label} is not a regular file")
    return details


def _confined(path: Path, root: Path, *, label: str) -> Path:
    try:
        resolved_root = root.resolve(strict=True)
        resolved = path.resolve(strict=False)
        resolved.relative_to(resolved_root)
    except (OSError, ValueError) as exc:
        raise TechnoeconomicExportError(f"{label} is outside the attempt directory") from exc
    return resolved


def _storage_key(path: Path) -> str:
    try:
        return path.resolve(strict=True).relative_to(
            config.OUTPUT_DIR.resolve(strict=True)
        ).as_posix()
    except (OSError, ValueError) as exc:
        raise TechnoeconomicExportError(
            "Published export is outside the configured output directory"
        ) from exc


def _safe_public_value(value: Any) -> Any:
    value = serializers._public_value(value)

    def normalize(item: Any) -> Any:
        if isinstance(item, Mapping):
            return {
                str(key): normalize(child)
                for key, child in sorted(item.items(), key=lambda pair: str(pair[0]))
            }
        if isinstance(item, (list, tuple)):
            return [normalize(child) for child in item]
        if isinstance(item, float) and not math.isfinite(item):
            return None
        return _numpy_scalar(item)

    return normalize(value)


def _load_sealed_calculation(
    *,
    attempt_directory: Path,
    sealed_calculation_path: Path,
    sealed_calculation_artifact: Mapping[str, Any],
    request_payload: Mapping[str, Any],
    source_snapshot: Mapping[str, Any],
    submission_provenance: Mapping[str, Any],
) -> _SealedCalculation:
    attempt_directory = attempt_directory.resolve(strict=True)
    path = _confined(
        sealed_calculation_path,
        attempt_directory,
        label="Sealed calculation payload",
    )
    details = _strict_regular_file(path, label="Sealed calculation payload")
    expected_bytes = sealed_calculation_artifact.get("byte_count")
    if not isinstance(expected_bytes, int) or isinstance(expected_bytes, bool):
        raise TechnoeconomicExportError("Sealed calculation byte count is invalid")
    if details.st_size != expected_bytes:
        raise TechnoeconomicExportError("Sealed calculation byte count changed")
    expected_sha256 = _require_digest(
        sealed_calculation_artifact.get("sha256"),
        "Sealed calculation digest",
    )
    if not secrets.compare_digest(_sha256_file(path), expected_sha256):
        raise TechnoeconomicExportError("Sealed calculation digest changed")

    try:
        with np.load(path, allow_pickle=False) as payload:
            metadata_bytes = payload["metadata_json_utf8"]
            if metadata_bytes.dtype != np.uint8 or metadata_bytes.ndim != 1:
                raise TechnoeconomicExportError(
                    "Sealed calculation metadata encoding is invalid"
                )
            metadata = json.loads(metadata_bytes.tobytes().decode("utf-8"))
            if not isinstance(metadata, Mapping):
                raise TechnoeconomicExportError(
                    "Sealed calculation metadata is not an object"
                )
            storage = metadata.get("realization_column_storage")
            names = metadata.get("realization_columns")
            if not isinstance(storage, list) or not isinstance(names, list):
                raise TechnoeconomicExportError(
                    "Sealed calculation column metadata is invalid"
                )
            if len(storage) != len(names) or not storage:
                raise TechnoeconomicExportError(
                    "Sealed calculation column metadata is inconsistent"
                )
            columns: list[np.ndarray] = []
            column_names: list[str] = []
            row_count: int | None = None
            for index, record in enumerate(storage):
                if not isinstance(record, Mapping):
                    raise TechnoeconomicExportError(
                        "Sealed calculation column record is invalid"
                    )
                column_name = record.get("column_name")
                storage_name = record.get("storage_name")
                if column_name != names[index] or not isinstance(storage_name, str):
                    raise TechnoeconomicExportError(
                        "Sealed calculation column identity is inconsistent"
                    )
                values = np.asarray(payload[storage_name])
                if values.ndim != 1 or values.dtype.kind not in "biufUS":
                    raise TechnoeconomicExportError(
                        "Sealed calculation contains an unsupported realization column"
                    )
                null_storage = record.get("null_storage_name")
                if null_storage is not None:
                    if not isinstance(null_storage, str):
                        raise TechnoeconomicExportError(
                            "Sealed calculation null-mask identity is invalid"
                        )
                    mask = np.asarray(payload[null_storage])
                    if mask.dtype != np.bool_ or mask.ndim != 1 or len(mask) != len(values):
                        raise TechnoeconomicExportError(
                            "Sealed calculation null mask is invalid"
                        )
                    restored = values.astype(object)
                    restored[mask] = None
                    values = restored
                if row_count is None:
                    row_count = len(values)
                elif len(values) != row_count:
                    raise TechnoeconomicExportError(
                        "Sealed calculation columns have inconsistent row counts"
                    )
                column_names.append(str(column_name))
                columns.append(values)
    except TechnoeconomicExportError:
        raise
    except (KeyError, OSError, ValueError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise TechnoeconomicExportError(
            "Sealed calculation payload cannot be decoded safely"
        ) from exc

    if row_count is None or row_count <= 0:
        raise TechnoeconomicExportError("Sealed calculation has no realization rows")
    expected_rows = sealed_calculation_artifact.get("row_count")
    expected_columns = sealed_calculation_artifact.get("column_count")
    if row_count != expected_rows or len(column_names) != expected_columns:
        raise TechnoeconomicExportError(
            "Sealed calculation dimensions differ from its durable identity"
        )
    expected_hashes = {
        "request_sha256": technoeconomic_api.canonical_json_sha256(request_payload),
        "source_snapshot_sha256": technoeconomic_api.canonical_json_sha256(source_snapshot),
        "submission_provenance_sha256": technoeconomic_api.canonical_json_sha256(
            submission_provenance
        ),
    }
    for key, expected in expected_hashes.items():
        actual = _require_digest(metadata.get(key), f"Sealed metadata {key}")
        if not secrets.compare_digest(actual, expected):
            raise TechnoeconomicExportError(
                f"Sealed calculation {key} does not match the frozen export input"
            )
    return _SealedCalculation(
        metadata=metadata,
        column_names=column_names,
        columns=columns,
        row_count=row_count,
    )


def _compact_cdf_points_for_binding(value: Any) -> Any:
    """Reproduce the durable routine-result projection of sealed CDF blocks."""

    if isinstance(value, Mapping):
        compacted: dict[str, Any] = {}
        for raw_key, item in value.items():
            key = str(raw_key)
            if key == "cdf" and isinstance(item, Mapping):
                values = item.get("values")
                compacted[key] = {
                    "population_count": item.get("population_count"),
                    "point_count": len(values) if isinstance(values, list) else 0,
                    "storage": "sealed_calculation_payload",
                }
            else:
                compacted[key] = _compact_cdf_points_for_binding(item)
        return compacted
    if isinstance(value, list):
        return [_compact_cdf_points_for_binding(item) for item in value]
    return value


def _standalone_headline_projection(
    summary: Mapping[str, Any],
    *,
    maximum_points: int = 1200,
) -> dict[str, Any]:
    """Rebuild the worker's bounded standalone-commercial ECDF projection."""

    cdf = summary.get("cdf")
    if not isinstance(cdf, Mapping):
        raise TechnoeconomicExportError(
            "Standalone commercial headline CDF is unavailable"
        )
    values = list(cdf.get("values") or [])
    probabilities = list(cdf.get("cumulative_probability") or [])
    if not values or len(values) != len(probabilities):
        raise TechnoeconomicExportError(
            "Standalone commercial headline CDF is invalid"
        )
    count = len(values)
    if count <= maximum_points:
        indexes = np.arange(count, dtype=np.int64)
    else:
        required = {0, count - 1}
        probability_vector = np.asarray(probabilities, dtype=np.float64)
        for quantile in (0.10, 0.50, 0.90):
            index = int(np.searchsorted(probability_vector, quantile, side="left"))
            for neighbor in (index - 1, index, index + 1):
                if 0 <= neighbor < count:
                    required.add(neighbor)
        if len(required) > maximum_points:
            raise TechnoeconomicExportError(
                "Standalone commercial CDF display cap is too small"
            )
        selected = set(required)
        remaining = maximum_points - len(required)
        if remaining:
            selected.update(
                np.linspace(0, count - 1, remaining, dtype=np.int64).tolist()
            )
        indexes = np.asarray(sorted(selected), dtype=np.int64)
    full_identity = {
        "values": values,
        "cumulative_count": list(cdf.get("cumulative_count") or []),
        "cumulative_probability": probabilities,
        "population_count": cdf.get("population_count"),
    }
    return {
        "population_count": cdf.get("population_count"),
        "source_point_count": count,
        "display_point_count": int(len(indexes)),
        "values": [values[int(index)] for index in indexes],
        "cumulative_probability": [
            probabilities[int(index)] for index in indexes
        ],
        "full_cdf_sha256": _canonical_json_sha256(full_identity),
        "full_storage": "sealed_calculation_payload",
    }


def _applied_capacity_authority(
    submission_provenance: Mapping[str, Any],
) -> dict[str, Mapping[str, Any]]:
    """Return the immutable v2 applied-capacity receipt, failing closed."""

    receipt = submission_provenance.get("normalization_receipt")
    if not isinstance(receipt, Mapping):
        raise TechnoeconomicExportError(
            "Applied-capacity normalization receipt is missing"
        )
    if receipt.get("capacity_normalization") != "annual_applied_capacity_v1":
        raise TechnoeconomicExportError(
            "Applied-capacity normalization receipt has the wrong method"
        )
    records = receipt.get("applied_capacities")
    if not isinstance(records, Mapping) or set(records) != {
        "solectria",
        "solaredge",
    }:
        raise TechnoeconomicExportError(
            "Applied-capacity normalization receipt is incomplete"
        )
    result: dict[str, Mapping[str, Any]] = {}
    for system in ("solectria", "solaredge"):
        record = records.get(system)
        if not isinstance(record, Mapping):
            raise TechnoeconomicExportError(
                "Applied-capacity normalization record is invalid"
            )
        applied_w = record.get("applied_capacity_w")
        if (
            isinstance(applied_w, bool)
            or not isinstance(applied_w, (int, float))
            or not math.isfinite(float(applied_w))
            or float(applied_w) <= 0
            or record.get("rating_basis")
            not in {"ac_operating_limit", "dc_installed_nameplate"}
            or record.get("selection_method")
            != "enabled_positive_annual_curtailment_else_installed_dc"
            or not isinstance(record.get("source_field"), str)
            or not record.get("source_field")
        ):
            raise TechnoeconomicExportError(
                "Applied-capacity normalization record is invalid"
            )
        result[system] = record
    return result


def _verify_routine_result(
    *,
    metadata: Mapping[str, Any],
    routine_result: Mapping[str, Any],
    request_payload: Mapping[str, Any],
    source_snapshot: Mapping[str, Any],
    submission_provenance: Mapping[str, Any],
    sealed_calculation_artifact: Mapping[str, Any],
) -> None:
    """Bind the complete durable worker projection to frozen export authority."""

    finance = request_payload.get("finance") or {}
    manifest = source_snapshot.get("capacity_manifest") or {}
    systems = manifest.get("systems") or {}
    capacities: dict[str, Any] = {}
    for system, record in sorted(systems.items()):
        if not isinstance(record, Mapping):
            raise TechnoeconomicExportError("Frozen capacity manifest is invalid")
        capacities[str(system)] = {
            "module_model": record.get("module_model"),
            "installed_wdc": record.get("installed_wdc"),
            "physics_version": record.get("calibration_physics_version"),
            "physics_fingerprint": record.get("calibration_physics_fingerprint"),
        }
    evidence_receipt = submission_provenance.get("evidence_receipt") or {}
    if not isinstance(evidence_receipt, Mapping):
        raise TechnoeconomicExportError("Frozen evidence receipt is invalid")
    eligible_rows = source_snapshot.get("eligible_paired_energy_rows") or []
    if not isinstance(eligible_rows, list):
        raise TechnoeconomicExportError("Frozen eligible energy rows are invalid")
    public_identity_fields = (
        "schema_version",
        "artifact_kind",
        "media_type",
        "sha256",
        "byte_count",
        "row_count",
        "column_count",
        "pickle_allowed",
        "public",
    )
    try:
        sealed_identity = {
            key: sealed_calculation_artifact[key] for key in public_identity_fields
        }
        eligible_years = [row["year"] for row in eligible_rows]
    except (KeyError, TypeError) as exc:
        raise TechnoeconomicExportError(
            "Frozen routine-result authority is incomplete"
        ) from exc
    calculation_contract_version = submission_provenance.get(
        "calculation_contract_version"
    )
    commercial_scaling_contract = (
        calculation_contract_version
        == technoeconomic_kernel.COMMERCIAL_SCALING_CALCULATION_CONTRACT_VERSION
    )
    standalone_commercial_contract = (
        calculation_contract_version
        == technoeconomic_kernel.STANDALONE_COMMERCIAL_CALCULATION_CONTRACT_VERSION
    )
    paired_commercial_contract = (
        calculation_contract_version
        == technoeconomic_kernel.PAIRED_COMMERCIAL_CALCULATION_CONTRACT_VERSION
    )
    applied_capacity_contract = calculation_contract_version in {
        technoeconomic_kernel.CALCULATION_CONTRACT_VERSION,
        technoeconomic_kernel.COMMERCIAL_SCALING_CALCULATION_CONTRACT_VERSION,
        technoeconomic_kernel.STANDALONE_COMMERCIAL_CALCULATION_CONTRACT_VERSION,
        technoeconomic_kernel.PAIRED_COMMERCIAL_CALCULATION_CONTRACT_VERSION,
    }
    expected = {
        "schema_version": (
            5
            if paired_commercial_contract
            else 4
            if standalone_commercial_contract
            else 3
            if commercial_scaling_contract
            else (2 if applied_capacity_contract else 1)
        ),
        "calculation_contract_version": submission_provenance.get(
            "calculation_contract_version"
        ),
        "sampling_version": submission_provenance.get("sampling_version"),
        "analysis_basis": request_payload.get("basis"),
        "realization_count": request_payload.get("n"),
        "seed": request_payload.get("seed"),
        "project_life_years": finance.get("project_life_years"),
        "cost_stack_completeness": request_payload.get("cost_stack_completeness"),
        "energy_available": metadata.get("energy_available"),
        "commercial_transfer_status": submission_provenance.get(
            "commercial_transfer_status"
        ),
        "commercial_reference_design": submission_provenance.get(
            "commercial_reference_design"
        ),
        "source_snapshot_sha256": submission_provenance.get(
            "source_snapshot_sha256"
        ),
        "eligible_weather_years": eligible_years,
        "capacity_basis": (
            "frozen_annual_applied_capacity_w"
            if applied_capacity_contract
            else "frozen_annual_module_dc_stc_wdc"
        ),
        "capacities": capacities,
        "input_status": evidence_receipt.get("status"),
        "evidence_class_counts": evidence_receipt.get("evidence_class_counts") or {},
        "common_cost_audit": metadata.get("common_cost_audit"),
        "summaries": _compact_cdf_points_for_binding(metadata.get("summaries")),
        "per_weather_year": _compact_cdf_points_for_binding(
            metadata.get("per_weather_year")
        ),
        "sensitivity": metadata.get("sensitivity"),
        "convergence": metadata.get("convergence"),
        "sealed_calculation": sealed_identity,
    }
    if applied_capacity_contract:
        authority = _applied_capacity_authority(submission_provenance)
        expected["applied_capacities"] = {
            system: {
                "applied_capacity_w": authority[system].get(
                    "applied_capacity_w"
                ),
                "rating_basis": authority[system].get("rating_basis"),
            }
            for system in ("solaredge", "solectria")
        }
    if commercial_scaling_contract:
        scaling = request_payload.get("commercial_scaling") or {}
        if not isinstance(scaling, Mapping):
            raise TechnoeconomicExportError(
                "Frozen commercial-scaling request is invalid"
            )
        unit_multiplier = {"kw": 1_000.0, "mw": 1_000_000.0}.get(
            scaling.get("target_capacity_unit")
        )
        if unit_multiplier is None:
            raise TechnoeconomicExportError(
                "Frozen commercial target-capacity unit is invalid"
            )
        expected["commercial_scaling"] = {
            "target_capacity_w": float(scaling.get("target_capacity"))
            * unit_multiplier,
            "target_rating_basis": scaling.get("target_rating_basis"),
            "marginal_cost_input_id": (
                technoeconomic_kernel.COMMERCIAL_MARGINAL_COST_DIFFERENCE_INPUT_ID
            ),
            "marginal_cost_timing": scaling.get("marginal_cost_timing"),
            "transfer_method": scaling.get("transfer_method"),
        }
    if standalone_commercial_contract:
        standalone = request_payload.get("standalone_commercial") or {}
        if not isinstance(standalone, Mapping):
            raise TechnoeconomicExportError(
                "Frozen standalone-commercial request is invalid"
            )
        unit_multiplier = {"kw": 1_000.0, "mw": 1_000_000.0}.get(
            standalone.get("target_capacity_unit")
        )
        if unit_multiplier is None:
            raise TechnoeconomicExportError(
                "Frozen standalone-commercial target unit is invalid"
            )
        authority = _applied_capacity_authority(submission_provenance)
        source = authority["solaredge"]
        target_capacity_w = float(standalone.get("target_capacity")) * unit_multiplier
        headline_metric_id = technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LCOE
        summaries = metadata.get("summaries") or {}
        headline = summaries.get(headline_metric_id) if isinstance(summaries, Mapping) else None
        if not isinstance(headline, Mapping) or not isinstance(
            headline.get("percentiles"), Mapping
        ):
            raise TechnoeconomicExportError(
                "Sealed standalone-commercial headline summary is invalid"
            )
        expected["standalone_commercial"] = {
            "technology": "solaredge",
            "target_capacity_w": target_capacity_w,
            "target_rating_basis": standalone.get("target_rating_basis"),
            "source_applied_capacity_w": source.get("applied_capacity_w"),
            "source_rating_basis": source.get("rating_basis"),
            "capacity_scale_factor": (
                target_capacity_w / float(source.get("applied_capacity_w"))
            ),
            "transfer_method": standalone.get("transfer_method"),
            "constant_dollar_cost_year": (
                (request_payload.get("finance") or {}).get(
                    "constant_dollar_cost_year"
                )
            ),
            "headline_metric_id": headline_metric_id,
            "unit": "constant_usd_per_kwh_ac",
            "percentiles": headline.get("percentiles"),
            "cdf": _standalone_headline_projection(headline),
            "commercial_cost_line_summaries": (
                summaries.get("commercial_cost_line_summaries") or []
            ),
        }
    if paired_commercial_contract:
        paired = request_payload.get("paired_commercial") or {}
        if not isinstance(paired, Mapping):
            raise TechnoeconomicExportError(
                "Frozen paired-commercial request is invalid"
            )
        unit_multiplier = {"kw": 1_000.0, "mw": 1_000_000.0}.get(
            paired.get("target_capacity_unit")
        )
        if unit_multiplier is None:
            raise TechnoeconomicExportError(
                "Frozen paired-commercial target unit is invalid"
            )
        target_capacity_w = float(paired.get("target_capacity")) * unit_multiplier
        request_systems = paired.get("systems") or []
        if not isinstance(request_systems, list):
            raise TechnoeconomicExportError(
                "Frozen paired-commercial systems are invalid"
            )
        request_system_map = {
            str(system.get("technology")): system
            for system in request_systems
            if isinstance(system, Mapping)
        }
        if set(request_system_map) != {"solectria", "solaredge"}:
            raise TechnoeconomicExportError(
                "Frozen paired-commercial systems are incomplete"
            )
        authority = _applied_capacity_authority(submission_provenance)
        summaries = metadata.get("summaries") or {}
        if not isinstance(summaries, Mapping):
            raise TechnoeconomicExportError(
                "Sealed paired-commercial summaries are invalid"
            )
        cost_summaries = summaries.get("paired_commercial_cost_line_summaries") or []
        if not isinstance(cost_summaries, list):
            raise TechnoeconomicExportError(
                "Sealed paired-commercial cost-line summaries are invalid"
            )
        field_contracts = {
            "solectria": technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_LCOE,
            "solaredge": technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LCOE,
        }
        paired_system_results: dict[str, Any] = {}
        for technology in ("solectria", "solaredge"):
            headline_metric_id = field_contracts[technology]
            headline = summaries.get(headline_metric_id)
            if not isinstance(headline, Mapping) or not isinstance(
                headline.get("percentiles"), Mapping
            ):
                raise TechnoeconomicExportError(
                    f"Sealed paired-commercial {technology} headline summary is invalid"
                )
            source = authority[technology]
            source_w = float(source.get("applied_capacity_w"))
            system_cost_summaries = [
                item
                for item in cost_summaries
                if isinstance(item, Mapping)
                and item.get("technology") == technology
            ]
            if len(system_cost_summaries) != len(
                request_system_map[technology].get("cost_lines") or []
            ):
                raise TechnoeconomicExportError(
                    f"Sealed paired-commercial {technology} cost summaries are incomplete"
                )
            paired_system_results[technology] = {
                "technology": technology,
                "source_applied_capacity_w": source.get("applied_capacity_w"),
                "source_rating_basis": source.get("rating_basis"),
                "capacity_scale_factor": target_capacity_w / source_w,
                "headline_metric_id": headline_metric_id,
                "unit": "constant_usd_per_kwh_ac",
                "percentiles": headline.get("percentiles"),
                "cdf": _standalone_headline_projection(headline),
                "commercial_cost_line_summaries": system_cost_summaries,
            }
        delta_metric_id = technoeconomic_kernel.COMMERCIAL_PAIRED_FIELD_LCOE_DELTA
        delta_headline = summaries.get(delta_metric_id)
        if not isinstance(delta_headline, Mapping) or not isinstance(
            delta_headline.get("percentiles"), Mapping
        ):
            raise TechnoeconomicExportError(
                "Sealed paired-commercial LCOE-delta summary is invalid"
            )
        expected["paired_commercial"] = {
            "target_capacity_w": target_capacity_w,
            "target_rating_basis": paired.get("target_rating_basis"),
            "transfer_method": paired.get("transfer_method"),
            "constant_dollar_cost_year": finance.get("constant_dollar_cost_year"),
            "systems": paired_system_results,
            "lcoe_delta_se_minus_sol": {
                "headline_metric_id": delta_metric_id,
                "unit": "constant_usd_per_kwh_ac",
                "percentiles": delta_headline.get("percentiles"),
                "cdf": _standalone_headline_projection(delta_headline),
            },
        }
    for key in expected:
        if key not in routine_result:
            raise TechnoeconomicExportError(
                f"Durable routine-result field {key} is missing"
            )
    actual_digest = _canonical_json_sha256(routine_result)
    expected_digest = _canonical_json_sha256(expected)
    if not secrets.compare_digest(actual_digest, expected_digest):
        raise TechnoeconomicExportError(
            "Durable routine result differs from frozen or sealed authority"
        )


def _distribution_columns(distribution: Mapping[str, Any]) -> tuple[Any, ...]:
    return tuple(
        distribution.get(key)
        for key in ("family", "value", "low", "mode", "high", "mean", "sd")
    )


def _evidence_columns(evidence: Mapping[str, Any] | None) -> tuple[Any, ...]:
    evidence = evidence if isinstance(evidence, Mapping) else {}
    citation = evidence.get("citation")
    citation = citation if isinstance(citation, Mapping) else {}
    return (
        evidence.get("evidence_class"),
        _safe_public_value(citation.get("title")),
        _safe_public_value(citation.get("organization")),
        _safe_public_value(citation.get("url")),
        _safe_public_value(citation.get("stable_reference")),
        citation.get("publication_or_as_of_date"),
        citation.get("accessed_date"),
        _safe_public_value(citation.get("excerpt_or_derivation_note")),
        citation.get("preservation_mode"),
        citation.get("user_supplied_content_sha256"),
        _safe_public_value(citation.get("metadata_only_rationale")),
        evidence.get("explicit_acceptance"),
        _safe_public_value(evidence.get("acceptance_rationale")),
    )


INPUT_COLUMNS = (
    "input_id",
    "input_category",
    "label",
    "ownership",
    "cost_type",
    "unit",
    "distribution_family",
    "fixed_value",
    "low",
    "mode",
    "high",
    "mean",
    "standard_deviation",
    "original_unit",
    "normalized_unit",
    "normalization_method",
    "constant_dollar_cost_year",
    "currency_normalization_method",
    "currency_source_cost_year",
    "currency_target_constant_dollar_cost_year",
    "currency_submitted_distribution_basis",
    "currency_index_identity",
    "currency_index_factor",
    "currency_derivation",
    "coverage_include_ids_json_part_1",
    "coverage_include_ids_json_part_2",
    "coverage_exclude_ids_json_part_1",
    "coverage_exclude_ids_json_part_2",
    "evidence_class",
    "citation_title",
    "citation_organization",
    "citation_url",
    "citation_stable_reference",
    "citation_publication_or_as_of_date",
    "citation_accessed_date",
    "citation_excerpt_or_derivation_note",
    "citation_preservation_mode",
    "citation_user_supplied_content_sha256",
    "citation_metadata_only_rationale",
    "evidence_explicit_acceptance",
    "evidence_acceptance_rationale",
    "analysis_basis",
    "cost_stack_completeness",
    "solectria_quantity",
    "solaredge_quantity",
    "quantity_unit",
    "normalization_derivation",
    "wdc_denominator_method",
    "solectria_wdc_denominator_applied",
    "solectria_wdc_denominator",
    "solectria_wdc_source_field",
    "solaredge_wdc_denominator_applied",
    "solaredge_wdc_denominator",
    "solaredge_wdc_source_field",
    "solectria_multiplier_to_intensity",
    "solaredge_multiplier_to_intensity",
)

APPLIED_INPUT_COLUMNS = INPUT_COLUMNS[:-9] + (
    "capacity_denominator_method",
    "solectria_capacity_denominator_applied",
    "solectria_applied_capacity_w",
    "solectria_applied_capacity_rating_basis",
    "solectria_applied_capacity_source_field",
    "solaredge_capacity_denominator_applied",
    "solaredge_applied_capacity_w",
    "solaredge_applied_capacity_rating_basis",
    "solaredge_applied_capacity_source_field",
    "solectria_multiplier_to_intensity",
    "solaredge_multiplier_to_intensity",
)


def _currency_normalization_columns(
    normalization: Mapping[str, Any] | None,
) -> tuple[Any, ...]:
    normalization = normalization if isinstance(normalization, Mapping) else {}
    return (
        normalization.get("method"),
        normalization.get("source_cost_year"),
        normalization.get("target_constant_dollar_cost_year"),
        normalization.get("submitted_distribution_basis"),
        _safe_public_value(normalization.get("index_identity")),
        normalization.get("index_factor"),
        _safe_public_value(normalization.get("derivation")),
    )


def _coverage_columns(values: Any) -> tuple[str, str]:
    encoded = _canonical_json_text(values if isinstance(values, list) else [])
    chunk_size = 30_000
    if len(encoded) > chunk_size * 2:
        raise TechnoeconomicExportError(
            "Frozen input coverage exceeds the lossless workbook schema"
        )
    return encoded[:chunk_size], encoded[chunk_size:]


def _input_contract_columns(
    request: Mapping[str, Any],
    line: Mapping[str, Any] | None = None,
) -> tuple[Any, ...]:
    line = line if isinstance(line, Mapping) else {}
    return (
        request.get("basis"),
        request.get("cost_stack_completeness"),
        line.get("solectria_quantity"),
        line.get("solaredge_quantity"),
        line.get("quantity_unit"),
        line.get("normalization_derivation"),
    )


def _input_normalization_receipt_columns(
    receipt_line: Mapping[str, Any] | None,
    *,
    applied_capacity_contract: bool = False,
) -> tuple[Any, ...]:
    receipt_line = receipt_line if isinstance(receipt_line, Mapping) else {}
    if applied_capacity_contract:
        denominator = receipt_line.get("capacity_denominator") or {}
        denominator = denominator if isinstance(denominator, Mapping) else {}
        solectria = denominator.get("solectria") or {}
        solaredge = denominator.get("solaredge") or {}
        solectria = solectria if isinstance(solectria, Mapping) else {}
        solaredge = solaredge if isinstance(solaredge, Mapping) else {}
        return (
            denominator.get("method"),
            solectria.get("applied"),
            solectria.get("applied_capacity_w"),
            solectria.get("rating_basis"),
            _safe_public_value(solectria.get("source_field")),
            solaredge.get("applied"),
            solaredge.get("applied_capacity_w"),
            solaredge.get("rating_basis"),
            _safe_public_value(solaredge.get("source_field")),
            receipt_line.get("solectria_multiplier_to_intensity"),
            receipt_line.get("solaredge_multiplier_to_intensity"),
        )
    denominator = receipt_line.get("wdc_denominator") or {}
    denominator = denominator if isinstance(denominator, Mapping) else {}
    solectria = denominator.get("solectria") or {}
    solaredge = denominator.get("solaredge") or {}
    solectria = solectria if isinstance(solectria, Mapping) else {}
    solaredge = solaredge if isinstance(solaredge, Mapping) else {}
    return (
        denominator.get("method"),
        solectria.get("applied"),
        solectria.get("installed_wdc"),
        _safe_public_value(solectria.get("source_field")),
        solaredge.get("applied"),
        solaredge.get("installed_wdc"),
        _safe_public_value(solaredge.get("source_field")),
        receipt_line.get("solectria_multiplier_to_intensity"),
        receipt_line.get("solaredge_multiplier_to_intensity"),
    )


def _input_rows(
    request: Mapping[str, Any],
    submission_provenance: Mapping[str, Any],
    *,
    applied_capacity_contract: bool = False,
) -> Iterator[tuple[Any, ...]]:
    cost_lines = request.get("cost_lines")
    if not isinstance(cost_lines, list):
        raise TechnoeconomicExportError("Frozen request cost lines are invalid")
    normalization_receipt = submission_provenance.get("normalization_receipt") or {}
    receipt_lines = normalization_receipt.get("lines") or []
    if not isinstance(receipt_lines, list):
        raise TechnoeconomicExportError("Frozen normalization receipt is invalid")
    receipts_by_input = {
        str(record.get("input_id")): record
        for record in receipt_lines
        if isinstance(record, Mapping)
    }
    for line in sorted(cost_lines, key=lambda item: str(item.get("input_id"))):
        distribution = line.get("distribution") or {}
        normalization = line.get("currency_year_normalization") or {}
        receipt_line = receipts_by_input.get(str(line.get("input_id")))
        yield (
            line.get("input_id"),
            "cost",
            line.get("label"),
            line.get("ownership"),
            line.get("cost_type"),
            line.get("normalized_unit"),
            *_distribution_columns(distribution),
            line.get("original_unit"),
            line.get("normalized_unit"),
            line.get("normalization_method"),
            line.get("constant_dollar_cost_year"),
            *_currency_normalization_columns(normalization),
            *_coverage_columns(line.get("coverage_include_ids") or []),
            *_coverage_columns(line.get("coverage_exclude_ids") or []),
            *_evidence_columns(line.get("evidence")),
            *_input_contract_columns(request, line),
            *_input_normalization_receipt_columns(
                receipt_line,
                applied_capacity_contract=applied_capacity_contract,
            ),
        )
        if normalization.get("method") == "price_index_adjustment":
            yield (
                f"currency-index::{line.get('input_id')}",
                "currency_year_normalization",
                normalization.get("index_identity"),
                "shared",
                None,
                "dimensionless_multiplier",
                "fixed",
                normalization.get("index_factor"),
                None,
                None,
                None,
                None,
                None,
                line.get("original_unit"),
                line.get("normalized_unit"),
                normalization.get("method"),
                normalization.get("target_constant_dollar_cost_year"),
                *_currency_normalization_columns(normalization),
                *_coverage_columns([]),
                *_coverage_columns([]),
                *_evidence_columns(normalization.get("index_source_evidence")),
                *_input_contract_columns(request, line),
                *_input_normalization_receipt_columns(
                    receipt_line,
                    applied_capacity_contract=applied_capacity_contract,
                ),
            )
    finance = request.get("finance") or {}
    discount = finance.get("real_discount_rate") or {}
    yield (
        "finance.discount-rate",
        "finance",
        "Real discount rate",
        "shared",
        None,
        discount.get("unit"),
        *_distribution_columns(discount.get("distribution") or {}),
        None,
        None,
        None,
        finance.get("constant_dollar_cost_year"),
        *_currency_normalization_columns(None),
        *_coverage_columns([]),
        *_coverage_columns([]),
        *_evidence_columns(discount.get("evidence")),
        *_input_contract_columns(request),
        *_input_normalization_receipt_columns(
            None,
            applied_capacity_contract=applied_capacity_contract,
        ),
    )
    yield (
        "finance.project-life",
        "finance",
        "Project life",
        "shared",
        None,
        "years",
        "fixed",
        finance.get("project_life_years"),
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        finance.get("treatment_key"),
        finance.get("constant_dollar_cost_year"),
        *_currency_normalization_columns(None),
        *_coverage_columns([]),
        *_coverage_columns([]),
        *_evidence_columns(finance.get("project_life_evidence")),
        *_input_contract_columns(request),
        *_input_normalization_receipt_columns(
            None,
            applied_capacity_contract=applied_capacity_contract,
        ),
    )
    degradation = request.get("shared_degradation") or {}
    annual_rate = degradation.get("annual_rate") or {}
    yield (
        "energy.shared-degradation",
        "degradation",
        "Shared annual module degradation",
        "shared",
        None,
        annual_rate.get("unit"),
        *_distribution_columns(annual_rate.get("distribution") or {}),
        None,
        None,
        degradation.get("degradation_model"),
        finance.get("constant_dollar_cost_year"),
        *_currency_normalization_columns(None),
        *_coverage_columns([]),
        *_coverage_columns([]),
        *_evidence_columns(annual_rate.get("evidence")),
        *_input_contract_columns(request),
        *_input_normalization_receipt_columns(
            None,
            applied_capacity_contract=applied_capacity_contract,
        ),
    )
    commercial_scaling = request.get("commercial_scaling")
    if isinstance(commercial_scaling, Mapping):
        marginal_unit = commercial_scaling.get("marginal_cost_unit")
        yield (
            technoeconomic_kernel.COMMERCIAL_MARGINAL_COST_DIFFERENCE_INPUT_ID,
            "commercial_marginal_cost",
            "Commercial marginal cost difference, SolarEdge minus Solectria",
            "se_minus_sol",
            commercial_scaling.get("marginal_cost_timing"),
            marginal_unit,
            *_distribution_columns(
                commercial_scaling.get("marginal_cost_difference") or {}
            ),
            marginal_unit,
            marginal_unit,
            commercial_scaling.get("transfer_method"),
            finance.get("constant_dollar_cost_year"),
            *_currency_normalization_columns(None),
            *_coverage_columns([]),
            *_coverage_columns([]),
            *_evidence_columns(commercial_scaling.get("evidence")),
            *_input_contract_columns(request),
            *_input_normalization_receipt_columns(
                None,
                applied_capacity_contract=applied_capacity_contract,
            ),
        )
    standalone = request.get("standalone_commercial")
    if isinstance(standalone, Mapping):
        commercial_lines = standalone.get("cost_lines") or []
        if not isinstance(commercial_lines, list):
            raise TechnoeconomicExportError(
                "Frozen standalone commercial cost lines are invalid"
            )
        for line in sorted(
            commercial_lines,
            key=lambda item: str(item.get("input_id")),
        ):
            if not isinstance(line, Mapping):
                raise TechnoeconomicExportError(
                    "A frozen standalone commercial cost line is invalid"
                )
            yield (
                line.get("input_id"),
                "standalone_commercial_cost",
                line.get("label"),
                "solaredge_only",
                line.get("cost_category"),
                line.get("unit"),
                *_distribution_columns(line.get("distribution") or {}),
                line.get("unit"),
                line.get("unit"),
                "direct_target_capacity_scaling",
                line.get("constant_dollar_cost_year"),
                *_currency_normalization_columns(None),
                *_coverage_columns(line.get("coverage_ids") or []),
                *_coverage_columns([]),
                *_evidence_columns(line.get("evidence")),
                *_input_contract_columns(request),
                *_input_normalization_receipt_columns(
                    None,
                    applied_capacity_contract=applied_capacity_contract,
                ),
            )
    paired = request.get("paired_commercial")
    if isinstance(paired, Mapping):
        paired_systems = paired.get("systems") or []
        if not isinstance(paired_systems, list):
            raise TechnoeconomicExportError(
                "Frozen paired commercial systems are invalid"
            )
        for system in sorted(
            paired_systems,
            key=lambda item: str(item.get("technology")),
        ):
            if not isinstance(system, Mapping):
                raise TechnoeconomicExportError(
                    "A frozen paired commercial system is invalid"
                )
            technology = str(system.get("technology"))
            commercial_lines = system.get("cost_lines") or []
            if not isinstance(commercial_lines, list):
                raise TechnoeconomicExportError(
                    f"Frozen paired commercial {technology} cost lines are invalid"
                )
            for line in sorted(
                commercial_lines,
                key=lambda item: str(item.get("input_id")),
            ):
                if not isinstance(line, Mapping):
                    raise TechnoeconomicExportError(
                        "A frozen paired commercial cost line is invalid"
                    )
                yield (
                    line.get("input_id"),
                    "paired_commercial_cost",
                    line.get("label"),
                    f"{technology}_only",
                    line.get("cost_category"),
                    line.get("unit"),
                    *_distribution_columns(line.get("distribution") or {}),
                    line.get("unit"),
                    line.get("unit"),
                    "direct_target_capacity_scaling",
                    line.get("constant_dollar_cost_year"),
                    *_currency_normalization_columns(None),
                    *_coverage_columns(line.get("coverage_ids") or []),
                    *_coverage_columns([]),
                    *_evidence_columns(line.get("evidence")),
                    *_input_contract_columns(request),
                    *_input_normalization_receipt_columns(
                        None,
                        applied_capacity_contract=applied_capacity_contract,
                    ),
                )
    transfer = request.get("commercial_transfer")
    reference_design = request.get("commercial_reference_design")
    if isinstance(reference_design, Mapping):
        yield (
            "commercial.reference-design",
            "commercial_reference_design",
            reference_design.get("design_id"),
            "shared",
            None,
            "Wdc",
            "fixed",
            reference_design.get("reference_wdc"),
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "module_count_times_module_stc_wdc",
            reference_design.get("constant_dollar_cost_year"),
            *_currency_normalization_columns(None),
            *_coverage_columns([]),
            *_coverage_columns([]),
            *_evidence_columns(reference_design.get("evidence")),
            *_input_contract_columns(request),
            *_input_normalization_receipt_columns(
                None,
                applied_capacity_contract=applied_capacity_contract,
            ),
        )
    if isinstance(transfer, Mapping):
        for input_id, label, field in (
            ("transfer.baseline", "Commercial baseline transfer factor", "baseline_factor"),
            ("transfer.incremental", "Commercial incremental transfer factor", "incremental_factor"),
        ):
            record = transfer.get(field) or {}
            yield (
                input_id,
                "commercial_transfer",
                label,
                "shared",
                None,
                record.get("unit"),
                *_distribution_columns(record.get("distribution") or {}),
                None,
                None,
                "commercial_transfer_attestation",
                finance.get("constant_dollar_cost_year"),
                *_currency_normalization_columns(None),
                *_coverage_columns([]),
                *_coverage_columns([]),
                *_evidence_columns(record.get("evidence")),
                *_input_contract_columns(request),
                *_input_normalization_receipt_columns(
                    None,
                    applied_capacity_contract=applied_capacity_contract,
                ),
            )
        for mechanism in sorted(
            transfer.get("mechanisms") or [],
            key=lambda item: str(item.get("mechanism")),
        ):
            yield (
                f"transfer-mechanism::{mechanism.get('mechanism')}",
                "commercial_transfer_mechanism",
                mechanism.get("mechanism"),
                "shared",
                None,
                "categorical_attestation",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                mechanism.get("status"),
                finance.get("constant_dollar_cost_year"),
                *_currency_normalization_columns(None),
                *_coverage_columns([]),
                *_coverage_columns([]),
                *_evidence_columns(mechanism.get("evidence")),
                *_input_contract_columns(request),
                *_input_normalization_receipt_columns(
                    None,
                    applied_capacity_contract=applied_capacity_contract,
                ),
            )


ENERGY_COLUMNS = (
    "eligibility_status",
    "year",
    "period_start",
    "period_end",
    "solectria_predicted_kwh_ac",
    "solaredge_predicted_kwh_ac",
    "exclusion_reason",
    "source_record_sha256",
    "source_record_provenance_path",
    "source_annual_job_id",
    "source_snapshot_sha256",
    "source_artifact_sha256",
    "source_artifact_bytes",
    "source_artifact_media_type",
)


def _first_present(record: Mapping[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in record and record[key] is not None:
            return record[key]
    return None


def _energy_rows(
    source_snapshot: Mapping[str, Any],
    source_snapshot_sha256: str,
) -> Iterator[tuple[Any, ...]]:
    artifact = source_snapshot.get("midc_source_artifact") or {}
    source_id = source_snapshot.get("source_annual_job_id")
    groups = (
        (
            "eligible",
            "eligible_paired_energy_rows",
            source_snapshot.get("eligible_paired_energy_rows") or [],
        ),
        (
            "excluded",
            "excluded_annual_energy_rows",
            source_snapshot.get("excluded_annual_energy_rows") or [],
        ),
    )
    for status_name, source_field, records in groups:
        if not isinstance(records, list):
            raise TechnoeconomicExportError("Frozen energy snapshot rows are invalid")

        def source_row(record: Mapping[str, Any]) -> Mapping[str, Any]:
            nested = record.get("row") if status_name == "excluded" else None
            return nested if isinstance(nested, Mapping) else record

        for source_index, record in sorted(
            enumerate(records),
            key=lambda pair: (
                int(source_row(pair[1]).get("year") or 0),
                _canonical_json_text(_safe_public_value(pair[1])),
            ),
        ):
            row = source_row(record)
            reasons = record.get("reasons") if status_name == "excluded" else None
            if isinstance(reasons, list):
                exclusion_reason = _canonical_json_text(reasons)
            else:
                exclusion_reason = row.get("reason") or row.get("exclusion_reason")
            yield (
                status_name,
                row.get("year"),
                row.get("period_start"),
                row.get("period_end"),
                _first_present(row, "sol_predicted_kwh", "sol_predicted_kwh_ac"),
                _first_present(row, "se_predicted_kwh", "se_predicted_kwh_ac"),
                exclusion_reason,
                _canonical_json_sha256(_safe_public_value(record)),
                f"source_snapshot.{source_field}[{source_index}]",
                source_id,
                source_snapshot_sha256,
                artifact.get("sha256"),
                artifact.get("byte_count"),
                artifact.get("media_type"),
            )


CAPACITY_COLUMNS = (
    "record_type",
    "system",
    "analysis_basis",
    "capacity_basis",
    "module_model",
    "module_stc_wdc",
    "module_count",
    "installed_wdc",
    "rating_basis",
    "strings",
    "bays_per_string",
    "modules_per_bay",
    "calibration_physics_version",
    "calibration_physics_fingerprint",
    "capacity_manifest_sha256",
    "capacity_manifest_source_json",
    "commercial_reference_design_id",
    "commercial_reference_wdc",
    "commercial_reference_module_model",
    "commercial_reference_module_stc_wdc",
    "commercial_reference_module_count",
    "commercial_reference_constant_dollar_cost_year",
    "commercial_reference_design_sha256",
    "commercial_transfer_status",
    "source_snapshot_sha256",
)

APPLIED_CAPACITY_COLUMNS = CAPACITY_COLUMNS[:8] + (
    "applied_capacity_w",
    "applied_capacity_rating_basis",
    "applied_capacity_selection_method",
    "applied_capacity_source_field",
) + CAPACITY_COLUMNS[8:]


def _capacity_rows(
    source_snapshot: Mapping[str, Any],
    submission_provenance: Mapping[str, Any],
    routine_result: Mapping[str, Any],
    *,
    applied_capacity_contract: bool = False,
) -> Iterator[tuple[Any, ...]]:
    manifest = source_snapshot.get("capacity_manifest") or {}
    systems = manifest.get("systems") or {}
    reference_design = submission_provenance.get("commercial_reference_design")
    reference_design = (
        reference_design if isinstance(reference_design, Mapping) else {}
    )
    applied_authority = (
        _applied_capacity_authority(submission_provenance)
        if applied_capacity_contract
        else {}
    )
    for system in ("solectria", "solaredge"):
        record = systems.get(system) or {}
        base = (
            "frozen_annual_capacity",
            system,
            routine_result.get("analysis_basis"),
            routine_result.get("capacity_basis"),
            record.get("module_model"),
            record.get("module_stc_wdc"),
            record.get("module_count"),
            record.get("installed_wdc"),
        )
        if applied_capacity_contract:
            applied = applied_authority[system]
            base += (
                applied.get("applied_capacity_w"),
                applied.get("rating_basis"),
                applied.get("selection_method"),
                _safe_public_value(applied.get("source_field")),
            )
        yield base + (
            record.get("rating_basis") or manifest.get("rating_basis"),
            record.get("strings"),
            record.get("bays_per_string"),
            record.get("modules_per_bay"),
            record.get("calibration_physics_version"),
            record.get("calibration_physics_fingerprint"),
            manifest.get("capacity_manifest_sha256"),
            _canonical_json_text(
                _safe_public_value(source_snapshot.get("capacity_manifest_source"))
            ),
            reference_design.get("design_id"),
            reference_design.get("reference_wdc"),
            reference_design.get("module_model"),
            reference_design.get("module_stc_wdc"),
            reference_design.get("module_count"),
            reference_design.get("constant_dollar_cost_year"),
            reference_design.get("design_sha256"),
            submission_provenance.get("commercial_transfer_status"),
            submission_provenance.get("source_snapshot_sha256"),
        )


COMMON_COST_COLUMNS = (
    "input_id",
    "label",
    "cost_type",
    "comparison_treatment",
    "reasons_json",
    "solectria_multiplier_to_intensity",
    "solaredge_multiplier_to_intensity",
    "solectria_treatment_key",
    "solaredge_treatment_key",
    "solectria_contribution_min",
    "solectria_contribution_max",
    "solaredge_contribution_min",
    "solaredge_contribution_max",
    "contribution_units",
    "delta_contribution_min_se_minus_sol",
    "delta_contribution_max_se_minus_sol",
    "delta_contribution_se_minus_sol_exactly_zero",
)


def _common_cost_rows(records: Sequence[Mapping[str, Any]]) -> Iterator[tuple[Any, ...]]:
    for record in sorted(records, key=lambda row: str(row.get("input_id"))):
        yield (
            record.get("input_id"),
            record.get("label"),
            record.get("cost_type"),
            record.get("comparison_treatment"),
            _canonical_json_text(record.get("reasons") or []),
            record.get("solectria_multiplier_to_intensity"),
            record.get("solaredge_multiplier_to_intensity"),
            record.get("solectria_treatment_key"),
            record.get("solaredge_treatment_key"),
            record.get("solectria_contribution_min"),
            record.get("solectria_contribution_max"),
            record.get("solaredge_contribution_min"),
            record.get("solaredge_contribution_max"),
            record.get("contribution_units"),
            record.get("delta_contribution_min_se_minus_sol"),
            record.get("delta_contribution_max_se_minus_sol"),
            record.get("delta_contribution_se_minus_sol_exactly_zero"),
        )


TRANSFER_COLUMNS = (
    "record_type",
    "status",
    "energy_available",
    "mechanism",
    "mechanism_status",
    "rationale_sha256",
    "mechanism_rationale",
    "mechanism_evidence_sha256",
    "mechanism_evidence_field_path",
    "mechanism_evidence_value_json",
    "explicit_attestation",
    "attested_by",
    "attested_at",
    "attestation_rationale_sha256",
    "baseline_factor_input_id",
    "incremental_factor_input_id",
    "all_mechanisms_resolved",
    "commercial_reference_design_sha256",
    "commercial_reference_design_field_path",
    "commercial_reference_design_value_json",
    "transfer_receipt_sha256",
)


def _transfer_rows(
    request_payload: Mapping[str, Any],
    submission_provenance: Mapping[str, Any],
) -> Iterator[tuple[Any, ...]]:
    receipt = submission_provenance.get("commercial_transfer_receipt") or {}
    common = {
        "status": receipt.get("status"),
        "energy_available": receipt.get("energy_available"),
        "explicit_attestation": receipt.get("explicit_attestation"),
        "attested_by": receipt.get("attested_by"),
        "attested_at": receipt.get("attested_at"),
        "attestation_rationale_sha256": receipt.get(
            "attestation_rationale_sha256"
        ),
        "baseline_factor_input_id": receipt.get("baseline_factor_input_id"),
        "incremental_factor_input_id": receipt.get("incremental_factor_input_id"),
        "all_mechanisms_resolved": receipt.get("all_mechanisms_resolved"),
        "commercial_reference_design_sha256": submission_provenance.get(
            "commercial_reference_design_sha256"
        ),
        "transfer_receipt_sha256": submission_provenance.get(
            "commercial_transfer_receipt_sha256"
        ),
    }

    def values(**updates: Any) -> tuple[Any, ...]:
        record = {column: None for column in TRANSFER_COLUMNS}
        record.update(common)
        record.update(updates)
        return tuple(record[column] for column in TRANSFER_COLUMNS)

    yield values(record_type="transfer_summary")
    transfer_request = request_payload.get("commercial_transfer") or {}
    request_mechanisms = transfer_request.get("mechanisms") or []
    request_by_mechanism = {
        str(record.get("mechanism")): record
        for record in request_mechanisms
        if isinstance(record, Mapping)
    }
    mechanisms = receipt.get("mechanisms") or []
    for record in sorted(mechanisms, key=lambda row: str(row.get("mechanism"))):
        mechanism = str(record.get("mechanism"))
        request_record = request_by_mechanism.get(mechanism) or {}
        evidence = request_record.get("evidence")
        evidence_sha256 = (
            _canonical_json_sha256(evidence) if isinstance(evidence, Mapping) else None
        )
        yield values(
            record_type="mechanism",
            mechanism=mechanism,
            mechanism_status=record.get("status"),
            rationale_sha256=record.get("rationale_sha256"),
            mechanism_rationale=_safe_public_value(request_record.get("rationale")),
            mechanism_evidence_sha256=evidence_sha256,
        )
        if isinstance(evidence, Mapping):
            for field_path, encoded in _flatten_leaves("", evidence):
                yield values(
                    record_type="mechanism_evidence_field",
                    mechanism=mechanism,
                    mechanism_status=record.get("status"),
                    rationale_sha256=record.get("rationale_sha256"),
                    mechanism_rationale=_safe_public_value(
                        request_record.get("rationale")
                    ),
                    mechanism_evidence_sha256=evidence_sha256,
                    mechanism_evidence_field_path=field_path,
                    mechanism_evidence_value_json=encoded,
                )
    reference_design = request_payload.get("commercial_reference_design")
    if isinstance(reference_design, Mapping):
        for field_path, encoded in _flatten_leaves("", reference_design):
            yield values(
                record_type="commercial_reference_design_field",
                commercial_reference_design_field_path=field_path,
                commercial_reference_design_value_json=encoded,
            )


CDF_COLUMNS = (
    "metric_id",
    "status",
    "reason",
    "population_count",
    "point_index",
    "value",
    "cumulative_count",
    "cumulative_probability",
    "p5",
    "p50",
    "p95",
)

STANDALONE_COMMERCIAL_CDF_COLUMNS = (
    "metric_id",
    "status",
    "reason",
    "population_count",
    "point_index",
    "value_constant_usd_per_kwh_ac",
    "cumulative_count",
    "cumulative_probability",
    "p10",
    "p50",
    "p90",
)

PAIRED_COMMERCIAL_CDF_COLUMNS = (
    "technology",
    *STANDALONE_COMMERCIAL_CDF_COLUMNS,
)

LIFECYCLE_CDF_COLUMNS = (
    "metric_id",
    "status",
    "reason",
    "population_count",
    "point_index",
    "value",
    "cumulative_count",
    "cumulative_probability",
    "p10",
    "p50",
    "p90",
)


def _metric_summaries(metadata: Mapping[str, Any]) -> Mapping[str, Any]:
    summaries = metadata.get("summaries")
    if not isinstance(summaries, Mapping):
        raise TechnoeconomicExportError("Sealed metric summaries are invalid")
    return summaries


def _cdf_rows(metadata: Mapping[str, Any]) -> Iterator[tuple[Any, ...]]:
    for metric_id, summary in sorted(_metric_summaries(metadata).items()):
        if not isinstance(summary, Mapping) or "percentiles" not in summary:
            continue
        percentiles = summary.get("percentiles") or {}
        cdf = summary.get("cdf")
        if not isinstance(cdf, Mapping):
            yield (
                metric_id,
                summary.get("status"),
                summary.get("reason"),
                summary.get("count", 0),
                None,
                None,
                None,
                None,
                percentiles.get("p5"),
                percentiles.get("p50"),
                percentiles.get("p95"),
            )
            continue
        values = cdf.get("values") or []
        counts = cdf.get("cumulative_count") or []
        probabilities = cdf.get("cumulative_probability") or []
        if not (len(values) == len(counts) == len(probabilities)):
            raise TechnoeconomicExportError("Sealed CDF arrays are inconsistent")
        for index, (value, count, probability) in enumerate(
            zip(values, counts, probabilities),
            start=1,
        ):
            yield (
                metric_id,
                summary.get("status"),
                summary.get("reason"),
                cdf.get("population_count"),
                index,
                value,
                count,
                probability,
                percentiles.get("p5"),
                percentiles.get("p50"),
                percentiles.get("p95"),
            )


def _lifecycle_cdf_rows(metadata: Mapping[str, Any]) -> Iterator[tuple[Any, ...]]:
    for metric_id, summary in sorted(_metric_summaries(metadata).items()):
        if not isinstance(summary, Mapping) or "percentiles" not in summary:
            continue
        percentiles = summary.get("percentiles") or {}
        cdf = summary.get("cdf")
        if not isinstance(cdf, Mapping):
            yield (
                metric_id,
                summary.get("status"),
                summary.get("reason"),
                summary.get("count", 0),
                None,
                None,
                None,
                None,
                percentiles.get("p10"),
                percentiles.get("p50"),
                percentiles.get("p90"),
            )
            continue
        values = cdf.get("values") or []
        counts = cdf.get("cumulative_count") or []
        probabilities = cdf.get("cumulative_probability") or []
        if not (len(values) == len(counts) == len(probabilities)):
            raise TechnoeconomicExportError("Sealed v6 CDF arrays are inconsistent")
        for index, (value, count, probability) in enumerate(
            zip(values, counts, probabilities), start=1
        ):
            yield (
                metric_id,
                summary.get("status"),
                summary.get("reason"),
                cdf.get("population_count"),
                index,
                value,
                count,
                probability,
                percentiles.get("p10"),
                percentiles.get("p50"),
                percentiles.get("p90"),
            )


def _standalone_commercial_cdf_rows(
    metadata: Mapping[str, Any],
) -> Iterator[tuple[Any, ...]]:
    metric_id = technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LCOE
    summary = _metric_summaries(metadata).get(metric_id)
    if not isinstance(summary, Mapping):
        raise TechnoeconomicExportError(
            "Standalone commercial LCOE summary is missing"
        )
    percentiles = summary.get("percentiles") or {}
    cdf = summary.get("cdf")
    if not isinstance(cdf, Mapping):
        yield (
            metric_id,
            summary.get("status"),
            summary.get("reason"),
            summary.get("count", 0),
            None,
            None,
            None,
            None,
            percentiles.get("p10"),
            percentiles.get("p50"),
            percentiles.get("p90"),
        )
        return
    values = cdf.get("values") or []
    counts = cdf.get("cumulative_count") or []
    probabilities = cdf.get("cumulative_probability") or []
    if not (len(values) == len(counts) == len(probabilities)):
        raise TechnoeconomicExportError(
            "Sealed standalone commercial CDF arrays are inconsistent"
        )
    for index, (value, count, probability) in enumerate(
        zip(values, counts, probabilities),
        start=1,
    ):
        yield (
            metric_id,
            summary.get("status"),
            summary.get("reason"),
            cdf.get("population_count"),
            index,
            value,
            count,
            probability,
            percentiles.get("p10"),
            percentiles.get("p50"),
            percentiles.get("p90"),
        )


def _paired_commercial_cdf_rows(
    metadata: Mapping[str, Any],
) -> Iterator[tuple[Any, ...]]:
    metric_contracts = (
        (
            "solectria",
            technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_LCOE,
        ),
        ("solaredge", technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LCOE),
    )
    summaries = _metric_summaries(metadata)
    for technology, metric_id in metric_contracts:
        summary = summaries.get(metric_id)
        if not isinstance(summary, Mapping):
            raise TechnoeconomicExportError(
                f"Paired commercial {technology} LCOE summary is missing"
            )
        percentiles = summary.get("percentiles") or {}
        cdf = summary.get("cdf")
        if not isinstance(cdf, Mapping):
            yield (
                technology,
                metric_id,
                summary.get("status"),
                summary.get("reason"),
                summary.get("count", 0),
                None,
                None,
                None,
                None,
                percentiles.get("p10"),
                percentiles.get("p50"),
                percentiles.get("p90"),
            )
            continue
        values = cdf.get("values") or []
        counts = cdf.get("cumulative_count") or []
        probabilities = cdf.get("cumulative_probability") or []
        if not (len(values) == len(counts) == len(probabilities)):
            raise TechnoeconomicExportError(
                f"Sealed paired commercial {technology} CDF arrays are inconsistent"
            )
        for index, (value, count, probability) in enumerate(
            zip(values, counts, probabilities),
            start=1,
        ):
            yield (
                technology,
                metric_id,
                summary.get("status"),
                summary.get("reason"),
                cdf.get("population_count"),
                index,
                value,
                count,
                probability,
                percentiles.get("p10"),
                percentiles.get("p50"),
                percentiles.get("p90"),
            )


def _per_year_base_columns(metadata: Mapping[str, Any]) -> tuple[str, ...]:
    rows = metadata.get("per_weather_year") or []
    applied_capacity_contract = any(
        isinstance(row, Mapping) and "solectria_applied_w" in row
        for row in rows
    )
    if applied_capacity_contract:
        return (
            "year",
            "source_sol_predicted_kwh_ac",
            "source_se_predicted_kwh_ac",
            "solectria_installed_wdc",
            "solaredge_installed_wdc",
            "solectria_applied_w",
            "solaredge_applied_w",
            "source_sol_specific_kwh_ac_per_applied_w_year",
            "source_se_specific_kwh_ac_per_applied_w_year",
            "source_delta_specific_se_minus_sol_kwh_ac_per_applied_w_year",
            "realization_count",
            "realization_share",
            "reason",
            "energy_class_counts_json",
            "energy_class_probabilities_json",
        )
    return (
        "year",
        "source_sol_predicted_kwh_ac",
        "source_se_predicted_kwh_ac",
        "solectria_installed_wdc",
        "solaredge_installed_wdc",
        "source_sol_specific_kwh_ac_per_wdc_year",
        "source_se_specific_kwh_ac_per_wdc_year",
        "source_delta_specific_se_minus_sol_kwh_ac_per_wdc_year",
        "realization_count",
        "realization_share",
        "reason",
        "energy_class_counts_json",
        "energy_class_probabilities_json",
    )


def _per_year_columns(metadata: Mapping[str, Any]) -> tuple[str, ...]:
    base = _per_year_base_columns(metadata)
    rows = metadata.get("per_weather_year") or []
    metric_ids = sorted(
        {
            str(metric_id)
            for row in rows
            if isinstance(row, Mapping)
            for metric_id in (row.get("metrics") or {})
        }
    )
    suffixes = ("status", "reason", "count", "p5", "p50", "p95")
    return base + tuple(
        f"{metric_id}::{suffix}" for metric_id in metric_ids for suffix in suffixes
    )


def _per_year_rows(metadata: Mapping[str, Any]) -> Iterator[tuple[Any, ...]]:
    base_columns = _per_year_base_columns(metadata)
    columns = _per_year_columns(metadata)
    metric_columns = columns[len(base_columns):]
    applied_capacity_contract = "solectria_applied_w" in base_columns
    rows = metadata.get("per_weather_year") or []
    for record in sorted(rows, key=lambda row: int(row.get("year"))):
        metrics = record.get("metrics") or {}
        flattened: dict[str, Any] = {}
        for metric_id, summary in metrics.items():
            percentiles = summary.get("percentiles") or {}
            flattened.update(
                {
                    f"{metric_id}::status": summary.get("status"),
                    f"{metric_id}::reason": summary.get("reason"),
                    f"{metric_id}::count": summary.get("count"),
                    f"{metric_id}::p5": percentiles.get("p5"),
                    f"{metric_id}::p50": percentiles.get("p50"),
                    f"{metric_id}::p95": percentiles.get("p95"),
                }
            )
        base = (
            record.get("year"),
            record.get("source_sol_predicted_kwh_ac"),
            record.get("source_se_predicted_kwh_ac"),
            record.get("solectria_installed_wdc"),
            record.get("solaredge_installed_wdc"),
        )
        if applied_capacity_contract:
            base += (
                record.get("solectria_applied_w"),
                record.get("solaredge_applied_w"),
                record.get("source_sol_specific_kwh_ac_per_applied_w_year"),
                record.get("source_se_specific_kwh_ac_per_applied_w_year"),
                record.get(
                    "source_delta_specific_se_minus_sol_kwh_ac_per_applied_w_year"
                ),
            )
        else:
            base += (
                record.get("source_sol_specific_kwh_ac_per_wdc_year"),
                record.get("source_se_specific_kwh_ac_per_wdc_year"),
                record.get(
                    "source_delta_specific_se_minus_sol_kwh_ac_per_wdc_year"
                ),
            )
        yield base + (
            record.get("realization_count"),
            record.get("realization_share"),
            record.get("reason"),
            _canonical_json_text(record.get("energy_class_counts") or {}),
            _canonical_json_text(record.get("energy_class_probabilities") or {}),
            *(flattened.get(name) for name in metric_columns),
        )


STANDALONE_COMMERCIAL_PER_YEAR_BASE_COLUMNS = (
    "year",
    "source_solaredge_predicted_kwh_ac",
    "source_solaredge_installed_wdc",
    "source_solaredge_applied_capacity_w",
    "source_solaredge_rating_basis",
    "source_solaredge_specific_kwh_ac_per_applied_w_year",
    "commercial_target_capacity_w",
    "commercial_target_rating_basis",
    "commercial_capacity_scale_factor_target_w_per_source_w",
    "commercial_transfer_method",
    "commercial_scaled_target_year1_energy_kwh_ac",
    "realization_count",
    "realization_share",
    "reason",
)


def _standalone_commercial_per_year_columns(
    metadata: Mapping[str, Any],
) -> tuple[str, ...]:
    rows = metadata.get("per_weather_year") or []
    metric_ids = sorted(
        {
            str(metric_id)
            for row in rows
            if isinstance(row, Mapping)
            for metric_id in (row.get("metrics") or {})
        }
    )
    suffixes = ("status", "reason", "count", "p10", "p50", "p90")
    return STANDALONE_COMMERCIAL_PER_YEAR_BASE_COLUMNS + tuple(
        f"{metric_id}::{suffix}" for metric_id in metric_ids for suffix in suffixes
    )


def _standalone_commercial_per_year_rows(
    metadata: Mapping[str, Any],
) -> Iterator[tuple[Any, ...]]:
    columns = _standalone_commercial_per_year_columns(metadata)
    metric_columns = columns[len(STANDALONE_COMMERCIAL_PER_YEAR_BASE_COLUMNS) :]
    rows = metadata.get("per_weather_year") or []
    for record in sorted(rows, key=lambda row: int(row.get("year"))):
        metrics = record.get("metrics") or {}
        flattened: dict[str, Any] = {}
        for metric_id, summary in metrics.items():
            percentiles = summary.get("percentiles") or {}
            flattened.update(
                {
                    f"{metric_id}::status": summary.get("status"),
                    f"{metric_id}::reason": summary.get("reason"),
                    f"{metric_id}::count": summary.get("count"),
                    f"{metric_id}::p10": percentiles.get("p10"),
                    f"{metric_id}::p50": percentiles.get("p50"),
                    f"{metric_id}::p90": percentiles.get("p90"),
                }
            )
        yield (
            record.get("year"),
            record.get("source_se_predicted_kwh_ac"),
            record.get("solaredge_installed_wdc"),
            record.get("solaredge_applied_w"),
            record.get("solaredge_source_rating_basis"),
            record.get("source_se_specific_kwh_ac_per_applied_w_year"),
            record.get("commercial_target_capacity_w"),
            record.get("commercial_target_rating_basis"),
            record.get("commercial_capacity_scale_factor_target_w_per_source_w"),
            record.get("commercial_transfer_method"),
            record.get("commercial_source_year1_energy_solaredge_kwh_ac"),
            record.get("realization_count"),
            record.get("realization_share"),
            record.get("reason"),
            *(flattened.get(name) for name in metric_columns),
        )


PAIRED_COMMERCIAL_PER_YEAR_BASE_COLUMNS = (
    "year",
    "commercial_target_capacity_w",
    "commercial_target_rating_basis",
    "commercial_transfer_method",
    "solectria_source_predicted_kwh_ac",
    "solectria_installed_wdc",
    "solectria_source_applied_capacity_w",
    "solectria_source_rating_basis",
    "solectria_source_specific_kwh_ac_per_applied_w_year",
    "solectria_capacity_scale_factor_target_w_per_source_w",
    "solectria_target_year1_energy_kwh_ac",
    "solaredge_source_predicted_kwh_ac",
    "solaredge_installed_wdc",
    "solaredge_source_applied_capacity_w",
    "solaredge_source_rating_basis",
    "solaredge_source_specific_kwh_ac_per_applied_w_year",
    "solaredge_capacity_scale_factor_target_w_per_source_w",
    "solaredge_target_year1_energy_kwh_ac",
    "realization_count",
    "realization_share",
    "reason",
)


def _paired_commercial_per_year_columns(
    metadata: Mapping[str, Any],
) -> tuple[str, ...]:
    rows = metadata.get("per_weather_year") or []
    metric_ids = sorted(
        {
            str(metric_id)
            for row in rows
            if isinstance(row, Mapping)
            for metric_id in (row.get("metrics") or {})
        }
    )
    suffixes = ("status", "reason", "count", "p10", "p50", "p90")
    return PAIRED_COMMERCIAL_PER_YEAR_BASE_COLUMNS + tuple(
        f"{metric_id}::{suffix}" for metric_id in metric_ids for suffix in suffixes
    )


def _paired_commercial_per_year_rows(
    metadata: Mapping[str, Any],
) -> Iterator[tuple[Any, ...]]:
    columns = _paired_commercial_per_year_columns(metadata)
    metric_columns = columns[len(PAIRED_COMMERCIAL_PER_YEAR_BASE_COLUMNS) :]
    rows = metadata.get("per_weather_year") or []
    for record in sorted(rows, key=lambda row: int(row.get("year"))):
        systems = record.get("systems") or {}
        if not isinstance(systems, Mapping) or set(systems) != {
            "solectria",
            "solaredge",
        }:
            raise TechnoeconomicExportError(
                "Sealed paired-commercial per-year systems are invalid"
            )
        metrics = record.get("metrics") or {}
        flattened: dict[str, Any] = {}
        for metric_id, summary in metrics.items():
            percentiles = summary.get("percentiles") or {}
            flattened.update(
                {
                    f"{metric_id}::status": summary.get("status"),
                    f"{metric_id}::reason": summary.get("reason"),
                    f"{metric_id}::count": summary.get("count"),
                    f"{metric_id}::p10": percentiles.get("p10"),
                    f"{metric_id}::p50": percentiles.get("p50"),
                    f"{metric_id}::p90": percentiles.get("p90"),
                }
            )
        system_values: list[Any] = []
        for technology in ("solectria", "solaredge"):
            system = systems.get(technology) or {}
            if not isinstance(system, Mapping):
                raise TechnoeconomicExportError(
                    "Sealed paired-commercial per-year system is invalid"
                )
            system_values.extend(
                (
                    system.get("source_predicted_kwh_ac"),
                    system.get("installed_wdc"),
                    system.get("source_applied_capacity_w"),
                    system.get("source_rating_basis"),
                    system.get("source_specific_kwh_ac_per_applied_w_year"),
                    system.get("capacity_scale_factor_target_w_per_source_w"),
                    system.get("target_year1_energy_kwh_ac"),
                )
            )
        yield (
            record.get("year"),
            record.get("commercial_target_capacity_w"),
            record.get("commercial_target_rating_basis"),
            record.get("commercial_transfer_method"),
            *system_values,
            record.get("realization_count"),
            record.get("realization_share"),
            record.get("reason"),
            *(flattened.get(name) for name in metric_columns),
        )


SENSITIVITY_COLUMNS = (
    "response_id",
    "record_type",
    "status",
    "reason",
    "sample_count",
    "minimum_sample_count",
    "candidate_predictor_count",
    "entered_predictor_count",
    "entry_order",
    "predictor_id",
    "incremental_r_squared",
    "cumulative_r_squared",
    "standardized_beta",
    "sign",
    "exclusion_reason",
    "exclusion_detail_json",
    "final_r_squared",
    "warnings_json",
)


def _sensitivity_rows(metadata: Mapping[str, Any]) -> Iterator[tuple[Any, ...]]:
    models = metadata.get("sensitivity") or {}
    for response_id, model in sorted(models.items()):
        if not isinstance(model, Mapping):
            yield (
                response_id,
                "diagnostic",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                _canonical_json_text({"value": model}),
                None,
                "[]",
            )
            continue
        if "steps" not in model and "status" not in model:
            yield (
                response_id,
                "diagnostic",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                _canonical_json_text(model),
                None,
                "[]",
            )
            continue
        common = (
            response_id,
            model.get("status"),
            model.get("reason"),
            model.get("sample_count"),
            model.get("minimum_sample_count"),
            model.get("candidate_predictor_count"),
            model.get("entered_predictor_count"),
            model.get("final_r_squared"),
            _canonical_json_text(model.get("warnings") or []),
        )
        yield (
            common[0],
            "model_summary",
            *common[1:7],
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "{}",
            common[7],
            common[8],
        )
        for step in model.get("steps") or []:
            yield (
                common[0],
                "entered_step",
                *common[1:7],
                step.get("entry_order"),
                step.get("predictor_id"),
                step.get("incremental_r_squared"),
                step.get("cumulative_r_squared"),
                step.get("standardized_beta"),
                step.get("sign"),
                None,
                "{}",
                common[7],
                common[8],
            )
        exclusions = model.get("exclusions") or {}
        for predictor_id, detail in sorted(exclusions.items()):
            if isinstance(detail, Mapping):
                reason = detail.get("reason")
                detail_json = _canonical_json_text(detail)
            else:
                reason = detail
                detail_json = _canonical_json_text({"reason": detail})
            yield (
                common[0],
                "exclusion",
                *common[1:7],
                None,
                predictor_id,
                None,
                None,
                None,
                None,
                reason,
                detail_json,
                common[7],
                common[8],
            )


CONVERGENCE_COLUMNS = (
    "record_type",
    "checkpoint_index",
    "realization_count",
    "metric_id",
    "category_or_year",
    "population_count",
    "p5",
    "p50",
    "p95",
    "p5_absolute_change",
    "p5_relative_change",
    "p50_absolute_change",
    "p50_relative_change",
    "p95_absolute_change",
    "p95_relative_change",
    "count",
    "share_or_probability",
    "convergence_status",
    "reasons_json",
    "metric_absolute_tolerance",
    "relative_change_threshold",
    "class_probability_change_threshold",
)
LIFECYCLE_CONVERGENCE_COLUMNS = (
    *CONVERGENCE_COLUMNS,
    "decision_probability_change_threshold",
)


def _convergence_rows(metadata: Mapping[str, Any]) -> Iterator[tuple[Any, ...]]:
    convergence = metadata.get("convergence") or {}
    tolerances = convergence.get("metric_absolute_tolerances") or {}
    common_tail = (
        convergence.get("status"),
        _canonical_json_text(convergence.get("reasons") or []),
        convergence.get("relative_change_threshold"),
        convergence.get("class_probability_change_threshold"),
    )
    for checkpoint_index, checkpoint in enumerate(
        convergence.get("checkpoints") or [],
        start=1,
    ):
        count = checkpoint.get("realization_count")
        for metric_id, metric in sorted((checkpoint.get("metrics") or {}).items()):
            percentiles = metric.get("percentiles") or {}
            changes = metric.get("change_from_previous") or {}
            yield (
                "metric",
                checkpoint_index,
                count,
                metric_id,
                None,
                metric.get("population_count"),
                percentiles.get("p5"),
                percentiles.get("p50"),
                percentiles.get("p95"),
                (changes.get("p5") or {}).get("absolute"),
                (changes.get("p5") or {}).get("relative"),
                (changes.get("p50") or {}).get("absolute"),
                (changes.get("p50") or {}).get("relative"),
                (changes.get("p95") or {}).get("absolute"),
                (changes.get("p95") or {}).get("relative"),
                None,
                None,
                common_tail[0],
                common_tail[1],
                tolerances.get(metric_id),
                common_tail[2],
                common_tail[3],
            )
        for record_type, probabilities in (
            ("energy_class", checkpoint.get("energy_class_probabilities")),
            ("tradeoff_class", checkpoint.get("tradeoff_probabilities")),
        ):
            for category, probability in sorted((probabilities or {}).items()):
                yield (
                    record_type,
                    checkpoint_index,
                    count,
                    None,
                    category,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    probability,
                    common_tail[0],
                    common_tail[1],
                    None,
                    common_tail[2],
                    common_tail[3],
                )
        decision_probabilities = checkpoint.get("decision_probabilities")
        if isinstance(decision_probabilities, Mapping):
            for metric_id, probabilities in sorted(
                decision_probabilities.items()
            ):
                if not isinstance(probabilities, Mapping):
                    continue
                for category, probability in sorted(probabilities.items()):
                    yield (
                        "decision_probability",
                        checkpoint_index,
                        count,
                        metric_id,
                        category,
                        count,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        probability,
                        common_tail[0],
                        common_tail[1],
                        None,
                        common_tail[2],
                        common_tail[3],
                    )
        year_counts = checkpoint.get("weather_year_counts") or {}
        year_shares = checkpoint.get("weather_year_shares") or {}
        for year, year_count in sorted(year_counts.items(), key=lambda pair: int(pair[0])):
            yield (
                "weather_year",
                checkpoint_index,
                count,
                None,
                year,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                year_count,
                year_shares.get(year, year_shares.get(str(year))),
                common_tail[0],
                common_tail[1],
                None,
                common_tail[2],
                common_tail[3],
            )


def _lifecycle_convergence_rows(
    metadata: Mapping[str, Any],
) -> Iterator[tuple[Any, ...]]:
    convergence = metadata.get("convergence") or {}
    threshold = (
        convergence.get("decision_probability_change_threshold")
        if isinstance(convergence, Mapping)
        else None
    )
    for row in _convergence_rows(metadata):
        yield (*row, threshold)


PROVENANCE_COLUMNS = ("section", "field_path", "value_json")

_LONG_TEXT_CHUNK_SIZE = 30_000
_LONG_TEXT_CHUNK_MARKER = ".__json_chunk__"


def _chunked_leaf_rows(prefix: str, encoded: str) -> Iterator[tuple[str, str]]:
    if len(encoded) <= _LONG_TEXT_CHUNK_SIZE:
        yield prefix, encoded
        return
    count = math.ceil(len(encoded) / _LONG_TEXT_CHUNK_SIZE)
    for index in range(count):
        start = index * _LONG_TEXT_CHUNK_SIZE
        suffix = f"{_LONG_TEXT_CHUNK_MARKER}[{index + 1:04d}-of-{count:04d}]"
        yield f"{prefix}{suffix}", encoded[start : start + _LONG_TEXT_CHUNK_SIZE]


def _flatten_leaves(prefix: str, value: Any) -> Iterator[tuple[str, str]]:
    value = _safe_public_value(value)
    if isinstance(value, Mapping):
        if not value:
            yield prefix, "{}"
        for key, item in sorted(value.items()):
            next_prefix = f"{prefix}.{key}" if prefix else str(key)
            yield from _flatten_leaves(next_prefix, item)
    elif isinstance(value, list):
        if not value:
            yield prefix, "[]"
        for index, item in enumerate(value):
            yield from _flatten_leaves(f"{prefix}[{index}]", item)
    else:
        yield from _chunked_leaf_rows(prefix, _canonical_json_text(value))


def _provenance_rows(
    request_payload: Mapping[str, Any],
    source_snapshot: Mapping[str, Any],
    submission_provenance: Mapping[str, Any],
    routine_result: Mapping[str, Any],
    metadata: Mapping[str, Any],
) -> Iterator[tuple[str, str, str]]:
    sections = {
        "request": request_payload,
        "source_snapshot": source_snapshot,
        "submission": submission_provenance,
        "routine_result": routine_result,
        "kernel": metadata.get("kernel_provenance") or {},
    }
    for section, value in sections.items():
        for field_path, encoded in _flatten_leaves("", value):
            yield section, field_path, encoded


def _lifecycle_provenance_rows(
    request_payload: Mapping[str, Any],
    source_snapshot: Mapping[str, Any],
    submission_provenance: Mapping[str, Any],
    routine_result: Mapping[str, Any],
    metadata: Mapping[str, Any],
    registry: Sequence[Mapping[str, Any]],
    checks: Sequence[Sequence[Any]],
) -> Iterator[tuple[str, str, str]]:
    """Emit compact v6 provenance without duplicating exported trace tables."""

    request_identity = {
        key: request_payload.get(key)
        for key in (
            "calculation_contract_version",
            "source_annual_job_id",
            "basis",
            "n",
            "seed",
            "cost_stack_completeness",
        )
    }
    request_identity["finance"] = request_payload.get("finance") or {}
    source_identity = {
        "source_annual_job_id": source_snapshot.get("source_annual_job_id"),
        "source_snapshot_sha256": submission_provenance.get(
            "source_snapshot_sha256"
        ),
        "midc_source_artifact": source_snapshot.get("midc_source_artifact") or {},
    }
    submission_identity = {
        key: submission_provenance.get(key)
        for key in (
            "request_sha256",
            "source_snapshot_sha256",
            "validated_kernel_request_sha256",
            "calculation_contract_version",
            "sampling_version",
            "source_annual_job_id",
            "commercial_transfer_status",
        )
    }
    routine_identity = {
        key: routine_result.get(key)
        for key in (
            "schema_version",
            "result_version",
            "calculation_contract_version",
            "sampling_version",
            "analysis_basis",
            "realization_count",
            "seed",
            "project_life_years",
            "source_snapshot_sha256",
            "capacity_basis",
            "energy_available",
        )
    }
    routine_identity["sealed_calculation"] = routine_result.get(
        "sealed_calculation"
    ) or {}
    reporting_identity = {
        "workbook_schema_version": LIFECYCLE_XLSX_SCHEMA_VERSION,
        "workbook_logical_hash_version": LIFECYCLE_XLSX_LOGICAL_HASH_VERSION,
        "csv_format_version": LIFECYCLE_CSV_FORMAT_VERSION,
        "formula_registry": {
            "version": getattr(
                technoeconomic_kernel,
                "FORMULA_REGISTRY_VERSION",
                "tea-formulas-v6",
            ),
            "count": len(registry),
            "sha256": _formula_registry_sha256(registry),
            "template_hash_version": LIFECYCLE_FORMULA_TEMPLATE_HASH_VERSION,
            "template_sha256": _formula_template_sha256(registry),
        },
        "decision_rule_version": LIFECYCLE_DECISION_RULE_VERSION,
        "audit_tie_outs": [
            {
                "check_id": row[0],
                "status": row[5],
                "actual": row[1],
                "expected": row[2],
                "difference": row[3],
                "tolerance": row[4],
            }
            for row in checks
        ],
        "chart_and_image_integrity": (
            "Native chart ranges/anchors and embedded image hashes are sealed "
            "in the export manifest."
        ),
    }
    sections = {
        "request_identity": request_identity,
        "source_identity": source_identity,
        "submission_identity": submission_identity,
        "routine_identity": routine_identity,
        "kernel": metadata.get("kernel_provenance") or {},
        "reporting": reporting_identity,
    }
    for section, value in sections.items():
        for field_path, encoded in _flatten_leaves("", value):
            yield section, field_path, encoded


CHECK_COLUMNS = (
    "check_id",
    "actual_authority",
    "expected_authority",
    "difference_authority",
    "tolerance",
    "status_authority",
    "notes",
)


def _finite_values(values: np.ndarray) -> np.ndarray:
    try:
        vector = np.asarray(values, dtype=np.float64)
    except (TypeError, ValueError) as exc:
        raise TechnoeconomicExportError("A numeric tie-out column is invalid") from exc
    return vector[np.isfinite(vector)]


def _metric_population(
    metric_id: str,
    calculation: _SealedCalculation,
) -> np.ndarray | None:
    if metric_id in calculation.by_name:
        return _finite_values(calculation.by_name[metric_id])
    lcoo_column = technoeconomic_kernel.FIELD_LCOO
    if metric_id == "headline_positive_gain_lcoo":
        if lcoo_column not in calculation.by_name or "energy_class" not in calculation.by_name:
            return None
        values = np.asarray(calculation.by_name[lcoo_column], dtype=np.float64)
        classes = np.asarray(calculation.by_name["energy_class"])
        return values[(classes == "positive_lifecycle_gain") & np.isfinite(values)]
    if metric_id == "signed_nonzero_lcoo":
        if lcoo_column not in calculation.by_name:
            return None
        return _finite_values(calculation.by_name[lcoo_column])
    return None


def _numeric_check(
    check_id: str,
    actual: float | int,
    expected: float | int,
    *,
    tolerance: float,
    notes: str,
) -> tuple[Any, ...]:
    difference = float(actual) - float(expected)
    return (
        check_id,
        actual,
        expected,
        difference,
        tolerance,
        "OK" if abs(difference) <= tolerance else "FAIL",
        notes,
    )


def _binary64_tie_out_tolerance(*values: np.ndarray) -> float:
    scale = 1.0
    for values_array in values:
        finite = np.asarray(values_array, dtype=np.float64)
        finite = finite[np.isfinite(finite)]
        if finite.size:
            scale = max(scale, float(np.max(np.abs(finite))))
    return float(16.0 * np.finfo(np.float64).eps * scale)


def _metric_fields_for_result(
    routine_result: Mapping[str, Any],
) -> Any:
    contract_version = routine_result.get("calculation_contract_version")
    if contract_version == technoeconomic_kernel.LEGACY_CALCULATION_CONTRACT_VERSION:
        return technoeconomic_kernel.LEGACY_METRIC_FIELDS
    if contract_version in {
        technoeconomic_kernel.CALCULATION_CONTRACT_VERSION,
        technoeconomic_kernel.COMMERCIAL_SCALING_CALCULATION_CONTRACT_VERSION,
        technoeconomic_kernel.STANDALONE_COMMERCIAL_CALCULATION_CONTRACT_VERSION,
        technoeconomic_kernel.PAIRED_COMMERCIAL_CALCULATION_CONTRACT_VERSION,
    }:
        return technoeconomic_kernel.APPLIED_METRIC_FIELDS
    raise TechnoeconomicExportError(
        f"Unsupported calculation contract in routine result: {contract_version!r}"
    )


def _build_checks(
    calculation: _SealedCalculation,
    source_snapshot: Mapping[str, Any],
    submission_provenance: Mapping[str, Any],
    routine_result: Mapping[str, Any],
) -> list[tuple[Any, ...]]:
    checks: list[tuple[Any, ...]] = []
    fields = _metric_fields_for_result(routine_result)
    applied_capacity_contract = routine_result.get(
        "calculation_contract_version"
    ) in {
        technoeconomic_kernel.CALCULATION_CONTRACT_VERSION,
        technoeconomic_kernel.COMMERCIAL_SCALING_CALCULATION_CONTRACT_VERSION,
        technoeconomic_kernel.STANDALONE_COMMERCIAL_CALCULATION_CONTRACT_VERSION,
        technoeconomic_kernel.PAIRED_COMMERCIAL_CALCULATION_CONTRACT_VERSION,
    }
    expected_rows = routine_result.get("realization_count")
    if not isinstance(expected_rows, int) or isinstance(expected_rows, bool):
        raise TechnoeconomicExportError("Routine result realization count is invalid")
    checks.append(
        _numeric_check(
            "realization_count",
            calculation.row_count,
            expected_rows,
            tolerance=0.0,
            notes="Sealed realization rows equal the durable routine-result count.",
        )
    )
    summaries = _metric_summaries(calculation.metadata)
    for metric_id, summary in sorted(summaries.items()):
        if not isinstance(summary, Mapping) or "percentiles" not in summary:
            continue
        population = _metric_population(metric_id, calculation)
        if population is None:
            continue
        checks.append(
            _numeric_check(
                f"summary_count::{metric_id}",
                len(population),
                int(summary.get("count") or 0),
                tolerance=0.0,
                notes="Finite realization population equals the frozen summary count.",
            )
        )
        if len(population):
            percentiles = summary.get("percentiles") or {}
            if {"p10", "p50", "p90"} <= set(percentiles):
                quantile_names = ("p10", "p50", "p90")
                quantile_values = (0.10, 0.50, 0.90)
            else:
                quantile_names = ("p5", "p50", "p95")
                quantile_values = (0.05, 0.50, 0.95)
            calculated = np.quantile(
                population,
                quantile_values,
                method="linear",
            )
            for index, quantile in enumerate(quantile_names):
                expected = percentiles.get(quantile)
                if expected is None:
                    checks.append(
                        (
                            f"percentile::{metric_id}::{quantile}",
                            float(calculated[index]),
                            None,
                            None,
                            1e-12,
                            "FAIL",
                            "Available finite population has a missing frozen percentile.",
                        )
                    )
                else:
                    scale = max(1.0, abs(float(expected)))
                    checks.append(
                        _numeric_check(
                            f"percentile::{metric_id}::{quantile}",
                            float(calculated[index]),
                            float(expected),
                            tolerance=1e-12 * scale,
                            notes="Type-7 percentile recomputed from sealed realization values.",
                        )
                    )
            cdf = summary.get("cdf")
            if isinstance(cdf, Mapping):
                unique_values, duplicate_counts = np.unique(
                    np.sort(population),
                    return_counts=True,
                )
                cumulative_count = np.cumsum(duplicate_counts, dtype=np.int64)
                recomputed_cdf = {
                    "values": [float(value) for value in unique_values],
                    "cumulative_count": [int(value) for value in cumulative_count],
                    "cumulative_probability": [
                        float(value / len(population)) for value in cumulative_count
                    ],
                    "population_count": int(len(population)),
                }
                actual_digest = _canonical_json_sha256(recomputed_cdf)
                expected_digest = _canonical_json_sha256(cdf)
                checks.append(
                    (
                        f"cdf_full_identity::{metric_id}",
                        actual_digest,
                        expected_digest,
                        None,
                        None,
                        "OK"
                        if secrets.compare_digest(actual_digest, expected_digest)
                        else "FAIL",
                        "Full right-continuous ECDF values, duplicate counts, and probabilities are recomputed from sealed realizations.",
                    )
                )
                terminal = (cdf.get("cumulative_probability") or [None])[-1]
                checks.append(
                    _numeric_check(
                        f"cdf_terminal_probability::{metric_id}",
                        float(terminal),
                        1.0,
                        tolerance=0.0,
                        notes="Right-continuous ECDF terminates at probability one.",
                    )
                )

    capacities = routine_result.get("capacities") or {}
    snapshot_systems = (source_snapshot.get("capacity_manifest") or {}).get("systems") or {}
    for system in ("solectria", "solaredge"):
        actual = (capacities.get(system) or {}).get("installed_wdc")
        expected = (snapshot_systems.get(system) or {}).get("installed_wdc")
        if actual is None or expected is None:
            raise TechnoeconomicExportError("Capacity tie-out evidence is missing")
        checks.append(
            _numeric_check(
                f"capacity_wdc::{system}",
                float(actual),
                float(expected),
                tolerance=0.0,
                notes="Durable result capacity equals the frozen Annual capacity manifest.",
            )
        )

    applied_capacity_authority: dict[str, Mapping[str, Any]] = {}
    if applied_capacity_contract:
        applied_capacity_authority = _applied_capacity_authority(
            submission_provenance
        )
        result_applied = routine_result.get("applied_capacities") or {}
        if not isinstance(result_applied, Mapping):
            raise TechnoeconomicExportError(
                "Durable applied-capacity result is invalid"
            )
        for system in ("solectria", "solaredge"):
            actual = result_applied.get(system) or {}
            expected = applied_capacity_authority[system]
            if not isinstance(actual, Mapping):
                raise TechnoeconomicExportError(
                    "Durable applied-capacity result is incomplete"
                )
            checks.append(
                _numeric_check(
                    f"applied_capacity_w::{system}",
                    float(actual.get("applied_capacity_w")),
                    float(expected.get("applied_capacity_w")),
                    tolerance=0.0,
                    notes=(
                        "Durable applied capacity equals the immutable "
                        "normalization receipt."
                    ),
                )
            )
            actual_basis = actual.get("rating_basis")
            expected_basis = expected.get("rating_basis")
            checks.append(
                (
                    f"applied_capacity_rating_basis::{system}",
                    actual_basis,
                    expected_basis,
                    None,
                    None,
                    "OK" if actual_basis == expected_basis else "FAIL",
                    "Durable applied-capacity rating basis matches its receipt.",
                )
            )

    source_hash_actual = routine_result.get("source_snapshot_sha256")
    source_hash_expected = submission_provenance.get("source_snapshot_sha256")
    checks.append(
        (
            "source_snapshot_sha256",
            source_hash_actual,
            source_hash_expected,
            None,
            None,
            "OK"
            if isinstance(source_hash_actual, str)
            and isinstance(source_hash_expected, str)
            and secrets.compare_digest(source_hash_actual, source_hash_expected)
            else "FAIL",
            "Durable result is bound to the submission's frozen source snapshot.",
        )
    )
    transfer_actual = routine_result.get("commercial_transfer_status")
    transfer_expected = submission_provenance.get("commercial_transfer_status")
    checks.append(
        (
            "commercial_transfer_status",
            transfer_actual,
            transfer_expected,
            None,
            None,
            "OK" if transfer_actual == transfer_expected else "FAIL",
            "Result and immutable transfer receipt use the same status.",
        )
    )

    common_records = calculation.metadata.get("common_cost_audit") or []
    for record in common_records:
        if record.get("comparison_treatment") != "common_cancelled":
            continue
        actual = bool(record.get("delta_contribution_se_minus_sol_exactly_zero"))
        checks.append(
            (
                f"common_cost_exact_cancellation::{record.get('input_id')}",
                actual,
                True,
                None,
                None,
                "OK" if actual else "FAIL",
                "Approved common-cost cancellation is exactly zero in the paired delta.",
            )
        )

    normalized_pairs = (
        (
            "lifecycle_cost_delta",
            fields.pv_cost_se,
            fields.pv_cost_sol,
            fields.delta_cost,
        ),
        (
            "equivalent_annual_cost_delta",
            fields.ea_cost_se,
            fields.ea_cost_sol,
            fields.delta_ea_cost,
        ),
        (
            "lifecycle_energy_delta",
            fields.pv_energy_se,
            fields.pv_energy_sol,
            fields.delta_energy,
        ),
        (
            "equivalent_annual_energy_delta",
            fields.ea_energy_se,
            fields.ea_energy_sol,
            fields.delta_ea_energy,
        ),
    )
    for check_name, se_name, sol_name, delta_name in normalized_pairs:
        if not all(name in calculation.by_name for name in (se_name, sol_name, delta_name)):
            continue
        se_values = np.asarray(calculation.by_name[se_name], dtype=np.float64)
        sol_values = np.asarray(calculation.by_name[sol_name], dtype=np.float64)
        delta_values = np.asarray(calculation.by_name[delta_name], dtype=np.float64)
        finite = np.isfinite(se_values) & np.isfinite(sol_values) & np.isfinite(delta_values)
        maximum = float(
            np.max(np.abs((se_values[finite] - sol_values[finite]) - delta_values[finite]))
        ) if finite.any() else 0.0
        checks.append(
            _numeric_check(
                check_name,
                maximum,
                0.0,
                tolerance=1e-12,
                notes="Signed SE-minus-SOL normalized delta ties to component columns.",
            )
        )

    delta_cost = np.asarray(
        calculation.by_name[fields.delta_cost],
        dtype=np.float64,
    )
    delta_energy = np.asarray(
        calculation.by_name[fields.delta_energy],
        dtype=np.float64,
    )
    delta_ea_cost = np.asarray(
        calculation.by_name[fields.delta_ea_cost],
        dtype=np.float64,
    )
    delta_ea_energy = np.asarray(
        calculation.by_name[fields.delta_ea_energy],
        dtype=np.float64,
    )
    lcoo = np.asarray(
        calculation.by_name[fields.lcoo],
        dtype=np.float64,
    )
    crf = np.asarray(
        calculation.by_name["CapitalRecoveryFactor_per_year"],
        dtype=np.float64,
    )
    lifecycle_ratio_mask = (
        np.isfinite(lcoo)
        & np.isfinite(delta_cost)
        & np.isfinite(delta_energy)
        & (delta_energy != 0)
    )
    lifecycle_ratio_error = (
        float(np.max(np.abs(lcoo[lifecycle_ratio_mask] - delta_cost[lifecycle_ratio_mask] / delta_energy[lifecycle_ratio_mask])))
        if lifecycle_ratio_mask.any()
        else 0.0
    )
    checks.append(
        _numeric_check(
            "lcoo_lifecycle_ratio",
            lifecycle_ratio_error,
            0.0,
            tolerance=1e-12,
            notes="Every finite LCOO equals lifecycle cost delta divided by lifecycle energy delta.",
        )
    )
    annual_ratio_mask = (
        np.isfinite(lcoo)
        & np.isfinite(delta_ea_cost)
        & np.isfinite(delta_ea_energy)
        & (delta_ea_energy != 0)
    )
    annual_ratio_error = (
        float(np.max(np.abs(lcoo[annual_ratio_mask] - delta_ea_cost[annual_ratio_mask] / delta_ea_energy[annual_ratio_mask])))
        if annual_ratio_mask.any()
        else 0.0
    )
    checks.append(
        _numeric_check(
            "lcoo_equivalent_annual_ratio",
            annual_ratio_error,
            0.0,
            tolerance=1e-12,
            notes="Every finite LCOO also equals the equivalent-annual cost/energy ratio.",
        )
    )
    for check_name, annual_values, lifecycle_values in (
        ("crf_cost_delta_transform", delta_ea_cost, delta_cost),
        ("crf_energy_delta_transform", delta_ea_energy, delta_energy),
    ):
        finite = np.isfinite(annual_values) & np.isfinite(lifecycle_values) & np.isfinite(crf)
        maximum = (
            float(np.max(np.abs(annual_values[finite] - crf[finite] * lifecycle_values[finite])))
            if finite.any()
            else 0.0
        )
        checks.append(
            _numeric_check(
                check_name,
                maximum,
                0.0,
                tolerance=1e-12,
                notes="Equivalent-annual signed delta equals CRF times its lifecycle delta.",
            )
        )

    for system, cost_name, energy_name, lcoe_name in (
        (
            "solectria",
            fields.pv_cost_sol,
            fields.pv_energy_sol,
            fields.lcoe_sol,
        ),
        (
            "solaredge",
            fields.pv_cost_se,
            fields.pv_energy_se,
            fields.lcoe_se,
        ),
    ):
        lifecycle_cost = np.asarray(calculation.by_name[cost_name], dtype=np.float64)
        lifecycle_energy = np.asarray(calculation.by_name[energy_name], dtype=np.float64)
        lcoe_values = np.asarray(calculation.by_name[lcoe_name], dtype=np.float64)
        ratio_mask = (
            np.isfinite(lifecycle_cost)
            & np.isfinite(lifecycle_energy)
            & np.isfinite(lcoe_values)
            & (lifecycle_energy != 0)
        )
        ratio_error = (
            float(
                np.max(
                    np.abs(
                        lcoe_values[ratio_mask]
                        - lifecycle_cost[ratio_mask] / lifecycle_energy[ratio_mask]
                    )
                )
            )
            if ratio_mask.any()
            else 0.0
        )
        checks.append(
            _numeric_check(
                f"lcoe_lifecycle_ratio::{system}",
                ratio_error,
                0.0,
                tolerance=_binary64_tie_out_tolerance(lcoe_values),
                notes="Standalone lifecycle LCOE equals lifecycle cost divided by lifecycle energy.",
            )
        )

    for check_name, annual_name, lifecycle_name in (
        (
            "crf_cost_transform::solectria",
            fields.ea_cost_sol,
            fields.pv_cost_sol,
        ),
        (
            "crf_cost_transform::solaredge",
            fields.ea_cost_se,
            fields.pv_cost_se,
        ),
        (
            "crf_energy_transform::solectria",
            fields.ea_energy_sol,
            fields.pv_energy_sol,
        ),
        (
            "crf_energy_transform::solaredge",
            fields.ea_energy_se,
            fields.pv_energy_se,
        ),
    ):
        annual_values = np.asarray(calculation.by_name[annual_name], dtype=np.float64)
        lifecycle_values = np.asarray(
            calculation.by_name[lifecycle_name], dtype=np.float64
        )
        finite = (
            np.isfinite(annual_values)
            & np.isfinite(lifecycle_values)
            & np.isfinite(crf)
        )
        maximum = (
            float(
                np.max(
                    np.abs(
                        annual_values[finite]
                        - crf[finite] * lifecycle_values[finite]
                    )
                )
            )
            if finite.any()
            else 0.0
        )
        checks.append(
            _numeric_check(
                check_name,
                maximum,
                0.0,
                tolerance=_binary64_tie_out_tolerance(annual_values),
                notes="Standalone equivalent-annual authority equals CRF times lifecycle authority.",
            )
        )

    energy_class = np.asarray(calculation.by_name["energy_class"])
    lcoo_reason = np.asarray(calculation.by_name["lcoo_unavailable_reason"])
    zero_mask = energy_class == "zero_lifecycle_gain"
    zero_violations = int(
        np.count_nonzero(
            zero_mask
            & (
                np.isfinite(lcoo)
                | (lcoo_reason != "zero_lifecycle_delta_energy")
            )
        )
    )
    checks.append(
        _numeric_check(
            "zero_energy_lcoo_null_and_reason",
            zero_violations,
            0,
            tolerance=0.0,
            notes="Every zero-energy-change row retains null LCOO and the explicit zero-delta reason.",
        )
    )
    negative_mask = energy_class == "negative_lifecycle_gain"
    retained_negative = int(
        np.count_nonzero(
            negative_mask & np.isfinite(lcoo) & (lcoo_reason == None)  # noqa: E711
        )
    )
    checks.append(
        _numeric_check(
            "negative_energy_signed_lcoo_retention",
            retained_negative,
            int(np.count_nonzero(negative_mask)),
            tolerance=0.0,
            notes="Every negative lifecycle-energy row retains its finite signed LCOO with no unavailable reason.",
        )
    )

    for summary_name in ("energy_classes", "tradeoff_classes"):
        class_summary = summaries.get(summary_name)
        if not isinstance(class_summary, Mapping) or "counts" not in class_summary:
            continue
        total = sum(int(value) for value in (class_summary.get("counts") or {}).values())
        checks.append(
            _numeric_check(
                f"class_count_total::{summary_name}",
                total,
                calculation.row_count,
                tolerance=0.0,
                notes="Every realization is retained in exactly one reported outcome class.",
            )
        )

    raw_pairs = (
        ("site_raw_cost_sol", "PVCost_SOL_USD", fields.pv_cost_sol, "solectria"),
        ("site_raw_cost_se", "PVCost_SE_USD", fields.pv_cost_se, "solaredge"),
        ("site_raw_energy_sol", "PVEnergy_SOL_kWh_AC", fields.pv_energy_sol, "solectria"),
        ("site_raw_energy_se", "PVEnergy_SE_kWh_AC", fields.pv_energy_se, "solaredge"),
    )
    for check_name, raw_name, normalized_name, system in raw_pairs:
        if raw_name not in calculation.by_name:
            continue
        denominator_w = float(
            applied_capacity_authority[system].get("applied_capacity_w")
            if applied_capacity_contract
            else (snapshot_systems.get(system) or {}).get("installed_wdc")
        )
        raw_values = np.asarray(calculation.by_name[raw_name], dtype=np.float64)
        normalized_values = np.asarray(calculation.by_name[normalized_name], dtype=np.float64)
        maximum = float(
            np.max(np.abs(raw_values / denominator_w - normalized_values))
        )
        checks.append(
            _numeric_check(
                check_name,
                maximum,
                0.0,
                tolerance=1e-12,
                notes=(
                    "Raw site total divided by the frozen applied capacity ties "
                    "to normalized authority."
                    if applied_capacity_contract
                    else "Raw site total divided by frozen system Wdc ties to normalized authority."
                ),
            )
        )

    for check_name, delta_name, se_name, sol_name in (
        (
            "site_raw_lifecycle_cost_delta",
            "DeltaPVCostUSD_se_minus_sol",
            "PVCost_SE_USD",
            "PVCost_SOL_USD",
        ),
        (
            "site_raw_lifecycle_energy_delta",
            "DeltaPVEnergyKWhAC_se_minus_sol",
            "PVEnergy_SE_kWh_AC",
            "PVEnergy_SOL_kWh_AC",
        ),
        (
            "reference_raw_lifecycle_cost_delta",
            "ReferenceDeltaPVCostUSD_se_minus_sol",
            "ReferencePVCost_SE_USD",
            "ReferencePVCost_SOL_USD",
        ),
        (
            "reference_raw_lifecycle_energy_delta",
            "ReferenceDeltaPVEnergyKWhAC_se_minus_sol",
            "ReferencePVEnergy_SE_kWh_AC",
            "ReferencePVEnergy_SOL_kWh_AC",
        ),
    ):
        if not all(name in calculation.by_name for name in (delta_name, se_name, sol_name)):
            continue
        delta_values = np.asarray(calculation.by_name[delta_name], dtype=np.float64)
        se_values = np.asarray(calculation.by_name[se_name], dtype=np.float64)
        sol_values = np.asarray(calculation.by_name[sol_name], dtype=np.float64)
        finite = np.isfinite(delta_values) & np.isfinite(se_values) & np.isfinite(sol_values)
        maximum = (
            float(np.max(np.abs(delta_values[finite] - (se_values[finite] - sol_values[finite]))))
            if finite.any()
            else 0.0
        )
        tolerance = _binary64_tie_out_tolerance(
            delta_values[finite],
            se_values[finite],
            sol_values[finite],
        )
        checks.append(
            _numeric_check(
                check_name,
                maximum,
                0.0,
                tolerance=tolerance,
                notes="Explicit raw SE-minus-SOL delta ties within a scale-aware binary64 roundoff bound.",
            )
        )

    reference_wdc = (submission_provenance.get("commercial_reference_design") or {}).get(
        "reference_wdc"
    )
    if reference_wdc is not None:
        for check_name, raw_name, normalized_name in (
            ("reference_raw_cost_sol", "ReferencePVCost_SOL_USD", fields.pv_cost_sol),
            ("reference_raw_cost_se", "ReferencePVCost_SE_USD", fields.pv_cost_se),
            ("reference_raw_energy_sol", "ReferencePVEnergy_SOL_kWh_AC", fields.pv_energy_sol),
            ("reference_raw_energy_se", "ReferencePVEnergy_SE_kWh_AC", fields.pv_energy_se),
        ):
            if raw_name not in calculation.by_name:
                continue
            raw_values = np.asarray(calculation.by_name[raw_name], dtype=np.float64)
            normalized_values = np.asarray(calculation.by_name[normalized_name], dtype=np.float64)
            maximum = float(
                np.max(np.abs(raw_values / float(reference_wdc) - normalized_values))
            )
            checks.append(
                _numeric_check(
                    check_name,
                    maximum,
                    0.0,
                    tolerance=1e-12,
                    notes="Commercial reference total divided by declared reference Wdc ties out.",
                )
            )
    if (
        routine_result.get("calculation_contract_version")
        == technoeconomic_kernel.COMMERCIAL_SCALING_CALCULATION_CONTRACT_VERSION
    ):
        commercial_names = (
            technoeconomic_kernel.COMMERCIAL_FIELD_TARGET_CAPACITY,
            technoeconomic_kernel.COMMERCIAL_FIELD_YEAR1_DELTA_ENERGY,
            technoeconomic_kernel.COMMERCIAL_FIELD_LIFECYCLE_DELTA_ENERGY,
            technoeconomic_kernel.COMMERCIAL_FIELD_EA_DELTA_ENERGY,
            technoeconomic_kernel.COMMERCIAL_FIELD_LIFECYCLE_MARGINAL_COST,
            technoeconomic_kernel.COMMERCIAL_FIELD_EA_MARGINAL_COST,
            technoeconomic_kernel.COMMERCIAL_FIELD_MARGINAL_LCOO,
            technoeconomic_kernel.COMMERCIAL_FIELD_MARGINAL_LCOO_REASON,
        )
        missing = [name for name in commercial_names if name not in calculation.by_name]
        if missing:
            raise TechnoeconomicExportError(
                f"Commercial-scaling realization fields are missing: {missing!r}"
            )
        scaling = routine_result.get("commercial_scaling") or {}
        receipt = submission_provenance.get("commercial_scaling_receipt") or {}
        if not isinstance(scaling, Mapping) or not isinstance(receipt, Mapping):
            raise TechnoeconomicExportError(
                "Commercial-scaling result or immutable receipt is invalid"
            )
        target_w = float(scaling.get("target_capacity_w"))
        checks.append(
            _numeric_check(
                "commercial_target_capacity_receipt",
                target_w,
                float(receipt.get("target_capacity_w")),
                tolerance=0.0,
                notes="Canonical target watts equal the immutable submission receipt.",
            )
        )
        checks.append(
            (
                "commercial_target_rating_basis_receipt",
                scaling.get("target_rating_basis"),
                receipt.get("target_rating_basis"),
                None,
                None,
                "OK"
                if scaling.get("target_rating_basis")
                == receipt.get("target_rating_basis")
                else "FAIL",
                "Commercial target and frozen source use the receipt-bound rating basis.",
            )
        )
        target_column = np.asarray(
            calculation.by_name[
                technoeconomic_kernel.COMMERCIAL_FIELD_TARGET_CAPACITY
            ],
            dtype=np.float64,
        )
        checks.append(
            _numeric_check(
                "commercial_target_capacity_realizations",
                float(np.max(np.abs(target_column - target_w))),
                0.0,
                tolerance=0.0,
                notes="Every realization uses the frozen commercial target capacity.",
            )
        )
        commercial_energy_pairs = (
            (
                "commercial_year1_energy_scaling",
                technoeconomic_kernel.COMMERCIAL_FIELD_YEAR1_DELTA_ENERGY,
                fields.year1_delta,
            ),
            (
                "commercial_lifecycle_energy_scaling",
                technoeconomic_kernel.COMMERCIAL_FIELD_LIFECYCLE_DELTA_ENERGY,
                fields.delta_energy,
            ),
            (
                "commercial_equivalent_annual_energy_scaling",
                technoeconomic_kernel.COMMERCIAL_FIELD_EA_DELTA_ENERGY,
                fields.delta_ea_energy,
            ),
        )
        for check_id, commercial_name, normalized_name in commercial_energy_pairs:
            commercial_values = np.asarray(
                calculation.by_name[commercial_name], dtype=np.float64
            )
            normalized_values = np.asarray(
                calculation.by_name[normalized_name], dtype=np.float64
            )
            error = float(
                np.max(np.abs(commercial_values - normalized_values * target_w))
            )
            checks.append(
                _numeric_check(
                    check_id,
                    error,
                    0.0,
                    tolerance=_binary64_tie_out_tolerance(commercial_values),
                    notes="Commercial energy equals normalized source authority times target watts.",
                )
            )
        commercial_lifecycle_cost = np.asarray(
            calculation.by_name[
                technoeconomic_kernel.COMMERCIAL_FIELD_LIFECYCLE_MARGINAL_COST
            ],
            dtype=np.float64,
        )
        commercial_ea_cost = np.asarray(
            calculation.by_name[
                technoeconomic_kernel.COMMERCIAL_FIELD_EA_MARGINAL_COST
            ],
            dtype=np.float64,
        )
        marginal_cost_sample_name = (
            "SampledInput::"
            f"{technoeconomic_kernel.COMMERCIAL_MARGINAL_COST_DIFFERENCE_INPUT_ID}"
        )
        if marginal_cost_sample_name not in calculation.by_name:
            raise TechnoeconomicExportError(
                "The sampled commercial marginal-cost authority is missing"
            )
        marginal_cost_sample = np.asarray(
            calculation.by_name[marginal_cost_sample_name], dtype=np.float64
        )
        marginal_cost_timing = receipt.get("marginal_cost_timing")
        if marginal_cost_timing == "lifecycle_present_value":
            authoritative_cost = commercial_lifecycle_cost
            authoritative_cost_field = (
                technoeconomic_kernel.COMMERCIAL_FIELD_LIFECYCLE_MARGINAL_COST
            )
        elif marginal_cost_timing == "equivalent_annual":
            authoritative_cost = commercial_ea_cost
            authoritative_cost_field = (
                technoeconomic_kernel.COMMERCIAL_FIELD_EA_MARGINAL_COST
            )
        else:
            raise TechnoeconomicExportError(
                "The commercial marginal-cost timing receipt is invalid"
            )
        sampled_cost_error = float(
            np.max(np.abs(authoritative_cost - marginal_cost_sample))
        )
        checks.append(
            _numeric_check(
                "commercial_marginal_cost_sampled_input_authority",
                sampled_cost_error,
                0.0,
                tolerance=_binary64_tie_out_tolerance(authoritative_cost),
                notes=(
                    "The sampled commercial marginal-cost input equals the receipt-"
                    f"selected authoritative {authoritative_cost_field} field."
                ),
            )
        )
        cost_transform_error = float(
            np.max(np.abs(commercial_ea_cost - crf * commercial_lifecycle_cost))
        )
        checks.append(
            _numeric_check(
                "commercial_marginal_cost_crf_transform",
                cost_transform_error,
                0.0,
                tolerance=_binary64_tie_out_tolerance(commercial_ea_cost),
                notes="Equivalent-annual commercial marginal cost equals CRF times lifecycle present value.",
            )
        )
        commercial_lifecycle_energy = np.asarray(
            calculation.by_name[
                technoeconomic_kernel.COMMERCIAL_FIELD_LIFECYCLE_DELTA_ENERGY
            ],
            dtype=np.float64,
        )
        commercial_ea_energy = np.asarray(
            calculation.by_name[
                technoeconomic_kernel.COMMERCIAL_FIELD_EA_DELTA_ENERGY
            ],
            dtype=np.float64,
        )
        commercial_lcoo = np.asarray(
            calculation.by_name[
                technoeconomic_kernel.COMMERCIAL_FIELD_MARGINAL_LCOO
            ],
            dtype=np.float64,
        )
        commercial_ratio_mask = np.isfinite(commercial_lcoo)
        lifecycle_ratio_error = (
            float(
                np.max(
                    np.abs(
                        commercial_lcoo[commercial_ratio_mask]
                        - commercial_lifecycle_cost[commercial_ratio_mask]
                        / commercial_lifecycle_energy[commercial_ratio_mask]
                    )
                )
            )
            if commercial_ratio_mask.any()
            else 0.0
        )
        annual_ratio_error = (
            float(
                np.max(
                    np.abs(
                        commercial_lcoo[commercial_ratio_mask]
                        - commercial_ea_cost[commercial_ratio_mask]
                        / commercial_ea_energy[commercial_ratio_mask]
                    )
                )
            )
            if commercial_ratio_mask.any()
            else 0.0
        )
        checks.append(
            _numeric_check(
                "commercial_marginal_lcoo_lifecycle_ratio",
                lifecycle_ratio_error,
                0.0,
                tolerance=_binary64_tie_out_tolerance(commercial_lcoo),
                notes="Commercial marginal LCOO equals lifecycle marginal cost divided by scaled lifecycle energy.",
            )
        )
        checks.append(
            _numeric_check(
                "commercial_marginal_lcoo_equivalent_annual_ratio",
                annual_ratio_error,
                0.0,
                tolerance=_binary64_tie_out_tolerance(commercial_lcoo),
                notes="Commercial marginal LCOO also equals the equivalent-annual ratio.",
            )
        )
        commercial_reason = np.asarray(
            calculation.by_name[
                technoeconomic_kernel.COMMERCIAL_FIELD_MARGINAL_LCOO_REASON
            ]
        )
        unavailable = (
            commercial_reason == technoeconomic_kernel.COMMERCIAL_ZERO_ENERGY_REASON
        )
        reason_violations = int(
            np.count_nonzero(
                (unavailable & np.isfinite(commercial_lcoo))
                | (~unavailable & ~np.isfinite(commercial_lcoo))
            )
        )
        zero_energy_class = (
            np.asarray(calculation.by_name["energy_class"])
            == "zero_lifecycle_gain"
        )
        reason_class_violations = int(
            np.count_nonzero(unavailable != zero_energy_class)
        )
        checks.append(
            _numeric_check(
                "commercial_zero_energy_lcoo_null_and_reason",
                reason_violations,
                0,
                tolerance=0.0,
                notes="Commercial LCOO is null exactly for normalized zero-energy rows with an explicit reason.",
            )
        )
        checks.append(
            _numeric_check(
                "commercial_zero_energy_reason_matches_energy_class",
                reason_class_violations,
                0,
                tolerance=0.0,
                notes=(
                    "The commercial unavailable-reason mask equals the authoritative "
                    "energy_class == zero_lifecycle_gain mask."
                ),
            )
        )
    if (
        routine_result.get("calculation_contract_version")
        == technoeconomic_kernel.STANDALONE_COMMERCIAL_CALCULATION_CONTRACT_VERSION
    ):
        standalone_names = (
            technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_TARGET_CAPACITY,
            technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_CAPACITY_SCALE_FACTOR,
            technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_YEAR1_ENERGY,
            technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LIFECYCLE_ENERGY,
            technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_EA_ENERGY,
            technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_INITIAL_COST,
            technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_RECURRING_PV_COST,
            technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_SCHEDULED_PV_COST,
            technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LIFECYCLE_COST,
            technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_EA_COST,
            technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LCOE,
        )
        missing = [name for name in standalone_names if name not in calculation.by_name]
        if missing:
            raise TechnoeconomicExportError(
                "Standalone-commercial realization fields are missing: "
                f"{missing!r}"
            )
        standalone = routine_result.get("standalone_commercial") or {}
        receipt = submission_provenance.get("standalone_commercial_receipt") or {}
        if not isinstance(standalone, Mapping) or not isinstance(receipt, Mapping):
            raise TechnoeconomicExportError(
                "Standalone-commercial result or immutable receipt is invalid"
            )
        target_w = float(standalone.get("target_capacity_w"))
        source_w = float(standalone.get("source_applied_capacity_w"))
        checks.extend(
            [
                _numeric_check(
                    "standalone_target_capacity_receipt",
                    target_w,
                    float(receipt.get("target_capacity_w")),
                    tolerance=0.0,
                    notes="Canonical target watts equal the immutable v4 receipt.",
                ),
                _numeric_check(
                    "standalone_source_capacity_receipt",
                    source_w,
                    float((receipt.get("source_capacity") or {}).get("applied_capacity_w")),
                    tolerance=0.0,
                    notes="SolarEdge source watts equal the immutable Annual-capacity receipt.",
                ),
                _numeric_check(
                    "standalone_capacity_scale_factor",
                    float(standalone.get("capacity_scale_factor")),
                    target_w / source_w,
                    tolerance=0.0,
                    notes="The commercial scale factor is target W divided by source applied W.",
                ),
            ]
        )
        checks.append(
            (
                "standalone_rating_basis_bridge",
                standalone.get("target_rating_basis"),
                standalone.get("source_rating_basis"),
                None,
                None,
                "OK"
                if standalone.get("target_rating_basis")
                == standalone.get("source_rating_basis")
                == receipt.get("target_rating_basis")
                else "FAIL",
                "Source and target use the same frozen rating basis.",
            )
        )

        target_column = np.asarray(
            calculation.by_name[
                technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_TARGET_CAPACITY
            ],
            dtype=np.float64,
        )
        checks.append(
            _numeric_check(
                "standalone_target_capacity_realizations",
                float(np.max(np.abs(target_column - target_w))),
                0.0,
                tolerance=0.0,
                notes="Every realization uses the frozen commercial target capacity.",
            )
        )
        scale_column = np.asarray(
            calculation.by_name[
                technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_CAPACITY_SCALE_FACTOR
            ],
            dtype=np.float64,
        )
        checks.append(
            _numeric_check(
                "standalone_capacity_scale_factor_realizations",
                float(np.max(np.abs(scale_column - target_w / source_w))),
                0.0,
                tolerance=0.0,
                notes=(
                    "Every realization records the exact target-W/source-W "
                    "capacity bridge."
                ),
            )
        )
        energy_pairs = (
            (
                "standalone_year1_energy_scaling",
                technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_YEAR1_ENERGY,
                fields.year1_se,
            ),
            (
                "standalone_lifecycle_energy_scaling",
                technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LIFECYCLE_ENERGY,
                fields.pv_energy_se,
            ),
            (
                "standalone_equivalent_annual_energy_scaling",
                technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_EA_ENERGY,
                fields.ea_energy_se,
            ),
        )
        for check_id, commercial_name, normalized_name in energy_pairs:
            actual = np.asarray(
                calculation.by_name[commercial_name],
                dtype=np.float64,
            )
            normalized = np.asarray(
                calculation.by_name[normalized_name],
                dtype=np.float64,
            )
            checks.append(
                _numeric_check(
                    check_id,
                    float(np.max(np.abs(actual - normalized * target_w))),
                    0.0,
                    tolerance=_binary64_tie_out_tolerance(actual),
                    notes="Commercial SolarEdge energy equals normalized source energy times target watts.",
                )
            )

        expected_initial = np.zeros(calculation.row_count, dtype=np.float64)
        expected_recurring = np.zeros(calculation.row_count, dtype=np.float64)
        expected_scheduled = np.zeros(calculation.row_count, dtype=np.float64)
        discount_column = "SampledInput::finance.discount-rate"
        if discount_column not in calculation.by_name:
            raise TechnoeconomicExportError(
                "Standalone-commercial discount-rate samples are missing"
            )
        discount_rate = np.asarray(
            calculation.by_name[discount_column],
            dtype=np.float64,
        )
        line_summaries = {
            str(item.get("input_id")): item
            for item in standalone.get("commercial_cost_line_summaries") or []
            if isinstance(item, Mapping)
        }
        receipt_lines = receipt.get("cost_lines") or []
        if not isinstance(receipt_lines, list) or not receipt_lines:
            raise TechnoeconomicExportError(
                "Standalone-commercial cost receipt is incomplete"
            )
        category_counts = {
            category: sum(
                isinstance(line, Mapping)
                and line.get("cost_category") == category
                for line in receipt_lines
            )
            for category in (
                "full_initial_capex",
                "full_annual_om",
                "scheduled_replacement",
            )
        }
        required_categories_ok = (
            receipt.get("cost_stack_completeness") == "full_system"
            and category_counts["full_initial_capex"] == 1
            and category_counts["full_annual_om"] == 1
            and receipt.get("cost_category_counts") == category_counts
        )
        checks.append(
            (
                "standalone_full_system_cost_categories",
                _canonical_json_text(category_counts),
                _canonical_json_text(
                    {
                        "full_initial_capex": 1,
                        "full_annual_om": 1,
                        "scheduled_replacement": category_counts[
                            "scheduled_replacement"
                        ],
                    }
                ),
                None,
                None,
                "OK" if required_categories_ok else "FAIL",
                "The immutable v4 receipt proves one full CAPEX and one full annual O&M category.",
            )
        )
        for line in receipt_lines:
            if not isinstance(line, Mapping):
                raise TechnoeconomicExportError(
                    "A standalone-commercial cost receipt line is invalid"
                )
            input_id = str(line.get("input_id"))
            sample_name = f"SampledInput::{input_id}"
            if sample_name not in calculation.by_name:
                raise TechnoeconomicExportError(
                    f"Standalone-commercial sampled input {input_id!r} is missing"
                )
            total_cost = np.asarray(
                calculation.by_name[sample_name],
                dtype=np.float64,
            ) * target_w
            timing = line.get("timing")
            if timing == "initial_t0":
                expected_initial += total_cost
            elif timing == "annual_year_end":
                expected_recurring += total_cost / crf
            elif timing == "scheduled_year_end":
                discount_sum = np.zeros(calculation.row_count, dtype=np.float64)
                for year in line.get("occurrence_years") or []:
                    discount_sum += np.exp(
                        -int(year) * np.log1p(discount_rate)
                    )
                expected_scheduled += total_cost * discount_sum
            else:
                raise TechnoeconomicExportError(
                    "Standalone-commercial cost timing receipt is invalid"
                )
            line_summary = line_summaries.get(input_id)
            if not isinstance(line_summary, Mapping):
                raise TechnoeconomicExportError(
                    f"Standalone-commercial summary for {input_id!r} is missing"
                )
            coverage_matches = (
                line_summary.get("cost_category") == line.get("cost_category")
                and list(line_summary.get("coverage_ids") or [])
                == list(line.get("coverage_ids") or [])
            )
            checks.append(
                (
                    f"standalone_cost_line_coverage::{input_id}",
                    _canonical_json_text(
                        {
                            "cost_category": line_summary.get("cost_category"),
                            "coverage_ids": list(
                                line_summary.get("coverage_ids") or []
                            ),
                        }
                    ),
                    _canonical_json_text(
                        {
                            "cost_category": line.get("cost_category"),
                            "coverage_ids": list(line.get("coverage_ids") or []),
                        }
                    ),
                    None,
                    None,
                    "OK" if coverage_matches else "FAIL",
                    "Sealed calculation summaries retain immutable category and coverage identity.",
                )
            )
            year_authorities = {
                "receipt": receipt.get("constant_dollar_cost_year"),
                "receipt_line": line.get("constant_dollar_cost_year"),
                "routine_result": standalone.get("constant_dollar_cost_year"),
                "sealed_summary": line_summary.get(
                    "constant_dollar_cost_year"
                ),
            }
            expected_year = receipt.get("constant_dollar_cost_year")
            year_matches = (
                isinstance(expected_year, int)
                and not isinstance(expected_year, bool)
                and all(value == expected_year for value in year_authorities.values())
            )
            checks.append(
                (
                    f"standalone_cost_line_constant_dollar_year::{input_id}",
                    _canonical_json_text(year_authorities),
                    _canonical_json_text(
                        {key: expected_year for key in year_authorities}
                    ),
                    None,
                    None,
                    "OK" if year_matches else "FAIL",
                    "The request, immutable receipt, durable result, and sealed "
                    "cost-line summary assert the same constant-dollar year.",
                )
            )
            expected_percentiles = np.quantile(
                total_cost,
                [0.10, 0.50, 0.90],
                method="linear",
            )
            recorded_percentiles = line_summary.get("percentiles") or {}
            for index, percentile in enumerate(("p10", "p50", "p90")):
                checks.append(
                    _numeric_check(
                        f"standalone_cost_line_percentile::{input_id}::{percentile}",
                        float(expected_percentiles[index]),
                        float(recorded_percentiles.get(percentile)),
                        tolerance=1e-12
                        * max(1.0, abs(float(expected_percentiles[index]))),
                        notes="Commercial cost-line percentile is recomputed from sampled target totals.",
                    )
                )

        cost_pairs = (
            (
                "standalone_initial_cost_stack",
                technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_INITIAL_COST,
                expected_initial,
            ),
            (
                "standalone_recurring_cost_stack",
                technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_RECURRING_PV_COST,
                expected_recurring,
            ),
            (
                "standalone_scheduled_cost_stack",
                technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_SCHEDULED_PV_COST,
                expected_scheduled,
            ),
        )
        for check_id, field_name, expected_values in cost_pairs:
            actual = np.asarray(calculation.by_name[field_name], dtype=np.float64)
            checks.append(
                _numeric_check(
                    check_id,
                    float(np.max(np.abs(actual - expected_values))),
                    0.0,
                    tolerance=_binary64_tie_out_tolerance(actual),
                    notes="Commercial lifecycle cost component ties to its sampled inputs and timing.",
                )
            )
        lifecycle_cost = np.asarray(
            calculation.by_name[
                technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LIFECYCLE_COST
            ],
            dtype=np.float64,
        )
        equivalent_cost = np.asarray(
            calculation.by_name[
                technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_EA_COST
            ],
            dtype=np.float64,
        )
        component_total = expected_initial + expected_recurring + expected_scheduled
        checks.extend(
            [
                _numeric_check(
                    "standalone_lifecycle_cost_component_sum",
                    float(np.max(np.abs(lifecycle_cost - component_total))),
                    0.0,
                    tolerance=_binary64_tie_out_tolerance(lifecycle_cost),
                    notes="Lifecycle cost equals initial, recurring, and scheduled present-value components.",
                ),
                _numeric_check(
                    "standalone_equivalent_annual_cost_crf",
                    float(np.max(np.abs(equivalent_cost - crf * lifecycle_cost))),
                    0.0,
                    tolerance=_binary64_tie_out_tolerance(equivalent_cost),
                    notes="Equivalent-annual cost equals CRF times lifecycle cost.",
                ),
            ]
        )
        lifecycle_energy = np.asarray(
            calculation.by_name[
                technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LIFECYCLE_ENERGY
            ],
            dtype=np.float64,
        )
        equivalent_energy = np.asarray(
            calculation.by_name[
                technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_EA_ENERGY
            ],
            dtype=np.float64,
        )
        lcoe = np.asarray(
            calculation.by_name[
                technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LCOE
            ],
            dtype=np.float64,
        )
        checks.extend(
            [
                _numeric_check(
                    "standalone_lcoe_lifecycle_ratio",
                    float(np.max(np.abs(lcoe - lifecycle_cost / lifecycle_energy))),
                    0.0,
                    tolerance=_binary64_tie_out_tolerance(lcoe),
                    notes="Standalone SolarEdge LCOE equals lifecycle cost divided by lifecycle energy.",
                ),
                _numeric_check(
                    "standalone_lcoe_equivalent_annual_ratio",
                    float(np.max(np.abs(lcoe - equivalent_cost / equivalent_energy))),
                    0.0,
                    tolerance=_binary64_tie_out_tolerance(lcoe),
                    notes="Standalone SolarEdge LCOE also equals the equivalent-annual ratio.",
                ),
            ]
        )
    if (
        routine_result.get("calculation_contract_version")
        == technoeconomic_kernel.PAIRED_COMMERCIAL_CALCULATION_CONTRACT_VERSION
    ):
        paired = routine_result.get("paired_commercial") or {}
        receipt = submission_provenance.get("paired_commercial_receipt") or {}
        if not isinstance(paired, Mapping) or not isinstance(receipt, Mapping):
            raise TechnoeconomicExportError(
                "Paired-commercial result or immutable receipt is invalid"
            )
        paired_systems = paired.get("systems") or {}
        receipt_systems = receipt.get("systems") or {}
        if (
            not isinstance(paired_systems, Mapping)
            or not isinstance(receipt_systems, Mapping)
            or set(paired_systems) != {"solectria", "solaredge"}
            or set(receipt_systems) != {"solectria", "solaredge"}
        ):
            raise TechnoeconomicExportError(
                "Paired-commercial system authority is incomplete"
            )
        target_w = float(paired.get("target_capacity_w"))
        checks.extend(
            [
                _numeric_check(
                    "paired_target_capacity_receipt",
                    target_w,
                    float(receipt.get("target_capacity_w")),
                    tolerance=0.0,
                    notes="Canonical target watts equal the immutable v5 receipt.",
                ),
                (
                    "paired_target_rating_basis_receipt",
                    paired.get("target_rating_basis"),
                    receipt.get("target_rating_basis"),
                    None,
                    None,
                    (
                        "OK"
                        if paired.get("target_rating_basis")
                        == receipt.get("target_rating_basis")
                        else "FAIL"
                    ),
                    "Both commercial systems share one frozen target rating basis.",
                ),
                (
                    "paired_transfer_method_receipt",
                    paired.get("transfer_method"),
                    receipt.get("transfer_method"),
                    None,
                    None,
                    (
                        "OK"
                        if paired.get("transfer_method")
                        == receipt.get("transfer_method")
                        else "FAIL"
                    ),
                    "The durable result retains the approved direct-scaling method.",
                ),
            ]
        )
        discount_column = "SampledInput::finance.discount-rate"
        if discount_column not in calculation.by_name:
            raise TechnoeconomicExportError(
                "Paired-commercial discount-rate samples are missing"
            )
        discount_rate = np.asarray(
            calculation.by_name[discount_column],
            dtype=np.float64,
        )
        annuity_factor = np.asarray(
            calculation.by_name["AnnuityFactor_years"],
            dtype=np.float64,
        )
        crf = np.asarray(
            calculation.by_name["CapitalRecoveryFactor_per_year"],
            dtype=np.float64,
        )
        system_field_contracts = {
            "solectria": {
                "target": technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_TARGET_CAPACITY,
                "scale": technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_CAPACITY_SCALE_FACTOR,
                "year1": technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_YEAR1_ENERGY,
                "lifecycle_energy": technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_LIFECYCLE_ENERGY,
                "ea_energy": technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_EA_ENERGY,
                "initial": technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_INITIAL_COST,
                "recurring": technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_RECURRING_PV_COST,
                "scheduled": technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_SCHEDULED_PV_COST,
                "lifecycle_cost": technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_LIFECYCLE_COST,
                "ea_cost": technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_EA_COST,
                "lcoe": technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_LCOE,
                "normalized_year1": fields.year1_sol,
                "normalized_lifecycle_energy": fields.pv_energy_sol,
                "normalized_ea_energy": fields.ea_energy_sol,
            },
            "solaredge": {
                "target": technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_TARGET_CAPACITY,
                "scale": technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_CAPACITY_SCALE_FACTOR,
                "year1": technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_YEAR1_ENERGY,
                "lifecycle_energy": technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LIFECYCLE_ENERGY,
                "ea_energy": technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_EA_ENERGY,
                "initial": technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_INITIAL_COST,
                "recurring": technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_RECURRING_PV_COST,
                "scheduled": technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_SCHEDULED_PV_COST,
                "lifecycle_cost": technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LIFECYCLE_COST,
                "ea_cost": technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_EA_COST,
                "lcoe": technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LCOE,
                "normalized_year1": fields.year1_se,
                "normalized_lifecycle_energy": fields.pv_energy_se,
                "normalized_ea_energy": fields.ea_energy_se,
            },
        }
        for technology in ("solectria", "solaredge"):
            system = paired_systems[technology]
            system_receipt = receipt_systems[technology]
            if not isinstance(system, Mapping) or not isinstance(
                system_receipt, Mapping
            ):
                raise TechnoeconomicExportError(
                    "Paired-commercial system authority is invalid"
                )
            source_receipt = system_receipt.get("source_capacity") or {}
            if not isinstance(source_receipt, Mapping):
                raise TechnoeconomicExportError(
                    "Paired-commercial source-capacity receipt is invalid"
                )
            source_w = float(system.get("source_applied_capacity_w"))
            scale_factor = target_w / source_w
            checks.extend(
                [
                    _numeric_check(
                        f"paired_{technology}_source_capacity_receipt",
                        source_w,
                        float(source_receipt.get("applied_capacity_w")),
                        tolerance=0.0,
                        notes="Source watts equal the frozen Annual-capacity receipt.",
                    ),
                    _numeric_check(
                        f"paired_{technology}_capacity_scale_factor",
                        float(system.get("capacity_scale_factor")),
                        scale_factor,
                        tolerance=0.0,
                        notes="Scale factor equals common target W divided by source W.",
                    ),
                    (
                        f"paired_{technology}_rating_basis_bridge",
                        system.get("source_rating_basis"),
                        paired.get("target_rating_basis"),
                        None,
                        None,
                        (
                            "OK"
                            if system.get("source_rating_basis")
                            == source_receipt.get("rating_basis")
                            == paired.get("target_rating_basis")
                            else "FAIL"
                        ),
                        "Source and target use the same frozen rating basis.",
                    ),
                ]
            )
            names = system_field_contracts[technology]
            missing = [
                name
                for name in names.values()
                if name not in calculation.by_name
            ]
            if missing:
                raise TechnoeconomicExportError(
                    f"Paired-commercial {technology} realization fields are "
                    f"missing: {missing!r}"
                )
            target_values = np.asarray(
                calculation.by_name[names["target"]], dtype=np.float64
            )
            scale_values = np.asarray(
                calculation.by_name[names["scale"]], dtype=np.float64
            )
            checks.extend(
                [
                    _numeric_check(
                        f"paired_{technology}_target_capacity_realizations",
                        float(np.max(np.abs(target_values - target_w))),
                        0.0,
                        tolerance=0.0,
                        notes="Every realization uses the common target capacity.",
                    ),
                    _numeric_check(
                        f"paired_{technology}_scale_realizations",
                        float(np.max(np.abs(scale_values - scale_factor))),
                        0.0,
                        tolerance=0.0,
                        notes="Every realization retains its system-specific capacity bridge.",
                    ),
                ]
            )
            for suffix, actual_name, normalized_name in (
                ("year1", names["year1"], names["normalized_year1"]),
                (
                    "lifecycle",
                    names["lifecycle_energy"],
                    names["normalized_lifecycle_energy"],
                ),
                ("equivalent_annual", names["ea_energy"], names["normalized_ea_energy"]),
            ):
                actual = np.asarray(
                    calculation.by_name[actual_name], dtype=np.float64
                )
                normalized = np.asarray(
                    calculation.by_name[normalized_name], dtype=np.float64
                )
                checks.append(
                    _numeric_check(
                        f"paired_{technology}_{suffix}_energy_scaling",
                        float(np.max(np.abs(actual - normalized * target_w))),
                        0.0,
                        tolerance=_binary64_tie_out_tolerance(actual),
                        notes="Commercial energy equals source-normalized energy times target watts.",
                    )
                )

            expected_initial = np.zeros(calculation.row_count, dtype=np.float64)
            expected_recurring = np.zeros(calculation.row_count, dtype=np.float64)
            expected_scheduled = np.zeros(calculation.row_count, dtype=np.float64)
            receipt_lines = system_receipt.get("cost_lines") or []
            if not isinstance(receipt_lines, list) or not receipt_lines:
                raise TechnoeconomicExportError(
                    f"Paired-commercial {technology} cost receipt is incomplete"
                )
            line_summaries = {
                str(item.get("input_id")): item
                for item in system.get("commercial_cost_line_summaries") or []
                if isinstance(item, Mapping)
            }
            for line in receipt_lines:
                if not isinstance(line, Mapping):
                    raise TechnoeconomicExportError(
                        "A paired-commercial cost receipt line is invalid"
                    )
                input_id = str(line.get("input_id"))
                sample_name = f"SampledInput::{input_id}"
                if sample_name not in calculation.by_name:
                    raise TechnoeconomicExportError(
                        f"Paired-commercial samples for {input_id!r} are missing"
                    )
                total_cost = (
                    np.asarray(calculation.by_name[sample_name], dtype=np.float64)
                    * target_w
                )
                timing = line.get("timing")
                if timing == "initial_t0":
                    expected_initial += total_cost
                elif timing == "annual_year_end":
                    expected_recurring += total_cost * annuity_factor
                elif timing == "scheduled_year_end":
                    discount_sum = np.zeros(calculation.row_count, dtype=np.float64)
                    for year in line.get("occurrence_years") or []:
                        discount_sum += np.exp(-int(year) * np.log1p(discount_rate))
                    expected_scheduled += total_cost * discount_sum
                else:
                    raise TechnoeconomicExportError(
                        "Paired-commercial cost timing receipt is invalid"
                    )
                line_summary = line_summaries.get(input_id)
                if not isinstance(line_summary, Mapping):
                    raise TechnoeconomicExportError(
                        f"Paired-commercial summary for {input_id!r} is missing"
                    )
                identity_matches = (
                    line_summary.get("technology") == technology
                    and line_summary.get("cost_category")
                    == line.get("cost_category")
                    and list(line_summary.get("coverage_ids") or [])
                    == list(line.get("coverage_ids") or [])
                    and line_summary.get("constant_dollar_cost_year")
                    == line.get("constant_dollar_cost_year")
                    == paired.get("constant_dollar_cost_year")
                    == receipt.get("constant_dollar_cost_year")
                )
                checks.append(
                    (
                        f"paired_{technology}_cost_line_identity::{input_id}",
                        _canonical_json_text(
                            {
                                "technology": line_summary.get("technology"),
                                "cost_category": line_summary.get("cost_category"),
                                "coverage_ids": list(
                                    line_summary.get("coverage_ids") or []
                                ),
                                "constant_dollar_cost_year": line_summary.get(
                                    "constant_dollar_cost_year"
                                ),
                            }
                        ),
                        _canonical_json_text(
                            {
                                "technology": technology,
                                "cost_category": line.get("cost_category"),
                                "coverage_ids": list(line.get("coverage_ids") or []),
                                "constant_dollar_cost_year": receipt.get(
                                    "constant_dollar_cost_year"
                                ),
                            }
                        ),
                        None,
                        None,
                        "OK" if identity_matches else "FAIL",
                        "Cost-line system, coverage, category, and dollar year retain immutable identity.",
                    )
                )
                expected_percentiles = np.quantile(
                    total_cost,
                    [0.10, 0.50, 0.90],
                    method="linear",
                )
                recorded_percentiles = line_summary.get("percentiles") or {}
                for index, percentile in enumerate(("p10", "p50", "p90")):
                    checks.append(
                        _numeric_check(
                            f"paired_{technology}_cost_line_percentile::{input_id}::{percentile}",
                            float(expected_percentiles[index]),
                            float(recorded_percentiles.get(percentile)),
                            tolerance=1e-12
                            * max(1.0, abs(float(expected_percentiles[index]))),
                            notes="Cost-line percentile is recomputed from sampled target totals.",
                        )
                    )
            for component, field_name, expected_values in (
                ("initial", names["initial"], expected_initial),
                ("recurring", names["recurring"], expected_recurring),
                ("scheduled", names["scheduled"], expected_scheduled),
            ):
                actual = np.asarray(
                    calculation.by_name[field_name], dtype=np.float64
                )
                checks.append(
                    _numeric_check(
                        f"paired_{technology}_{component}_cost_stack",
                        float(np.max(np.abs(actual - expected_values))),
                        0.0,
                        tolerance=_binary64_tie_out_tolerance(actual),
                        notes="Cost component ties to sampled inputs and timing.",
                    )
                )
            lifecycle_cost = np.asarray(
                calculation.by_name[names["lifecycle_cost"]], dtype=np.float64
            )
            equivalent_cost = np.asarray(
                calculation.by_name[names["ea_cost"]], dtype=np.float64
            )
            lifecycle_energy = np.asarray(
                calculation.by_name[names["lifecycle_energy"]], dtype=np.float64
            )
            equivalent_energy = np.asarray(
                calculation.by_name[names["ea_energy"]], dtype=np.float64
            )
            lcoe = np.asarray(calculation.by_name[names["lcoe"]], dtype=np.float64)
            component_total = expected_initial + expected_recurring + expected_scheduled
            checks.extend(
                [
                    _numeric_check(
                        f"paired_{technology}_lifecycle_cost_component_sum",
                        float(np.max(np.abs(lifecycle_cost - component_total))),
                        0.0,
                        tolerance=_binary64_tie_out_tolerance(lifecycle_cost),
                        notes="Lifecycle cost equals initial, recurring, and scheduled present values.",
                    ),
                    _numeric_check(
                        f"paired_{technology}_equivalent_annual_cost_crf",
                        float(np.max(np.abs(equivalent_cost - crf * lifecycle_cost))),
                        0.0,
                        tolerance=_binary64_tie_out_tolerance(equivalent_cost),
                        notes="Equivalent-annual cost equals CRF times lifecycle cost.",
                    ),
                    _numeric_check(
                        f"paired_{technology}_lcoe_lifecycle_ratio",
                        float(np.max(np.abs(lcoe - lifecycle_cost / lifecycle_energy))),
                        0.0,
                        tolerance=_binary64_tie_out_tolerance(lcoe),
                        notes="Lifecycle LCOE equals lifecycle cost divided by lifecycle energy.",
                    ),
                    _numeric_check(
                        f"paired_{technology}_lcoe_equivalent_annual_ratio",
                        float(np.max(np.abs(lcoe - equivalent_cost / equivalent_energy))),
                        0.0,
                        tolerance=_binary64_tie_out_tolerance(lcoe),
                        notes="Lifecycle LCOE also equals the equivalent-annual ratio.",
                    ),
                ]
            )
        solectria_lcoe = np.asarray(
            calculation.by_name[
                technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_LCOE
            ],
            dtype=np.float64,
        )
        solaredge_lcoe = np.asarray(
            calculation.by_name[technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LCOE],
            dtype=np.float64,
        )
        lcoe_delta = np.asarray(
            calculation.by_name[technoeconomic_kernel.COMMERCIAL_PAIRED_FIELD_LCOE_DELTA],
            dtype=np.float64,
        )
        checks.append(
            _numeric_check(
                "paired_lcoe_delta_se_minus_sol",
                float(np.max(np.abs(lcoe_delta - (solaredge_lcoe - solectria_lcoe)))),
                0.0,
                tolerance=_binary64_tie_out_tolerance(lcoe_delta),
                notes="The comparison diagnostic is SolarEdge LCOE minus Solectria LCOE per realization.",
            )
        )

        energy_rows = {
            int(row.get("year")): row
            for row in source_snapshot.get("eligible_paired_energy_rows") or []
            if isinstance(row, Mapping)
        }
        per_year_rows = calculation.metadata.get("per_weather_year") or []
        assigned_count = 0
        for record in per_year_rows:
            if not isinstance(record, Mapping):
                raise TechnoeconomicExportError(
                    "Paired-commercial per-year summary is invalid"
                )
            year = int(record.get("year"))
            assigned_count += int(record.get("realization_count") or 0)
            source_row = energy_rows.get(year)
            systems = record.get("systems") or {}
            if source_row is None or not isinstance(systems, Mapping):
                raise TechnoeconomicExportError(
                    "Paired-commercial per-year source authority is incomplete"
                )
            checks.extend(
                [
                    _numeric_check(
                        f"paired_per_year_{year}_target_capacity",
                        float(record.get("commercial_target_capacity_w")),
                        target_w,
                        tolerance=0.0,
                        notes="Per-year rows retain the common commercial target.",
                    ),
                    (
                        f"paired_per_year_{year}_target_rating_basis",
                        record.get("commercial_target_rating_basis"),
                        paired.get("target_rating_basis"),
                        None,
                        None,
                        (
                            "OK"
                            if record.get("commercial_target_rating_basis")
                            == paired.get("target_rating_basis")
                            else "FAIL"
                        ),
                        "Per-year rows retain the common target rating basis.",
                    ),
                    (
                        f"paired_per_year_{year}_transfer_method",
                        record.get("commercial_transfer_method"),
                        paired.get("transfer_method"),
                        None,
                        None,
                        (
                            "OK"
                            if record.get("commercial_transfer_method")
                            == paired.get("transfer_method")
                            else "FAIL"
                        ),
                        "Per-year rows retain the approved transfer method.",
                    ),
                ]
            )
            for technology, source_key in (
                ("solectria", "sol_predicted_kwh"),
                ("solaredge", "se_predicted_kwh"),
            ):
                system = systems.get(technology) or {}
                source_w = float(
                    paired_systems[technology].get("source_applied_capacity_w")
                )
                source_energy = float(source_row.get(source_key))
                checks.extend(
                    [
                        _numeric_check(
                            f"paired_per_year_{year}_{technology}_source_capacity",
                            float(system.get("source_applied_capacity_w")),
                            source_w,
                            tolerance=0.0,
                            notes="Per-year source capacity equals the frozen system authority.",
                        ),
                        _numeric_check(
                            f"paired_per_year_{year}_{technology}_scale_factor",
                            float(
                                system.get(
                                    "capacity_scale_factor_target_w_per_source_w"
                                )
                            ),
                            target_w / source_w,
                            tolerance=0.0,
                            notes="Per-year scale factor is target W divided by source W.",
                        ),
                        (
                            f"paired_per_year_{year}_{technology}_rating_basis",
                            system.get("source_rating_basis"),
                            paired_systems[technology].get("source_rating_basis"),
                            None,
                            None,
                            (
                                "OK"
                                if system.get("source_rating_basis")
                                == paired_systems[technology].get(
                                    "source_rating_basis"
                                )
                                else "FAIL"
                            ),
                            "Per-year source rating basis equals the frozen bridge.",
                        ),
                        _numeric_check(
                            f"paired_per_year_{year}_{technology}_source_energy",
                            float(system.get("source_predicted_kwh_ac")),
                            source_energy,
                            tolerance=0.0,
                            notes="Per-year source energy equals the frozen paired row.",
                        ),
                        _numeric_check(
                            f"paired_per_year_{year}_{technology}_specific_energy",
                            float(
                                system.get(
                                    "source_specific_kwh_ac_per_applied_w_year"
                                )
                            ),
                            source_energy / source_w,
                            tolerance=0.0,
                            notes="Per-year specific energy uses the system's source capacity.",
                        ),
                        _numeric_check(
                            f"paired_per_year_{year}_{technology}_target_energy",
                            float(system.get("target_year1_energy_kwh_ac")),
                            source_energy / source_w * target_w,
                            tolerance=0.0,
                            notes="Per-year target energy uses direct system-specific capacity scaling.",
                        ),
                    ]
                )
        checks.append(
            _numeric_check(
                "paired_per_year_realization_partition",
                assigned_count,
                calculation.row_count,
                tolerance=0.0,
                notes="Per-year realization counts partition the paired sample exactly once.",
            )
        )
    return checks


STANDALONE_COMMERCIAL_SUMMARY_COLUMNS = (
    "record_type",
    "technology",
    "input_id",
    "label",
    "cost_category",
    "coverage_ids_json",
    "timing",
    "input_unit",
    "occurrence_years_json",
    "target_capacity_w",
    "target_rating_basis",
    "source_applied_capacity_w",
    "source_rating_basis",
    "capacity_scale_factor",
    "transfer_method",
    "constant_dollar_cost_year",
    "metric_id",
    "metric_unit",
    "p10",
    "p50",
    "p90",
    "evidence_subject",
)


def _standalone_commercial_summary_rows(
    request_payload: Mapping[str, Any],
    submission_provenance: Mapping[str, Any],
    routine_result: Mapping[str, Any],
) -> Iterator[tuple[Any, ...]]:
    request = request_payload.get("standalone_commercial") or {}
    receipt = submission_provenance.get("standalone_commercial_receipt") or {}
    result = routine_result.get("standalone_commercial") or {}
    if not all(isinstance(item, Mapping) for item in (request, receipt, result)):
        raise TechnoeconomicExportError(
            "Standalone-commercial summary authority is invalid"
        )
    percentiles = result.get("percentiles") or {}
    yield (
        "headline",
        result.get("technology"),
        None,
        "Commercial SolarEdge lifecycle LCOE",
        None,
        "[]",
        None,
        None,
        "[]",
        result.get("target_capacity_w"),
        result.get("target_rating_basis"),
        result.get("source_applied_capacity_w"),
        result.get("source_rating_basis"),
        result.get("capacity_scale_factor"),
        result.get("transfer_method"),
        receipt.get("constant_dollar_cost_year"),
        result.get("headline_metric_id"),
        result.get("unit"),
        percentiles.get("p10"),
        percentiles.get("p50"),
        percentiles.get("p90"),
        receipt.get("energy_scaling_evidence_subject"),
    )
    request_lines = {
        str(line.get("input_id")): line
        for line in request.get("cost_lines") or []
        if isinstance(line, Mapping)
    }
    for summary in result.get("commercial_cost_line_summaries") or []:
        if not isinstance(summary, Mapping):
            raise TechnoeconomicExportError(
                "Standalone-commercial cost-line summary is invalid"
            )
        input_id = str(summary.get("input_id"))
        line = request_lines.get(input_id) or {}
        line_percentiles = summary.get("percentiles") or {}
        yield (
            "cost_line",
            result.get("technology"),
            input_id,
            summary.get("label"),
            summary.get("cost_category"),
            _canonical_json_text(list(summary.get("coverage_ids") or [])),
            summary.get("timing"),
            line.get("unit"),
            _canonical_json_text(list(summary.get("occurrence_years") or [])),
            result.get("target_capacity_w"),
            result.get("target_rating_basis"),
            result.get("source_applied_capacity_w"),
            result.get("source_rating_basis"),
            result.get("capacity_scale_factor"),
            result.get("transfer_method"),
            line.get("constant_dollar_cost_year"),
            None,
            summary.get("total_unit"),
            line_percentiles.get("p10"),
            line_percentiles.get("p50"),
            line_percentiles.get("p90"),
            f"standalone-commercial-cost:{input_id}",
        )


PAIRED_COMMERCIAL_SUMMARY_COLUMNS = STANDALONE_COMMERCIAL_SUMMARY_COLUMNS


def _paired_commercial_summary_rows(
    request_payload: Mapping[str, Any],
    submission_provenance: Mapping[str, Any],
    routine_result: Mapping[str, Any],
) -> Iterator[tuple[Any, ...]]:
    request = request_payload.get("paired_commercial") or {}
    receipt = submission_provenance.get("paired_commercial_receipt") or {}
    result = routine_result.get("paired_commercial") or {}
    if not all(isinstance(item, Mapping) for item in (request, receipt, result)):
        raise TechnoeconomicExportError(
            "Paired-commercial summary authority is invalid"
        )
    request_systems = {
        str(system.get("technology")): system
        for system in request.get("systems") or []
        if isinstance(system, Mapping)
    }
    result_systems = result.get("systems") or {}
    receipt_systems = receipt.get("systems") or {}
    if set(request_systems) != {"solectria", "solaredge"} or not isinstance(
        result_systems, Mapping
    ) or set(result_systems) != {"solectria", "solaredge"} or not isinstance(
        receipt_systems, Mapping
    ) or set(receipt_systems) != {"solectria", "solaredge"}:
        raise TechnoeconomicExportError(
            "Paired-commercial summary systems are incomplete"
        )
    for technology in ("solectria", "solaredge"):
        system_result = result_systems[technology]
        system_receipt = receipt_systems[technology]
        if not isinstance(system_result, Mapping) or not isinstance(
            system_receipt, Mapping
        ):
            raise TechnoeconomicExportError(
                "Paired-commercial summary system is invalid"
            )
        percentiles = system_result.get("percentiles") or {}
        label = "Solectria" if technology == "solectria" else "SolarEdge"
        yield (
            "headline",
            technology,
            None,
            f"Commercial {label} lifecycle LCOE",
            None,
            "[]",
            None,
            None,
            "[]",
            result.get("target_capacity_w"),
            result.get("target_rating_basis"),
            system_result.get("source_applied_capacity_w"),
            system_result.get("source_rating_basis"),
            system_result.get("capacity_scale_factor"),
            result.get("transfer_method"),
            result.get("constant_dollar_cost_year"),
            system_result.get("headline_metric_id"),
            system_result.get("unit"),
            percentiles.get("p10"),
            percentiles.get("p50"),
            percentiles.get("p90"),
            system_receipt.get("system_evidence_subject"),
        )
        request_lines = {
            str(line.get("input_id")): line
            for line in request_systems[technology].get("cost_lines") or []
            if isinstance(line, Mapping)
        }
        receipt_lines = {
            str(line.get("input_id")): line
            for line in system_receipt.get("cost_lines") or []
            if isinstance(line, Mapping)
        }
        for summary in system_result.get("commercial_cost_line_summaries") or []:
            if not isinstance(summary, Mapping):
                raise TechnoeconomicExportError(
                    "Paired-commercial cost-line summary is invalid"
                )
            input_id = str(summary.get("input_id"))
            line = request_lines.get(input_id) or {}
            line_percentiles = summary.get("percentiles") or {}
            yield (
                "cost_line",
                technology,
                input_id,
                summary.get("label"),
                summary.get("cost_category"),
                _canonical_json_text(list(summary.get("coverage_ids") or [])),
                summary.get("timing"),
                line.get("unit"),
                _canonical_json_text(list(summary.get("occurrence_years") or [])),
                result.get("target_capacity_w"),
                result.get("target_rating_basis"),
                system_result.get("source_applied_capacity_w"),
                system_result.get("source_rating_basis"),
                system_result.get("capacity_scale_factor"),
                result.get("transfer_method"),
                line.get("constant_dollar_cost_year"),
                None,
                summary.get("total_unit"),
                line_percentiles.get("p10"),
                line_percentiles.get("p50"),
                line_percentiles.get("p90"),
                (receipt_lines.get(input_id) or {}).get("evidence_subject"),
            )
    delta = result.get("lcoe_delta_se_minus_sol") or {}
    if not isinstance(delta, Mapping):
        raise TechnoeconomicExportError(
            "Paired-commercial LCOE-delta summary is invalid"
        )
    percentiles = delta.get("percentiles") or {}
    yield (
        "diagnostic",
        "se_minus_sol",
        None,
        "Commercial lifecycle LCOE delta, SolarEdge minus Solectria",
        None,
        "[]",
        None,
        None,
        "[]",
        result.get("target_capacity_w"),
        result.get("target_rating_basis"),
        None,
        None,
        None,
        result.get("transfer_method"),
        result.get("constant_dollar_cost_year"),
        delta.get("headline_metric_id"),
        delta.get("unit"),
        percentiles.get("p10"),
        percentiles.get("p50"),
        percentiles.get("p90"),
        "paired-commercial:lcoe-delta",
    )


def _lifecycle_request(request_payload: Mapping[str, Any]) -> Mapping[str, Any]:
    lifecycle = request_payload.get("paired_lifecycle")
    if isinstance(lifecycle, Mapping):
        return lifecycle
    paired = request_payload.get("paired_commercial")
    paired = paired if isinstance(paired, Mapping) else {}
    nested = paired.get("lifecycle")
    if not isinstance(nested, Mapping):
        raise TechnoeconomicExportError(
            "The v6 export requires paired_commercial.lifecycle"
        )
    unit = str(paired.get("target_capacity_unit") or "w").lower()
    multipliers = {"w": 1.0, "kw": 1_000.0, "mw": 1_000_000.0}
    if unit not in multipliers:
        raise TechnoeconomicExportError("The v6 target capacity unit is invalid")
    normalized = dict(nested)
    normalized["target_capacity_w"] = (
        float(paired.get("target_capacity") or 0.0) * multipliers[unit]
    )
    normalized["target_rating_basis"] = paired.get("target_rating_basis")
    normalized["npv_absolute_tolerance_usd_per_w"] = nested.get(
        "decision_npv_tolerance_usd_per_target_w"
    )
    return normalized


def _lifecycle_input_records(
    request_payload: Mapping[str, Any],
) -> tuple[Mapping[str, Any], ...]:
    records: list[Mapping[str, Any]] = []
    for section, value in (
        ("finance", request_payload.get("finance")),
        ("paired_lifecycle", _lifecycle_request(request_payload)),
    ):
        for field_path, encoded in _flatten_leaves(
            "", value if value is not None else {}
        ):
            records.append(
                {
                    "section": section,
                    "field_path": field_path,
                    "value_json": encoded,
                }
            )
    return tuple(records)


def _lifecycle_target_design_records(
    request_payload: Mapping[str, Any],
) -> tuple[Mapping[str, Any], ...]:
    lifecycle = _lifecycle_request(request_payload)
    systems = lifecycle.get("systems")
    if not isinstance(systems, list):
        raise TechnoeconomicExportError("The v6 target systems are invalid")
    records: list[Mapping[str, Any]] = []
    for system in systems:
        if not isinstance(system, Mapping):
            raise TechnoeconomicExportError("A v6 target system is invalid")
        records.append(
            {
                "technology": system.get("technology"),
                "target_capacity_w": lifecycle.get("target_capacity_w"),
                "target_rating_basis": lifecycle.get("target_rating_basis"),
                "source_energy_basis": lifecycle.get("source_energy_basis"),
                "reliability_mode": lifecycle.get("reliability_mode"),
                "degradation_json": system.get("degradation"),
                "base_availability_json": system.get("base_availability"),
                "base_om_cost_per_w_year_json": system.get(
                    "base_om_cost_per_w_year"
                ),
                "base_om_real_growth_json": system.get("base_om_real_growth"),
                "initial_cost_lines_json": system.get("initial_cost_lines") or [],
                "scheduled_costs_json": system.get("scheduled_costs") or [],
                "decommissioning_cost_json": system.get("decommissioning_cost"),
                "salvage_value_json": system.get("salvage_value"),
                "base_om_coverage_ids_json": system.get("base_om_coverage_ids") or [],
                "evidence_json": system.get("evidence") or {},
            }
        )
    return tuple(records)


def _lifecycle_reliability_input_records(
    request_payload: Mapping[str, Any],
) -> tuple[Mapping[str, Any], ...]:
    lifecycle = _lifecycle_request(request_payload)
    records: list[Mapping[str, Any]] = []
    systems = lifecycle.get("systems")
    if not isinstance(systems, list):
        raise TechnoeconomicExportError("The v6 target systems are invalid")
    for system in systems:
        if not isinstance(system, Mapping):
            raise TechnoeconomicExportError("A v6 target system is invalid")
        components = system.get("components") or []
        if not isinstance(components, list):
            raise TechnoeconomicExportError("A v6 component list is invalid")
        for component in components:
            if not isinstance(component, Mapping):
                raise TechnoeconomicExportError("A v6 component input is invalid")
            record = {
                "record_type": "component",
                "technology": system.get("technology"),
            }
            record.update(
                {
                    str(key): _safe_public_value(value)
                    for key, value in component.items()
                }
            )
            records.append(record)
    events = lifecycle.get("common_cause_events") or []
    if not isinstance(events, list):
        raise TechnoeconomicExportError("The v6 common-cause inputs are invalid")
    for event in events:
        if not isinstance(event, Mapping):
            raise TechnoeconomicExportError("A v6 common-cause event is invalid")
        record = {"record_type": "common_cause", "technology": "shared"}
        record.update(
            {str(key): _safe_public_value(value) for key, value in event.items()}
        )
        records.append(record)
    return tuple(records)


def _lifecycle_weather_records(
    metadata: Mapping[str, Any],
) -> tuple[Mapping[str, Any], ...]:
    summaries = _lifecycle_summaries(metadata)
    if "per_weather_year" in metadata:
        return _mapping_sequence(
            metadata.get("per_weather_year"), label="The v6 weather summary"
        )
    for key in ("weather_summary", "annual_weather_summary"):
        if key in summaries:
            return _mapping_sequence(
                summaries.get(key), label="The v6 weather summary"
            )
    counts: dict[tuple[Any, Any], int] = {}
    seen: set[tuple[Any, Any, Any]] = set()
    for record in _representative_trace_records(metadata):
        if record.get("record_type") != "annual":
            continue
        identity = (
            record.get("realization_index"),
            record.get("project_year"),
            record.get("weather_year"),
        )
        if identity in seen:
            continue
        seen.add(identity)
        key = (record.get("project_year"), record.get("weather_year"))
        counts[key] = counts.get(key, 0) + 1
    return tuple(
        {
            "project_year": project_year,
            "weather_year": weather_year,
            "representative_path_count": count,
            "allocation_method": "seeded_balanced_iid_per_project_year",
            "paired_system_weather": True,
        }
        for (project_year, weather_year), count in sorted(counts.items())
    )


def _lifecycle_capacity_records(
    request_payload: Mapping[str, Any],
    submission_provenance: Mapping[str, Any],
) -> tuple[Mapping[str, Any], ...]:
    lifecycle = _lifecycle_request(request_payload)
    records: list[Mapping[str, Any]] = [
        {
            "record_type": "target",
            "technology": "paired",
            "capacity_w": lifecycle.get("target_capacity_w"),
            "rating_basis": lifecycle.get("target_rating_basis"),
            "source_field": "paired_lifecycle.target_capacity_w",
        }
    ]
    receipt = submission_provenance.get("normalization_receipt") or {}
    capacities = (
        receipt.get("applied_capacities") if isinstance(receipt, Mapping) else {}
    )
    if isinstance(capacities, Mapping):
        for technology, record in sorted(capacities.items()):
            if isinstance(record, Mapping):
                records.append(
                    {
                        "record_type": "source",
                        "technology": technology,
                        "capacity_w": record.get("applied_capacity_w"),
                        "rating_basis": record.get("rating_basis"),
                        "source_field": record.get("source_field"),
                    }
                )
    return tuple(records)


def _lifecycle_lcoe_records(
    calculation: _SealedCalculation,
) -> tuple[Mapping[str, Any], ...]:
    fields = (
        ("solectria", "LifecycleLCOE_SOL_USD_per_kWh_AC"),
        ("solaredge", "LifecycleLCOE_SE_USD_per_kWh_AC"),
        (
            "delta_se_minus_sol",
            "DeltaLifecycleLCOE_se_minus_sol_USD_per_kWh_AC",
        ),
        ("incremental", "IncrementalLCOO_se_minus_sol_USD_per_kWh_AC"),
    )
    records: list[Mapping[str, Any]] = []
    for technology, field_name in fields:
        if field_name not in calculation.by_name:
            continue
        values = _finite_values(calculation.by_name[field_name])
        percentiles = (
            np.quantile(values, (0.10, 0.50, 0.90), method="linear")
            if len(values)
            else (None, None, None)
        )
        records.append(
            {
                "technology": technology,
                "metric_id": field_name,
                "unit": "constant USD/kWh_AC",
                "population_count": len(values),
                "p10": percentiles[0],
                "p50": percentiles[1],
                "p90": percentiles[2],
            }
        )
    return tuple(records)


def _lifecycle_economic_npv_tolerance(
    request_payload: Mapping[str, Any],
) -> float:
    lifecycle = _lifecycle_request(request_payload)
    target_w = float(lifecycle.get("target_capacity_w") or 0.0)
    per_w = float(lifecycle.get("npv_absolute_tolerance_usd_per_w") or 0.0)
    return target_w * per_w


def _lifecycle_realization_value_map(
    calculation: _SealedCalculation,
    field_name: str,
) -> dict[int, float]:
    """Return one-based realization keys matching representative trace rows."""

    values = calculation.by_name.get(field_name)
    if values is None:
        return {}
    realization_values = calculation.by_name.get("Realization")
    if realization_values is None:
        realization_values = np.arange(calculation.row_count, dtype=np.int64)
    result: dict[int, float] = {}
    for raw_index, raw_value in zip(realization_values, values):
        if isinstance(raw_value, bool):
            continue
        try:
            value = float(raw_value)
            trace_index = int(raw_index) + 1
        except (TypeError, ValueError, OverflowError):
            continue
        if math.isfinite(value):
            result[trace_index] = value
    return result


def _registry_formula_id(
    registry: Sequence[Mapping[str, Any]],
    *needles: str,
) -> str:
    lowered_needles = tuple(needle.lower() for needle in needles)
    for record in registry:
        haystack = " ".join(
            str(record.get(field) or "")
            for field in ("name", "equation", "output")
        ).lower()
        if all(needle in haystack for needle in lowered_needles):
            return str(record["formula_id"])
    return str(registry[0]["formula_id"])


def _lifecycle_audit_records(
    trace_records: Sequence[Mapping[str, Any]],
    registry: Sequence[Mapping[str, Any]],
    *,
    economic_decision_tolerance: float,
    economic_decision_tolerances: Mapping[int, float] | None = None,
) -> tuple[Mapping[str, Any], ...]:
    column_index = {
        name: openpyxl.utils.get_column_letter(index)
        for index, name in enumerate(LIFECYCLE_TRACE_COLUMNS, start=1)
    }
    records: list[Mapping[str, Any]] = []

    def add(
        *,
        trace: Mapping[str, Any],
        formula_id: str,
        output_field: str,
        formula: str,
        replica: float,
        inputs: Sequence[float],
        suffix: str,
    ) -> None:
        frozen = trace.get(output_field)
        if isinstance(frozen, bool) or not isinstance(frozen, (int, float)):
            return
        frozen_number = float(frozen)
        scale_values = np.asarray([frozen_number, replica, *inputs], dtype=np.float64)
        tolerance = _binary64_tie_out_tolerance(scale_values)
        difference = replica - frozen_number
        raw_realization = trace.get("realization_index")
        try:
            realization_index = int(raw_realization)
        except (TypeError, ValueError, OverflowError):
            realization_index = -1
        decision_tolerance = float(
            (economic_decision_tolerances or {}).get(
                realization_index,
                economic_decision_tolerance,
            )
        )
        records.append(
            {
                "audit_id": (
                    f"{trace.get('selection_label')}::{trace.get('system')}::"
                    f"{trace.get('project_year')}::{suffix}::"
                    f"{trace.get('component_id') or ''}::"
                    f"{trace.get('cohort_age') if trace.get('cohort_age') is not None else ''}"
                ),
                "formula_id": formula_id,
                "selection_label": trace.get("selection_label"),
                "realization_index": trace.get("realization_index"),
                "system": trace.get("system"),
                "project_year": trace.get("project_year"),
                "component_id": trace.get("component_id"),
                "frozen_authority": frozen_number,
                "formula_replica": formula,
                "difference": difference,
                "binary64_tolerance": tolerance,
                "status": "OK" if abs(difference) <= tolerance else "FAIL",
                "economic_decision_tolerance": decision_tolerance,
                "notes": (
                    "Binary64 audit tolerance tests the replica tie-out only; "
                    "the economic tolerance is a separate decision materiality threshold."
                ),
            }
        )

    def cell(field: str, row: int) -> str:
        return f"'Representative Event Traces'!{column_index[field]}{row}"

    for trace_row, trace in enumerate(trace_records, start=2):
        if trace.get("record_type") == "annual":
            base = float(trace.get("base_availability") or 0.0)
            component = float(trace.get("component_availability") or 0.0)
            common = float(trace.get("common_cause_availability") or 0.0)
            add(
                trace=trace,
                formula_id=_registry_formula_id(registry, "target", "availability"),
                output_field="target_availability",
                formula=(
                    f"={cell('base_availability', trace_row)}*"
                    f"{cell('component_availability', trace_row)}*"
                    f"{cell('common_cause_availability', trace_row)}"
                ),
                replica=base * component * common,
                inputs=(base, component, common),
                suffix="target_availability",
            )
            target = float(trace.get("target_availability") or 0.0)
            source = trace.get("source_availability")
            source_number = (
                float(source)
                if isinstance(source, (int, float)) and not isinstance(source, bool)
                else None
            )
            adjustment = target if source_number is None else target / source_number
            source_inputs: tuple[float, ...] = (
                () if source_number is None else (source_number,)
            )
            add(
                trace=trace,
                formula_id=_registry_formula_id(registry, "availability", "source"),
                output_field="availability_adjustment",
                formula=(
                    f"=IF({cell('source_availability', trace_row)}=\"\","
                    f"{cell('target_availability', trace_row)},"
                    f"{cell('target_availability', trace_row)}/"
                    f"{cell('source_availability', trace_row)})"
                ),
                replica=adjustment,
                inputs=(target, *source_inputs),
                suffix="availability_adjustment",
            )
            target_energy = float(trace.get("target_source_energy_kwh") or 0.0)
            degradation = float(trace.get("degradation_factor") or 0.0)
            availability_adjustment = float(
                trace.get("availability_adjustment") or 0.0
            )
            add(
                trace=trace,
                formula_id=_registry_formula_id(registry, "delivered", "energy"),
                output_field="delivered_energy_kwh",
                formula=(
                    f"={cell('target_source_energy_kwh', trace_row)}*"
                    f"{cell('degradation_factor', trace_row)}*"
                    f"{cell('availability_adjustment', trace_row)}"
                ),
                replica=target_energy * degradation * availability_adjustment,
                inputs=(target_energy, degradation, availability_adjustment),
                suffix="delivered_energy",
            )
            cost_fields = (
                "base_om_cost_usd",
                "scheduled_cost_usd",
                "preventive_cost_usd",
                "corrective_cost_usd",
                "common_cause_cost_usd",
            )
            costs = tuple(float(trace.get(field) or 0.0) for field in cost_fields)
            add(
                trace=trace,
                formula_id=_registry_formula_id(registry, "annual", "cost"),
                output_field="annual_cost_usd",
                formula="=SUM("
                + ",".join(cell(field, trace_row) for field in cost_fields)
                + ")",
                replica=sum(costs),
                inputs=costs,
                suffix="annual_cost",
            )
            cashflow = float(trace.get("incremental_cashflow_usd") or 0.0)
            discount_factor = float(trace.get("discount_factor") or 0.0)
            add(
                trace=trace,
                formula_id=_registry_formula_id(registry, "upgrade", "npv"),
                output_field="pv_incremental_cashflow_usd",
                formula=(
                    f"={cell('incremental_cashflow_usd', trace_row)}*"
                    f"{cell('discount_factor', trace_row)}"
                ),
                replica=cashflow * discount_factor,
                inputs=(cashflow, discount_factor),
                suffix="pv_incremental_cashflow",
            )
        elif trace.get("record_type") == "component":
            start = float(
                trace.get("expected_start_count")
                if trace.get("expected_start_count") is not None
                else trace.get("start_count")
                or 0.0
            )
            probability = float(trace.get("annual_failure_probability") or 0.0)
            add(
                trace=trace,
                formula_id=_registry_formula_id(registry, "expected", "failure"),
                output_field="expected_failures",
                formula=(
                    f"={cell('expected_start_count', trace_row)}*"
                    f"{cell('annual_failure_probability', trace_row)}"
                ),
                replica=start * probability,
                inputs=(start, probability),
                suffix="expected_failures",
            )
            if trace.get("component_year_total_row") is not True:
                continue
            spares_start = float(trace.get("spares_start") or 0.0)
            stocked = float(trace.get("stocked_replacements") or 0.0)
            restock = float(trace.get("restock_quantity") or 0.0)
            add(
                trace=trace,
                formula_id=_registry_formula_id(registry, "spare"),
                output_field="spares_end",
                formula=(
                    f"={cell('spares_start', trace_row)}-"
                    f"{cell('stocked_replacements', trace_row)}+"
                    f"{cell('restock_quantity', trace_row)}"
                ),
                replica=spares_start - stocked + restock,
                inputs=(spares_start, stocked, restock),
                suffix="spares_end",
            )
            cost_fields = (
                "hardware_cost_usd",
                "labor_cost_usd",
                "mobilization_cost_usd",
            )
            costs = tuple(float(trace.get(field) or 0.0) for field in cost_fields)
            warranty = float(trace.get("warranty_credit_usd") or 0.0)
            add(
                trace=trace,
                formula_id=_registry_formula_id(registry, "corrective", "cost"),
                output_field="corrective_cost_usd",
                formula="=SUM("
                + ",".join(cell(field, trace_row) for field in cost_fields)
                + f")-{cell('warranty_credit_usd', trace_row)}",
                replica=sum(costs) - warranty,
                inputs=(*costs, warranty),
                suffix="corrective_cost",
            )
    return tuple(records)


def _build_lifecycle_checks(
    calculation: _SealedCalculation,
    routine_result: Mapping[str, Any],
    registry: Sequence[Mapping[str, Any]],
    trace_records: Sequence[Mapping[str, Any]],
    audit_records: Sequence[Mapping[str, Any]],
) -> list[tuple[Any, ...]]:
    expected_rows = routine_result.get("realization_count")
    if isinstance(expected_rows, bool) or not isinstance(expected_rows, int):
        raise TechnoeconomicExportError("Routine v6 realization count is invalid")
    checks = [
        _numeric_check(
            "realization_count",
            calculation.row_count,
            expected_rows,
            tolerance=0.0,
            notes="Sealed v6 realization rows equal the durable result count.",
        ),
        _numeric_check(
            "formula_registry_count",
            len(registry),
            int(
                (
                    (calculation.metadata.get("kernel_provenance") or {}).get(
                        "formula_registry"
                    )
                    or {}
                ).get("count")
                or len(registry)
            ),
            tolerance=0.0,
            notes="Every canonical formula-registry entry is exported.",
        ),
    ]
    required_fields = {
        "LifecyclePVCost_SOL_USD",
        "LifecyclePVCost_SE_USD",
        "LifecyclePVEnergy_SOL_kWh_AC",
        "LifecyclePVEnergy_SE_kWh_AC",
        "LifecycleLCOE_SOL_USD_per_kWh_AC",
        "LifecycleLCOE_SE_USD_per_kWh_AC",
        "DeltaLifecycleCost_se_minus_sol_USD",
        "DeltaLifecycleEnergy_se_minus_sol_kWh_AC",
        "DeltaLifecycleLCOE_se_minus_sol_USD_per_kWh_AC",
        "IncrementalLCOO_se_minus_sol_USD_per_kWh_AC",
        "UpgradeNPV_se_minus_sol_USD",
    }
    missing = sorted(required_fields - set(calculation.column_names))
    checks.append(
        (
            "required_v6_realization_fields",
            ",".join(missing),
            "",
            None,
            0.0,
            "OK" if not missing else "FAIL",
            "Required lifecycle decision fields are present in the sealed payload.",
        )
    )
    trace_labels = sorted(
        {
            str(record.get("selection_label"))
            for record in trace_records
            if record.get("selection_label") is not None
        }
    )
    checks.append(
        (
            "representative_npv_trace_labels",
            ",".join(trace_labels),
            "NPV-P10,NPV-P50,NPV-P90",
            None,
            0.0,
            "OK"
            if trace_labels == ["NPV-P10", "NPV-P50", "NPV-P90"]
            else "FAIL",
            "Representative traces use explicit Upgrade-NPV quantile labels.",
        )
    )
    failed_audits = [
        record for record in audit_records if record.get("status") != "OK"
    ]
    checks.append(
        _numeric_check(
            "representative_formula_audits",
            len(failed_audits),
            0,
            tolerance=0.0,
            notes=(
                "Representative formula replicas tie to frozen kernel values at "
                "the binary64 audit tolerance, not the economic decision tolerance."
            ),
        )
    )
    provenance_registry = (calculation.metadata.get("kernel_provenance") or {}).get(
        "formula_registry"
    )
    provenance_registry = (
        provenance_registry if isinstance(provenance_registry, Mapping) else {}
    )
    provenance_digest = provenance_registry.get("sha256")
    if provenance_digest is not None:
        actual_digest = _formula_registry_sha256(registry)
        checks.append(
            (
                "formula_registry_hash",
                actual_digest,
                provenance_digest,
                None,
                0.0,
                "OK"
                if secrets.compare_digest(str(actual_digest), str(provenance_digest))
                else "FAIL",
                "Kernel and exporter consume one canonical v6 formula registry.",
            )
        )
    return checks


def _verify_lifecycle_routine_result(
    *,
    metadata: Mapping[str, Any],
    routine_result: Mapping[str, Any],
    request_payload: Mapping[str, Any],
    source_snapshot: Mapping[str, Any],
    submission_provenance: Mapping[str, Any],
    sealed_calculation_artifact: Mapping[str, Any],
) -> None:
    """Bind the complete durable v6 projection to sealed export authority."""

    lifecycle_version = getattr(
        technoeconomic_kernel,
        "LIFECYCLE_CALCULATION_CONTRACT_VERSION",
        "tea-calculation-v6",
    )
    sampling_version = getattr(
        technoeconomic_kernel,
        "LIFECYCLE_SAMPLING_VERSION",
        "tea-lhs-v2",
    )
    result_version = getattr(
        technoeconomic_kernel,
        "LIFECYCLE_RESULT_VERSION",
        "tea-result-v6",
    )
    if (
        submission_provenance.get("calculation_contract_version")
        != lifecycle_version
        or submission_provenance.get("sampling_version") != sampling_version
    ):
        raise TechnoeconomicExportError(
            "Frozen v6 submission provenance has the wrong contract identity"
        )
    kernel_provenance = metadata.get("kernel_provenance")
    summaries = metadata.get("summaries")
    if not isinstance(kernel_provenance, Mapping) or not isinstance(
        summaries, Mapping
    ):
        raise TechnoeconomicExportError(
            "Sealed v6 result authority is incomplete"
        )
    if (
        kernel_provenance.get("calculation_contract_version")
        != lifecycle_version
        or kernel_provenance.get("sampling_version") != sampling_version
        or kernel_provenance.get("result_version") != result_version
        or summaries.get("result_version") != result_version
    ):
        raise TechnoeconomicExportError(
            "Sealed v6 result authority has the wrong contract identity"
        )

    finance = request_payload.get("finance")
    finance = finance if isinstance(finance, Mapping) else {}
    manifest = source_snapshot.get("capacity_manifest")
    manifest = manifest if isinstance(manifest, Mapping) else {}
    manifest_systems = manifest.get("systems")
    if not isinstance(manifest_systems, Mapping):
        raise TechnoeconomicExportError("Frozen v6 capacity manifest is invalid")
    capacities: dict[str, Any] = {}
    for system, record in sorted(manifest_systems.items()):
        if not isinstance(record, Mapping):
            raise TechnoeconomicExportError("Frozen v6 capacity record is invalid")
        capacities[str(system)] = {
            "module_model": record.get("module_model"),
            "installed_wdc": record.get("installed_wdc"),
            "physics_version": record.get("calibration_physics_version"),
            "physics_fingerprint": record.get(
                "calibration_physics_fingerprint"
            ),
        }
    eligible_rows = source_snapshot.get("eligible_paired_energy_rows")
    if not isinstance(eligible_rows, list):
        raise TechnoeconomicExportError("Frozen v6 eligible energy rows are invalid")
    try:
        eligible_years = [row["year"] for row in eligible_rows]
    except (KeyError, TypeError) as exc:
        raise TechnoeconomicExportError(
            "Frozen v6 eligible energy rows are incomplete"
        ) from exc
    evidence_receipt = submission_provenance.get("evidence_receipt")
    if not isinstance(evidence_receipt, Mapping):
        raise TechnoeconomicExportError("Frozen v6 evidence receipt is invalid")
    public_identity_fields = (
        "schema_version",
        "artifact_kind",
        "media_type",
        "sha256",
        "byte_count",
        "row_count",
        "column_count",
        "pickle_allowed",
        "public",
    )
    try:
        sealed_identity = {
            key: sealed_calculation_artifact[key]
            for key in public_identity_fields
        }
    except KeyError as exc:
        raise TechnoeconomicExportError(
            "Frozen v6 sealed-calculation identity is incomplete"
        ) from exc
    authority = _applied_capacity_authority(submission_provenance)
    lifecycle = _lifecycle_request(request_payload)

    lcoo_summary = summaries.get("lcoo")
    headline = summaries.get("headline_decision")
    probabilities = summaries.get("probability_counts")
    if not all(
        isinstance(value, Mapping)
        for value in (lcoo_summary, headline, probabilities)
    ):
        raise TechnoeconomicExportError(
            "Sealed v6 decision summaries are invalid"
        )
    reason_codes = list(headline.get("reason_codes") or ())
    if lcoo_summary.get("status") != "available":
        reason = lcoo_summary.get("reason")
        if isinstance(reason, str) and reason and reason not in reason_codes:
            reason_codes.append(reason)

    required_summary_fields = (
        "upgrade_npv",
        "delta_lcoe",
        "annual_lifecycle",
        "reliability_summary",
        "representative_event_traces",
        "cost_coverage_audit",
        "warnings",
    )
    if any(field not in summaries for field in required_summary_fields):
        raise TechnoeconomicExportError("Sealed v6 summaries are incomplete")
    expected = {
        "schema_version": 6,
        "calculation_contract_version": lifecycle_version,
        "sampling_version": sampling_version,
        "analysis_basis": request_payload.get("basis"),
        "realization_count": request_payload.get("n"),
        "seed": request_payload.get("seed"),
        "project_life_years": finance.get("project_life_years"),
        "cost_stack_completeness": request_payload.get(
            "cost_stack_completeness"
        ),
        "energy_available": metadata.get("energy_available"),
        "commercial_transfer_status": submission_provenance.get(
            "commercial_transfer_status"
        ),
        "commercial_reference_design": submission_provenance.get(
            "commercial_reference_design"
        ),
        "source_snapshot_sha256": submission_provenance.get(
            "source_snapshot_sha256"
        ),
        "eligible_weather_years": eligible_years,
        "capacity_basis": "frozen_annual_applied_capacity_w",
        "capacities": capacities,
        "input_status": evidence_receipt.get("status"),
        "evidence_class_counts": evidence_receipt.get(
            "evidence_class_counts"
        )
        or {},
        "common_cost_audit": metadata.get("common_cost_audit"),
        "summaries": _compact_cdf_points_for_binding(summaries),
        "per_weather_year": _compact_cdf_points_for_binding(
            metadata.get("per_weather_year")
        ),
        "sensitivity": metadata.get("sensitivity"),
        "convergence": metadata.get("convergence"),
        "sealed_calculation": sealed_identity,
        "applied_capacities": {
            system: {
                "applied_capacity_w": authority[system].get(
                    "applied_capacity_w"
                ),
                "rating_basis": authority[system].get("rating_basis"),
            }
            for system in ("solaredge", "solectria")
        },
        "result_version": result_version,
        "paired_lifecycle": {
            "target_capacity_w": lifecycle.get("target_capacity_w"),
            "target_rating_basis": lifecycle.get("target_rating_basis"),
            "source_energy_basis": lifecycle.get("source_energy_basis"),
            "reliability_mode": lifecycle.get("reliability_mode"),
            "constant_dollar_cost_year": finance.get(
                "constant_dollar_cost_year"
            ),
            "headline_metric_id": "upgrade_npv",
            "headline_decision": headline,
            "probability_counts": probabilities,
            "upgrade_npv": _compact_cdf_points_for_binding(
                summaries["upgrade_npv"]
            ),
            "delta_lcoe": _compact_cdf_points_for_binding(
                summaries["delta_lcoe"]
            ),
            "lcoo": _compact_cdf_points_for_binding(lcoo_summary),
            "reason_codes": reason_codes,
            "annual_lifecycle": summaries["annual_lifecycle"],
            "reliability_summary": summaries["reliability_summary"],
            "representative_event_traces": summaries[
                "representative_event_traces"
            ],
            "cost_coverage_audit": summaries["cost_coverage_audit"],
            "warnings": summaries["warnings"],
            "formula_registry": kernel_provenance.get("formula_registry") or {},
            "formula_catalog_endpoint": "/api/technoeconomic/formulas/v6",
            "admission": kernel_provenance.get("admission") or {},
        },
    }
    actual_digest = _canonical_json_sha256(routine_result)
    expected_digest = _canonical_json_sha256(expected)
    if not secrets.compare_digest(actual_digest, expected_digest):
        raise TechnoeconomicExportError(
            "Durable v6 routine result differs from frozen or sealed authority"
        )


def _build_lifecycle_tables(
    calculation: _SealedCalculation,
    request_payload: Mapping[str, Any],
    source_snapshot: Mapping[str, Any],
    submission_provenance: Mapping[str, Any],
    routine_result: Mapping[str, Any],
    checks: Sequence[Sequence[Any]],
    registry: Sequence[Mapping[str, Any]],
    trace_records: Sequence[Mapping[str, Any]],
    audit_records: Sequence[Mapping[str, Any]],
) -> tuple[_Table, ...]:
    summaries = _lifecycle_summaries(calculation.metadata)
    annual_records = _mapping_sequence(
        summaries.get("annual_lifecycle"), label="The v6 annual lifecycle summary"
    )
    reliability_records = _mapping_sequence(
        summaries.get("reliability_summary"),
        label="The v6 reliability summary",
    )
    coverage_records = _mapping_sequence(
        summaries.get("cost_coverage_audit") or (),
        label="The v6 cost-coverage audit",
    )
    reliability_input_records = _lifecycle_reliability_input_records(request_payload)
    common_cause_records = tuple(
        record
        for record in reliability_input_records
        if record.get("record_type") == "common_cause"
    )
    registry_rows = tuple(
        tuple(
            _canonical_json_text(list(record[column]))
            if column == "inputs"
            else record[column]
            for column in LIFECYCLE_FORMULA_CATALOG_COLUMNS
        )
        for record in registry
    )
    audit_rows = tuple(
        tuple(record.get(column) for column in LIFECYCLE_AUDIT_COLUMNS)
        for record in audit_records
    )
    trace_rows = tuple(
        tuple(record.get(column) for column in LIFECYCLE_TRACE_COLUMNS)
        for record in trace_records
    )

    dynamic: dict[str, tuple[tuple[str, ...], tuple[tuple[Any, ...], ...]]] = {}
    for name, records, preferred in (
        ("annual", annual_records, ("system", "project_year", "statistic", "value", "unit")),
        (
            "reliability",
            reliability_records,
            ("system", "component_id", "project_year", "statistic", "value", "unit"),
        ),
        (
            "input",
            _lifecycle_input_records(request_payload),
            ("section", "field_path", "value_json"),
        ),
        (
            "target",
            _lifecycle_target_design_records(request_payload),
            ("technology", "target_capacity_w", "target_rating_basis"),
        ),
        (
            "reliability_input",
            reliability_input_records,
            ("record_type", "technology", "component_id", "category", "count"),
        ),
        (
            "weather",
            _lifecycle_weather_records(calculation.metadata),
            ("project_year", "weather_year"),
        ),
        (
            "capacity",
            _lifecycle_capacity_records(request_payload, submission_provenance),
            ("record_type", "technology", "capacity_w", "rating_basis"),
        ),
        (
            "common",
            common_cause_records,
            ("record_type", "technology", "event_id", "affected_systems"),
        ),
        (
            "coverage",
            coverage_records,
            ("coverage_id", "status", "owner", "notes"),
        ),
        (
            "lcoe",
            _lifecycle_lcoe_records(calculation),
            ("technology", "metric_id", "unit", "population_count", "p10", "p50", "p90"),
        ),
    ):
        dynamic[name] = _records_to_table(records, preferred_columns=preferred)

    source_snapshot_sha256 = str(submission_provenance.get("source_snapshot_sha256"))
    transfer_records = tuple(
        {
            "section": "submission_provenance",
            "field_path": field_path,
            "value_json": encoded,
        }
        for field_path, encoded in _flatten_leaves(
            "",
            submission_provenance.get("paired_lifecycle_receipt")
            or submission_provenance.get("normalization_receipt")
            or {},
        )
    )
    transfer_columns, transfer_rows = _records_to_table(
        transfer_records,
        preferred_columns=("section", "field_path", "value_json"),
    )
    provenance_rows = tuple(
        _lifecycle_provenance_rows(
            request_payload,
            source_snapshot,
            submission_provenance,
            routine_result,
            calculation.metadata,
            registry,
            checks,
        )
    )
    return (
        _Table(
            "formula-catalog.csv",
            "Formula Catalog",
            LIFECYCLE_FORMULA_CATALOG_COLUMNS,
            lambda: iter(registry_rows),
        ),
        _Table(
            "calculation-audit.csv",
            "Calculation Audit",
            LIFECYCLE_AUDIT_COLUMNS,
            lambda: iter(audit_rows),
        ),
        _Table(
            "realizations.csv",
            "Realizations",
            calculation.column_names,
            calculation.rows,
        ),
        _Table(
            "annual-lifecycle.csv",
            "Annual Lifecycle",
            dynamic["annual"][0],
            lambda: iter(dynamic["annual"][1]),
        ),
        _Table(
            "reliability-summary.csv",
            "Reliability Summary",
            dynamic["reliability"][0],
            lambda: iter(dynamic["reliability"][1]),
        ),
        _Table(
            "representative-event-traces.csv",
            "Representative Event Traces",
            LIFECYCLE_TRACE_COLUMNS,
            lambda: iter(trace_rows),
        ),
        _Table(
            "input-specifications.csv",
            "Input Specifications",
            dynamic["input"][0],
            lambda: iter(dynamic["input"][1]),
        ),
        _Table(
            "target-design.csv",
            "Target Design",
            dynamic["target"][0],
            lambda: iter(dynamic["target"][1]),
        ),
        _Table(
            "reliability-inputs.csv",
            "Reliability Inputs",
            dynamic["reliability_input"][0],
            lambda: iter(dynamic["reliability_input"][1]),
        ),
        _Table(
            "energy-snapshot.csv",
            "Energy Snapshot",
            ENERGY_COLUMNS,
            lambda: _energy_rows(source_snapshot, source_snapshot_sha256),
        ),
        _Table(
            "weather-summary.csv",
            "Weather Summary",
            dynamic["weather"][0],
            lambda: iter(dynamic["weather"][1]),
        ),
        _Table(
            "capacity-and-basis.csv",
            "Capacity and Basis",
            dynamic["capacity"][0],
            lambda: iter(dynamic["capacity"][1]),
        ),
        _Table(
            "common-cost-audit.csv",
            "Common-Cost Audit",
            dynamic["common"][0],
            lambda: iter(dynamic["common"][1]),
        ),
        _Table(
            "cost-coverage-audit.csv",
            "Cost-Coverage Audit",
            dynamic["coverage"][0],
            lambda: iter(dynamic["coverage"][1]),
        ),
        _Table(
            "commercial-transfer.csv",
            "Commercial Transfer",
            transfer_columns,
            lambda: iter(transfer_rows),
        ),
        _Table(
            "commercial-lcoe.csv",
            "Commercial LCOE",
            dynamic["lcoe"][0],
            lambda: iter(dynamic["lcoe"][1]),
        ),
        _Table(
            "metric-cdfs.csv",
            "Metric CDFs",
            LIFECYCLE_CDF_COLUMNS,
            lambda: _lifecycle_cdf_rows(calculation.metadata),
        ),
        _Table(
            "sensitivity.csv",
            "Sensitivity",
            SENSITIVITY_COLUMNS,
            lambda: _sensitivity_rows(calculation.metadata),
        ),
        _Table(
            "convergence.csv",
            "Convergence",
            LIFECYCLE_CONVERGENCE_COLUMNS,
            lambda: _lifecycle_convergence_rows(calculation.metadata),
        ),
        _Table(
            "provenance.csv",
            "Provenance",
            PROVENANCE_COLUMNS,
            lambda: iter(provenance_rows),
        ),
        _Table(
            "checks.csv",
            "Checks",
            CHECK_COLUMNS,
            lambda: iter(checks),
        ),
    )


def _build_tables(
    calculation: _SealedCalculation,
    request_payload: Mapping[str, Any],
    source_snapshot: Mapping[str, Any],
    submission_provenance: Mapping[str, Any],
    routine_result: Mapping[str, Any],
    checks: Sequence[Sequence[Any]],
) -> tuple[_Table, ...]:
    source_snapshot_sha256 = str(submission_provenance.get("source_snapshot_sha256"))
    common_records = calculation.metadata.get("common_cost_audit") or []
    applied_capacity_contract = routine_result.get(
        "calculation_contract_version"
    ) in {
        technoeconomic_kernel.CALCULATION_CONTRACT_VERSION,
        technoeconomic_kernel.COMMERCIAL_SCALING_CALCULATION_CONTRACT_VERSION,
        technoeconomic_kernel.STANDALONE_COMMERCIAL_CALCULATION_CONTRACT_VERSION,
        technoeconomic_kernel.PAIRED_COMMERCIAL_CALCULATION_CONTRACT_VERSION,
    }
    standalone_contract = (
        routine_result.get("calculation_contract_version")
        == technoeconomic_kernel.STANDALONE_COMMERCIAL_CALCULATION_CONTRACT_VERSION
    )
    paired_contract = (
        routine_result.get("calculation_contract_version")
        == technoeconomic_kernel.PAIRED_COMMERCIAL_CALCULATION_CONTRACT_VERSION
    )
    leading_tables = (
        _Table(
            "realizations.csv",
            "Realizations",
            calculation.column_names,
            calculation.rows,
        ),
        _Table(
            "input-specifications.csv",
            "Input Specifications",
            APPLIED_INPUT_COLUMNS if applied_capacity_contract else INPUT_COLUMNS,
            lambda: _input_rows(
                request_payload,
                submission_provenance,
                applied_capacity_contract=applied_capacity_contract,
            ),
        ),
        _Table(
            "energy-snapshot.csv",
            "Energy Snapshot",
            ENERGY_COLUMNS,
            lambda: _energy_rows(source_snapshot, source_snapshot_sha256),
        ),
        _Table(
            "capacity-and-basis.csv",
            "Capacity and Basis",
            (
                APPLIED_CAPACITY_COLUMNS
                if applied_capacity_contract
                else CAPACITY_COLUMNS
            ),
            lambda: _capacity_rows(
                source_snapshot,
                submission_provenance,
                routine_result,
                applied_capacity_contract=applied_capacity_contract,
            ),
        ),
        _Table(
            "common-cost-audit.csv",
            "Common-Cost Audit",
            COMMON_COST_COLUMNS,
            lambda: _common_cost_rows(common_records),
        ),
        _Table(
            "commercial-transfer.csv",
            "Commercial Transfer",
            TRANSFER_COLUMNS,
            lambda: _transfer_rows(request_payload, submission_provenance),
        ),
    )
    standalone_tables = (
        (
            _Table(
                "standalone-commercial-summary.csv",
                "Commercial LCOE",
                STANDALONE_COMMERCIAL_SUMMARY_COLUMNS,
                lambda: _standalone_commercial_summary_rows(
                    request_payload,
                    submission_provenance,
                    routine_result,
                ),
            ),
        )
        if standalone_contract
        else ()
    )
    paired_tables = (
        (
            _Table(
                "paired-commercial-summary.csv",
                "Commercial LCOE",
                PAIRED_COMMERCIAL_SUMMARY_COLUMNS,
                lambda: _paired_commercial_summary_rows(
                    request_payload,
                    submission_provenance,
                    routine_result,
                ),
            ),
        )
        if paired_contract
        else ()
    )
    trailing_tables = (
        _Table(
            "metric-cdfs.csv",
            "Metric CDFs",
            (
                PAIRED_COMMERCIAL_CDF_COLUMNS
                if paired_contract
                else STANDALONE_COMMERCIAL_CDF_COLUMNS
                if standalone_contract
                else CDF_COLUMNS
            ),
            (
                (lambda: _paired_commercial_cdf_rows(calculation.metadata))
                if paired_contract
                else (lambda: _standalone_commercial_cdf_rows(calculation.metadata))
                if standalone_contract
                else (lambda: _cdf_rows(calculation.metadata))
            ),
        ),
        _Table(
            "per-year-summary.csv",
            "Per-Year Summary",
            (
                _paired_commercial_per_year_columns(calculation.metadata)
                if paired_contract
                else _standalone_commercial_per_year_columns(calculation.metadata)
                if standalone_contract
                else _per_year_columns(calculation.metadata)
            ),
            (
                (lambda: _paired_commercial_per_year_rows(calculation.metadata))
                if paired_contract
                else (lambda: _standalone_commercial_per_year_rows(calculation.metadata))
                if standalone_contract
                else (lambda: _per_year_rows(calculation.metadata))
            ),
        ),
        _Table(
            "sensitivity.csv",
            "Sensitivity",
            SENSITIVITY_COLUMNS,
            lambda: _sensitivity_rows(calculation.metadata),
        ),
        _Table(
            "convergence.csv",
            "Convergence",
            CONVERGENCE_COLUMNS,
            lambda: _convergence_rows(calculation.metadata),
        ),
        _Table(
            "provenance.csv",
            "Provenance",
            PROVENANCE_COLUMNS,
            lambda: _provenance_rows(
                request_payload,
                source_snapshot,
                submission_provenance,
                routine_result,
                calculation.metadata,
            ),
        ),
        _Table(
            "checks.csv",
            "Checks",
            CHECK_COLUMNS,
            lambda: iter(checks),
        ),
    )
    return leading_tables + standalone_tables + paired_tables + trailing_tables


def _csv_scalar(value: Any) -> str:
    value = _numpy_scalar(value)
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        return repr(value)
    if isinstance(value, (Mapping, list, tuple)):
        return _canonical_json_text(_safe_public_value(value))
    text = str(value)
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _write_csv_table(
    path: Path,
    table: _Table,
    cancellation_check: Callable[[], None],
) -> dict[str, Any]:
    row_count = 0
    with path.open("x", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n", quoting=csv.QUOTE_MINIMAL)
        writer.writerow(table.columns)
        for row_count, row in enumerate(table.rows_factory(), start=1):
            if len(row) != len(table.columns):
                raise TechnoeconomicExportError(
                    f"Export table {table.filename} produced an inconsistent row width"
                )
            if row_count % _CANCEL_INTERVAL == 0:
                cancellation_check()
            writer.writerow([_csv_scalar(value) for value in row])
        handle.flush()
        os.fsync(handle.fileno())
    return {
        "filename": table.filename,
        "row_count": row_count,
        "column_count": len(table.columns),
        "sha256": _sha256_file(path),
    }


def _zip_info(filename: str, *, nested: bool = False) -> zipfile.ZipInfo:
    if nested:
        member = PurePosixPath(filename)
        if (
            member.is_absolute()
            or not member.parts
            or any(part in {"", ".", ".."} for part in member.parts)
            or "\\" in filename
            or ":" in filename
        ):
            raise TechnoeconomicExportError("Unsafe archive member name")
        safe_name = member.as_posix()
    else:
        safe_name = _safe_filename(filename)
    info = zipfile.ZipInfo(safe_name, date_time=_FIXED_ZIP_TIME)
    info.compress_type = zipfile.ZIP_DEFLATED
    info.create_system = 3
    info.external_attr = 0o600 << 16
    info.flag_bits |= 0x800
    return info


def _copy_with_cancellation(
    source: Any,
    destination: Any,
    cancellation_check: Callable[[], None],
    *,
    block_size: int = 1024 * 1024,
) -> None:
    while True:
        block = source.read(block_size)
        if not block:
            return
        destination.write(block)
        cancellation_check()


def _write_csv_bundle(
    target: Path,
    staging_directory: Path,
    tables: Sequence[_Table],
    cancellation_check: Callable[[], None],
    *,
    schema_versions: Mapping[str, str] | None = None,
    bundle_provenance: Mapping[str, Any] | None = None,
) -> tuple[list[dict[str, Any]], int]:
    schema_versions = schema_versions or export_contract_versions(
        technoeconomic_kernel.LEGACY_CALCULATION_CONTRACT_VERSION
    )
    staging_directory.mkdir(parents=False, exist_ok=False)
    metadata: list[dict[str, Any]] = []
    try:
        for table in tables:
            cancellation_check()
            metadata.append(
                _write_csv_table(
                    staging_directory / _safe_filename(table.filename),
                    table,
                    cancellation_check,
                )
            )
        bundle_metadata = {
            "schema_version": schema_versions["csv_bundle"],
            "csv_format_version": schema_versions["csv_format"],
            "encoding": "UTF-8",
            "line_terminator": "LF",
            "float_format": "Python finite-float repr (shortest round-trip binary64)",
            "null_format": "empty field",
            "table_count": len(metadata),
            "tables": metadata,
        }
        if bundle_provenance is not None:
            bundle_metadata["provenance"] = _safe_public_value(bundle_provenance)
        manifest_bytes = (
            _canonical_json_text(bundle_metadata) + "\n"
        ).encode("utf-8")
        with zipfile.ZipFile(
            target,
            mode="x",
            compression=zipfile.ZIP_DEFLATED,
            compresslevel=9,
            strict_timestamps=True,
        ) as archive:
            archive.writestr(
                _zip_info(schema_versions["csv_bundle_manifest_filename"]),
                manifest_bytes,
            )
            for table_metadata in metadata:
                cancellation_check()
                source = staging_directory / table_metadata["filename"]
                with source.open("rb") as source_handle, archive.open(
                    _zip_info(table_metadata["filename"]), "w"
                ) as destination:
                    _copy_with_cancellation(
                        source_handle,
                        destination,
                        cancellation_check,
                    )
    finally:
        shutil.rmtree(staging_directory, ignore_errors=True)
    return metadata, sum(int(record["row_count"]) for record in metadata)


_THIN_GREY = Side(style="thin", color="D8DEE5")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
_PASS_FILL = PatternFill("solid", fgColor="DDEBF7")
_FAIL_FILL = PatternFill("solid", fgColor="FCE4D6")
_V6_SEALED_FILL = PatternFill("solid", fgColor="E7E6E6")
_V6_WARNING_FILL = PatternFill("solid", fgColor="FFF2CC")
_V6_FAIL_FILL = PatternFill("solid", fgColor="F4CCCC")
_V6_REFERENCE_FONT = Font(name="Aptos", size=10, color="008000")
_V6_REFERENCE_SHEETS = frozenset(
    {
        "Input Specifications",
        "Target Design",
        "Reliability Inputs",
        "Energy Snapshot",
        "Capacity and Basis",
        "Commercial Transfer",
    }
)
_FAST_STREAMING_SHEETS = frozenset({"Realizations", "Metric CDFs"})


def _excel_value(value: Any) -> Any:
    value = _numpy_scalar(value)
    if isinstance(value, float) and not math.isfinite(value):
        return None
    if isinstance(value, (Mapping, list, tuple)):
        return _canonical_json_text(_safe_public_value(value))
    return value


def _write_only_cell(
    sheet: Any,
    value: Any,
    *,
    header: bool = False,
    section: bool = False,
    status: str | None = None,
    number_format: str | None = None,
) -> WriteOnlyCell:
    excel_value = _excel_value(value)
    if isinstance(excel_value, str) and len(excel_value) > 32_767:
        raise TechnoeconomicExportError(
            "An export value exceeds the lossless XLSX cell limit"
        )
    cell = WriteOnlyCell(sheet, value=excel_value)
    cell.alignment = Alignment(
        vertical="top",
        wrap_text=isinstance(value, str) and len(value) > 35,
    )
    if header:
        cell.fill = _HEADER_FILL
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=_THIN_GREY)
    elif section:
        cell.fill = _SECTION_FILL
        cell.font = Font(name="Aptos", size=10, bold=True, color="1F1F1F")
    else:
        cell.font = Font(name="Aptos", size=10, color="000000")
        cell.border = Border(bottom=Side(style="hair", color="E7EBEF"))
    if status == "OK":
        cell.fill = _PASS_FILL
        cell.font = Font(name="Aptos", size=10, bold=True, color="1F1F1F")
    elif status == "FAIL":
        cell.fill = _FAIL_FILL
        cell.font = Font(name="Aptos", size=10, bold=True, color="9C0006")
    if number_format:
        cell.number_format = number_format
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        cell.data_type = "s"
    return cell


def _fast_streaming_value(sheet: Any, value: Any) -> Any:
    """Return an allocation-light value for very large write-only sheets."""

    excel_value = _excel_value(value)
    if isinstance(excel_value, str) and len(excel_value) > 32_767:
        raise TechnoeconomicExportError(
            "An export value exceeds the lossless XLSX cell limit"
        )
    if isinstance(excel_value, str) and excel_value.startswith(("=", "+", "-", "@")):
        cell = WriteOnlyCell(sheet, value=excel_value)
        cell.data_type = "s"
        return cell
    return excel_value


def _apply_fast_streaming_formats(sheet: Any, columns: Sequence[str]) -> None:
    """Apply numeric display formats once per column, not once per body cell."""

    sheet.sheet_view.showGridLines = True
    for index, header in enumerate(columns, start=1):
        number_format = _number_format_for_header(str(header), 0.0)
        if number_format:
            letter = openpyxl.utils.get_column_letter(index)
            sheet.column_dimensions[letter].number_format = number_format


def _number_format_for_header(header: str, value: Any) -> str | None:
    if not isinstance(value, (int, float, np.integer, np.floating)) or isinstance(
        value, (bool, np.bool_)
    ):
        return None
    lowered = header.lower()
    if (
        lowered.endswith(("probability", "share", "rate"))
        or "_probability" in lowered
        or "_share" in lowered
        or "_relative_change" in lowered
        or "discount-rate" in lowered
        or "degradation" in lowered
    ):
        return "0.0000%"
    if "count" in lowered or header in {"year", "realization_index", "entry_order"}:
        return "#,##0"
    return "0.###############"


def _set_sheet_layout(sheet: Any, columns: Sequence[str]) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    for index, header in enumerate(columns, start=1):
        letter = openpyxl.utils.get_column_letter(index)
        normalized = str(header).lower()
        if normalized == "check_id" or normalized.endswith("field_path"):
            width = 60.0
        elif normalized.endswith("_provenance_path"):
            width = 58.0
        elif normalized in {"value_json", "notes"} or normalized.endswith(
            "value_json"
        ):
            width = 58.0
        elif normalized == "metric_id":
            width = 54.0
        elif normalized in {"label", "response_id", "predictor_id"}:
            width = 42.0
        elif normalized in {"reason", "exclusion_reason"}:
            width = 36.0
        elif normalized.endswith("_sha256"):
            width = 68.0
        elif normalized == "record_type":
            width = 34.0
        elif normalized in {"analysis_basis", "module_model"}:
            width = 28.0
        elif normalized in {"capacity_basis", "rating_basis"}:
            width = 34.0
        elif normalized.endswith("_json"):
            width = 45.0
        elif normalized in {
            "value",
            "p5",
            "p50",
            "p95",
            "p5_absolute_change",
            "p5_relative_change",
            "p50_absolute_change",
            "p50_relative_change",
            "p95_absolute_change",
            "p95_relative_change",
        }:
            width = 22.0
        elif normalized in {
            "actual_authority",
            "expected_authority",
            "difference_authority",
        }:
            width = 34.0
        elif normalized == "display_formula_status":
            width = 24.0
        else:
            width = min(42.0, max(16.0, len(str(header)) * 0.9 + 2.0))
        sheet.column_dimensions[letter].width = width


def _append_header(sheet: Any, columns: Sequence[str]) -> None:
    sheet.append([_write_only_cell(sheet, value, header=True) for value in columns])


def _update_logical_sheet_hash(
    digest: Any,
    values: Sequence[Any],
    *,
    formula_indexes: frozenset[int] = frozenset(),
) -> None:
    row = {
        "values": [_excel_value(value) for value in values],
        "formula_indexes": sorted(formula_indexes),
    }
    digest.update((_canonical_json_text(row) + "\n").encode("utf-8"))


def _new_logical_sheet_hash() -> Any:
    digest = hashlib.sha256()
    digest.update((XLSX_LOGICAL_HASH_VERSION + "\n").encode("ascii"))
    return digest


def _new_lifecycle_logical_sheet_hash() -> Any:
    digest = hashlib.sha256()
    digest.update((LIFECYCLE_XLSX_LOGICAL_HASH_VERSION + "\n").encode("ascii"))
    return digest


def _update_lifecycle_logical_sheet_hash(
    digest: Any,
    values: Sequence[Any],
    *,
    formula_identities: Mapping[int, str] | None = None,
) -> None:
    """Hash v6 formula semantics without binding identity to A1 row numbers.

    The normalized XLSX physical SHA still seals the exact formula text.  This
    logical hash uses stable semantic roles, so inserting a trace row does not
    misrepresent an otherwise unchanged audit formula as a contract change.
    """

    identities = dict(formula_identities or {})
    normalized_values = [
        (
            {"formula_identity": identities[index]}
            if index in identities
            else _excel_value(value)
        )
        for index, value in enumerate(values)
    ]
    row = {
        "values": normalized_values,
        "formula_indexes": sorted(identities),
    }
    digest.update((_canonical_json_text(row) + "\n").encode("utf-8"))


def _write_standalone_summary_sheet(
    workbook: openpyxl.Workbook,
    routine_result: Mapping[str, Any],
    checks: Sequence[Sequence[Any]],
) -> tuple[int, int, str]:
    sheet = workbook.create_sheet("Summary")
    logical_digest = _new_logical_sheet_hash()
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"
    for letter, width in {"A": 52, "B": 36, "C": 28, "D": 66}.items():
        sheet.column_dimensions[letter].width = width
    title_values = ("Standalone Commercial SolarEdge LCOE", None, None, None)
    title = _write_only_cell(sheet, title_values[0], section=True)
    title.font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    title.fill = _HEADER_FILL
    sheet.append(
        [title, _write_only_cell(sheet, None), _write_only_cell(sheet, None), _write_only_cell(sheet, None)]
    )
    _update_logical_sheet_hash(logical_digest, title_values)
    note_values = (
        "Frozen numeric values are authoritative; formulas are display aids.",
        None,
        None,
        "P10/P50/P90 are statistical quantiles of P(LCOE <= x).",
    )
    sheet.append([_write_only_cell(sheet, value) for value in note_values])
    _update_logical_sheet_hash(logical_digest, note_values)
    blank = (None, None, None, None)
    sheet.append([_write_only_cell(sheet, None) for _ in range(4)])
    _update_logical_sheet_hash(logical_digest, blank)
    headers = ("Frozen authority", "Value", "Display formula/status", "Notes")
    _append_header(sheet, headers)
    _update_logical_sheet_hash(logical_digest, headers)
    standalone = routine_result.get("standalone_commercial") or {}
    percentiles = standalone.get("percentiles") or {}
    all_passed = all(row[5] == "OK" for row in checks)
    check_end_row = len(checks) + 1
    realization_end_row = int(routine_result.get("realization_count") or 0) + 1
    rows = [
        ("Model status", "OK" if all_passed else "FAIL", f"=IF(COUNTIF('Checks'!F2:F{check_end_row},\"FAIL\")=0,\"OK\",\"FAIL\")", "Formula recalculates visible check status."),
        ("Technology", standalone.get("technology"), None, "Standalone commercial technology."),
        ("Target capacity (W)", standalone.get("target_capacity_w"), None, standalone.get("target_rating_basis")),
        ("Source applied capacity (W)", standalone.get("source_applied_capacity_w"), None, standalone.get("source_rating_basis")),
        ("Capacity scale factor", standalone.get("capacity_scale_factor"), None, "Target W divided by frozen source W."),
        ("Transfer method", standalone.get("transfer_method"), None, "Direct scaling of verified SolarEdge specific energy."),
        ("Realizations", routine_result.get("realization_count"), f"=ROWS('Realizations'!A2:A{realization_end_row})", "Sealed LHS realization count."),
        ("Seed", routine_result.get("seed"), None, "Unsigned deterministic seed."),
        ("Project life (years)", routine_result.get("project_life_years"), None, "Constant-real lifecycle horizon."),
        ("Calculation contract", routine_result.get("calculation_contract_version"), None, "Pinned standalone calculation semantics."),
        ("Sampling version", routine_result.get("sampling_version"), None, "Pinned LHS and weather allocation semantics."),
    ]
    for label, value, formula, notes in rows:
        status = value if label == "Model status" else None
        cells = [
            _write_only_cell(sheet, label),
            _write_only_cell(sheet, value, status=status),
            _write_only_cell(sheet, formula),
            _write_only_cell(sheet, notes),
        ]
        if isinstance(formula, str) and formula.startswith("="):
            cells[2].data_type = "f"
        sheet.append(cells)
        _update_logical_sheet_hash(
            logical_digest,
            (label, value, formula, notes),
            formula_indexes=frozenset({2}) if formula else frozenset(),
        )
    sheet.append([_write_only_cell(sheet, None) for _ in range(4)])
    _update_logical_sheet_hash(logical_digest, blank)
    section_values = ("Commercial SolarEdge LCOE percentiles", None, None, None)
    section = _write_only_cell(sheet, section_values[0], section=True)
    sheet.append(
        [section, _write_only_cell(sheet, None), _write_only_cell(sheet, None), _write_only_cell(sheet, None)]
    )
    _update_logical_sheet_hash(logical_digest, section_values)
    percentile_headers = ("Metric", "P10", "P50", "P90")
    _append_header(sheet, percentile_headers)
    _update_logical_sheet_hash(logical_digest, percentile_headers)
    percentile_values = (
        "Commercial SolarEdge lifecycle LCOE (constant USD/kWh_AC)",
        percentiles.get("p10"),
        percentiles.get("p50"),
        percentiles.get("p90"),
    )
    sheet.append(
        [
            _write_only_cell(
                sheet,
                value,
                number_format=_number_format_for_header(header, value),
            )
            for header, value in zip(percentile_headers, percentile_values)
        ]
    )
    _update_logical_sheet_hash(logical_digest, percentile_values)
    return len(rows) + 1, 4, logical_digest.hexdigest()


def _write_paired_summary_sheet(
    workbook: openpyxl.Workbook,
    routine_result: Mapping[str, Any],
    checks: Sequence[Sequence[Any]],
) -> tuple[int, int, str]:
    sheet = workbook.create_sheet("Summary")
    logical_digest = _new_logical_sheet_hash()
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"
    for letter, width in {"A": 52, "B": 36, "C": 28, "D": 66}.items():
        sheet.column_dimensions[letter].width = width
    row_count = 0

    def append(values: Sequence[Any], *, header: bool = False) -> None:
        nonlocal row_count
        if header:
            _append_header(sheet, values)
        else:
            sheet.append([_write_only_cell(sheet, value) for value in values])
        _update_logical_sheet_hash(logical_digest, values)
        row_count += 1

    title_values = ("Paired Commercial Lifecycle LCOE", None, None, None)
    title = _write_only_cell(sheet, title_values[0], section=True)
    title.font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    title.fill = _HEADER_FILL
    sheet.append(
        [
            title,
            _write_only_cell(sheet, None),
            _write_only_cell(sheet, None),
            _write_only_cell(sheet, None),
        ]
    )
    _update_logical_sheet_hash(logical_digest, title_values)
    row_count += 1
    append(
        (
            "Frozen numeric values are authoritative.",
            None,
            None,
            "P10/P50/P90 are statistical quantiles of P(LCOE <= x).",
        )
    )
    append((None, None, None, None))
    append(("Frozen authority", "Value", "Status", "Notes"), header=True)

    paired = routine_result.get("paired_commercial") or {}
    systems = paired.get("systems") or {}
    all_passed = all(row[5] == "OK" for row in checks)
    authority_rows = [
        ("Model status", "OK" if all_passed else "FAIL", None, "All export tie-outs."),
        ("Target capacity (W)", paired.get("target_capacity_w"), None, paired.get("target_rating_basis")),
        ("Transfer method", paired.get("transfer_method"), None, "Independent direct scaling from each frozen source."),
        ("Constant-dollar year", paired.get("constant_dollar_cost_year"), None, "Shared by both cost stacks."),
        ("Realizations", routine_result.get("realization_count"), None, "Sealed LHS realization count."),
        ("Seed", routine_result.get("seed"), None, "Unsigned deterministic seed."),
        ("Project life (years)", routine_result.get("project_life_years"), None, "Constant-real lifecycle horizon."),
        ("Calculation contract", routine_result.get("calculation_contract_version"), None, "Pinned paired calculation semantics."),
    ]
    for technology in ("solectria", "solaredge"):
        system = systems.get(technology) or {}
        label = "Solectria" if technology == "solectria" else "SolarEdge"
        authority_rows.extend(
            (
                (
                    f"{label} source applied capacity (W)",
                    system.get("source_applied_capacity_w"),
                    None,
                    system.get("source_rating_basis"),
                ),
                (
                    f"{label} capacity scale factor",
                    system.get("capacity_scale_factor"),
                    None,
                    "Target W divided by the same-rated source W.",
                ),
            )
        )
    for row in authority_rows:
        append(row)
    append((None, None, None, None))
    append(("Commercial lifecycle LCOE percentiles", None, None, None))
    percentile_headers = ("Metric", "P10", "P50", "P90")
    append(percentile_headers, header=True)
    for technology in ("solectria", "solaredge"):
        system = systems.get(technology) or {}
        percentiles = system.get("percentiles") or {}
        label = "Solectria" if technology == "solectria" else "SolarEdge"
        append(
            (
                f"Commercial {label} lifecycle LCOE (constant USD/kWh_AC)",
                percentiles.get("p10"),
                percentiles.get("p50"),
                percentiles.get("p90"),
            )
        )
    delta = paired.get("lcoe_delta_se_minus_sol") or {}
    delta_percentiles = delta.get("percentiles") or {}
    append(
        (
            "LCOE delta, SolarEdge minus Solectria (constant USD/kWh_AC)",
            delta_percentiles.get("p10"),
            delta_percentiles.get("p50"),
            delta_percentiles.get("p90"),
        )
    )
    return row_count, 4, logical_digest.hexdigest()


def _write_summary_sheet(
    workbook: openpyxl.Workbook,
    routine_result: Mapping[str, Any],
    metadata: Mapping[str, Any],
    checks: Sequence[Sequence[Any]],
) -> tuple[int, int, str]:
    if (
        routine_result.get("calculation_contract_version")
        == technoeconomic_kernel.PAIRED_COMMERCIAL_CALCULATION_CONTRACT_VERSION
    ):
        return _write_paired_summary_sheet(
            workbook,
            routine_result,
            checks,
        )
    if (
        routine_result.get("calculation_contract_version")
        == technoeconomic_kernel.STANDALONE_COMMERCIAL_CALCULATION_CONTRACT_VERSION
    ):
        return _write_standalone_summary_sheet(
            workbook,
            routine_result,
            checks,
        )
    sheet = workbook.create_sheet("Summary")
    logical_digest = _new_logical_sheet_hash()
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"
    for letter, width in {"A": 68, "B": 68, "C": 28, "D": 58}.items():
        sheet.column_dimensions[letter].width = width
    title = _write_only_cell(sheet, "Probabilistic Technoeconomic Analysis", section=True)
    title.font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    title.fill = _HEADER_FILL
    sheet.append([title, _write_only_cell(sheet, None), _write_only_cell(sheet, None), _write_only_cell(sheet, None)])
    _update_logical_sheet_hash(
        logical_digest,
        ("Probabilistic Technoeconomic Analysis", None, None, None),
    )
    sheet.append(
        [
            _write_only_cell(sheet, "Frozen numeric values are authoritative; formulas are display aids."),
            _write_only_cell(sheet, None),
            _write_only_cell(sheet, None),
            _write_only_cell(sheet, "All signed metrics use SolarEdge minus Solectria."),
        ]
    )
    _update_logical_sheet_hash(
        logical_digest,
        (
            "Frozen numeric values are authoritative; formulas are display aids.",
            None,
            None,
            "All signed metrics use SolarEdge minus Solectria.",
        ),
    )
    sheet.append([_write_only_cell(sheet, None) for _ in range(4)])
    _update_logical_sheet_hash(logical_digest, (None, None, None, None))
    _append_header(sheet, ("Frozen authority", "Value", "Display formula/status", "Notes"))
    _update_logical_sheet_hash(
        logical_digest,
        ("Frozen authority", "Value", "Display formula/status", "Notes"),
    )
    all_passed = all(row[5] == "OK" for row in checks)
    check_end_row = len(checks) + 1
    realization_end_row = int(routine_result.get("realization_count") or 0) + 1
    summary_rows = [
        ("Model status", "OK" if all_passed else "FAIL", f"=IF(COUNTIF('Checks'!F2:F{check_end_row},\"FAIL\")=0,\"OK\",\"FAIL\")", "Formula recalculates visible check status."),
        ("Analysis basis", routine_result.get("analysis_basis"), None, "Cost/energy interpretation basis."),
        ("Realizations", routine_result.get("realization_count"), f"=ROWS('Realizations'!A2:A{realization_end_row})", "Sealed LHS realization count."),
        ("Seed", routine_result.get("seed"), None, "Unsigned deterministic seed."),
        ("Project life (years)", routine_result.get("project_life_years"), None, "Constant-real lifecycle horizon."),
        ("Energy status", "available" if routine_result.get("energy_available") else "cost_only", None, "Commercial energy requires approved transfer."),
        ("Commercial transfer status", routine_result.get("commercial_transfer_status"), None, "Bound to immutable submission provenance."),
        ("Source snapshot SHA-256", routine_result.get("source_snapshot_sha256"), None, "Frozen Annual Simulation evidence identity."),
        ("Calculation contract", routine_result.get("calculation_contract_version"), None, "Pinned calculation semantics."),
        ("Sampling version", routine_result.get("sampling_version"), None, "Pinned LHS and weather allocation semantics."),
    ]
    for label, value, formula, notes in summary_rows:
        status = value if label == "Model status" else None
        row = [
            _write_only_cell(sheet, label),
            _write_only_cell(sheet, value, status=status),
            _write_only_cell(sheet, formula),
            _write_only_cell(sheet, notes),
        ]
        if isinstance(formula, str) and formula.startswith("="):
            row[2].data_type = "f"
        sheet.append(row)
        _update_logical_sheet_hash(
            logical_digest,
            (label, value, formula, notes),
            formula_indexes=frozenset({2}) if formula else frozenset(),
        )
    sheet.append([_write_only_cell(sheet, None) for _ in range(4)])
    _update_logical_sheet_hash(logical_digest, (None, None, None, None))
    section = _write_only_cell(sheet, "Metric percentiles (frozen authority)", section=True)
    sheet.append([section, _write_only_cell(sheet, None), _write_only_cell(sheet, None), _write_only_cell(sheet, None)])
    _update_logical_sheet_hash(
        logical_digest,
        ("Metric percentiles (frozen authority)", None, None, None),
    )
    _append_header(sheet, ("Metric", "P5", "P50", "P95"))
    _update_logical_sheet_hash(logical_digest, ("Metric", "P5", "P50", "P95"))
    metric_rows = 0
    for metric_id, summary in sorted(_metric_summaries(metadata).items()):
        if not isinstance(summary, Mapping) or "percentiles" not in summary:
            continue
        percentiles = summary.get("percentiles") or {}
        values = (
            _human_metric(metric_id),
            percentiles.get("p5"),
            percentiles.get("p50"),
            percentiles.get("p95"),
        )
        sheet.append(
            [
                _write_only_cell(
                    sheet,
                    value,
                    number_format=_number_format_for_header(header, value),
                )
                for header, value in zip(("Metric", "P5", "P50", "P95"), values)
            ]
        )
        _update_logical_sheet_hash(logical_digest, values)
        metric_rows += 1
    return len(summary_rows) + metric_rows, 4, logical_digest.hexdigest()


def _write_workbook(
    raw_path: Path,
    tables: Sequence[_Table],
    routine_result: Mapping[str, Any],
    metadata: Mapping[str, Any],
    checks: Sequence[Sequence[Any]],
    cancellation_check: Callable[[], None],
) -> tuple[list[dict[str, Any]], int]:
    workbook = openpyxl.Workbook(write_only=True)
    workbook.properties.creator = "SBE PV technoeconomic reporting"
    workbook.properties.title = "Probabilistic Technoeconomic Analysis"
    workbook.properties.subject = "Frozen simulation results and export tie-outs"
    fixed_time = datetime(1980, 1, 1)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    sheets: list[dict[str, Any]] = []
    summary_rows, summary_columns, summary_logical_sha256 = _write_summary_sheet(
        workbook,
        routine_result,
        metadata,
        checks,
    )
    sheets.append(
        {
            "sheet_name": "Summary",
            "row_count": summary_rows,
            "column_count": summary_columns,
            "logical_sha256": summary_logical_sha256,
            "logical_hash_version": XLSX_LOGICAL_HASH_VERSION,
        }
    )
    for table in tables:
        cancellation_check()
        sheet = workbook.create_sheet(table.sheet_name)
        workbook_columns = (
            table.columns + ("display_formula_status",)
            if table.sheet_name == "Checks"
            else table.columns
        )
        _set_sheet_layout(sheet, workbook_columns)
        fast_streaming = table.sheet_name in _FAST_STREAMING_SHEETS
        if fast_streaming:
            _apply_fast_streaming_formats(sheet, workbook_columns)
        _append_header(sheet, workbook_columns)
        logical_digest = _new_logical_sheet_hash()
        _update_logical_sheet_hash(logical_digest, workbook_columns)
        row_count = 0
        for row_count, raw_row in enumerate(table.rows_factory(), start=1):
            if len(raw_row) != len(table.columns):
                raise TechnoeconomicExportError(
                    f"Workbook table {table.sheet_name} produced an inconsistent row width"
                )
            if row_count % _CANCEL_INTERVAL == 0:
                cancellation_check()
            if fast_streaming:
                cells = [
                    _fast_streaming_value(sheet, value) for value in raw_row
                ]
            else:
                cells = [
                    _write_only_cell(
                        sheet,
                        value,
                        status=(value if header == "status_authority" else None),
                        number_format=_number_format_for_header(header, value),
                    )
                    for header, value in zip(table.columns, raw_row)
                ]
            if table.sheet_name == "Checks":
                excel_row = row_count + 1
                formula = (
                    f'=IF(OR(ISTEXT(B{excel_row}),ISTEXT(C{excel_row})),'
                    f'IF(B{excel_row}=C{excel_row},"OK","FAIL"),'
                    f'IF(ABS(B{excel_row}-C{excel_row})<=E{excel_row},"OK","FAIL"))'
                )
                formula_cell = _write_only_cell(sheet, formula)
                formula_cell.data_type = "f"
                cells.append(formula_cell)
            sheet.append(cells)
            logical_values = list(raw_row)
            formula_indexes = frozenset()
            if table.sheet_name == "Checks":
                logical_values.append(formula)
                formula_indexes = frozenset({len(logical_values) - 1})
            _update_logical_sheet_hash(
                logical_digest,
                logical_values,
                formula_indexes=formula_indexes,
            )
        if len(workbook_columns) <= 200:
            sheet.auto_filter.ref = (
                f"A1:{openpyxl.utils.get_column_letter(len(workbook_columns))}{row_count + 1}"
            )
        sheets.append(
            {
                "sheet_name": table.sheet_name,
                "row_count": row_count,
                "column_count": len(workbook_columns),
                "logical_sha256": logical_digest.hexdigest(),
                "logical_hash_version": XLSX_LOGICAL_HASH_VERSION,
            }
        )
    workbook.save(raw_path)
    return sheets, sum(int(record["row_count"]) for record in sheets)


def _lifecycle_percentiles(
    calculation: _SealedCalculation,
    field_name: str,
) -> tuple[float | None, float | None, float | None]:
    if field_name not in calculation.by_name:
        return None, None, None
    values = _finite_values(calculation.by_name[field_name])
    if not len(values):
        return None, None, None
    percentiles = np.quantile(values, (0.10, 0.50, 0.90), method="linear")
    return tuple(float(value) for value in percentiles)  # type: ignore[return-value]


def _lifecycle_decision_values(
    calculation: _SealedCalculation,
    request_payload: Mapping[str, Any],
    metadata: Mapping[str, Any],
) -> dict[str, Any]:
    summaries = _lifecycle_summaries(metadata)
    decision = summaries.get("headline_decision") or summaries.get("decision")
    decision = decision if isinstance(decision, Mapping) else {}
    probability_counts = summaries.get("probability_counts")
    probability_counts = (
        probability_counts if isinstance(probability_counts, Mapping) else {}
    )
    npv_counts = probability_counts.get("upgrade_npv")
    npv_counts = npv_counts if isinstance(npv_counts, Mapping) else {}
    delta_lcoe_counts = probability_counts.get("delta_lcoe")
    delta_lcoe_counts = (
        delta_lcoe_counts if isinstance(delta_lcoe_counts, Mapping) else {}
    )
    absolute_tolerance = _lifecycle_economic_npv_tolerance(request_payload)
    if "UpgradeNPV_se_minus_sol_USD" in calculation.by_name:
        raw_npv = np.asarray(
            calculation.by_name["UpgradeNPV_se_minus_sol_USD"],
            dtype=np.float64,
        )
    else:
        raw_npv = np.asarray([], dtype=np.float64)
    raw_tolerances = np.asarray(
        calculation.by_name.get(
            "NPVTolerance_USD",
            np.full(len(raw_npv), absolute_tolerance, dtype=np.float64),
        ),
        dtype=np.float64,
    )
    if raw_tolerances.shape != raw_npv.shape:
        raise TechnoeconomicExportError(
            "The v6 NPV decision-tolerance vector is inconsistent"
        )
    finite_mask = np.isfinite(raw_npv) & np.isfinite(raw_tolerances)
    npv = raw_npv[finite_mask]
    tolerances = raw_tolerances[finite_mask]
    positive = int(
        npv_counts.get("positive")
        if npv_counts.get("positive") is not None
        else np.count_nonzero(npv > tolerances)
    )
    negative = int(
        npv_counts.get("negative")
        if npv_counts.get("negative") is not None
        else np.count_nonzero(npv < -tolerances)
    )
    tie = int(
        npv_counts.get("tie")
        if npv_counts.get("tie") is not None
        else np.count_nonzero((npv >= -tolerances) & (npv <= tolerances))
    )
    denominator = int(npv_counts.get("denominator") or len(npv))
    threshold = float(
        _lifecycle_request(request_payload).get("decision_probability_threshold")
        or 0.75
    )
    if denominator and positive / denominator >= threshold:
        derived_decision = "SolarEdge preferred"
    elif denominator and negative / denominator >= threshold:
        derived_decision = "Solectria preferred"
    else:
        derived_decision = "No decisive winner"
    headline = (
        decision.get("headline_decision") or decision.get("headline") or decision.get("decision")
        or derived_decision
    )
    warnings = summaries.get("warnings") or metadata.get("warnings") or []
    if isinstance(warnings, str):
        warnings = [warnings]
    if not isinstance(warnings, (list, tuple)):
        warnings = []
    warning_text = tuple(
        _canonical_json_text(value)
        if isinstance(value, (Mapping, list, tuple))
        else str(value)
        for value in warnings
    )
    warning_codes = tuple(
        str(value.get("code"))
        for value in warnings
        if isinstance(value, Mapping) and value.get("code") is not None
    )
    quadrants = summaries.get("cost_energy_quadrants")
    quadrants = quadrants if isinstance(quadrants, Mapping) else {}
    quadrant_rows = [
        (
            str(name),
            int(record.get("count") or 0),
            float(record.get("probability") or 0.0),
        )
        for name, record in quadrants.items()
        if isinstance(record, Mapping)
    ]
    dominant_quadrant = (
        sorted(quadrant_rows, key=lambda row: (-row[2], row[0]))[0]
        if quadrant_rows
        else ("unavailable", 0, 0.0)
    )
    reliability_rows = _mapping_sequence(
        summaries.get("reliability_summary"),
        label="The v6 reliability summary",
    )
    corrective_by_component: dict[str, float] = {}
    for row in reliability_rows:
        if (
            row.get("mode") != "event"
            or row.get("metric") != "corrective_cost"
            or row.get("statistic") != "p50"
        ):
            continue
        identity = f"{row.get('system')}::{row.get('component_id')}"
        corrective_by_component[identity] = corrective_by_component.get(
            identity, 0.0
        ) + float(row.get("value") or 0.0)
    reliability_contributor = (
        sorted(
            corrective_by_component.items(),
            key=lambda item: (-item[1], item[0]),
        )[0]
        if corrective_by_component
        else ("unavailable", 0.0)
    )
    convergence = metadata.get("convergence")
    convergence = convergence if isinstance(convergence, Mapping) else {}
    tolerance_percentiles = (
        tuple(
            float(value)
            for value in np.quantile(
                tolerances,
                (0.10, 0.50, 0.90),
                method="linear",
            )
        )
        if len(tolerances)
        else (absolute_tolerance, absolute_tolerance, absolute_tolerance)
    )
    return {
        "headline": headline,
        "threshold": threshold,
        "tolerance": tolerance_percentiles[1],
        "tolerance_percentiles": tolerance_percentiles,
        "positive_count": positive,
        "negative_count": negative,
        "tie_count": tie,
        "denominator": denominator,
        "positive_probability": positive / denominator if denominator else None,
        "negative_probability": negative / denominator if denominator else None,
        "tie_probability": tie / denominator if denominator else None,
        "npv_percentiles": _lifecycle_percentiles(
            calculation, "UpgradeNPV_se_minus_sol_USD"
        ),
        "delta_lcoe_positive_count": int(delta_lcoe_counts.get("positive") or 0),
        "delta_lcoe_negative_count": int(delta_lcoe_counts.get("negative") or 0),
        "delta_lcoe_tie_count": int(delta_lcoe_counts.get("tie") or 0),
        "delta_lcoe_denominator": int(delta_lcoe_counts.get("denominator") or 0),
        "warnings": warning_text,
        "warning_codes": warning_codes,
        "dominant_quadrant": dominant_quadrant,
        "reliability_contributor": reliability_contributor,
        "convergence_status": convergence.get("status") or "unavailable",
        "suppressed": decision.get("status") == "suppressed"
        or bool(decision.get("suppressed")),
        "reason_codes": _safe_public_value(decision.get("reason_codes") or []),
    }


def _lifecycle_chart_rows(
    calculation: _SealedCalculation,
    request_payload: Mapping[str, Any],
    metadata: Mapping[str, Any],
) -> tuple[list[tuple[Any, ...]], list[tuple[Any, ...]], list[tuple[Any, ...]]]:
    decision = _lifecycle_decision_values(calculation, request_payload, metadata)
    denominator = int(decision["denominator"] or 0)
    delta_lcoe_denominator = int(decision["delta_lcoe_denominator"] or 0)
    decision_rows = [
        (
            label,
            npv_count,
            delta_lcoe_count,
            npv_count / denominator if denominator else 0.0,
            (
                delta_lcoe_count / delta_lcoe_denominator
                if delta_lcoe_denominator
                else 0.0
            ),
        )
        for label, npv_count, delta_lcoe_count in (
            (
                "Positive / higher",
                decision["positive_count"],
                decision["delta_lcoe_positive_count"],
            ),
            ("Tie", decision["tie_count"], decision["delta_lcoe_tie_count"]),
            (
                "Negative / lower",
                decision["negative_count"],
                decision["delta_lcoe_negative_count"],
            ),
        )
    ]

    annual_by_year: dict[int, dict[str, Any]] = {}
    reliability_by_component: dict[str, dict[str, float]] = {}
    for record in _representative_trace_records(metadata):
        if record.get("selection_label") != "NPV-P50":
            continue
        if record.get("record_type") == "annual":
            year = int(record.get("project_year") or 0)
            system = str(record.get("system") or "").lower()
            key = "se" if "edge" in system or system == "se" else "so"
            row = annual_by_year.setdefault(year, {"year": year})
            row[f"{key}_energy"] = record.get("delivered_energy_kwh")
            row[f"{key}_availability"] = record.get("target_availability")
            row[f"{key}_cost"] = record.get("annual_cost_usd")
            if record.get("cumulative_upgrade_npv_usd") is not None:
                row["cumulative_npv"] = record.get("cumulative_upgrade_npv_usd")
        elif record.get("record_type") == "component":
            component = f"{record.get('system')}::{record.get('component_id')}"
            row = reliability_by_component.setdefault(
                component,
                {"failures": 0.0, "downtime": 0.0, "corrective": 0.0},
            )
            row["failures"] += float(record.get("event_failures") or 0.0)
            if record.get("component_year_total_row") is True:
                row["downtime"] += float(record.get("downtime_fraction") or 0.0)
                row["corrective"] += float(
                    record.get("corrective_cost_usd") or 0.0
                )
    annual_rows = [
        (
            year,
            row.get("so_energy"),
            row.get("se_energy"),
            row.get("so_availability"),
            row.get("se_availability"),
            row.get("so_cost"),
            row.get("se_cost"),
            row.get("cumulative_npv"),
        )
        for year, row in sorted(annual_by_year.items())
    ]
    reliability_rows = [
        (component, values["failures"], values["downtime"], values["corrective"])
        for component, values in sorted(reliability_by_component.items())
    ]
    return decision_rows, annual_rows, reliability_rows


def _write_lifecycle_summary_sheet(
    sheet: Any,
    calculation: _SealedCalculation,
    request_payload: Mapping[str, Any],
    routine_result: Mapping[str, Any],
    metadata: Mapping[str, Any],
    checks: Sequence[Sequence[Any]],
) -> tuple[int, int, str]:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"
    for letter, width in {"A": 56, "B": 25, "C": 25, "D": 72}.items():
        sheet.column_dimensions[letter].width = width
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 32
    digest = _new_lifecycle_logical_sheet_hash()

    title_values = ("TEA v6 — Upgrade NPV Decision", None, None, None)
    title = _write_only_cell(sheet, title_values[0], section=True)
    title.font = Font(name="Aptos Display", size=18, bold=True, color="FFFFFF")
    title.fill = _HEADER_FILL
    title.alignment = Alignment(vertical="center", wrap_text=False)
    sheet.append(
        [title, _write_only_cell(sheet, None), _write_only_cell(sheet, None), _write_only_cell(sheet, None)]
    )
    _update_lifecycle_logical_sheet_hash(digest, title_values)
    note = (
        "Kernel-written frozen values are authoritative; Excel formulas are audit replicas.",
        None,
        None,
        "All signed quantities are SolarEdge minus Solectria; positive Upgrade NPV favors SolarEdge.",
    )
    sheet.append([_write_only_cell(sheet, value) for value in note])
    _update_lifecycle_logical_sheet_hash(digest, note)
    blank = (None, None, None, None)
    sheet.append([_write_only_cell(sheet, None) for _ in blank])
    _update_lifecycle_logical_sheet_hash(digest, blank)
    headers = ("Decision item", "Frozen authority", "Formula/status", "Interpretation")
    _append_header(sheet, headers)
    _update_lifecycle_logical_sheet_hash(digest, headers)

    decision = _lifecycle_decision_values(calculation, request_payload, metadata)
    p10, p50, p90 = decision["npv_percentiles"]
    tolerance_p10, tolerance_p50, tolerance_p90 = decision[
        "tolerance_percentiles"
    ]
    _, so_lcoe_p50, _ = _lifecycle_percentiles(
        calculation,
        "LifecycleLCOE_SOL_USD_per_kWh_AC",
    )
    _, se_lcoe_p50, _ = _lifecycle_percentiles(
        calculation,
        "LifecycleLCOE_SE_USD_per_kWh_AC",
    )
    _, lcoo_p50, _ = _lifecycle_percentiles(
        calculation,
        "IncrementalLCOO_se_minus_sol_USD_per_kWh_AC",
    )
    dominant_quadrant, quadrant_count, quadrant_probability = decision[
        "dominant_quadrant"
    ]
    reliability_component, reliability_cost = decision[
        "reliability_contributor"
    ]
    check_end = len(checks) + 1
    rows = (
        (
            "Model status",
            "OK" if all(row[5] == "OK" for row in checks) else "FAIL",
            f'=IF(COUNTIF(\'Checks\'!F2:F{check_end},"FAIL")=0,"OK","FAIL")',
            "Failed export checks suppress the recommendation.",
        ),
        (
            "Headline decision",
            decision["headline"],
            None,
            f"Requires at least {float(decision['threshold']):.0%} probability beyond the economic NPV tolerance.",
        ),
        ("P(NPV > tolerance)", decision["positive_probability"], None, "SolarEdge-favorable probability."),
        ("P(NPV < -tolerance)", decision["negative_probability"], None, "Solectria-favorable probability."),
        ("P(|NPV| <= tolerance)", decision["tie_probability"], None, "Economically tied probability."),
        ("Economic NPV tolerance P10 (USD)", tolerance_p10, None, "Scale-aware decision materiality threshold; never used for formula tie-outs."),
        ("Economic NPV tolerance P50 (USD)", tolerance_p50, None, "Median scale-aware decision materiality threshold."),
        ("Economic NPV tolerance P90 (USD)", tolerance_p90, None, "Upper scale-aware decision materiality threshold."),
        ("Upgrade NPV P10 (USD)", p10, None, "10th percentile."),
        ("Upgrade NPV P50 (USD)", p50, None, "Median."),
        ("Upgrade NPV P90 (USD)", p90, None, "90th percentile."),
        ("Solectria standalone LCOE P50 (USD/kWh_AC)", so_lcoe_p50, None, "Supporting standalone metric."),
        ("SolarEdge standalone LCOE P50 (USD/kWh_AC)", se_lcoe_p50, None, "Supporting standalone metric."),
        ("Incremental LCOO P50 (USD/kWh_AC)", lcoo_p50, None, "Diagnostic only; undefined near zero incremental energy."),
        (
            "Dominant cost/energy quadrant",
            dominant_quadrant,
            None,
            f"{quadrant_count} realizations ({quadrant_probability:.1%}); signed SolarEdge-minus-Solectria classification.",
        ),
        (
            "Largest P50 event corrective-cost contributor",
            reliability_component,
            None,
            f"Summed annual P50 corrective cost: {reliability_cost:,.2f} constant USD.",
        ),
        (
            "Convergence condition",
            decision["convergence_status"],
            None,
            "An unstable condition suppresses the headline decision.",
        ),
        (
            "Provisional-input condition",
            "present" if "accepted_provisional_inputs" in decision["warning_codes"] else "none",
            None,
            "Accepted provisional evidence remains visibly flagged and requires review.",
        ),
        (
            "Sensitivity interpretation",
            "association_not_causation",
            None,
            "Rank sensitivity describes association, not causation.",
        ),
        ("Realizations", routine_result.get("realization_count"), None, "Complete sealed realization population."),
        ("Calculation contract", routine_result.get("calculation_contract_version"), None, "Additive v6 lifecycle path."),
        (
            "Weather-path method",
            _lifecycle_request(request_payload).get("weather_path_method"),
            None,
            "V6 balances weather years across realizations and samples project years independently; versus v5's one-weather-year-for-life method, this methodological change generally narrows interannual-weather uncertainty and is not a causal performance claim.",
        ),
        ("Sampling contract", routine_result.get("sampling_version"), None, "Balanced weather and domain-separated random streams."),
        ("Decision rule", LIFECYCLE_DECISION_RULE_VERSION, None, _canonical_json_text(decision["reason_codes"])),
        ("Warnings", "; ".join(decision["warnings"]), None, "Review all warnings and provisional inputs before use."),
    )
    for label, value, formula, notes in rows:
        status = value if label == "Model status" else None
        frozen = _write_only_cell(sheet, value, status=status)
        if label != "Model status":
            frozen.fill = _V6_SEALED_FILL
        elif value == "FAIL":
            frozen.fill = _V6_FAIL_FILL
        formula_cell = _write_only_cell(sheet, formula)
        if isinstance(formula, str) and formula.startswith("="):
            formula_cell.data_type = "f"
            formula_cell.font = Font(name="Aptos", size=10, color="000000")
        cells = [
            _write_only_cell(sheet, label),
            frozen,
            formula_cell,
            _write_only_cell(sheet, notes),
        ]
        warning_row = (
            (label == "Warnings" and bool(value))
            or (label == "Provisional-input condition" and value == "present")
            or (label == "Convergence condition" and value != "stable")
            or (label == "Headline decision" and decision["suppressed"])
        )
        if warning_row:
            for cell in cells:
                cell.fill = _V6_WARNING_FILL
        sheet.append(cells)
        _update_lifecycle_logical_sheet_hash(
            digest,
            (label, value, formula, notes),
            formula_identities={2: "summary_check_status"} if formula else {},
        )
    return len(rows) + 4, 4, digest.hexdigest()


def _write_lifecycle_chart_sheet(
    sheet: Any,
    *,
    title: str,
    interpretation: str,
    columns: Sequence[str],
    rows: Sequence[Sequence[Any]],
    charts: Sequence[Mapping[str, Any]],
    image_path: Path | None = None,
    image_anchor: str | None = None,
) -> tuple[int, int, str, list[dict[str, Any]], list[dict[str, Any]]]:
    _set_sheet_layout(sheet, columns)
    sheet.freeze_panes = "A4"
    sheet.column_dimensions["A"].width = max(
        float(sheet.column_dimensions["A"].width or 0),
        52.0,
    )
    digest = _new_lifecycle_logical_sheet_hash()
    title_values = (title, *([None] * (len(columns) - 1)))
    title_cell = _write_only_cell(sheet, title, section=True)
    title_cell.font = Font(name="Aptos Display", size=16, bold=True, color="FFFFFF")
    title_cell.fill = _HEADER_FILL
    title_cell.alignment = Alignment(vertical="center", wrap_text=False)
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 48
    sheet.append(
        [title_cell, *[_write_only_cell(sheet, None) for _ in columns[1:]]]
    )
    _update_lifecycle_logical_sheet_hash(digest, title_values)
    interpretation_values = (
        interpretation,
        *([None] * (len(columns) - 1)),
    )
    interpretation_cells = [
        _write_only_cell(sheet, value) for value in interpretation_values
    ]
    interpretation_cells[0].alignment = Alignment(vertical="top", wrap_text=True)
    sheet.append(interpretation_cells)
    _update_lifecycle_logical_sheet_hash(digest, interpretation_values)
    blank = tuple(None for _ in columns)
    sheet.append([_write_only_cell(sheet, None) for _ in columns])
    _update_lifecycle_logical_sheet_hash(digest, blank)
    _append_header(sheet, columns)
    _update_lifecycle_logical_sheet_hash(digest, columns)
    for row in rows:
        sheet.append(
            [
                _write_only_cell(
                    sheet,
                    value,
                    number_format=_number_format_for_header(header, value),
                )
                for header, value in zip(columns, row)
            ]
        )
        _update_lifecycle_logical_sheet_hash(digest, row)
    max_row = max(5, len(rows) + 4)
    chart_records: list[dict[str, Any]] = []
    for specification in charts:
        chart = BarChart() if specification.get("type") == "bar" else LineChart()
        chart.title = str(specification["title"])
        chart.style = 10
        chart.height = 7.2
        chart.width = 11.8
        chart.y_axis.title = str(specification.get("y_title") or "")
        chart.x_axis.title = str(specification.get("x_title") or "")
        if specification.get("y_number_format"):
            chart.y_axis.numFmt = str(specification["y_number_format"])
        if specification.get("y_min") is not None:
            chart.y_axis.scaling.min = float(specification["y_min"])
        if specification.get("y_max") is not None:
            chart.y_axis.scaling.max = float(specification["y_max"])
        data = Reference(
            sheet,
            min_col=int(specification["min_col"]),
            max_col=int(specification.get("max_col") or specification["min_col"]),
            min_row=4,
            max_row=max_row,
        )
        categories = Reference(
            sheet,
            min_col=int(specification.get("category_col") or 1),
            min_row=5,
            max_row=max_row,
        )
        chart.add_data(data, titles_from_data=True)
        first_series_column = int(specification["min_col"])
        for offset, series in enumerate(chart.series):
            source_column = first_series_column + offset
            if source_column <= len(columns):
                series_title = str(columns[source_column - 1]).replace("_", " ").title()
                for source, target in (
                    ("Npv", "NPV"),
                    ("Lcoe", "LCOE"),
                    ("Kwh", "kWh"),
                    ("Usd", "USD"),
                ):
                    series_title = series_title.replace(source, target)
                series.tx = SeriesLabel(
                    v=series_title
                )
        chart.set_categories(categories)
        chart.legend.position = "b"
        anchor = str(specification["anchor"])
        sheet.add_chart(chart, anchor)
        chart_records.append(
            {
                "sheet_name": sheet.title,
                "title": str(specification["title"]),
                "source_range": (
                    f"{openpyxl.utils.get_column_letter(int(specification['min_col']))}4:"
                    f"{openpyxl.utils.get_column_letter(int(specification.get('max_col') or specification['min_col']))}{max_row}"
                ),
                "category_range": (
                    f"{openpyxl.utils.get_column_letter(int(specification.get('category_col') or 1))}5:"
                    f"{openpyxl.utils.get_column_letter(int(specification.get('category_col') or 1))}{max_row}"
                ),
                "anchor": anchor,
            }
        )
    image_records: list[dict[str, Any]] = []
    if image_path is not None and image_anchor is not None:
        image = XLImage(str(image_path))
        original_width = float(image.width)
        original_height = float(image.height)
        image.width = min(original_width, 820.0)
        if original_width:
            image.height = image.width * original_height / original_width
        sheet.add_image(image, image_anchor)
        image_records.append(
            {
                "sheet_name": sheet.title,
                "anchor": image_anchor,
                "sha256": _sha256_file(image_path),
                "filename": image_path.name,
            }
        )
    sheet.auto_filter.ref = (
        f"A4:{openpyxl.utils.get_column_letter(len(columns))}{max_row}"
    )
    return (
        len(rows) + 4,
        len(columns),
        digest.hexdigest(),
        chart_records,
        image_records,
    )


def _write_lifecycle_table_sheet(
    sheet: Any,
    table: _Table,
    cancellation_check: Callable[[], None],
    *,
    image_path: Path | None = None,
    image_anchor: str = "J2",
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    workbook_columns = (
        table.columns + ("display_formula_status",)
        if table.sheet_name == "Checks"
        else table.columns
    )
    _set_sheet_layout(sheet, workbook_columns)
    if table.sheet_name == "Formula Catalog":
        for letter, width in {
            "A": 16,
            "B": 30,
            "C": 62,
            "D": 62,
            "E": 40,
            "F": 24,
            "G": 18,
            "H": 38,
            "I": 26,
            "J": 20,
        }.items():
            sheet.column_dimensions[letter].width = width
    elif table.sheet_name == "Calculation Audit":
        for letter, width in {
            "A": 48,
            "B": 16,
            "C": 16,
            "D": 16,
            "E": 16,
            "F": 16,
            "G": 26,
            "H": 22,
            "I": 54,
            "J": 20,
            "K": 22,
            "L": 16,
            "M": 26,
            "N": 72,
        }.items():
            sheet.column_dimensions[letter].width = width
    fast_streaming = table.sheet_name in _FAST_STREAMING_SHEETS
    if fast_streaming:
        _apply_fast_streaming_formats(sheet, workbook_columns)
    _append_header(sheet, workbook_columns)
    digest = _new_lifecycle_logical_sheet_hash()
    _update_lifecycle_logical_sheet_hash(digest, workbook_columns)
    row_count = 0
    for row_count, raw_row in enumerate(table.rows_factory(), start=1):
        if len(raw_row) != len(table.columns):
            raise TechnoeconomicExportError(
                f"Workbook table {table.sheet_name} produced an inconsistent row width"
            )
        if row_count % _CANCEL_INTERVAL == 0:
            cancellation_check()
        if fast_streaming:
            cells = [_fast_streaming_value(sheet, value) for value in raw_row]
        else:
            cells = [
                _write_only_cell(
                    sheet,
                    value,
                    status=(
                        value
                        if header in {"status", "status_authority"}
                        else None
                    ),
                    number_format=_number_format_for_header(header, value),
                )
                for header, value in zip(table.columns, raw_row)
            ]
            if table.sheet_name in _V6_REFERENCE_SHEETS:
                for cell, value in zip(cells, raw_row):
                    if value is not None:
                        cell.font = _V6_REFERENCE_FONT
            for cell, value in zip(cells, raw_row):
                if "provisional" in str(value).lower():
                    cell.fill = _V6_WARNING_FILL
        logical_values = list(raw_row)
        formula_identities: dict[int, str] = {}
        if table.sheet_name == "Calculation Audit":
            excel_row = row_count + 1
            formula_index = table.columns.index("formula_replica")
            difference_index = table.columns.index("difference")
            status_index = table.columns.index("status")
            formula = str(raw_row[formula_index])
            cells[formula_index] = _write_only_cell(sheet, formula)
            cells[formula_index].data_type = "f"
            cells[formula_index].font = Font(name="Aptos", size=10, color="000000")
            difference_formula = f"=I{excel_row}-H{excel_row}"
            cells[difference_index] = _write_only_cell(sheet, difference_formula)
            cells[difference_index].data_type = "f"
            cells[difference_index].font = Font(name="Aptos", size=10, color="000000")
            status_formula = (
                f'=IF(ABS(J{excel_row})<=K{excel_row},"OK","FAIL")'
            )
            cells[status_index] = _write_only_cell(sheet, status_formula)
            cells[status_index].data_type = "f"
            cells[status_index].font = Font(name="Aptos", size=10, color="000000")
            cells[status_index].fill = (
                _PASS_FILL if raw_row[status_index] == "OK" else _V6_FAIL_FILL
            )
            logical_values[difference_index] = difference_formula
            logical_values[status_index] = status_formula
            formula_identifier = str(
                raw_row[table.columns.index("formula_id")]
            )
            formula_identities.update(
                {
                    formula_index: f"registry:{formula_identifier}",
                    difference_index: "audit_difference",
                    status_index: "audit_binary64_status",
                }
            )
            cells[table.columns.index("frozen_authority")].fill = _V6_SEALED_FILL
        if table.sheet_name == "Checks":
            excel_row = row_count + 1
            formula = (
                f'=IF(OR(ISTEXT(B{excel_row}),ISTEXT(C{excel_row})),'
                f'IF(B{excel_row}=C{excel_row},"OK","FAIL"),'
                f'IF(ABS(B{excel_row}-C{excel_row})<=E{excel_row},"OK","FAIL"))'
            )
            formula_cell = _write_only_cell(sheet, formula)
            formula_cell.data_type = "f"
            formula_cell.font = Font(name="Aptos", size=10, color="000000")
            cells.append(formula_cell)
            logical_values.append(formula)
            formula_identities[len(logical_values) - 1] = "check_status"
            status_index = table.columns.index("status_authority")
            if raw_row[status_index] == "FAIL":
                cells[status_index].fill = _V6_FAIL_FILL
        sheet.append(cells)
        _update_lifecycle_logical_sheet_hash(
            digest,
            logical_values,
            formula_identities=formula_identities,
        )
    if len(workbook_columns) <= 200:
        sheet.auto_filter.ref = (
            f"A1:{openpyxl.utils.get_column_letter(len(workbook_columns))}{row_count + 1}"
        )
    image_records: list[dict[str, Any]] = []
    if image_path is not None:
        image = XLImage(str(image_path))
        original_width = float(image.width)
        original_height = float(image.height)
        image.width = min(original_width, 900.0)
        if original_width:
            image.height = image.width * original_height / original_width
        sheet.add_image(image, image_anchor)
        image_records.append(
            {
                "sheet_name": sheet.title,
                "anchor": image_anchor,
                "sha256": _sha256_file(image_path),
                "filename": image_path.name,
            }
        )
    return (
        {
            "sheet_name": table.sheet_name,
            "row_count": row_count,
            "column_count": len(workbook_columns),
            "logical_sha256": digest.hexdigest(),
            "logical_hash_version": LIFECYCLE_XLSX_LOGICAL_HASH_VERSION,
        },
        image_records,
    )


def _write_lifecycle_workbook(
    raw_path: Path,
    tables: Sequence[_Table],
    calculation: _SealedCalculation,
    request_payload: Mapping[str, Any],
    routine_result: Mapping[str, Any],
    metadata: Mapping[str, Any],
    checks: Sequence[Sequence[Any]],
    registry: Sequence[Mapping[str, Any]],
    cancellation_check: Callable[[], None],
    *,
    image_paths: Mapping[str, Path],
) -> tuple[list[dict[str, Any]], int, dict[str, Any]]:
    workbook = openpyxl.Workbook(write_only=True)
    workbook.properties.creator = "SBE PV technoeconomic reporting"
    workbook.properties.title = "TEA v6 Sealed-Audit Workbook"
    workbook.properties.subject = "Frozen lifecycle results and formula replicas"
    fixed_time = datetime(1980, 1, 1)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    sheets_by_name = {
        name: workbook.create_sheet(name) for name in LIFECYCLE_WORKBOOK_SHEET_ORDER
    }
    if tuple(sheets_by_name) != LIFECYCLE_WORKBOOK_SHEET_ORDER:
        raise TechnoeconomicExportError("The v6 workbook sheet order is invalid")

    sheet_records: list[dict[str, Any]] = []
    summary_rows, summary_columns, summary_hash = _write_lifecycle_summary_sheet(
        sheets_by_name["Summary"],
        calculation,
        request_payload,
        routine_result,
        metadata,
        checks,
    )
    sheet_records.append(
        {
            "sheet_name": "Summary",
            "row_count": summary_rows,
            "column_count": summary_columns,
            "logical_sha256": summary_hash,
            "logical_hash_version": LIFECYCLE_XLSX_LOGICAL_HASH_VERSION,
        }
    )

    decision_rows, annual_rows, reliability_rows = _lifecycle_chart_rows(
        calculation, request_payload, metadata
    )
    chart_contracts: list[dict[str, Any]] = []
    image_contracts: list[dict[str, Any]] = []
    chart_specs = (
        (
            "Decision Charts",
            "Upgrade NPV and ΔLCOE decision probabilities",
            "Positive NPV favors SolarEdge; lower ΔLCOE favors SolarEdge; each tie uses its own scale-aware economic tolerance.",
            (
                "outcome class",
                "upgrade_npv_count",
                "delta_lcoe_count",
                "upgrade_npv_probability",
                "delta_lcoe_probability",
            ),
            decision_rows,
            (
                {"type": "bar", "title": "Decision outcome counts", "min_col": 2, "max_col": 3, "category_col": 1, "anchor": "G4", "y_title": "Realizations"},
                {"type": "bar", "title": "Decision outcome probabilities", "min_col": 4, "max_col": 5, "category_col": 1, "anchor": "G19", "y_title": "Probability", "y_number_format": "0%", "y_min": 0.0, "y_max": 1.0},
            ),
            image_paths.get("cdf_plot"),
            "A40",
        ),
        (
            "Lifecycle Charts",
            "P50 representative lifecycle",
            "These native charts use the sealed NPV-P50 trace; they are representative, not expected-value forecasts.",
            ("project_year", "solectria_energy_kwh", "solaredge_energy_kwh", "solectria_availability", "solaredge_availability", "solectria_cost_usd", "solaredge_cost_usd", "cumulative_upgrade_npv_usd"),
            annual_rows,
            (
                {"type": "line", "title": "Annual delivered energy", "min_col": 2, "max_col": 3, "category_col": 1, "anchor": "J4", "y_title": "kWh_AC"},
                {"type": "line", "title": "Target availability", "min_col": 4, "max_col": 5, "category_col": 1, "anchor": "J19", "y_title": "Availability", "y_number_format": "0.0%"},
                {"type": "line", "title": "Annual lifecycle cost", "min_col": 6, "max_col": 7, "category_col": 1, "anchor": "S4", "y_title": "constant USD"},
                {"type": "line", "title": "Cumulative Upgrade NPV", "min_col": 8, "category_col": 1, "anchor": "S19", "y_title": "constant USD"},
            ),
            None,
            None,
        ),
        (
            "Reliability Charts",
            "P50 representative reliability contribution",
            "Event failures are the headline reliability result; expected failures remain a diagnostic.",
            ("system_component", "event_failures", "downtime_fraction", "corrective_cost_usd"),
            reliability_rows,
            (
                {"type": "bar", "title": "Event failures by component", "min_col": 2, "category_col": 1, "anchor": "F4", "y_title": "Failures"},
                {"type": "bar", "title": "Downtime by component", "min_col": 3, "category_col": 1, "anchor": "F19", "y_title": "Fraction", "y_number_format": "0.000%"},
                {"type": "bar", "title": "Corrective cost by component", "min_col": 4, "category_col": 1, "anchor": "O4", "y_title": "constant USD"},
            ),
            None,
            None,
        ),
    )
    for (
        sheet_name,
        title,
        interpretation,
        columns,
        rows,
        charts,
        image_path,
        image_anchor,
    ) in chart_specs:
        result = _write_lifecycle_chart_sheet(
            sheets_by_name[sheet_name],
            title=title,
            interpretation=interpretation,
            columns=columns,
            rows=rows,
            charts=charts,
            image_path=image_path,
            image_anchor=image_anchor,
        )
        row_count, column_count, logical_hash, chart_rows, image_rows = result
        sheet_records.append(
            {
                "sheet_name": sheet_name,
                "row_count": row_count,
                "column_count": column_count,
                "logical_sha256": logical_hash,
                "logical_hash_version": LIFECYCLE_XLSX_LOGICAL_HASH_VERSION,
            }
        )
        chart_contracts.extend(chart_rows)
        image_contracts.extend(image_rows)

    table_by_sheet = {table.sheet_name: table for table in tables}
    expected_table_sheets = LIFECYCLE_WORKBOOK_SHEET_ORDER[4:]
    if set(table_by_sheet) != set(expected_table_sheets):
        raise TechnoeconomicExportError(
            "The v6 workbook table set does not match its sealed-audit schema"
        )
    for sheet_name in expected_table_sheets:
        cancellation_check()
        image_path = None
        image_anchor = "J2"
        if sheet_name == "Sensitivity":
            image_path = image_paths.get("sensitivity_plot")
            image_anchor = "X2"
        elif sheet_name == "Convergence":
            image_path = image_paths.get("convergence_plot")
            image_anchor = "X2"
        record, image_rows = _write_lifecycle_table_sheet(
            sheets_by_name[sheet_name],
            table_by_sheet[sheet_name],
            cancellation_check,
            image_path=image_path,
            image_anchor=image_anchor,
        )
        sheet_records.append(record)
        image_contracts.extend(image_rows)
    if tuple(record["sheet_name"] for record in sheet_records) != LIFECYCLE_WORKBOOK_SHEET_ORDER:
        raise TechnoeconomicExportError("The v6 workbook sheet manifest is out of order")
    workbook.save(raw_path)
    registry_hash = _formula_registry_sha256(registry)
    return (
        sheet_records,
        sum(int(record["row_count"]) for record in sheet_records),
        {
            "formula_registry_version": getattr(
                technoeconomic_kernel, "FORMULA_REGISTRY_VERSION", "tea-formulas-v6"
            ),
            "formula_registry_count": len(registry),
            "formula_registry_sha256": registry_hash,
            "formula_template_hash_version": LIFECYCLE_FORMULA_TEMPLATE_HASH_VERSION,
            "formula_template_sha256": _formula_template_sha256(registry),
            "native_charts": chart_contracts,
            "embedded_images": image_contracts,
            "decision_rule_version": LIFECYCLE_DECISION_RULE_VERSION,
        },
    )


def _normalize_xlsx_archive(
    source: Path,
    target: Path,
    cancellation_check: Callable[[], None],
    *,
    canonicalize_core_properties: bool = False,
) -> None:
    with zipfile.ZipFile(source, "r") as incoming, zipfile.ZipFile(
        target,
        "x",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=9,
        strict_timestamps=True,
    ) as outgoing:
        for name in sorted(incoming.namelist()):
            info = _zip_info(name, nested=True)
            with incoming.open(name, "r") as source_handle, outgoing.open(
                info, "w"
            ) as target_handle:
                if canonicalize_core_properties and name == "docProps/core.xml":
                    cancellation_check()
                    normalized, replacement_count = re.subn(
                        rb"(<dcterms:modified\b[^>]*>)[^<]*(</dcterms:modified>)",
                        rb"\g<1>1980-01-01T00:00:00Z\g<2>",
                        source_handle.read(),
                        count=1,
                    )
                    if replacement_count != 1:
                        raise TechnoeconomicExportError(
                            "The v6 workbook core-properties timestamp is invalid"
                        )
                    target_handle.write(normalized)
                else:
                    _copy_with_cancellation(
                        source_handle,
                        target_handle,
                        cancellation_check,
                    )


def _validate_lifecycle_xlsx(path: Path) -> dict[str, Any]:
    """Fail closed on unsafe or broken v6 workbook formula structures."""

    formula_count = 0
    forbidden_functions = re.compile(
        r"\b(?:INDIRECT|OFFSET|RAND|RANDBETWEEN|NOW|TODAY)\s*\(",
        re.IGNORECASE,
    )
    with zipfile.ZipFile(path, "r") as archive:
        names = archive.namelist()
        if any(name.startswith("xl/externalLinks/") for name in names):
            raise TechnoeconomicExportError(
                "The v6 workbook contains an external workbook link"
            )
        for name in names:
            if not name.startswith("xl/worksheets/") or not name.endswith(".xml"):
                continue
            xml = archive.read(name).decode("utf-8")
            if any(token in xml for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?")):
                raise TechnoeconomicExportError(
                    "The v6 workbook contains a formula error token"
                )
            formulas = re.findall(r"<f(?:\s[^>]*)?>(.*?)</f>", xml, flags=re.DOTALL)
            formula_count += len(formulas)
            for formula in formulas:
                normalized = (
                    formula.replace("&apos;", "'")
                    .replace("&quot;", '"')
                    .replace("&amp;", "&")
                    .replace("&lt;", "<")
                    .replace("&gt;", ">")
                )
                if "[" in normalized or forbidden_functions.search(normalized):
                    raise TechnoeconomicExportError(
                        "The v6 workbook contains an external or volatile formula"
                    )
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        if tuple(workbook.sheetnames) != LIFECYCLE_WORKBOOK_SHEET_ORDER:
            raise TechnoeconomicExportError(
                "The normalized v6 workbook sheet order changed"
            )
    finally:
        workbook.close()
    if formula_count <= 0:
        raise TechnoeconomicExportError("The v6 workbook contains no audit formulas")
    return {"formula_count": formula_count, "formula_scan_status": "passed"}


def _human_metric(
    value: str,
    *,
    applied_capacity_contract: bool = False,
) -> str:
    normalized_capacity_label = (
        "applied W" if applied_capacity_contract else "Wdc"
    )
    replacements = {
        technoeconomic_kernel.FIELD_LCOE_SOL: "Solectria lifecycle LCOE (USD/kWh_AC)",
        technoeconomic_kernel.FIELD_LCOE_SE: "SolarEdge lifecycle LCOE (USD/kWh_AC)",
        technoeconomic_kernel.FIELD_DELTA_COST: "Lifecycle cost delta, SE − SOL (USD/Wdc)",
        technoeconomic_kernel.FIELD_DELTA_ENERGY: "Lifecycle energy delta, SE − SOL (kWh_AC/Wdc)",
        technoeconomic_kernel.FIELD_DELTA_EA_COST: "Equivalent-annual cost delta, SE − SOL (USD/Wdc-year)",
        technoeconomic_kernel.FIELD_DELTA_EA_ENERGY: "Equivalent-annual energy delta, SE − SOL (kWh_AC/Wdc-year)",
        technoeconomic_kernel.APPLIED_FIELD_DELTA_COST: "Lifecycle cost delta, SE − SOL (USD/applied W)",
        technoeconomic_kernel.APPLIED_FIELD_DELTA_ENERGY: "Lifecycle energy delta, SE − SOL (kWh_AC/applied W)",
        technoeconomic_kernel.APPLIED_FIELD_DELTA_EA_COST: "Equivalent-annual cost delta, SE − SOL (USD/applied W-year)",
        technoeconomic_kernel.APPLIED_FIELD_DELTA_EA_ENERGY: "Equivalent-annual energy delta, SE − SOL (kWh_AC/applied W-year)",
        technoeconomic_kernel.COMMERCIAL_FIELD_TARGET_CAPACITY: "Commercial target capacity (W; explicit rating basis)",
        technoeconomic_kernel.COMMERCIAL_FIELD_YEAR1_DELTA_ENERGY: "Commercial first-year energy delta, SE − SOL (kWh_AC)",
        technoeconomic_kernel.COMMERCIAL_FIELD_LIFECYCLE_DELTA_ENERGY: "Commercial lifecycle energy delta, SE − SOL (kWh_AC)",
        technoeconomic_kernel.COMMERCIAL_FIELD_EA_DELTA_ENERGY: "Commercial equivalent-annual energy delta, SE − SOL (kWh_AC/year)",
        technoeconomic_kernel.COMMERCIAL_FIELD_LIFECYCLE_MARGINAL_COST: "Commercial lifecycle marginal cost delta, SE − SOL (constant USD)",
        technoeconomic_kernel.COMMERCIAL_FIELD_EA_MARGINAL_COST: "Commercial equivalent-annual marginal cost delta, SE − SOL (constant USD/year)",
        technoeconomic_kernel.COMMERCIAL_FIELD_MARGINAL_LCOO: "Commercial marginal LCOO, SE − SOL (constant USD/kWh_AC)",
        technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_TARGET_CAPACITY: "Commercial SolarEdge target capacity (W)",
        technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_CAPACITY_SCALE_FACTOR: "Commercial SolarEdge capacity scale factor (target W/source W)",
        technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_YEAR1_ENERGY: "Commercial SolarEdge first-year energy (kWh_AC)",
        technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LIFECYCLE_ENERGY: "Commercial SolarEdge lifecycle energy (kWh_AC)",
        technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_EA_ENERGY: "Commercial SolarEdge equivalent-annual energy (kWh_AC/year)",
        technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_INITIAL_COST: "Commercial SolarEdge initial cost (constant USD)",
        technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_RECURRING_PV_COST: "Commercial SolarEdge recurring lifecycle cost (constant USD)",
        technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_SCHEDULED_PV_COST: "Commercial SolarEdge scheduled lifecycle cost (constant USD)",
        technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LIFECYCLE_COST: "Commercial SolarEdge lifecycle cost (constant USD)",
        technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_EA_COST: "Commercial SolarEdge equivalent-annual cost (constant USD/year)",
        technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LCOE: "Commercial SolarEdge lifecycle LCOE (constant USD/kWh_AC)",
        technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_TARGET_CAPACITY: "Commercial Solectria target capacity (W)",
        technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_CAPACITY_SCALE_FACTOR: "Commercial Solectria capacity scale factor (target W/source W)",
        technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_YEAR1_ENERGY: "Commercial Solectria first-year energy (kWh_AC)",
        technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_LIFECYCLE_ENERGY: "Commercial Solectria lifecycle energy (kWh_AC)",
        technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_EA_ENERGY: "Commercial Solectria equivalent-annual energy (kWh_AC/year)",
        technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_INITIAL_COST: "Commercial Solectria initial cost (constant USD)",
        technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_RECURRING_PV_COST: "Commercial Solectria recurring lifecycle cost (constant USD)",
        technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_SCHEDULED_PV_COST: "Commercial Solectria scheduled lifecycle cost (constant USD)",
        technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_LIFECYCLE_COST: "Commercial Solectria lifecycle cost (constant USD)",
        technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_EA_COST: "Commercial Solectria equivalent-annual cost (constant USD/year)",
        technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_LCOE: "Commercial Solectria lifecycle LCOE (constant USD/kWh_AC)",
        technoeconomic_kernel.COMMERCIAL_PAIRED_FIELD_LCOE_DELTA: "Commercial lifecycle LCOE delta, SolarEdge minus Solectria (constant USD/kWh_AC)",
        "headline_positive_gain_lcoo": "Headline LCOO, SE − SOL (USD/kWh_AC; positive gain)",
        "signed_nonzero_lcoo": "Signed LCOO diagnostic, SE − SOL (USD/kWh_AC)",
        "lifecycle_lcoe_solectria": "Solectria lifecycle LCOE (USD/kWh_AC)",
        "lifecycle_lcoe_solaredge": "SolarEdge lifecycle LCOE (USD/kWh_AC)",
        "lifecycle_cost_delta_se_minus_sol": (
            "Lifecycle cost delta, SE − SOL "
            f"(USD/{normalized_capacity_label})"
        ),
        "lifecycle_energy_delta_se_minus_sol": (
            "Lifecycle energy delta, SE − SOL "
            f"(kWh_AC/{normalized_capacity_label})"
        ),
        "headline_positive_gain_lcoo_se_minus_sol": "Headline LCOO, SE − SOL (USD/kWh_AC)",
    }
    return replacements.get(value, value.replace("_", " "))


def _plot_grid(count: int) -> tuple[int, int]:
    columns = 2 if count <= 6 else 3
    rows = max(1, math.ceil(max(1, count) / columns))
    return rows, columns


def _figure_axes(
    count: int,
    *,
    title: str,
    subtitle: str,
    single_panel: bool = False,
) -> tuple[Any, list[Any]]:
    from matplotlib import pyplot as plt

    rows, columns = (1, 1) if single_panel else _plot_grid(count)
    figure, axes = plt.subplots(rows, columns, figsize=(16, 10), squeeze=False)
    figure.patch.set_facecolor("white")
    figure.suptitle(title, x=0.04, y=0.98, ha="left", fontsize=17, fontweight="bold", color=_INK)
    figure.text(0.04, 0.945, subtitle, ha="left", va="top", fontsize=10, color="#55616D")
    flattened = list(axes.flat)
    for axis in flattened:
        axis.set_facecolor("white")
        axis.grid(axis="y", color=_GRID, linewidth=0.8)
        axis.spines[["top", "right"]].set_visible(False)
        axis.spines[["left", "bottom"]].set_color("#8B96A1")
        axis.tick_params(labelsize=8, colors=_INK)
    return figure, flattened


def _save_figure(
    figure: Any,
    path: Path,
    *,
    layout_bottom: float = 0.035,
) -> tuple[int, int]:
    from matplotlib import pyplot as plt

    figure.tight_layout(
        rect=(0.035, layout_bottom, 0.98, 0.91),
        h_pad=2.4,
        w_pad=1.8,
    )
    figure.savefig(
        path,
        dpi=100,
        facecolor="white",
        metadata={"Software": "SBE PV technoeconomic reporting v1"},
    )
    plt.close(figure)
    return _png_dimensions(path)


def _paired_cdf_subtitle(
    available: Sequence[tuple[str, Mapping[str, Any]]],
) -> str:
    populations: list[int] = []
    outcome_counts: list[int] = []
    system_details: list[str] = []
    labels = {
        technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_LCOE: "Solectria",
        technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LCOE: "SolarEdge",
    }
    for metric_id, summary in available:
        cdf = summary.get("cdf")
        if not isinstance(cdf, Mapping):
            continue
        values = cdf.get("values") or []
        population = int(cdf.get("population_count") or len(values))
        outcome_count = len(values)
        populations.append(population)
        outcome_counts.append(outcome_count)
        label = labels.get(metric_id, _human_metric(metric_id))
        system_details.append(
            f"{label}: {population:,} runs, {outcome_count:,} outcomes"
        )
    if (
        len(populations) == 2
        and len(set(populations)) == 1
        and len(set(outcome_counts)) == 1
    ):
        return (
            f"{populations[0]:,} runs per system • "
            f"{outcome_counts[0]:,} distinct outcomes per system • "
            "Exact empirical CDF"
        )
    if system_details:
        return " • ".join((*system_details, "Exact empirical CDF"))
    return "Exact empirical CDF"


def _cdf_value_text(value: Any) -> str:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return "—"
    if not math.isfinite(number):
        return "—"
    absolute = abs(number)
    if absolute >= 10:
        return f"{number:.2f}"
    if absolute >= 1:
        return f"{number:.3f}"
    return f"{number:.4f}"


def _paired_cdf_constant_dollar_year(metadata: Mapping[str, Any]) -> int | None:
    provenance = metadata.get("kernel_provenance")
    if not isinstance(provenance, Mapping):
        return None
    paired = provenance.get("commercial_paired")
    if not isinstance(paired, Mapping):
        return None
    year = paired.get("constant_dollar_cost_year")
    if (
        not isinstance(year, int)
        or isinstance(year, bool)
        or year < 1900
        or year > 3000
    ):
        return None
    return year


def _paired_cdf_display_lcoe(value: Any) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return math.nan
    if not math.isfinite(number):
        return math.nan
    return number * 1_000.0


def _cdf_display_indices(
    probabilities: np.ndarray,
    *,
    maximum: int = 1200,
    quantiles: tuple[float, ...] = (0.05, 0.5, 0.95),
) -> np.ndarray:
    count = len(probabilities)
    if count <= maximum:
        return np.arange(count, dtype=np.int64)
    required = {0, count - 1}
    for quantile in quantiles:
        index = int(np.searchsorted(probabilities, quantile, side="left"))
        for neighbor in (index - 1, index, index + 1):
            if 0 <= neighbor < count:
                required.add(neighbor)
    if len(required) > maximum:
        raise TechnoeconomicExportError("CDF display cap is too small")
    selected = set(required)
    remaining = maximum - len(required)
    if remaining:
        selected.update(
            np.linspace(0, count - 1, remaining, dtype=np.int64).tolist()
        )
    return np.asarray(sorted(selected), dtype=np.int64)


def _render_cdf_plot(
    metadata: Mapping[str, Any],
    path: Path,
    *,
    headline_only: bool = False,
    paired_headlines_only: bool = False,
) -> tuple[int, int, int, int]:
    paired_metric_ids = (
        technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_LCOE,
        technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LCOE,
    )
    available: list[tuple[str, Mapping[str, Any]]] = []
    for metric_id, summary in sorted(_metric_summaries(metadata).items()):
        if paired_headlines_only and metric_id not in paired_metric_ids:
            continue
        if (
            headline_only
            and metric_id
            != technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LCOE
        ):
            continue
        if isinstance(summary, Mapping) and isinstance(summary.get("cdf"), Mapping):
            available.append((metric_id, summary))
    if paired_headlines_only:
        from matplotlib.ticker import PercentFormatter

        available.sort(key=lambda item: paired_metric_ids.index(item[0]))
        figure, axes = _figure_axes(
            1,
            title="Lifecycle LCOE comparison",
            subtitle=_paired_cdf_subtitle(available),
            single_panel=True,
        )
        axis = axes[0]
        point_count = 0
        display_point_count = 0
        styles = {
            technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_LCOE: (
                "Solectria",
                _GOLD,
                "--",
                0.90,
            ),
            technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LCOE: (
                "SolarEdge",
                _BLUE,
                "-",
                0.75,
            ),
        }
        if not available:
            axis.text(
                0.5,
                0.5,
                "No finite paired LCOE populations",
                ha="center",
                va="center",
                color=_INK,
            )
            axis.set_axis_off()
        for metric_id, summary in available:
            cdf = summary["cdf"]
            values = np.asarray(cdf.get("values") or [], dtype=np.float64)
            probability = np.asarray(
                cdf.get("cumulative_probability") or [],
                dtype=np.float64,
            )
            point_count += len(values)
            display_indices = _cdf_display_indices(
                probability,
                quantiles=(0.10, 0.50, 0.90),
            )
            display_values = values[display_indices] * 1_000.0
            display_probability = probability[display_indices]
            display_point_count += len(display_indices)
            label, color, linestyle, direct_label_probability = styles[metric_id]
            axis.step(
                display_values,
                display_probability,
                where="post",
                color=color,
                linestyle=linestyle,
                linewidth=3.0,
            )
            percentiles = summary.get("percentiles") or {}
            p50 = _paired_cdf_display_lcoe(percentiles.get("p50"))
            if math.isfinite(p50):
                axis.vlines(
                    p50,
                    0.0,
                    0.5,
                    color=color,
                    linewidth=1.2,
                    linestyle=":",
                    alpha=0.9,
                )
                axis.scatter(
                    [p50],
                    [0.5],
                    s=60,
                    marker="o",
                    facecolor="white",
                    edgecolor=color,
                    linewidth=2.2,
                    zorder=4,
                )
            if len(display_values):
                label_index = int(
                    np.searchsorted(
                        display_probability,
                        direct_label_probability,
                        side="left",
                    )
                )
                label_index = min(label_index, len(display_values) - 1)
                axis.annotate(
                    label,
                    xy=(
                        display_values[label_index],
                        display_probability[label_index],
                    ),
                    xytext=(12, 0),
                    textcoords="offset points",
                    color=color,
                    fontsize=10,
                    fontweight="bold",
                    ha="left",
                    va="center",
                    annotation_clip=False,
                    bbox={
                        "facecolor": "white",
                        "edgecolor": "none",
                        "pad": 1.0,
                    },
                )
        axis.axhline(0.5, color=_INK, linewidth=0.9, linestyle=":", alpha=0.7)
        axis.set_ylim(0, 1.02)
        axis.set_yticks(np.linspace(0.0, 1.0, 6))
        axis.yaxis.set_major_formatter(PercentFormatter(xmax=1.0, decimals=0))
        axis.margins(x=0.08)
        constant_dollar_year = _paired_cdf_constant_dollar_year(metadata)
        dollar_basis = (
            f"real {constant_dollar_year} USD"
            if constant_dollar_year is not None
            else "constant USD"
        )
        axis.set_xlabel(
            f"Lifecycle LCOE ({dollar_basis}/MWh AC)",
            fontsize=9,
            color=_INK,
        )
        axis.set_ylabel(
            "Probability at or below this LCOE",
            fontsize=9,
            color=_INK,
        )
        summary_by_metric = dict(available)
        percentile_rows = (
            (
                technoeconomic_kernel.COMMERCIAL_STANDALONE_FIELD_LCOE,
                0.095,
            ),
            (
                technoeconomic_kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_LCOE,
                0.055,
            ),
        )
        medians: dict[str, float] = {}
        for metric_id, row_y in percentile_rows:
            summary = summary_by_metric.get(metric_id)
            if not isinstance(summary, Mapping):
                continue
            label, color, _linestyle, _direct_probability = styles[metric_id]
            percentiles = summary.get("percentiles") or {}
            figure.text(
                0.07,
                row_y,
                label,
                ha="left",
                va="center",
                fontsize=10,
                fontweight="bold",
                color=color,
            )
            figure.text(
                0.19,
                row_y,
                (
                    "P10  "
                    f"{_cdf_value_text(_paired_cdf_display_lcoe(percentiles.get('p10')))}"
                    "      P50  "
                    f"{_cdf_value_text(_paired_cdf_display_lcoe(percentiles.get('p50')))}"
                    "      P90  "
                    f"{_cdf_value_text(_paired_cdf_display_lcoe(percentiles.get('p90')))}"
                    "   USD/MWh AC"
                ),
                ha="left",
                va="center",
                fontsize=10,
                color=_INK,
            )
            median = _paired_cdf_display_lcoe(percentiles.get("p50"))
            if math.isfinite(median):
                medians[label] = median
        if len(medians) == 2:
            median_values = list(medians.values())
            if math.isclose(
                median_values[0],
                median_values[1],
                rel_tol=0.0,
                abs_tol=1e-15,
            ):
                median_message = "Same median LCOE"
            else:
                lower_label = min(medians, key=lambda label: medians[label])
                higher_label = max(medians, key=lambda label: medians[label])
                median_gap = medians[higher_label] - medians[lower_label]
                gap_text = (
                    f"{median_gap:.2f}"
                    if max(abs(value) for value in median_values) >= 10
                    else _cdf_value_text(median_gap)
                )
                median_message = (
                    f"Lower median: {lower_label} by {gap_text} USD/MWh AC"
                )
            figure.text(
                0.69,
                0.075,
                median_message,
                ha="left",
                va="center",
                fontsize=10,
                fontweight="bold",
                color=_INK,
                bbox={
                    "boxstyle": "round,pad=0.45",
                    "facecolor": "#EEF4F2",
                    "edgecolor": "none",
                },
            )
        for unused in axes[1:]:
            unused.set_visible(False)
        width, height = _save_figure(figure, path, layout_bottom=0.16)
        return width, height, point_count, display_point_count
    figure, axes = _figure_axes(
        len(available),
        title=(
            "Commercial SolarEdge lifecycle LCOE empirical CDF"
            if headline_only
            else "Technoeconomic metric empirical CDFs"
        ),
        subtitle=(
            "Standalone commercial SolarEdge; P(X <= x), with P10/P50/P90 reported separately."
            if headline_only
            else "Finite metric-specific populations; each panel states its denominator. Signed zero and negative values are retained."
        ),
    )
    point_count = 0
    display_point_count = 0
    if not available:
        axes[0].text(0.5, 0.5, "No finite CDF populations", ha="center", va="center", color=_INK)
        axes[0].set_axis_off()
    for axis, (metric_id, summary) in zip(axes, available):
        cdf = summary["cdf"]
        values = np.asarray(cdf.get("values") or [], dtype=np.float64)
        probability = np.asarray(cdf.get("cumulative_probability") or [], dtype=np.float64)
        point_count += len(values)
        display_indices = _cdf_display_indices(
            probability,
            quantiles=(0.10, 0.50, 0.90) if headline_only else (0.05, 0.50, 0.95),
        )
        display_values = values[display_indices]
        display_probability = probability[display_indices]
        display_point_count += len(display_indices)
        axis.step(display_values, display_probability, where="post", color=_BLUE, linewidth=1.8)
        if len(display_values) <= 100:
            axis.scatter(
                display_values,
                display_probability,
                s=12,
                color="white",
                edgecolor=_BLUE,
                linewidth=0.8,
                zorder=3,
            )
        axis.axhline(0.5, color=_INK, linewidth=0.8, linestyle="--", alpha=0.7)
        axis.set_ylim(0, 1.02)
        axis.set_title(
            f"{_human_metric(metric_id)}\nRight-continuous ECDF • n={cdf.get('population_count')} • plotted {len(display_indices):,} points",
            loc="left",
            fontsize=9.5,
            color=_INK,
        )
        axis.set_xlabel("Metric value", fontsize=8, color=_INK)
        axis.set_ylabel("P(X ≤ x)", fontsize=8, color=_INK)
    for axis in axes[len(available):]:
        axis.set_visible(False)
    width, height = _save_figure(figure, path)
    return width, height, point_count, display_point_count


def _render_sensitivity_plot(
    metadata: Mapping[str, Any],
    path: Path,
) -> tuple[int, int, int, int]:
    models = metadata.get("sensitivity") or {}
    kernel_provenance = metadata.get("kernel_provenance") or {}
    applied_capacity_contract = (
        isinstance(kernel_provenance, Mapping)
        and isinstance(kernel_provenance.get("capacity_normalization"), Mapping)
    )
    entries = [
        (response_id, model)
        for response_id, model in sorted(models.items())
        if isinstance(model, Mapping)
        and ("steps" in model or "status" in model)
    ]
    figure, axes = _figure_axes(
        len(entries),
        title="Forward stepwise rank sensitivity",
        subtitle="Incremental R² by entered predictor; panel subtitles state the response-specific sample denominator.",
    )
    step_count = 0
    display_step_count = 0
    for axis, (response_id, model) in zip(axes, entries):
        steps = model.get("steps") or []
        step_count += len(steps)
        display_steps = sorted(
            steps,
            key=lambda step: (
                -float(step.get("incremental_r_squared") or 0.0),
                int(step.get("entry_order") or 0),
                str(step.get("predictor_id")),
            ),
        )[:15]
        display_step_count += len(display_steps)
        axis.set_title(
            f"{_human_metric(response_id, applied_capacity_contract=applied_capacity_contract)}\nstatus={model.get('status')} • n={model.get('sample_count', 0)} • shown {len(display_steps)} of {len(steps)}",
            loc="left",
            fontsize=9.5,
            color=_INK,
        )
        if not display_steps:
            axis.text(
                0.5,
                0.5,
                f"Unavailable: {model.get('reason') or 'no predictors entered'}",
                ha="center",
                va="center",
                fontsize=9,
                color="#55616D",
                wrap=True,
            )
            axis.set_axis_off()
            continue
        labels = [str(step.get("predictor_id")) for step in display_steps]
        values = [float(step.get("incremental_r_squared") or 0.0) for step in display_steps]
        signs = [step.get("sign") for step in display_steps]
        positions = np.arange(len(labels))
        bars = axis.barh(positions, values, color=_BLUE_LIGHT, edgecolor=_BLUE, linewidth=1.0)
        axis.set_yticks(positions, labels=labels)
        axis.invert_yaxis()
        axis.set_xlabel("Incremental R²", fontsize=8, color=_INK)
        axis.set_xlim(left=0)
        for bar, value, sign in zip(bars, values, signs):
            glyph = "+" if sign == "positive" else "−" if sign == "negative" else "0"
            axis.text(
                bar.get_width(),
                bar.get_y() + bar.get_height() / 2,
                f" {value:.4f} ({glyph}β)",
                va="center",
                fontsize=7.5,
                color=_INK,
            )
    for axis in axes[len(entries):]:
        axis.set_visible(False)
    width, height = _save_figure(figure, path)
    return width, height, step_count, display_step_count


def _render_convergence_plot(metadata: Mapping[str, Any], path: Path) -> tuple[int, int, int]:
    convergence = metadata.get("convergence") or {}
    checkpoints = convergence.get("checkpoints") or []
    metric_ids = sorted(
        {
            metric_id
            for checkpoint in checkpoints
            for metric_id in (checkpoint.get("metrics") or {})
        }
    )
    figure, axes = _figure_axes(
        len(metric_ids),
        title="Cumulative convergence diagnostics",
        subtitle=f"Deterministic LHS prefixes; final status={convergence.get('status')}. Bands show cumulative P5–P95.",
    )
    point_count = 0
    for axis, metric_id in zip(axes, metric_ids):
        x_values: list[int] = []
        p5_values: list[float] = []
        p50_values: list[float] = []
        p95_values: list[float] = []
        for checkpoint in checkpoints:
            metric = (checkpoint.get("metrics") or {}).get(metric_id) or {}
            percentiles = metric.get("percentiles") or {}
            if any(percentiles.get(name) is None for name in ("p5", "p50", "p95")):
                continue
            x_values.append(int(checkpoint.get("realization_count")))
            p5_values.append(float(percentiles["p5"]))
            p50_values.append(float(percentiles["p50"]))
            p95_values.append(float(percentiles["p95"]))
        point_count += len(x_values)
        axis.set_title(
            f"{_human_metric(metric_id)}\nfinite population at each cumulative checkpoint",
            loc="left",
            fontsize=9.5,
            color=_INK,
        )
        if x_values:
            axis.fill_between(x_values, p5_values, p95_values, color=_BLUE_LIGHT, alpha=0.45, label="P5–P95")
            axis.plot(x_values, p50_values, color=_BLUE, linewidth=1.8, marker="o", markersize=4, label="P50")
            axis.legend(frameon=False, fontsize=7.5, loc="best")
        else:
            axis.text(0.5, 0.5, "No finite checkpoints", ha="center", va="center", color=_INK)
        axis.set_xlabel("Cumulative realizations", fontsize=8, color=_INK)
        axis.set_ylabel("Metric value", fontsize=8, color=_INK)
    for axis in axes[len(metric_ids):]:
        axis.set_visible(False)
    width, height = _save_figure(figure, path)
    return width, height, point_count


def _png_dimensions(path: Path) -> tuple[int, int]:
    with path.open("rb") as handle:
        header = handle.read(24)
    if len(header) != 24 or header[:8] != b"\x89PNG\r\n\x1a\n" or header[12:16] != b"IHDR":
        raise TechnoeconomicExportError("Generated plot is not a valid PNG")
    return struct.unpack(">II", header[16:24])


def _artifact_contract_by_id(artifact_id: str) -> Mapping[str, Any]:
    matches = [
        specification
        for specification in api_artifacts.TECHNOECONOMIC_PUBLIC_ARTIFACT_CONTRACT.values()
        if specification.get("artifact_id") == artifact_id
    ]
    if len(matches) != 1:
        raise TechnoeconomicExportError(
            f"Public artifact contract is missing {artifact_id!r}"
        )
    return matches[0]


def _publish_artifact(
    *,
    artifact_id: str,
    pending_path: Path,
    attempt_directory: Path,
    job_id: str,
    cancellation_check: Callable[[], None],
    publish_check: Callable[[], None],
    extra: Mapping[str, Any],
) -> dict[str, Any]:
    contract = _artifact_contract_by_id(artifact_id)
    filename = _safe_filename(str(contract.get("filename")))
    target = attempt_directory / filename
    if target.exists() or target.is_symlink():
        raise TechnoeconomicExportError("An export target already exists")
    with pending_path.open("r+b") as pending_handle:
        pending_handle.flush()
        os.fsync(pending_handle.fileno())
    details = _strict_regular_file(pending_path, label=f"Pending {artifact_id} export")
    byte_count = int(details.st_size)
    if byte_count <= 0:
        raise TechnoeconomicExportError(f"Generated {artifact_id} export is empty")
    digest = _sha256_file(pending_path)
    cancellation_check()
    publish_check()
    pending_path.replace(target)
    published = _strict_regular_file(target, label=f"Published {artifact_id} export")
    if published.st_size != byte_count or not secrets.compare_digest(
        _sha256_file(target), digest
    ):
        target.unlink(missing_ok=True)
        raise TechnoeconomicExportError(
            f"Published {artifact_id} export changed during publication"
        )
    return {
        "schema_version": extra.get("schema_version"),
        "artifact_id": artifact_id,
        "artifact_kind": contract.get("artifact_kind"),
        "owner_workflow": "technoeconomic",
        "owner_job_id": job_id,
        "storage_key": _storage_key(target),
        "filename": filename,
        "media_type": contract.get("media_type"),
        "sha256": digest,
        "byte_count": byte_count,
        "public": True,
        **{key: value for key, value in extra.items() if key != "schema_version"},
    }


def _signed_metric_counts(calculation: _SealedCalculation) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for column_name, raw_values in zip(calculation.column_names, calculation.columns):
        if "Delta" not in column_name and column_name not in {
            technoeconomic_kernel.FIELD_LCOO,
            technoeconomic_kernel.COMMERCIAL_FIELD_MARGINAL_LCOO,
        }:
            continue
        try:
            values = np.asarray(raw_values, dtype=np.float64)
        except (TypeError, ValueError):
            continue
        finite = values[np.isfinite(values)]
        result[column_name] = {
            "negative_count": int(np.count_nonzero(finite < 0)),
            "zero_count": int(np.count_nonzero(finite == 0)),
            "positive_count": int(np.count_nonzero(finite > 0)),
            "null_count": int(len(values) - len(finite)),
            "row_count": int(len(values)),
        }
    return result


def generate_technoeconomic_exports(
    *,
    job_id: str,
    attempt_directory: Path,
    sealed_calculation_path: Path,
    sealed_calculation_artifact: Mapping[str, Any],
    request_payload: Mapping[str, Any],
    source_snapshot: Mapping[str, Any],
    submission_provenance: Mapping[str, Any],
    routine_result: Mapping[str, Any],
    cancellation_check: Callable[[], None],
    publish_check: Callable[[], None],
) -> dict[str, Any]:
    """Generate, verify, and lease-fence the complete Phase-4 artifact set.

    All filenames are server-owned constants from ``api.artifacts``.  Long tables
    are streamed, and the XLSX ``Realizations`` sheet uses openpyxl write-only mode
    so the allowed 100,000-realization job size does not materialize worksheet
    objects per cell.
    """

    if not isinstance(job_id, str) or not job_id:
        raise TechnoeconomicExportError("Export owner job ID is invalid")
    if not callable(cancellation_check) or not callable(publish_check):
        raise TechnoeconomicExportError("Export lease callbacks are required")
    schema_versions = export_contract_versions(
        str(routine_result.get("calculation_contract_version"))
    )
    standalone_contract = (
        routine_result.get("calculation_contract_version")
        == technoeconomic_kernel.STANDALONE_COMMERCIAL_CALCULATION_CONTRACT_VERSION
    )
    paired_contract = (
        routine_result.get("calculation_contract_version")
        == technoeconomic_kernel.PAIRED_COMMERCIAL_CALCULATION_CONTRACT_VERSION
    )
    lifecycle_contract = _is_lifecycle_contract(
        routine_result.get("calculation_contract_version")
    )
    cancellation_check()
    attempt_directory = attempt_directory.resolve(strict=True)
    _confined(attempt_directory, config.OUTPUT_DIR.resolve(strict=True), label="Attempt directory")
    calculation = _load_sealed_calculation(
        attempt_directory=attempt_directory,
        sealed_calculation_path=sealed_calculation_path,
        sealed_calculation_artifact=sealed_calculation_artifact,
        request_payload=request_payload,
        source_snapshot=source_snapshot,
        submission_provenance=submission_provenance,
    )
    registry: tuple[Mapping[str, Any], ...] = ()
    trace_records: tuple[Mapping[str, Any], ...] = ()
    audit_records: tuple[Mapping[str, Any], ...] = ()
    if lifecycle_contract:
        _verify_lifecycle_routine_result(
            metadata=calculation.metadata,
            routine_result=routine_result,
            request_payload=request_payload,
            source_snapshot=source_snapshot,
            submission_provenance=submission_provenance,
            sealed_calculation_artifact=sealed_calculation_artifact,
        )
        registry = _formula_registry_records()
        trace_records = _representative_trace_records(calculation.metadata)
        audit_records = _lifecycle_audit_records(
            trace_records,
            registry,
            economic_decision_tolerance=_lifecycle_economic_npv_tolerance(
                request_payload
            ),
            economic_decision_tolerances=_lifecycle_realization_value_map(
                calculation,
                "NPVTolerance_USD",
            ),
        )
        checks = _build_lifecycle_checks(
            calculation,
            routine_result,
            registry,
            trace_records,
            audit_records,
        )
    else:
        _verify_routine_result(
            metadata=calculation.metadata,
            routine_result=routine_result,
            request_payload=request_payload,
            source_snapshot=source_snapshot,
            submission_provenance=submission_provenance,
            sealed_calculation_artifact=sealed_calculation_artifact,
        )
        checks = _build_checks(
            calculation,
            source_snapshot,
            submission_provenance,
            routine_result,
        )
    failed_checks = [str(row[0]) for row in checks if row[5] != "OK"]
    if failed_checks:
        raise TechnoeconomicExportError(
            "Export tie-outs failed: " + ", ".join(failed_checks[:5])
        )
    if lifecycle_contract:
        tables = _build_lifecycle_tables(
            calculation,
            request_payload,
            source_snapshot,
            submission_provenance,
            routine_result,
            checks,
            registry,
            trace_records,
            audit_records,
        )
    else:
        tables = _build_tables(
            calculation,
            request_payload,
            source_snapshot,
            submission_provenance,
            routine_result,
            checks,
        )

    artifacts: dict[str, dict[str, Any]] = {}
    csv_pending = attempt_directory / ".pending-csv.zip"
    workbook_raw = attempt_directory / ".raw-workbook.xlsx"
    workbook_pending = attempt_directory / ".pending-workbook.xlsx"
    cdf_pending = attempt_directory / ".pending-cdf.png"
    sensitivity_pending = attempt_directory / ".pending-sensitivity.png"
    convergence_pending = attempt_directory / ".pending-convergence.png"
    staging = attempt_directory / ".csv-staging"
    temporary_paths = (
        csv_pending,
        workbook_raw,
        workbook_pending,
        cdf_pending,
        sensitivity_pending,
        convergence_pending,
    )
    if any(path.exists() or path.is_symlink() for path in (*temporary_paths, staging)):
        raise TechnoeconomicExportError("The export staging area is not empty")
    try:
        csv_tables, csv_row_count = _write_csv_bundle(
            csv_pending,
            staging,
            tables,
            cancellation_check,
            schema_versions=schema_versions,
            bundle_provenance=(
                {
                    "calculation_contract_version": routine_result.get(
                        "calculation_contract_version"
                    ),
                    "sampling_version": routine_result.get("sampling_version"),
                    "result_version": routine_result.get("result_version"),
                    "formula_registry_version": getattr(
                        technoeconomic_kernel,
                        "FORMULA_REGISTRY_VERSION",
                        "tea-formulas-v6",
                    ),
                    "formula_registry_count": len(registry),
                    "formula_registry_sha256": _formula_registry_sha256(registry),
                    "formula_template_hash_version": (
                        LIFECYCLE_FORMULA_TEMPLATE_HASH_VERSION
                    ),
                    "formula_template_sha256": _formula_template_sha256(registry),
                    "decision_rule_version": LIFECYCLE_DECISION_RULE_VERSION,
                }
                if lifecycle_contract
                else None
            ),
        )
        artifacts["csv_bundle"] = _publish_artifact(
            artifact_id="csv_bundle",
            pending_path=csv_pending,
            attempt_directory=attempt_directory,
            job_id=job_id,
            cancellation_check=cancellation_check,
            publish_check=publish_check,
            extra={
                "schema_version": schema_versions["csv_bundle"],
                "table_count": len(csv_tables),
                "row_count": csv_row_count,
                "tables": csv_tables,
            },
        )

        lifecycle_plot_metadata: dict[str, tuple[Any, ...]] = {}
        workbook_audit: dict[str, Any] = {}
        if lifecycle_contract:
            lifecycle_plot_metadata["cdf_plot"] = _render_cdf_plot(
                calculation.metadata,
                cdf_pending,
            )
            lifecycle_plot_metadata["sensitivity_plot"] = _render_sensitivity_plot(
                calculation.metadata,
                sensitivity_pending,
            )
            lifecycle_plot_metadata["convergence_plot"] = _render_convergence_plot(
                calculation.metadata,
                convergence_pending,
            )
            (
                workbook_sheets,
                workbook_row_count,
                workbook_audit,
            ) = _write_lifecycle_workbook(
                workbook_raw,
                tables,
                calculation,
                request_payload,
                routine_result,
                calculation.metadata,
                checks,
                registry,
                cancellation_check,
                image_paths={
                    "cdf_plot": cdf_pending,
                    "sensitivity_plot": sensitivity_pending,
                    "convergence_plot": convergence_pending,
                },
            )
        else:
            workbook_sheets, workbook_row_count = _write_workbook(
                workbook_raw,
                tables,
                routine_result,
                calculation.metadata,
                checks,
                cancellation_check,
            )
        cancellation_check()
        _normalize_xlsx_archive(
            workbook_raw,
            workbook_pending,
            cancellation_check,
            canonicalize_core_properties=lifecycle_contract,
        )
        if lifecycle_contract:
            workbook_audit.update(_validate_lifecycle_xlsx(workbook_pending))
        workbook_raw.unlink(missing_ok=True)
        artifacts["xlsx_workbook"] = _publish_artifact(
            artifact_id="xlsx_workbook",
            pending_path=workbook_pending,
            attempt_directory=attempt_directory,
            job_id=job_id,
            cancellation_check=cancellation_check,
            publish_check=publish_check,
            extra={
                "schema_version": schema_versions["xlsx"],
                "sheet_count": len(workbook_sheets),
                "row_count": workbook_row_count,
                "sheets": workbook_sheets,
                "write_only_streaming": True,
                **workbook_audit,
            },
        )

        if lifecycle_contract:
            width, height, row_count, display_count = lifecycle_plot_metadata[
                "cdf_plot"
            ]
        else:
            width, height, row_count, display_count = _render_cdf_plot(
                calculation.metadata,
                cdf_pending,
                headline_only=standalone_contract,
                paired_headlines_only=paired_contract,
            )
        artifacts["cdf_plot"] = _publish_artifact(
            artifact_id="cdf_plot",
            pending_path=cdf_pending,
            attempt_directory=attempt_directory,
            job_id=job_id,
            cancellation_check=cancellation_check,
            publish_check=publish_check,
            extra={
                "schema_version": schema_versions["png"],
                "row_count": row_count,
                "source_point_count": row_count,
                "display_point_count": display_count,
                "width_px": width,
                "height_px": height,
                "chart_contract_id": (
                    LIFECYCLE_CDF_CHART_CONTRACT_ID
                    if lifecycle_contract
                    else PAIRED_COMMERCIAL_CDF_CHART_CONTRACT_ID
                    if paired_contract
                    else STANDALONE_COMMERCIAL_CDF_CHART_CONTRACT_ID
                    if standalone_contract
                    else "cdf_v1"
                ),
            },
        )
        if lifecycle_contract:
            width, height, row_count, display_count = lifecycle_plot_metadata[
                "sensitivity_plot"
            ]
        else:
            width, height, row_count, display_count = _render_sensitivity_plot(
                calculation.metadata,
                sensitivity_pending,
            )
        artifacts["sensitivity_plot"] = _publish_artifact(
            artifact_id="sensitivity_plot",
            pending_path=sensitivity_pending,
            attempt_directory=attempt_directory,
            job_id=job_id,
            cancellation_check=cancellation_check,
            publish_check=publish_check,
            extra={
                "schema_version": schema_versions["png"],
                "row_count": row_count,
                "source_step_count": row_count,
                "display_step_count": display_count,
                "width_px": width,
                "height_px": height,
                "chart_contract_id": "sensitivity_v1",
            },
        )
        if lifecycle_contract:
            width, height, row_count = lifecycle_plot_metadata[
                "convergence_plot"
            ]
        else:
            width, height, row_count = _render_convergence_plot(
                calculation.metadata,
                convergence_pending,
            )
        artifacts["convergence_plot"] = _publish_artifact(
            artifact_id="convergence_plot",
            pending_path=convergence_pending,
            attempt_directory=attempt_directory,
            job_id=job_id,
            cancellation_check=cancellation_check,
            publish_check=publish_check,
            extra={
                "schema_version": schema_versions["png"],
                "row_count": row_count,
                "width_px": width,
                "height_px": height,
                "chart_contract_id": "convergence_v1",
            },
        )
    finally:
        shutil.rmtree(staging, ignore_errors=True)
        for path in temporary_paths:
            path.unlink(missing_ok=True)

    expected_order = (
        "csv_bundle",
        "xlsx_workbook",
        "cdf_plot",
        "sensitivity_plot",
        "convergence_plot",
    )
    artifacts = {artifact_id: artifacts[artifact_id] for artifact_id in expected_order}
    request_sha256 = technoeconomic_api.canonical_json_sha256(request_payload)
    source_snapshot_sha256 = technoeconomic_api.canonical_json_sha256(source_snapshot)
    submission_provenance_sha256 = technoeconomic_api.canonical_json_sha256(
        submission_provenance
    )
    manifest: dict[str, Any] = {
        "schema_version": schema_versions["manifest"],
        "csv_format_version": schema_versions["csv_format"],
        "owner_workflow": "technoeconomic",
        "owner_job_id": job_id,
        "request_sha256": request_sha256,
        "source_snapshot_sha256": source_snapshot_sha256,
        "submission_provenance_sha256": submission_provenance_sha256,
        "sealed_calculation_sha256": sealed_calculation_artifact["sha256"],
        "calculation_contract_version": routine_result.get(
            "calculation_contract_version"
        ),
        "sampling_version": routine_result.get("sampling_version"),
        "artifact_count": len(artifacts),
        "artifacts": artifacts,
        "tie_outs": {
            "status": "passed",
            "check_count": len(checks),
            "failed_check_ids": [],
            "realization_row_count": calculation.row_count,
            "csv_table_row_counts": {
                record["filename"]: record["row_count"] for record in csv_tables
            },
            "xlsx_sheet_row_counts": {
                record["sheet_name"]: record["row_count"]
                for record in workbook_sheets
            },
            "signed_metric_value_counts": _signed_metric_counts(calculation),
        },
        "chart_contracts": (
            LIFECYCLE_CHART_CONTRACTS
            if lifecycle_contract
            else PAIRED_COMMERCIAL_CHART_CONTRACTS
            if paired_contract
            else STANDALONE_COMMERCIAL_CHART_CONTRACTS
            if standalone_contract
            else CHART_CONTRACTS
        ),
    }
    if lifecycle_contract:
        manifest["formula_registry"] = {
            "version": workbook_audit["formula_registry_version"],
            "count": workbook_audit["formula_registry_count"],
            "sha256": workbook_audit["formula_registry_sha256"],
            "template_hash_version": workbook_audit[
                "formula_template_hash_version"
            ],
            "template_sha256": workbook_audit["formula_template_sha256"],
        }
        manifest["workbook_audit"] = {
            "native_charts": workbook_audit["native_charts"],
            "embedded_images": workbook_audit["embedded_images"],
            "decision_rule_version": workbook_audit["decision_rule_version"],
            "formula_count": workbook_audit["formula_count"],
            "formula_scan_status": workbook_audit["formula_scan_status"],
        }
    manifest["manifest_sha256"] = _canonical_json_sha256(manifest)
    return manifest


__all__ = [
    "APPLIED_CSV_BUNDLE_SCHEMA_VERSION",
    "APPLIED_CSV_FORMAT_VERSION",
    "APPLIED_EXPORT_MANIFEST_SCHEMA_VERSION",
    "APPLIED_XLSX_SCHEMA_VERSION",
    "COMMERCIAL_SCALING_CSV_BUNDLE_SCHEMA_VERSION",
    "COMMERCIAL_SCALING_CSV_FORMAT_VERSION",
    "COMMERCIAL_SCALING_EXPORT_MANIFEST_SCHEMA_VERSION",
    "COMMERCIAL_SCALING_XLSX_SCHEMA_VERSION",
    "CHART_CONTRACTS",
    "CSV_FORMAT_VERSION",
    "EXPORT_MANIFEST_SCHEMA_VERSION",
    "LIFECYCLE_CDF_CHART_CONTRACT_ID",
    "LIFECYCLE_CHART_CONTRACTS",
    "LIFECYCLE_CSV_BUNDLE_SCHEMA_VERSION",
    "LIFECYCLE_CSV_FORMAT_VERSION",
    "LIFECYCLE_DECISION_RULE_VERSION",
    "LIFECYCLE_EXPORT_MANIFEST_SCHEMA_VERSION",
    "LIFECYCLE_FORMULA_TEMPLATE_HASH_VERSION",
    "LIFECYCLE_WORKBOOK_SHEET_ORDER",
    "LIFECYCLE_XLSX_LOGICAL_HASH_VERSION",
    "LIFECYCLE_XLSX_SCHEMA_VERSION",
    "PAIRED_COMMERCIAL_CDF_CHART_CONTRACT_ID",
    "PAIRED_COMMERCIAL_CHART_CONTRACTS",
    "PAIRED_COMMERCIAL_CSV_BUNDLE_SCHEMA_VERSION",
    "PAIRED_COMMERCIAL_CSV_FORMAT_VERSION",
    "PAIRED_COMMERCIAL_EXPORT_MANIFEST_SCHEMA_VERSION",
    "PAIRED_COMMERCIAL_XLSX_SCHEMA_VERSION",
    "STANDALONE_COMMERCIAL_CHART_CONTRACTS",
    "STANDALONE_COMMERCIAL_CSV_BUNDLE_SCHEMA_VERSION",
    "STANDALONE_COMMERCIAL_CSV_FORMAT_VERSION",
    "STANDALONE_COMMERCIAL_EXPORT_MANIFEST_SCHEMA_VERSION",
    "STANDALONE_COMMERCIAL_XLSX_SCHEMA_VERSION",
    "TechnoeconomicExportError",
    "XLSX_LOGICAL_HASH_VERSION",
    "export_contract_versions",
    "generate_technoeconomic_exports",
]
