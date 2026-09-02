"""Standalone Bazefield collection and selection-aware quality reporting.

This module deliberately does not call the model-oriented ``run_historian``
entrypoint.  It reuses the same Bazefield client, target map, normalization, and
CSV writer while retaining source quality flags that the model CSV omits.
"""

from __future__ import annotations

from collections import Counter, defaultdict
import csv
from datetime import datetime, timezone
import math
import os
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence
from zoneinfo import ZoneInfo

from sbepv.ingest import bazefield


DATA_GROUP_ORDER = ("solaredge", "solectria", "weather")
DATA_GROUP_COLUMNS = {
    "solaredge": ("solaredge_measured_power",),
    "solectria": ("solectria_measured_power",),
    "weather": ("dni", "ghi", "dhi", "temp_air", "wind_speed"),
}
DATA_GROUP_LABELS = {
    "solaredge": "SolarEdge power",
    "solectria": "Solectria power",
    "weather": "Site weather",
}
SERIES_METADATA = {
    "solaredge_measured_power": {
        "label": "SolarEdge measured power",
        "unit": "W",
        "group": "solaredge",
    },
    "solectria_measured_power": {
        "label": "Solectria measured power",
        "unit": "W",
        "group": "solectria",
    },
    "dni": {"label": "Direct normal irradiance", "unit": "W/m²", "group": "weather"},
    "ghi": {"label": "Global horizontal irradiance", "unit": "W/m²", "group": "weather"},
    "dhi": {"label": "Diffuse horizontal irradiance", "unit": "W/m²", "group": "weather"},
    "temp_air": {"label": "Ambient temperature", "unit": "°C", "group": "weather"},
    "wind_speed": {"label": "Wind speed", "unit": "m/s", "group": "weather"},
}

# These are screening bounds for the collection report, not cleaning rules.
# Power is checked after the established kW-to-W conversion.
DOMAIN_BOUNDS = {
    "solaredge_measured_power": (-1.0, 200_000.0),
    "solectria_measured_power": (-1.0, 200_000.0),
    "dni": (-5.0, 1_500.0),
    "ghi": (-5.0, 1_400.0),
    "dhi": (-5.0, 1_400.0),
    "temp_air": (-40.0, 60.0),
    "wind_speed": (-0.1, 50.0),
}

MEASUREMENT_PLOT_METADATA = {
    "measured_ac_power": {
        "title": "Measured AC power",
        "alt": "Measured AC power for the selected SolarEdge and Solectria series",
    },
    "cumulative_energy": {
        "title": "Measured cumulative energy",
        "alt": "Cumulative measured energy for the selected SolarEdge and Solectria series",
    },
}
_MEASURED_POWER_SERIES = (
    ("solaredge_measured_power", "SolarEdge measured", "#dc2626"),
    ("solectria_measured_power", "Solectria measured", "#2563eb"),
)
_MAX_PLOT_POINTS = 4_000

_PRIMARY_QUALITY_MASK = 0xC0
_PRIMARY_GOOD = 0xC0
_PRIMARY_UNCERTAIN = 0x40
_PRIMARY_BAD = 0x00


def _quality_int(value: Any) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError, OverflowError):
        return None
    if not math.isfinite(numeric) or not numeric.is_integer():
        return None
    return int(numeric)


def primary_quality_label(value: Any) -> str:
    """Return the Bazefield primary-quality label for a composite flag."""

    numeric = _quality_int(value)
    decoded = None if numeric is None else numeric & _PRIMARY_QUALITY_MASK
    if decoded == _PRIMARY_GOOD:
        return "good"
    if decoded == _PRIMARY_UNCERTAIN:
        return "uncertain"
    if decoded == _PRIMARY_BAD:
        return "bad"
    return "unknown"


def _finite_number(value: Any) -> bool:
    try:
        return math.isfinite(float(value))
    except (TypeError, ValueError, OverflowError):
        return False


def _utc_milliseconds(value: str) -> int:
    parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    else:
        parsed = parsed.astimezone(timezone.utc)
    return int(parsed.timestamp() * 1_000)


def _unique(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        if value not in seen:
            seen.add(value)
            ordered.append(value)
    return ordered


def selected_series(data_groups: Sequence[str]) -> list[str]:
    """Resolve a group selection to the canonical historian column order."""

    requested = set(data_groups)
    unknown = sorted(requested.difference(DATA_GROUP_ORDER))
    if unknown:
        raise bazefield.BazefieldError(
            "Unknown data group(s): " + ", ".join(unknown)
        )
    if not requested:
        raise bazefield.BazefieldError("Select at least one data group.")
    included = {
        column
        for group in DATA_GROUP_ORDER
        if group in requested
        for column in DATA_GROUP_COLUMNS[group]
    }
    return [
        column
        for _object_id, _point_name, column in bazefield.COLUMN_MAP
        if column in included
    ]


def _issue(
    code: str,
    title: str,
    message: str,
    count: int,
    *,
    series: str | None = None,
) -> dict[str, Any]:
    item: dict[str, Any] = {
        "code": code,
        "severity": "warning",
        "title": title,
        "message": message,
        "count": int(count),
    }
    if series is not None:
        item["series"] = series
        item["series_label"] = SERIES_METADATA[series]["label"]
    return item


def _quality_report(
    buckets: Mapping[int, Mapping[str, list[Mapping[str, Any]]]],
    columns: Sequence[str],
    *,
    expected_timestamp_count: int,
    invalid_timestamp_sample_count: int,
    outside_window_sample_count: int,
    off_grid_sample_count: int,
) -> dict[str, Any]:
    issues: list[dict[str, Any]] = []
    observed_timestamp_count = len(buckets)
    present_sample_count = 0
    usable_sample_count = 0
    series_completeness: list[dict[str, Any]] = []
    missing_timestamp_count = max(
        0, expected_timestamp_count - observed_timestamp_count
    )
    if missing_timestamp_count:
        issues.append(
            _issue(
                "missing_timestamps",
                "Requested timestamps are missing",
                (
                    f"Bazefield returned no selected data for "
                    f"{missing_timestamp_count:,} expected timestamp(s)."
                ),
                missing_timestamp_count,
            )
        )

    timestamp_anomaly_count = (
        invalid_timestamp_sample_count
        + outside_window_sample_count
        + off_grid_sample_count
    )
    if timestamp_anomaly_count:
        issues.append(
            _issue(
                "timestamp_anomalies",
                "Some source timestamps were excluded",
                (
                    f"Excluded {timestamp_anomaly_count:,} sample(s): "
                    f"{invalid_timestamp_sample_count:,} invalid, "
                    f"{outside_window_sample_count:,} outside the requested window, "
                    f"and {off_grid_sample_count:,} off the requested interval grid."
                ),
                timestamp_anomaly_count,
            )
        )

    totals = Counter()
    for column in columns:
        samples_by_timestamp = [
            samples
            for timestamp_columns in buckets.values()
            if (samples := timestamp_columns.get(column))
        ]
        present_for_series = len(samples_by_timestamp)
        chosen_samples = [samples[-1] for samples in samples_by_timestamp]
        usable_for_series = sum(
            _finite_number(sample.get("value")) for sample in chosen_samples
        )
        present_sample_count += present_for_series
        usable_sample_count += usable_for_series
        missing_samples = max(0, expected_timestamp_count - present_for_series)
        series_completeness.append(
            {
                "name": column,
                "label": SERIES_METADATA[column]["label"],
                "expected_sample_count": int(expected_timestamp_count),
                "present_sample_count": int(present_for_series),
                "usable_sample_count": int(usable_for_series),
                "missing_sample_count": int(missing_samples),
                "sample_presence_percent": (
                    100.0
                    if expected_timestamp_count == 0
                    else round(
                        100.0 * present_for_series / expected_timestamp_count,
                        1,
                    )
                ),
                "usable_value_completeness_percent": (
                    100.0
                    if expected_timestamp_count == 0
                    else round(
                        100.0 * usable_for_series / expected_timestamp_count,
                        1,
                    )
                ),
            }
        )
        duplicate_samples = sum(max(0, len(samples) - 1) for samples in samples_by_timestamp)
        flattened_samples = [sample for samples in samples_by_timestamp for sample in samples]
        nonfinite_values = sum(
            not _finite_number(sample.get("value")) for sample in flattened_samples
        )
        quality_counts = Counter(
            primary_quality_label(sample.get("quality"))
            for sample in flattened_samples
        )
        non_good_quality = sum(
            count for label, count in quality_counts.items() if label != "good"
        )
        lower, upper = DOMAIN_BOUNDS[column]
        bound_violations = 0
        for sample in flattened_samples:
            if not _finite_number(sample.get("value")):
                continue
            numeric = float(sample["value"])
            if column in bazefield.KW_TO_W_COLUMNS:
                numeric *= 1_000.0
            if numeric < lower or numeric > upper:
                bound_violations += 1

        totals.update(
            {
                "missing_sample_count": missing_samples,
                "duplicate_sample_count": duplicate_samples,
                "nonfinite_value_count": nonfinite_values,
                "non_good_quality_count": non_good_quality,
                "domain_violation_count": bound_violations,
            }
        )
        if missing_samples:
            issues.append(
                _issue(
                    "missing_samples",
                    "Series has missing samples",
                    f"{SERIES_METADATA[column]['label']} is missing {missing_samples:,} expected sample(s).",
                    missing_samples,
                    series=column,
                )
            )
        if duplicate_samples:
            issues.append(
                _issue(
                    "duplicate_samples",
                    "Series has duplicate samples",
                    f"{SERIES_METADATA[column]['label']} has {duplicate_samples:,} extra sample(s) at duplicate timestamps; the last returned value is in the CSV.",
                    duplicate_samples,
                    series=column,
                )
            )
        if nonfinite_values:
            issues.append(
                _issue(
                    "nonfinite_values",
                    "Series has non-finite values",
                    f"{SERIES_METADATA[column]['label']} has {nonfinite_values:,} value(s) that are blank, infinite, or not numeric.",
                    nonfinite_values,
                    series=column,
                )
            )
        if non_good_quality:
            breakdown = ", ".join(
                f"{label} {quality_counts[label]:,}"
                for label in ("uncertain", "bad", "unknown")
                if quality_counts[label]
            )
            issues.append(
                _issue(
                    "non_good_source_quality",
                    "Bazefield quality flags need attention",
                    f"{SERIES_METADATA[column]['label']} has {non_good_quality:,} non-good source flag(s) ({breakdown}).",
                    non_good_quality,
                    series=column,
                )
            )
        if bound_violations:
            unit = SERIES_METADATA[column]["unit"]
            issues.append(
                _issue(
                    "domain_bound_violations",
                    "Values fall outside screening bounds",
                    f"{SERIES_METADATA[column]['label']} has {bound_violations:,} value(s) outside {lower:g} to {upper:g} {unit}.",
                    bound_violations,
                    series=column,
                )
            )

    timestamp_coverage = (
        100.0
        if expected_timestamp_count == 0
        else round(100.0 * observed_timestamp_count / expected_timestamp_count, 1)
    )
    expected_sample_count = expected_timestamp_count * len(columns)
    sample_presence = (
        100.0
        if expected_sample_count == 0
        else round(100.0 * present_sample_count / expected_sample_count, 1)
    )
    usable_value_completeness = (
        100.0
        if expected_sample_count == 0
        else round(100.0 * usable_sample_count / expected_sample_count, 1)
    )
    summary = {
        "expected_timestamp_count": int(expected_timestamp_count),
        "observed_timestamp_count": int(observed_timestamp_count),
        "missing_timestamp_count": int(missing_timestamp_count),
        "timestamp_coverage_percent": timestamp_coverage,
        "expected_sample_count": int(expected_sample_count),
        "present_sample_count": int(present_sample_count),
        "usable_sample_count": int(usable_sample_count),
        "sample_presence_percent": sample_presence,
        "usable_value_completeness_percent": usable_value_completeness,
        "series_completeness": series_completeness,
        "invalid_timestamp_sample_count": int(invalid_timestamp_sample_count),
        "outside_window_sample_count": int(outside_window_sample_count),
        "off_grid_sample_count": int(off_grid_sample_count),
        **{key: int(value) for key, value in sorted(totals.items())},
    }
    return {
        "status": "issues_detected" if issues else "clean",
        "issue_count": len(issues),
        "summary": summary,
        "issues": issues,
    }


def _measurement_plot_frame(
    *,
    csv_path: str | os.PathLike[str],
    interval_seconds: int,
    to_time: str,
) -> tuple[Any, list[tuple[str, str, str]]]:
    """Return bounded measured-power and cumulative-energy chart data.

    The collection interval represents an interval-average power sample. Missing
    or non-finite values remain gaps rather than being interpreted as zero energy.
    """

    import pandas as pd

    interval_seconds = int(interval_seconds)
    if interval_seconds <= 0:
        raise ValueError("The plot interval must be positive.")

    frame = pd.read_csv(csv_path)
    if "timestamp" not in frame.columns:
        raise ValueError("Collected CSV is missing the timestamp column.")
    frame["_plot_time"] = pd.to_datetime(
        frame["timestamp"], errors="coerce", utc=True
    )
    frame = frame.loc[frame["_plot_time"].notna()].copy()
    if frame.empty:
        return frame, []
    frame.sort_values("_plot_time", inplace=True, kind="stable")
    end_timestamp = pd.to_datetime(to_time, errors="raise", utc=True)
    interval_durations = (
        (end_timestamp - frame["_plot_time"])
        .dt.total_seconds()
        .clip(lower=0.0, upper=float(interval_seconds))
    )
    frame["_plot_time"] = frame["_plot_time"].dt.tz_convert(
        ZoneInfo("America/Denver")
    )

    available: list[tuple[str, str, str]] = []
    for column, label, color in _MEASURED_POWER_SERIES:
        if column not in frame.columns:
            continue
        frame[column] = pd.to_numeric(frame[column], errors="coerce")
        if frame[column].notna().any():
            available.append((column, label, color))
            frame[f"_{column}_energy_kwh"] = (
                frame[column] * interval_durations / 3_600_000.0
            ).cumsum(skipna=True)
    if not available:
        return frame, []

    plot_frame = frame
    if len(plot_frame.index) > _MAX_PLOT_POINTS:
        last = len(plot_frame.index) - 1
        positions = sorted(
            {
                round(index * last / (_MAX_PLOT_POINTS - 1))
                for index in range(_MAX_PLOT_POINTS)
            }
        )
        plot_frame = plot_frame.iloc[positions]
    return plot_frame, available


def render_measurement_plots(
    *,
    csv_path: str | os.PathLike[str],
    output_paths: Mapping[str, str | os.PathLike[str]],
    interval_seconds: int,
    to_time: str,
) -> dict[str, dict[str, Any]]:
    """Render measured-only AC power and cumulative-energy PNGs."""

    import matplotlib.dates as mdates
    from matplotlib.backends.backend_agg import FigureCanvasAgg
    from matplotlib.figure import Figure

    required_outputs = set(MEASUREMENT_PLOT_METADATA)
    if not required_outputs.issubset(output_paths):
        raise ValueError("Measurement plot output paths are incomplete.")

    plot_frame, available = _measurement_plot_frame(
        csv_path=csv_path,
        interval_seconds=interval_seconds,
        to_time=to_time,
    )
    if not available:
        return {}

    mountain_tz = ZoneInfo("America/Denver")

    def render_one(
        *,
        plot_key: str,
        title: str,
        ylabel: str,
        value_for: Any,
    ) -> None:
        destination = Path(output_paths[plot_key])
        destination.parent.mkdir(parents=True, exist_ok=True)
        figure = Figure(figsize=(14, 6), constrained_layout=True)
        FigureCanvasAgg(figure)
        axes = figure.subplots()
        for column, label, color in available:
            axes.plot(
                plot_frame["_plot_time"],
                value_for(column),
                color=color,
                linewidth=1.8,
                label=label,
            )
        axes.set_title(title)
        axes.set_xlabel("Time (Mountain)")
        axes.set_ylabel(ylabel)
        axes.grid(True, alpha=0.25)
        axes.legend(loc="best")
        axes.xaxis.set_major_formatter(
            mdates.DateFormatter("%m-%d-%Y %H:%M", tz=mountain_tz)
        )
        figure.autofmt_xdate()
        figure.savefig(destination, dpi=160)
        figure.clear()

    render_one(
        plot_key="measured_ac_power",
        title="Measured AC Power",
        ylabel="Measured Power (kW)",
        value_for=lambda column: plot_frame[column] / 1_000.0,
    )
    render_one(
        plot_key="cumulative_energy",
        title="Measured Cumulative Energy",
        ylabel="Cumulative Energy (kWh)",
        value_for=lambda column: plot_frame[f"_{column}_energy_kwh"],
    )
    series = [column for column, _label, _color in available]
    return {
        plot_key: {**metadata, "series": series}
        for plot_key, metadata in MEASUREMENT_PLOT_METADATA.items()
    }


def write_collection_workbook(
    *,
    csv_path: str | os.PathLike[str],
    output_path: str | os.PathLike[str],
    interval_seconds: int,
    to_time: str,
) -> dict[str, Any]:
    """Create an XLSX containing the collected data and two measured charts."""

    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.axis import DateAxis
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    source = Path(csv_path)
    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    plot_frame, available = _measurement_plot_frame(
        csv_path=source,
        interval_seconds=interval_seconds,
        to_time=to_time,
    )

    workbook = Workbook(write_only=True)
    workbook.properties.creator = "SBE PV Operations Dashboard"
    workbook.properties.title = "Collected Bazefield data"
    data_sheet = workbook.create_sheet("Collected data")
    data_sheet.freeze_panes = "A2"
    data_sheet.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)

    with source.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        try:
            headers = next(reader)
        except StopIteration:
            headers = []
        if not headers:
            raise ValueError("Collected CSV is missing its header row.")
        for index, header in enumerate(headers, start=1):
            width = 22 if header == "timestamp" else 18 if "quality" in header else 28
            data_sheet.column_dimensions[get_column_letter(index)].width = width

        styled_headers = []
        for header in headers:
            cell = WriteOnlyCell(data_sheet, value=header)
            cell.fill = header_fill
            cell.font = header_font
            styled_headers.append(cell)
        data_sheet.append(styled_headers)

        row_count = 0
        for raw_row in reader:
            row_count += 1
            padded = [*raw_row, *([""] * max(0, len(headers) - len(raw_row)))]
            cells = []
            for column, raw_value in zip(headers, padded[: len(headers)]):
                value: Any = raw_value
                number_format = None
                if not raw_value:
                    value = None
                elif column == "timestamp":
                    try:
                        value = datetime.fromisoformat(raw_value)
                        number_format = "yyyy-mm-dd hh:mm:ss"
                    except ValueError:
                        value = raw_value
                elif not column.endswith("_quality"):
                    try:
                        numeric = float(raw_value)
                    except (TypeError, ValueError, OverflowError):
                        value = raw_value
                    else:
                        if not math.isfinite(numeric):
                            value = None
                        else:
                            value = int(numeric) if numeric.is_integer() else numeric
                            number_format = (
                                "0" if isinstance(value, int) else "0.000"
                            )
                cell = WriteOnlyCell(data_sheet, value=value)
                if number_format:
                    cell.number_format = number_format
                cells.append(cell)
            data_sheet.append(cells)

    last_column = get_column_letter(len(headers))
    data_sheet.auto_filter.ref = f"A1:{last_column}{row_count + 1}"

    charts_sheet = workbook.create_sheet("Measured charts")
    charts_sheet.sheet_view.showGridLines = False
    system_names = " and ".join(
        label.removesuffix(" measured") for _column, label, _color in available
    )
    title = WriteOnlyCell(
        charts_sheet,
        value=(
            f"Measured {system_names} performance"
            if system_names
            else "Measured system performance"
        ),
    )
    title.font = Font(color="0F4C45", bold=True, size=16)
    charts_sheet.append([title])
    charts_sheet.append(
        [
            "Both charts contain measured values only. Cumulative energy is "
            "integrated from the selected interval-average AC power."
        ]
    )

    chart_data_sheet = workbook.create_sheet("Chart data")
    chart_data_sheet.sheet_state = "hidden"
    chart_data_sheet.sheet_view.showGridLines = False
    chart_headers = ["Timestamp (Mountain)"]
    chart_headers.extend(f"{label} power (kW)" for _column, label, _color in available)
    chart_headers.extend(
        f"{label} cumulative energy (kWh)"
        for _column, label, _color in available
    )
    chart_data_sheet.append(chart_headers)

    def finite_or_none(value: Any) -> float | None:
        try:
            numeric = float(value)
        except (TypeError, ValueError, OverflowError):
            return None
        return numeric if math.isfinite(numeric) else None

    for _, row in plot_frame.iterrows():
        timestamp = row["_plot_time"].to_pydatetime().replace(tzinfo=None)
        chart_row: list[Any] = [timestamp]
        chart_row.extend(
            finite_or_none(row[column] / 1_000.0)
            for column, _label, _color in available
        )
        chart_row.extend(
            finite_or_none(row[f"_{column}_energy_kwh"])
            for column, _label, _color in available
        )
        chart_data_sheet.append(chart_row)

    chart_count = 0
    if available and not plot_frame.empty:
        max_row = len(plot_frame.index) + 1
        categories = Reference(
            chart_data_sheet,
            min_col=1,
            min_row=2,
            max_row=max_row,
        )

        def add_line_chart(
            *,
            title_text: str,
            y_axis_title: str,
            min_col: int,
            max_col: int,
            anchor: str,
        ) -> None:
            nonlocal chart_count
            chart = LineChart()
            chart.title = title_text
            chart.style = 13
            chart.height = 8
            chart.width = 18
            chart.legend.position = "b"
            chart.display_blanks = "gap"
            chart.y_axis.title = y_axis_title
            chart.x_axis = DateAxis()
            chart.x_axis.title = "Time (Mountain)"
            chart.x_axis.number_format = "mm-dd-yyyy hh:mm"
            chart.add_data(
                Reference(
                    chart_data_sheet,
                    min_col=min_col,
                    min_row=1,
                    max_col=max_col,
                    max_row=max_row,
                ),
                titles_from_data=True,
            )
            chart.set_categories(categories)
            for series, (_column, label, color) in zip(chart.series, available):
                series.tx = SeriesLabel(v=label)
                series.graphicalProperties.line.solidFill = color.removeprefix("#")
                series.graphicalProperties.line.width = 28_575
            charts_sheet.add_chart(chart, anchor)
            chart_count += 1

        power_first = 2
        power_last = power_first + len(available) - 1
        energy_first = power_last + 1
        energy_last = energy_first + len(available) - 1
        add_line_chart(
            title_text=f"Measured AC Power: {system_names}",
            y_axis_title="Power (kW)",
            min_col=power_first,
            max_col=power_last,
            anchor="A4",
        )
        add_line_chart(
            title_text=f"Measured Cumulative Energy: {system_names}",
            y_axis_title="Cumulative Energy (kWh)",
            min_col=energy_first,
            max_col=energy_last,
            anchor="A21",
        )
    else:
        charts_sheet.append(
            ["No SolarEdge or Solectria measured-power values were available."]
        )

    workbook.save(destination)
    return {
        "sheet_count": 3,
        "chart_count": chart_count,
        "series": [column for column, _label, _color in available],
    }


def collect_historian_data(
    *,
    from_time: str,
    to_time: str,
    interval_seconds: int,
    data_groups: Sequence[str],
    output_csv: str | os.PathLike[str],
    api_key: str | None = None,
    base_url: str | None = None,
) -> dict[str, Any]:
    """Collect selected STAC1 series, retain quality flags, and write a CSV."""

    interval_seconds = int(interval_seconds)
    if interval_seconds <= 0:
        raise bazefield.BazefieldError("The collection interval must be positive.")
    start_ms = _utc_milliseconds(from_time)
    end_ms = _utc_milliseconds(to_time)
    if start_ms >= end_ms:
        raise bazefield.BazefieldError(
            "The collection start must be before the collection end."
        )
    step_ms = interval_seconds * 1_000
    expected_timestamp_count = math.ceil((end_ms - start_ms) / step_ms)

    columns = selected_series(data_groups)
    selected_column_set = set(columns)
    selected_map = [
        item for item in bazefield.COLUMN_MAP if item[2] in selected_column_set
    ]
    object_ids = _unique(item[0] for item in selected_map)
    points = _unique(item[1] for item in selected_map)
    column_by_source = {
        (object_id, point_name): column
        for object_id, point_name, column in selected_map
    }

    client = bazefield.BazefieldClient(
        base_url or os.environ.get("BAZEFIELD_BASE_URL") or bazefield.DEFAULT_BASE_URL,
        api_key or os.environ.get("BAZEFIELD_API_KEY"),
    )
    response = client.get_historian(
        object_ids=",".join(object_ids),
        points=",".join(points),
        aggregates=bazefield.normalize_aggregates(bazefield.AGGREGATE),
        frm=from_time,
        to=to_time,
        interval=str(interval_seconds),
    )
    reply_info = response.get("replyInfo") if isinstance(response, dict) else None
    reply_error = (reply_info or {}).get("error")
    if reply_error:
        raise bazefield.BazefieldError(
            f"Historian response reported an error: {reply_error}"
        )

    buckets: defaultdict[int, defaultdict[str, list[Mapping[str, Any]]]] = (
        defaultdict(lambda: defaultdict(list))
    )
    invalid_timestamp_sample_count = 0
    outside_window_sample_count = 0
    off_grid_sample_count = 0
    for sample in bazefield.flatten(response):
        column = column_by_source.get(
            (sample.get("objectId"), sample.get("pointName"))
        )
        if column is None:
            continue
        try:
            timestamp_ms = int(sample.get("t_ms"))
        except (TypeError, ValueError, OverflowError):
            invalid_timestamp_sample_count += 1
            continue
        if not start_ms <= timestamp_ms < end_ms:
            outside_window_sample_count += 1
            continue
        if (timestamp_ms - start_ms) % step_ms:
            off_grid_sample_count += 1
            continue
        buckets[timestamp_ms][column].append(sample)

    rows: list[dict[str, Any]] = []
    output_columns = ["timestamp"]
    for column in columns:
        output_columns.extend(
            (column, f"{column}_quality_code", f"{column}_quality")
        )
    for timestamp_ms in sorted(buckets):
        timestamp_columns = buckets[timestamp_ms]
        row: dict[str, Any] = {"timestamp": bazefield._utc_stamp(timestamp_ms)}
        for column in columns:
            samples = timestamp_columns.get(column) or []
            sample = samples[-1] if samples else None
            row[column] = (
                bazefield._fmt_value(column, sample.get("value"))
                if sample is not None
                else None
            )
            quality = sample.get("quality") if sample is not None else None
            row[f"{column}_quality_code"] = quality
            row[f"{column}_quality"] = (
                primary_quality_label(quality) if sample is not None else "missing"
            )
        rows.append(row)

    destination = Path(output_csv)
    bazefield.write_csv(rows, str(destination), output_columns)
    measured_energy_kwh: dict[str, float | None] = {}
    ordered_timestamps = sorted(buckets)
    for column, _label, _color in _MEASURED_POWER_SERIES:
        if column not in selected_column_set:
            continue
        total_kwh = 0.0
        usable_sample_count = 0
        for timestamp_ms, row in zip(ordered_timestamps, rows, strict=True):
            value = row.get(column)
            if not _finite_number(value):
                continue
            duration_seconds = min(
                float(interval_seconds),
                max(0.0, (end_ms - timestamp_ms) / 1_000.0),
            )
            total_kwh += float(value) * duration_seconds / 3_600_000.0
            usable_sample_count += 1
        measured_energy_kwh[column] = total_kwh if usable_sample_count else None
    quality = _quality_report(
        buckets,
        columns,
        expected_timestamp_count=expected_timestamp_count,
        invalid_timestamp_sample_count=invalid_timestamp_sample_count,
        outside_window_sample_count=outside_window_sample_count,
        off_grid_sample_count=off_grid_sample_count,
    )
    selected_groups = [group for group in DATA_GROUP_ORDER if group in set(data_groups)]
    return {
        "row_count": len(rows),
        "column_count": len(output_columns),
        "data_groups": [
            {"id": group, "label": DATA_GROUP_LABELS[group]}
            for group in selected_groups
        ],
        "series": [
            {"name": column, **SERIES_METADATA[column]}
            for column in columns
        ],
        "measured_energy_kwh": measured_energy_kwh,
        "quality": quality,
    }
