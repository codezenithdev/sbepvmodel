"""Deterministic baseline-versus-scenario comparison and reporting.

This module deliberately contains no language-model logic.  It reads the model
workbooks produced by :mod:`sbe_pv_model`, verifies source provenance for
like-for-like comparisons, calculates all engineering deltas in Python, and
creates the comparison workbook and charts consumed by the scenario-agent UI.
"""

from __future__ import annotations

from collections.abc import Iterable, Mapping
from dataclasses import dataclass
from datetime import date, datetime
from hashlib import sha256
import json
import math
from pathlib import Path
from typing import Any, Literal
from urllib.parse import quote

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


COMPARISON_SCHEMA_VERSION = 1
SAME_INPUT = "same_input"
CROSS_RUN = "cross_run"
COMPARISON_TYPES = frozenset({SAME_INPUT, CROSS_RUN})

CROSS_RUN_CAVEAT = (
    "Non-like-for-like comparison: the source data, timestamps, or run window "
    "differs. Aggregate differences are descriptive only and must not be "
    "attributed causally to parameter changes."
)

PREDICTED_POWER_COLUMNS = (
    "se_predicted_power_w",
    "sol_predicted_power_w",
)
MEASURED_POWER_COLUMNS = (
    "se_measured_power_w",
    "sol_measured_power_w",
)
TIMESTAMP_COLUMNS = ("timestamp_utc_naive", "timestamp_local_naive")

FORMULAS = {
    "predicted_delta_kwh": "candidate_predicted_kwh - baseline_predicted_kwh",
    "predicted_delta_pct": (
        "predicted_delta_kwh / baseline_predicted_kwh * 100; null when baseline is zero"
    ),
    "validation_residual_pct": (
        "(predicted_kwh - measured_kwh) / measured_kwh * 100; null when measured is zero"
    ),
    "absolute_error_improvement_pp": (
        "abs(baseline_residual_pct) - abs(candidate_residual_pct); positive means improved fit"
    ),
    "predicted_system_gap_kwh": (
        "solaredge_predicted_kwh - solectria_predicted_kwh"
    ),
    "predicted_system_gap_pct": (
        "predicted_system_gap_kwh / solectria_predicted_kwh * 100; null when Solectria is zero"
    ),
}


class ComparisonInvariantError(ValueError):
    """Raised when a requested same-input comparison is not like-for-like."""


class SourceFingerprintMismatch(ComparisonInvariantError):
    """Raised when a source file no longer matches its recorded SHA-256."""


@dataclass(frozen=True)
class ModelWorkbook:
    """Normalized contents of one model-output workbook."""

    path: Path | None
    time_series: pd.DataFrame
    run_info: dict[str, Any]
    monthly_energy: pd.DataFrame | None
    mode: Literal["validation", "annual"]


def sha256_file(path: str | Path, chunk_size: int = 1024 * 1024) -> str:
    """Return the lowercase SHA-256 hex digest of *path*.

    Reading in fixed-size chunks keeps this deterministic without loading a
    historian or annual weather file wholly into memory.
    """

    source = Path(path)
    if not source.is_file():
        raise FileNotFoundError(f"Source file does not exist: {source}")
    if chunk_size <= 0:
        raise ValueError("chunk_size must be positive.")

    digest = sha256()
    with source.open("rb") as handle:
        while chunk := handle.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest()


def _normalize_sha256(value: str) -> str:
    normalized = str(value).strip().lower()
    if len(normalized) != 64 or any(char not in "0123456789abcdef" for char in normalized):
        raise ValueError("SHA-256 fingerprints must be 64 hexadecimal characters.")
    return normalized


def verify_source_sha256(path: str | Path, expected_sha256: str) -> str:
    """Hash *path* and raise if it differs from *expected_sha256*."""

    expected = _normalize_sha256(expected_sha256)
    actual = sha256_file(path)
    if actual != expected:
        raise SourceFingerprintMismatch(
            f"Source fingerprint mismatch for {Path(path).name}: "
            f"expected {expected}, calculated {actual}."
        )
    return actual


def _as_bool(value: Any) -> bool:
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "on"}
    return bool(value)


def load_model_workbook(path: str | Path) -> ModelWorkbook:
    """Load and validate a workbook written by ``sbe_pv_model.write_excel``."""

    workbook_path = Path(path)
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Model workbook does not exist: {workbook_path}")

    with pd.ExcelFile(workbook_path) as workbook:
        if "time_series" not in workbook.sheet_names:
            raise ValueError(f"{workbook_path.name} has no time_series sheet.")
        time_series = pd.read_excel(workbook, sheet_name="time_series")
        run_info: dict[str, Any] = {}
        if "run_info" in workbook.sheet_names:
            raw_info = pd.read_excel(workbook, sheet_name="run_info")
            if {"parameter", "value"}.issubset(raw_info.columns):
                run_info = {
                    str(row["parameter"]): row["value"]
                    for _, row in raw_info.iterrows()
                    if not pd.isna(row["parameter"])
                }
        monthly = (
            pd.read_excel(workbook, sheet_name="monthly_energy")
            if "monthly_energy" in workbook.sheet_names
            else None
        )

    if time_series.empty:
        raise ValueError(f"{workbook_path.name} contains no model time-series rows.")
    missing = [column for column in PREDICTED_POWER_COLUMNS if column not in time_series]
    if missing:
        raise ValueError(
            f"{workbook_path.name} is missing required columns: {', '.join(missing)}."
        )
    if not any(column in time_series for column in TIMESTAMP_COLUMNS):
        raise ValueError(f"{workbook_path.name} contains no canonical timestamp column.")

    for column in TIMESTAMP_COLUMNS:
        if column in time_series:
            time_series[column] = pd.to_datetime(time_series[column], errors="coerce")

    mode: Literal["validation", "annual"] = (
        "annual" if _as_bool(run_info.get("annual_mode", False)) else "validation"
    )
    return ModelWorkbook(
        path=workbook_path,
        time_series=time_series,
        run_info=run_info,
        monthly_energy=monthly,
        mode=mode,
    )


def _coerce_workbook(
    value: ModelWorkbook | pd.DataFrame | str | Path,
    *,
    mode: str | None,
) -> ModelWorkbook:
    if isinstance(value, ModelWorkbook):
        return value
    if isinstance(value, pd.DataFrame):
        selected_mode = mode or "validation"
        if selected_mode not in {"validation", "annual"}:
            raise ValueError("mode must be validation or annual.")
        return ModelWorkbook(
            path=None,
            time_series=value.copy(),
            run_info={"annual_mode": selected_mode == "annual"},
            monthly_energy=None,
            mode=selected_mode,  # type: ignore[arg-type]
        )
    return load_model_workbook(value)


def _timestamp_series(frame: pd.DataFrame) -> pd.Series:
    for column in TIMESTAMP_COLUMNS:
        if column in frame:
            values = pd.to_datetime(frame[column], errors="coerce")
            if values.isna().any():
                raise ValueError(f"Column {column} contains invalid timestamps.")
            return values.reset_index(drop=True)
    if isinstance(frame.index, pd.DatetimeIndex):
        values = pd.Series(frame.index)
        if values.isna().any():
            raise ValueError("Time-series index contains invalid timestamps.")
        return values.reset_index(drop=True)
    raise ValueError("Time series has no canonical timestamp column or DatetimeIndex.")


def _timestamps_aligned(baseline: pd.DataFrame, candidate: pd.DataFrame) -> bool:
    baseline_ts = _timestamp_series(baseline)
    candidate_ts = _timestamp_series(candidate)
    return len(baseline_ts) == len(candidate_ts) and baseline_ts.equals(candidate_ts)


def _plot_timestamps(frame: pd.DataFrame) -> tuple[pd.Series, str]:
    """Prefer human-readable local timestamps while retaining UTC for invariants."""

    if "timestamp_local_naive" in frame:
        values = pd.to_datetime(frame["timestamp_local_naive"], errors="coerce")
        if not values.isna().any():
            return values.reset_index(drop=True), "Local time"
    return _timestamp_series(frame), "Time"


def _last_finite(series: pd.Series) -> float | None:
    values = pd.to_numeric(series, errors="coerce")
    values = values[np.isfinite(values)]
    return float(values.iloc[-1]) if len(values) else None


def _energy_total(frame: pd.DataFrame, system: str, kind: str) -> float | None:
    cumulative = f"{system}_{kind}_energy_kwh"
    if cumulative in frame:
        return _last_finite(frame[cumulative])

    power = f"{system}_{kind}_power_w"
    if power not in frame:
        return None
    if "dt_hours" in frame:
        dt_hours = pd.to_numeric(frame["dt_hours"], errors="coerce")
    else:
        timestamps = _timestamp_series(frame)
        dt_hours = timestamps.diff().dt.total_seconds() / 3600.0
        median = dt_hours.dropna().median()
        dt_hours = dt_hours.fillna(float(median) if pd.notna(median) else 0.0)
    watts = pd.to_numeric(frame[power], errors="coerce")
    value = (watts * dt_hours / 1000.0).sum(min_count=1)
    return float(value) if pd.notna(value) and np.isfinite(value) else None


def _safe_number(value: Any, digits: int = 8) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    rounded = round(number, digits)
    return 0.0 if rounded == 0 else rounded


def _delta_pct(candidate: float | None, baseline: float | None) -> float | None:
    if candidate is None or baseline is None or baseline == 0:
        return None
    return _safe_number((candidate - baseline) / baseline * 100.0)


def _residual_pct(predicted: float | None, measured: float | None) -> float | None:
    if predicted is None or measured is None or measured == 0:
        return None
    return _safe_number((predicted - measured) / measured * 100.0)


def _numbers_match(first: float | None, second: float | None) -> bool:
    if first is None or second is None:
        return first is second
    return math.isclose(first, second, rel_tol=0.0, abs_tol=1e-9)


def _series_match(first: pd.Series, second: pd.Series) -> bool:
    if len(first) != len(second):
        return False
    a = pd.to_numeric(first, errors="coerce").to_numpy(dtype=float)
    b = pd.to_numeric(second, errors="coerce").to_numpy(dtype=float)
    return bool(np.allclose(a, b, rtol=0.0, atol=1e-9, equal_nan=True))


def _flatten_snapshot(value: Mapping[str, Any] | None) -> dict[str, Any]:
    flattened: dict[str, Any] = {}

    def visit(prefix: str, item: Any) -> None:
        if isinstance(item, Mapping):
            if not item and prefix:
                flattened[prefix] = {}
            for key in sorted(item, key=str):
                name = f"{prefix}.{key}" if prefix else str(key)
                visit(name, item[key])
        else:
            flattened[prefix] = _json_safe(item)

    if value:
        visit("", value)
    return flattened


def parameter_changes(
    baseline_request: Mapping[str, Any] | None,
    candidate_request: Mapping[str, Any] | None,
) -> list[dict[str, Any]]:
    """Return deterministic leaf-level request differences."""

    baseline = _flatten_snapshot(baseline_request)
    candidate = _flatten_snapshot(candidate_request)
    changes: list[dict[str, Any]] = []
    for field in sorted(set(baseline) | set(candidate)):
        before = baseline.get(field)
        after = candidate.get(field)
        if json.dumps(before, sort_keys=True) != json.dumps(after, sort_keys=True):
            changes.append(
                {"field": field, "baseline_value": before, "candidate_value": after}
            )
    return changes


def compute_comparison(
    baseline: ModelWorkbook | pd.DataFrame | str | Path,
    candidate: ModelWorkbook | pd.DataFrame | str | Path,
    *,
    mode: Literal["validation", "annual"] | None = None,
    comparison_type: Literal["same_input", "cross_run"] | None = None,
    baseline_source_sha256: str | None = None,
    candidate_source_sha256: str | None = None,
    baseline_request: Mapping[str, Any] | None = None,
    candidate_request: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """Calculate a JSON-safe, deterministic model comparison.

    ``same_input`` is intentionally strict.  It is rejected unless both source
    hashes match, timestamps align, and measured observations are unchanged.
    ``cross_run`` never permits causal attribution.
    """

    baseline_book = _coerce_workbook(baseline, mode=mode)
    candidate_book = _coerce_workbook(candidate, mode=mode)
    selected_mode = mode or candidate_book.mode
    if selected_mode not in {"validation", "annual"}:
        raise ValueError("mode must be validation or annual.")
    if baseline_book.mode != candidate_book.mode and mode is None:
        raise ValueError("Baseline and candidate workbooks use different run modes.")

    baseline_hash = (
        _normalize_sha256(baseline_source_sha256)
        if baseline_source_sha256 is not None
        else None
    )
    candidate_hash = (
        _normalize_sha256(candidate_source_sha256)
        if candidate_source_sha256 is not None
        else None
    )
    hashes_present = baseline_hash is not None and candidate_hash is not None
    hashes_match = hashes_present and baseline_hash == candidate_hash

    baseline_frame = baseline_book.time_series
    candidate_frame = candidate_book.time_series
    timestamps_aligned = _timestamps_aligned(baseline_frame, candidate_frame)

    measured_totals: dict[str, dict[str, float | None]] = {}
    measured_total_flags: list[bool] = []
    measured_series_flags: list[bool] = []
    for system in ("se", "sol"):
        baseline_measured = _energy_total(baseline_frame, system, "measured")
        candidate_measured = _energy_total(candidate_frame, system, "measured")
        measured_totals[system] = {
            "baseline": baseline_measured,
            "candidate": candidate_measured,
        }
        measured_total_flags.append(_numbers_match(baseline_measured, candidate_measured))
        column = f"{system}_measured_power_w"
        measured_series_flags.append(
            column in baseline_frame
            and column in candidate_frame
            and _series_match(baseline_frame[column], candidate_frame[column])
        )

    measured_totals_match = all(measured_total_flags)
    measured_series_match = all(measured_series_flags)

    if comparison_type is None:
        comparison_type = (
            SAME_INPUT
            if hashes_match
            and timestamps_aligned
            and measured_totals_match
            and measured_series_match
            else CROSS_RUN
        )
    if comparison_type not in COMPARISON_TYPES:
        raise ValueError("comparison_type must be same_input or cross_run.")
    if comparison_type == SAME_INPUT:
        failures = []
        if not hashes_present:
            failures.append("both source SHA-256 fingerprints are required")
        elif not hashes_match:
            failures.append("source SHA-256 fingerprints differ")
        if not timestamps_aligned:
            failures.append("timestamps do not align")
        if not measured_totals_match:
            failures.append("measured energy totals differ")
        if not measured_series_match:
            failures.append("measured power observations differ")
        if failures:
            raise ComparisonInvariantError(
                "Cannot create a same-input comparison: " + "; ".join(failures) + "."
            )

    systems: dict[str, Any] = {}
    predicted_totals: dict[str, tuple[float | None, float | None]] = {}
    for system, public_name in (("se", "solaredge"), ("sol", "solectria")):
        baseline_predicted = _energy_total(baseline_frame, system, "predicted")
        candidate_predicted = _energy_total(candidate_frame, system, "predicted")
        predicted_totals[system] = (baseline_predicted, candidate_predicted)
        delta = (
            candidate_predicted - baseline_predicted
            if baseline_predicted is not None and candidate_predicted is not None
            else None
        )

        validation = None
        if selected_mode == "validation":
            baseline_measured = measured_totals[system]["baseline"]
            candidate_measured = measured_totals[system]["candidate"]
            baseline_residual = _residual_pct(baseline_predicted, baseline_measured)
            candidate_residual = _residual_pct(candidate_predicted, candidate_measured)
            improvement = (
                abs(baseline_residual) - abs(candidate_residual)
                if baseline_residual is not None and candidate_residual is not None
                else None
            )
            validation = {
                "baseline_measured_kwh": _safe_number(baseline_measured),
                "candidate_measured_kwh": _safe_number(candidate_measured),
                "baseline_residual_pct": _safe_number(baseline_residual),
                "candidate_residual_pct": _safe_number(candidate_residual),
                "absolute_error_improvement_pp": _safe_number(improvement),
                "positive_improvement_means": "candidate fit is closer to measured energy",
            }

        systems[public_name] = {
            "baseline_predicted_kwh": _safe_number(baseline_predicted),
            "candidate_predicted_kwh": _safe_number(candidate_predicted),
            "delta_kwh": _safe_number(delta),
            "delta_pct": _delta_pct(candidate_predicted, baseline_predicted),
            "validation": validation,
        }

    baseline_se, candidate_se = predicted_totals["se"]
    baseline_sol, candidate_sol = predicted_totals["sol"]
    baseline_gap = (
        baseline_se - baseline_sol
        if baseline_se is not None and baseline_sol is not None
        else None
    )
    candidate_gap = (
        candidate_se - candidate_sol
        if candidate_se is not None and candidate_sol is not None
        else None
    )
    baseline_gap_pct = _delta_pct(baseline_se, baseline_sol)
    candidate_gap_pct = _delta_pct(candidate_se, candidate_sol)
    gap_change = (
        candidate_gap - baseline_gap
        if candidate_gap is not None and baseline_gap is not None
        else None
    )
    gap_pct_point_change = (
        candidate_gap_pct - baseline_gap_pct
        if candidate_gap_pct is not None and baseline_gap_pct is not None
        else None
    )

    changes = parameter_changes(baseline_request, candidate_request)
    like_for_like = comparison_type == SAME_INPUT
    if not like_for_like:
        attribution_scope = "descriptive_only"
    elif len(changes) == 1:
        attribution_scope = "single_parameter"
    elif len(changes) > 1:
        attribution_scope = "combined_configuration"
    else:
        attribution_scope = "no_parameter_change"

    result = {
        "schema_version": COMPARISON_SCHEMA_VERSION,
        "mode": selected_mode,
        "comparison_type": comparison_type,
        "like_for_like": like_for_like,
        "caveat": None if like_for_like else CROSS_RUN_CAVEAT,
        "systems": systems,
        "cross_system_gap": {
            "baseline_kwh": _safe_number(baseline_gap),
            "candidate_kwh": _safe_number(candidate_gap),
            "change_kwh": _safe_number(gap_change),
            "baseline_pct_of_solectria": _safe_number(baseline_gap_pct),
            "candidate_pct_of_solectria": _safe_number(candidate_gap_pct),
            "change_pct_points": _safe_number(gap_pct_point_change),
        },
        "invariants": {
            "source_hashes_present": hashes_present,
            "source_hashes_match": bool(hashes_match),
            "timestamps_aligned": timestamps_aligned,
            "measured_totals_match": measured_totals_match,
            "measured_series_match": measured_series_match,
        },
        "parameter_changes": changes,
        "attribution": {
            "scope": attribution_scope,
            "causal_attribution_allowed": like_for_like and bool(changes),
            "individual_parameter_attribution_allowed": (
                like_for_like and len(changes) == 1
            ),
            "changed_field_count": len(changes),
        },
        "formulas": FORMULAS,
    }
    return _json_safe(result)


def _resolve_source_hash(
    book: ModelWorkbook,
    source_path: str | Path | None,
    expected_hash: str | None,
) -> str | None:
    if source_path is not None:
        return (
            verify_source_sha256(source_path, expected_hash)
            if expected_hash is not None
            else sha256_file(source_path)
        )
    if expected_hash is not None:
        return _normalize_sha256(expected_hash)

    recorded_path = book.run_info.get("input_csv")
    if recorded_path and Path(str(recorded_path)).is_file():
        return sha256_file(Path(str(recorded_path)))
    return None


def _warning_list(value: Any) -> list[str]:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return []
    if isinstance(value, (list, tuple, set)):
        return [str(item) for item in value if str(item).strip()]
    text = str(value).strip()
    if not text or text.lower() == "none":
        return []
    return [item.strip() for item in text.split("|") if item.strip()]


def _build_provenance(
    baseline: ModelWorkbook,
    candidate: ModelWorkbook,
    comparison: Mapping[str, Any],
    *,
    baseline_job_id: str,
    candidate_job_id: str,
    baseline_request: Mapping[str, Any] | None,
    candidate_request: Mapping[str, Any] | None,
    baseline_hash: str | None,
    candidate_hash: str | None,
    extra_warnings: Iterable[str],
) -> dict[str, Any]:
    warnings = (
        _warning_list(baseline.run_info.get("data_quality_warnings"))
        + _warning_list(candidate.run_info.get("data_quality_warnings"))
        + [str(item) for item in extra_warnings if str(item).strip()]
    )
    if comparison.get("caveat"):
        warnings.append(str(comparison["caveat"]))
    warnings = list(dict.fromkeys(warnings))

    return _json_safe(
        {
            "schema_version": COMPARISON_SCHEMA_VERSION,
            "generated_at_utc": pd.Timestamp.now(tz="UTC").isoformat(),
            "comparison_type": comparison["comparison_type"],
            "like_for_like": comparison["like_for_like"],
            "baseline": {
                "job_id": baseline_job_id,
                "workbook": str(baseline.path) if baseline.path else None,
                "request": dict(baseline_request or {}),
                "source_sha256": baseline_hash,
                "run_timestamp_utc": baseline.run_info.get("run_timestamp_utc"),
                "model_version": baseline.run_info.get("version"),
                "dhi_source": baseline.run_info.get("dhi_source"),
                "curtailment_scope": baseline.run_info.get("curtailment_scope"),
            },
            "candidate": {
                "job_id": candidate_job_id,
                "workbook": str(candidate.path) if candidate.path else None,
                "request": dict(candidate_request or {}),
                "source_sha256": candidate_hash,
                "run_timestamp_utc": candidate.run_info.get("run_timestamp_utc"),
                "model_version": candidate.run_info.get("version"),
                "dhi_source": candidate.run_info.get("dhi_source"),
                "curtailment_scope": candidate.run_info.get("curtailment_scope"),
            },
            "formulas": FORMULAS,
            "warnings": warnings,
        }
    )


def _aligned_delta_frame(baseline: pd.DataFrame, candidate: pd.DataFrame) -> pd.DataFrame:
    data: dict[str, Any] = {}
    for timestamp_column in ("timestamp_local_naive", "timestamp_utc_naive"):
        if timestamp_column in candidate:
            data[timestamp_column] = candidate[timestamp_column].reset_index(drop=True)
    for system in ("se", "sol"):
        measured = f"{system}_measured_power_w"
        if measured in baseline:
            data[measured] = baseline[measured].reset_index(drop=True)
        for quantity in ("power_w", "energy_kwh"):
            column = f"{system}_predicted_{quantity}"
            if column not in baseline or column not in candidate:
                continue
            baseline_values = pd.to_numeric(baseline[column], errors="coerce").reset_index(drop=True)
            candidate_values = pd.to_numeric(candidate[column], errors="coerce").reset_index(drop=True)
            data[f"baseline_{column}"] = baseline_values
            data[f"candidate_{column}"] = candidate_values
            data[f"delta_{column}"] = candidate_values - baseline_values
    return pd.DataFrame(data)


def _monthly_from_book(book: ModelWorkbook) -> pd.DataFrame | None:
    if book.monthly_energy is not None and not book.monthly_energy.empty:
        monthly = book.monthly_energy.copy()
        required = {"SolarEdge_predicted_kWh", "Solectria_predicted_kWh"}
        if required.issubset(monthly.columns):
            if "month_start" in monthly:
                monthly["month_start"] = pd.to_datetime(monthly["month_start"], errors="coerce")
            elif "month" in monthly:
                monthly["month_start"] = pd.to_datetime(monthly["month"], errors="coerce")
            return monthly[
                ["month_start", "SolarEdge_predicted_kWh", "Solectria_predicted_kWh"]
            ].dropna(subset=["month_start"])

    frame = book.time_series
    if "dt_hours" not in frame:
        return None
    try:
        timestamps, _ = _plot_timestamps(frame)
    except ValueError:
        return None
    month_start = timestamps.dt.to_period("M").dt.to_timestamp()
    result = pd.DataFrame({"month_start": month_start})
    for system, name in (
        ("se", "SolarEdge_predicted_kWh"),
        ("sol", "Solectria_predicted_kWh"),
    ):
        power = pd.to_numeric(frame[f"{system}_predicted_power_w"], errors="coerce")
        dt_hours = pd.to_numeric(frame["dt_hours"], errors="coerce")
        result[name] = power * dt_hours / 1000.0
    return result.groupby("month_start", as_index=False).sum(numeric_only=True)


def _monthly_comparison_frame(
    baseline: ModelWorkbook, candidate: ModelWorkbook
) -> pd.DataFrame | None:
    baseline_monthly = _monthly_from_book(baseline)
    candidate_monthly = _monthly_from_book(candidate)
    if baseline_monthly is None or candidate_monthly is None:
        return None
    merged = baseline_monthly.merge(
        candidate_monthly,
        on="month_start",
        how="outer",
        suffixes=("_baseline", "_scenario"),
    ).sort_values("month_start")
    merged.insert(1, "month", merged["month_start"].dt.strftime("%b %Y"))
    for system in ("SolarEdge", "Solectria"):
        baseline_column = f"{system}_predicted_kWh_baseline"
        scenario_column = f"{system}_predicted_kWh_scenario"
        merged[f"{system}_delta_kWh"] = merged[scenario_column] - merged[baseline_column]
        merged[f"{system}_delta_pct"] = np.where(
            merged[baseline_column].notna() & (merged[baseline_column] != 0),
            merged[f"{system}_delta_kWh"] / merged[baseline_column] * 100.0,
            np.nan,
        )
    return merged


def _summary_frame(comparison: Mapping[str, Any]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for system_name, label in (("solaredge", "SolarEdge"), ("solectria", "Solectria")):
        system = comparison["systems"][system_name]
        rows.append(
            {
                "section": "predicted_energy",
                "system": label,
                "metric": "predicted_energy",
                "baseline_value": system["baseline_predicted_kwh"],
                "candidate_value": system["candidate_predicted_kwh"],
                "change": system["delta_kwh"],
                "change_pct_or_pp": system["delta_pct"],
                "unit": "kWh / %",
                "note": "Signed candidate minus baseline",
            }
        )
        validation = system.get("validation")
        if validation:
            rows.append(
                {
                    "section": "validation",
                    "system": label,
                    "metric": "model_residual",
                    "baseline_value": validation["baseline_residual_pct"],
                    "candidate_value": validation["candidate_residual_pct"],
                    "change": validation["absolute_error_improvement_pp"],
                    "change_pct_or_pp": validation["absolute_error_improvement_pp"],
                    "unit": "% / percentage points",
                    "note": "Change is absolute-error improvement; positive means better fit",
                }
            )
    gap = comparison["cross_system_gap"]
    rows.append(
        {
            "section": "cross_system",
            "system": "SolarEdge - Solectria",
            "metric": "predicted_energy_gap",
            "baseline_value": gap["baseline_kwh"],
            "candidate_value": gap["candidate_kwh"],
            "change": gap["change_kwh"],
            "change_pct_or_pp": gap["change_pct_points"],
            "unit": "kWh / percentage points",
            "note": "Percent gap denominator is Solectria predicted energy",
        }
    )
    return pd.DataFrame(rows)


def _provenance_frame(provenance: Mapping[str, Any]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []

    def visit(section: str, key: str, value: Any) -> None:
        if isinstance(value, Mapping):
            for child_key in sorted(value, key=str):
                child_name = f"{key}.{child_key}" if key else str(child_key)
                visit(section, child_name, value[child_key])
        else:
            display = json.dumps(value, sort_keys=True) if isinstance(value, (list, dict)) else value
            rows.append({"section": section, "key": key, "value": display})

    for top_key in sorted(provenance, key=str):
        top_value = provenance[top_key]
        if isinstance(top_value, Mapping):
            visit(str(top_key), "", top_value)
        else:
            visit("comparison", str(top_key), top_value)
    return pd.DataFrame(rows, columns=["section", "key", "value"])


def _format_workbook(path: Path, summary_header_row: int = 6) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="17365D", bold=True, size=16)
    caveat_fill = PatternFill("solid", fgColor="FFF2CC")

    for worksheet in workbook.worksheets:
        worksheet.sheet_view.showGridLines = False
        header_row = summary_header_row if worksheet.title == "summary" else 1
        worksheet.freeze_panes = f"A{header_row + 1}"
        if worksheet.max_row >= header_row:
            for cell in worksheet[header_row]:
                if cell.value is not None:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            worksheet.auto_filter.ref = (
                f"A{header_row}:{worksheet.cell(worksheet.max_row, worksheet.max_column).coordinate}"
            )

        for column_cells in worksheet.iter_cols(
            min_row=1,
            max_row=min(worksheet.max_row, 250),
            max_col=worksheet.max_column,
        ):
            width = max(
                (len(str(cell.value)) for cell in column_cells if cell.value is not None),
                default=8,
            )
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(
                max(width + 2, 11), 45
            )

        for cell in worksheet[header_row]:
            header = str(cell.value or "").lower()
            if "timestamp" in header or header in {"month_start"}:
                for data_cell in worksheet.iter_rows(
                    min_row=header_row + 1,
                    max_row=worksheet.max_row,
                    min_col=cell.column,
                    max_col=cell.column,
                ):
                    data_cell[0].number_format = "yyyy-mm-dd hh:mm"
            elif any(token in header for token in ("kwh", "power", "delta", "pct", "eff")):
                for data_cell in worksheet.iter_rows(
                    min_row=header_row + 1,
                    max_row=worksheet.max_row,
                    min_col=cell.column,
                    max_col=cell.column,
                ):
                    data_cell[0].number_format = "#,##0.00"

    summary = workbook["summary"]
    summary["A1"].font = title_font
    summary["A4"].fill = caveat_fill
    summary["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    summary["B4"].fill = caveat_fill
    summary["B4"].alignment = Alignment(wrap_text=True, vertical="top")
    summary.column_dimensions["B"].width = max(summary.column_dimensions["B"].width or 0, 38)
    summary.row_dimensions[4].height = 42
    workbook.save(path)


def _write_comparison_workbook(
    path: Path,
    baseline: ModelWorkbook,
    candidate: ModelWorkbook,
    comparison: Mapping[str, Any],
    provenance: Mapping[str, Any],
    monthly: pd.DataFrame | None,
) -> None:
    summary = _summary_frame(comparison)
    changes = pd.DataFrame(
        comparison["parameter_changes"],
        columns=["field", "baseline_value", "candidate_value"],
    )
    if changes.empty:
        changes = pd.DataFrame(
            [{"field": "(none)", "baseline_value": None, "candidate_value": None}]
        )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="summary", index=False, startrow=5)
        changes.to_excel(writer, sheet_name="parameter_changes", index=False)
        baseline.time_series.to_excel(
            writer, sheet_name="baseline_time_series", index=False
        )
        candidate.time_series.to_excel(
            writer, sheet_name="scenario_time_series", index=False
        )
        _provenance_frame(provenance).to_excel(
            writer, sheet_name="provenance", index=False
        )
        if comparison["comparison_type"] == SAME_INPUT:
            _aligned_delta_frame(baseline.time_series, candidate.time_series).to_excel(
                writer, sheet_name="aligned_delta", index=False
            )
        if monthly is not None:
            monthly.to_excel(writer, sheet_name="monthly_comparison", index=False)

        worksheet = writer.book["summary"]
        worksheet["A1"] = "Solar Scenario Comparison"
        worksheet["A2"] = "Comparison type"
        worksheet["B2"] = comparison["comparison_type"]
        worksheet["A3"] = "Like for like"
        worksheet["B3"] = bool(comparison["like_for_like"])
        worksheet["A4"] = "Caveat"
        worksheet["B4"] = comparison.get("caveat") or "Like-for-like source and measured-data invariants passed."

    _format_workbook(path)


def _plot_same_input(
    baseline: pd.DataFrame,
    candidate: pd.DataFrame,
    path: Path,
    *,
    quantity: Literal["power", "energy"],
) -> None:
    timestamps, x_label = _plot_timestamps(candidate)
    suffix = "power_w" if quantity == "power" else "energy_kwh"
    divisor = 1000.0 if quantity == "power" else 1.0
    unit = "kW" if quantity == "power" else "kWh"
    title = "Predicted AC Power Comparison" if quantity == "power" else "Cumulative Predicted Energy Comparison"
    colors = {"se": "#2563EB", "sol": "#D97706"}
    labels = {"se": "SolarEdge", "sol": "Solectria"}

    fig, axis = plt.subplots(figsize=(14, 6.5))
    for system in ("se", "sol"):
        column = f"{system}_predicted_{suffix}"
        axis.plot(
            timestamps,
            pd.to_numeric(baseline[column], errors="coerce") / divisor,
            color=colors[system],
            linestyle="--",
            linewidth=1.4,
            label=f"{labels[system]} baseline",
        )
        axis.plot(
            timestamps,
            pd.to_numeric(candidate[column], errors="coerce") / divisor,
            color=colors[system],
            linestyle="-",
            linewidth=1.7,
            label=f"{labels[system]} scenario",
        )
    axis.set_title(f"{title}\nLike-for-like inputs; dashed = baseline, solid = scenario")
    axis.set_xlabel(x_label)
    axis.set_ylabel(unit)
    axis.grid(True, color="#D1D5DB", alpha=0.55, linewidth=0.7)
    axis.legend(loc="best", ncol=2)
    fig.autofmt_xdate()
    fig.tight_layout()
    fig.savefig(path, dpi=180, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def _plot_cross_run(
    baseline: pd.DataFrame,
    candidate: pd.DataFrame,
    path: Path,
    *,
    quantity: Literal["power", "energy"],
) -> None:
    suffix = "power_w" if quantity == "power" else "energy_kwh"
    divisor = 1000.0 if quantity == "power" else 1.0
    unit = "kW" if quantity == "power" else "kWh"
    title = "Predicted AC Power" if quantity == "power" else "Cumulative Predicted Energy"
    colors = {"se": "#2563EB", "sol": "#D97706"}
    labels = {"se": "SolarEdge", "sol": "Solectria"}

    fig, axes = plt.subplots(1, 2, figsize=(15, 6.5), sharey=True)
    for axis, frame, panel_name in (
        (axes[0], baseline, "Baseline run"),
        (axes[1], candidate, "Scenario run"),
    ):
        timestamps, x_label = _plot_timestamps(frame)
        for system in ("se", "sol"):
            column = f"{system}_predicted_{suffix}"
            axis.plot(
                timestamps,
                pd.to_numeric(frame[column], errors="coerce") / divisor,
                color=colors[system],
                linewidth=1.5,
                label=labels[system],
            )
        axis.set_title(panel_name)
        axis.set_xlabel(x_label)
        axis.grid(True, color="#D1D5DB", alpha=0.55, linewidth=0.7)
        axis.legend(loc="best")
        axis.tick_params(axis="x", rotation=25)
    axes[0].set_ylabel(unit)
    fig.suptitle(f"{title} — Separate Run Windows", fontsize=14)
    fig.text(
        0.5,
        0.015,
        "Non-like-for-like: panels are not pointwise aligned; aggregate differences are descriptive only.",
        ha="center",
        color="#6B7280",
        fontsize=9,
    )
    fig.tight_layout(rect=(0, 0.04, 1, 0.95))
    fig.savefig(path, dpi=180, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def _plot_monthly(monthly: pd.DataFrame, path: Path, comparison_type: str) -> None:
    labels = monthly["month"].astype(str).tolist()
    positions = np.arange(len(labels))
    width = 0.38
    fig, axes = plt.subplots(2, 1, figsize=(14, 9), sharex=True)
    for axis, system, color in (
        (axes[0], "SolarEdge", "#2563EB"),
        (axes[1], "Solectria", "#D97706"),
    ):
        baseline_values = monthly[f"{system}_predicted_kWh_baseline"].to_numpy(dtype=float)
        scenario_values = monthly[f"{system}_predicted_kWh_scenario"].to_numpy(dtype=float)
        axis.bar(
            positions - width / 2,
            baseline_values,
            width,
            color="#D1D5DB",
            edgecolor="#4B5563",
            label="Baseline",
        )
        axis.bar(
            positions + width / 2,
            scenario_values,
            width,
            color=color,
            edgecolor="#1F2937",
            label="Scenario",
        )
        axis.set_title(system)
        axis.set_ylabel("Predicted energy (kWh)")
        axis.grid(True, axis="y", color="#D1D5DB", alpha=0.55, linewidth=0.7)
        axis.legend(loc="best")
    axes[1].set_xticks(positions)
    axes[1].set_xticklabels(labels, rotation=35, ha="right")
    subtitle = (
        "Like-for-like monthly totals"
        if comparison_type == SAME_INPUT
        else "Non-like-for-like monthly totals; descriptive comparison only"
    )
    fig.suptitle(f"Monthly Predicted Energy Comparison\n{subtitle}", fontsize=14)
    fig.tight_layout(rect=(0, 0, 1, 0.94))
    fig.savefig(path, dpi=180, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def _artifact_reference(path: Path, output_url_prefix: str | None) -> dict[str, str | None]:
    absolute = path.resolve()
    url = None
    if output_url_prefix is not None:
        url = f"{output_url_prefix.rstrip('/')}/{quote(path.name)}"
    return {"path": str(absolute), "url": url}


def generate_comparison_artifacts(
    baseline_workbook: str | Path,
    candidate_workbook: str | Path,
    output_base: str | Path,
    *,
    baseline_job_id: str,
    candidate_job_id: str,
    baseline_request: Mapping[str, Any],
    candidate_request: Mapping[str, Any],
    baseline_source_path: str | Path | None = None,
    candidate_source_path: str | Path | None = None,
    baseline_source_sha256: str | None = None,
    candidate_source_sha256: str | None = None,
    comparison_type: Literal["same_input", "cross_run"] | None = None,
    mode: Literal["validation", "annual"] | None = None,
    output_url_prefix: str | None = "/outputs",
    extra_warnings: Iterable[str] = (),
) -> dict[str, Any]:
    """Generate the comparison workbook/charts and return JSON-safe metadata."""

    baseline = load_model_workbook(baseline_workbook)
    candidate = load_model_workbook(candidate_workbook)
    baseline_hash = _resolve_source_hash(
        baseline, baseline_source_path, baseline_source_sha256
    )
    candidate_hash = _resolve_source_hash(
        candidate, candidate_source_path, candidate_source_sha256
    )

    comparison = compute_comparison(
        baseline,
        candidate,
        mode=mode,
        comparison_type=comparison_type,
        baseline_source_sha256=baseline_hash,
        candidate_source_sha256=candidate_hash,
        baseline_request=baseline_request,
        candidate_request=candidate_request,
    )
    provenance = _build_provenance(
        baseline,
        candidate,
        comparison,
        baseline_job_id=baseline_job_id,
        candidate_job_id=candidate_job_id,
        baseline_request=baseline_request,
        candidate_request=candidate_request,
        baseline_hash=baseline_hash,
        candidate_hash=candidate_hash,
        extra_warnings=extra_warnings,
    )

    prefix = Path(output_base)
    if prefix.suffix:
        prefix = prefix.with_suffix("")
    prefix.parent.mkdir(parents=True, exist_ok=True)
    workbook_path = Path(f"{prefix}_comparison.xlsx")
    power_path = Path(f"{prefix}_comparison_power.png")
    energy_path = Path(f"{prefix}_comparison_energy.png")
    monthly_path = Path(f"{prefix}_comparison_monthly.png")

    monthly = (
        _monthly_comparison_frame(baseline, candidate)
        if comparison["mode"] == "annual"
        else None
    )
    _write_comparison_workbook(
        workbook_path, baseline, candidate, comparison, provenance, monthly
    )
    if comparison["comparison_type"] == SAME_INPUT:
        _plot_same_input(
            baseline.time_series,
            candidate.time_series,
            power_path,
            quantity="power",
        )
        _plot_same_input(
            baseline.time_series,
            candidate.time_series,
            energy_path,
            quantity="energy",
        )
    else:
        _plot_cross_run(
            baseline.time_series,
            candidate.time_series,
            power_path,
            quantity="power",
        )
        _plot_cross_run(
            baseline.time_series,
            candidate.time_series,
            energy_path,
            quantity="energy",
        )
    if monthly is not None and not monthly.empty:
        _plot_monthly(monthly, monthly_path, comparison["comparison_type"])

    artifacts = {
        "workbook": _artifact_reference(workbook_path, output_url_prefix),
        "power_png": _artifact_reference(power_path, output_url_prefix),
        "energy_png": _artifact_reference(energy_path, output_url_prefix),
        "monthly_png": (
            _artifact_reference(monthly_path, output_url_prefix)
            if monthly is not None and not monthly.empty
            else None
        ),
    }
    return _json_safe(
        {"comparison": comparison, "provenance": provenance, "artifacts": artifacts}
    )


# A concise alias for orchestration code.
build_comparison_report = generate_comparison_artifacts


def _json_safe(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_json_safe(item) for item in value]
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, (pd.Timestamp, datetime, date)):
        return value.isoformat()
    if value is pd.NA:
        return None
    if isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    return value


__all__ = [
    "COMPARISON_SCHEMA_VERSION",
    "CROSS_RUN",
    "CROSS_RUN_CAVEAT",
    "SAME_INPUT",
    "ComparisonInvariantError",
    "ModelWorkbook",
    "SourceFingerprintMismatch",
    "build_comparison_report",
    "compute_comparison",
    "generate_comparison_artifacts",
    "load_model_workbook",
    "parameter_changes",
    "sha256_file",
    "verify_source_sha256",
]
