from __future__ import annotations

from copy import deepcopy
import csv
import hashlib
import io
import json
from pathlib import Path
import shutil
import unittest
import uuid
from unittest.mock import patch
import zipfile

import numpy as np
import openpyxl

from sbepv import model
from sbepv import technoeconomic as kernel
from sbepv import technoeconomic_reporting as reporting
from sbepv.api import artifacts as api_artifacts
from sbepv.api import config
from sbepv.api import technoeconomic as technoeconomic_api
from sbepv.api.schemas import TechnoeconomicSubmissionRequest
from sbepv.worker import run_technoeconomic
from tests.test_technoeconomic_api import (
    _applied_site_request_payload,
    _commercial_request_payload,
    _commercial_scaling_request_payload,
    _paired_commercial_request_payload,
    _site_request_payload,
    _standalone_commercial_request_payload,
    _v6_lifecycle_request_payload,
)


class TechnoeconomicExportTests(unittest.TestCase):
    def setUp(self) -> None:
        self.test_root = Path(__file__).resolve().parent / f".tea-export-{uuid.uuid4().hex}"
        self.output = self.test_root / "outputs"
        self.output.mkdir(parents=True)
        self.output_patch = patch.object(config, "OUTPUT_DIR", self.output)
        self.output_patch.start()
        self.addCleanup(self._cleanup)

        self.job_id = "tea_export_fixture"
        self.lease_token = "lease_export_fixture"
        payload = _site_request_payload(source_id="annual-export-source", n=24)
        payload["cost_lines"][0]["label"] = '=HYPERLINK("https://example.com","x")'
        payload["cost_lines"][0]["evidence"].update(
            {
                "evidence_class": "engineering_judgment",
                "explicit_acceptance": True,
                "acceptance_rationale": "Accepted explicitly for export testing.",
            }
        )
        payload["cost_lines"][0]["evidence"]["citation"].update(
            {
                "stable_reference": r"C:\private\evidence-reference.txt",
                "user_supplied_content_sha256": "b" * 64,
            }
        )
        payload["cost_lines"][1]["distribution"] = {
            "family": "bounded_normal",
            "low": 85_000.0,
            "high": 95_000.0,
            "mean": 90_000.0,
            "sd": 2_000.0,
        }
        shared = deepcopy(payload["cost_lines"][0])
        shared.update(
            {
                "input_id": "cost.shared.capex",
                "label": "Shared interconnection CAPEX",
                "ownership": "paired_shared",
                "distribution": {"family": "fixed", "value": 10_000.0},
                "coverage_include_ids": ["interconnection.shared"],
                "solectria_quantity": 1.0,
                "solaredge_quantity": 1.0,
                "normalization_derivation": (
                    "Divide the common project total by each frozen system Wdc."
                ),
            }
        )
        payload["cost_lines"].append(shared)
        parsed = TechnoeconomicSubmissionRequest.model_validate(payload)
        self.request_payload = parsed.model_dump(mode="json", exclude_none=False)
        self.request_payload.pop("capacity_normalization", None)
        self.request_payload.pop("standalone_commercial", None)
        self.snapshot = {
            "schema_version": 1,
            "eligibility_version": technoeconomic_api.ANNUAL_SOURCE_ELIGIBILITY_VERSION,
            "source_annual_job_id": "annual-export-source",
            "eligible_paired_energy_rows": [
                {
                    "year": 2024,
                    "period_start": "2024-01-01",
                    "period_end": "2024-12-31",
                    "sol_predicted_kwh": 200_000.0,
                    "se_predicted_kwh": 200_000.0,
                }
            ],
            "excluded_annual_energy_rows": [
                {
                    "row": {
                        "year": 2023,
                        "period_start": "2023-01-01",
                        "period_end": "2023-12-31",
                        "sol_predicted_kwh": 0.0,
                        "se_predicted_kwh": 0.0,
                        "diagnostic_blob": "x" * 40_000,
                    },
                    "reasons": ["incomplete_annual_coverage"],
                }
            ],
            "capacity_manifest": model.capacity_manifest(),
            "capacity_manifest_source": "phase4_test_fixture",
            "midc_source_artifact": {
                "sha256": "a" * 64,
                "byte_count": 17,
                "media_type": "text/csv",
                "storage_key": "private/source.csv",
            },
            "calibration_lineage": {"promotion": "fixture"},
            "diagnostic_note": (
                r"internal source read from C:\private\annual-source.csv"
            ),
        }
        self.source_snapshot_sha256 = technoeconomic_api.canonical_json_sha256(
            self.snapshot
        )
        self.request = technoeconomic_api.build_technoeconomic_kernel_request(
            self.request_payload,
            self.snapshot,
        )
        self.submission_provenance = (
            technoeconomic_api.build_technoeconomic_submission_provenance(
                self.request_payload,
                {
                    "source_snapshot": self.snapshot,
                    "source_snapshot_sha256": self.source_snapshot_sha256,
                },
                self.request,
            )
        )
        self.calculation = kernel.run_technoeconomic(self.request)
        self.sealed_artifact = run_technoeconomic._write_sealed_calculation_payload(
            self.job_id,
            self.lease_token,
            self.calculation,
            request_sha256=technoeconomic_api.canonical_json_sha256(
                self.request_payload
            ),
            source_snapshot_sha256=self.source_snapshot_sha256,
            submission_provenance_sha256=technoeconomic_api.canonical_json_sha256(
                self.submission_provenance
            ),
            publish_check=lambda: None,
        )
        self.sealed_path = self.output / Path(self.sealed_artifact["storage_key"])
        self.attempt_directory = self.sealed_path.parent
        self.routine_result = run_technoeconomic._routine_result(
            self.request,
            self.calculation,
            self.sealed_artifact,
            self.submission_provenance,
        )

    def _cleanup(self) -> None:
        self.output_patch.stop()
        shutil.rmtree(self.test_root, ignore_errors=True)

    def _generate(self, **overrides):
        arguments = {
            "job_id": self.job_id,
            "attempt_directory": self.attempt_directory,
            "sealed_calculation_path": self.sealed_path,
            "sealed_calculation_artifact": self.sealed_artifact,
            "request_payload": self.request_payload,
            "source_snapshot": self.snapshot,
            "submission_provenance": self.submission_provenance,
            "routine_result": self.routine_result,
            "cancellation_check": lambda: None,
            "publish_check": lambda: None,
        }
        arguments.update(overrides)
        return reporting.generate_technoeconomic_exports(**arguments)

    def test_complete_manifest_hashes_and_shared_artifact_contract(self) -> None:
        manifest = self._generate()
        self.assertEqual("technoeconomic-exports-manifest-v1", manifest["schema_version"])
        self.assertEqual("technoeconomic-csv-v1", manifest["csv_format_version"])
        self.assertEqual(5, manifest["artifact_count"])
        self.assertEqual(
            [
                "csv_bundle",
                "xlsx_workbook",
                "cdf_plot",
                "sensitivity_plot",
                "convergence_plot",
            ],
            list(manifest["artifacts"]),
        )
        expected_contracts = {
            item["artifact_id"]: item
            for item in api_artifacts.TECHNOECONOMIC_PUBLIC_ARTIFACT_CONTRACT.values()
        }
        for artifact_id, artifact in manifest["artifacts"].items():
            contract = expected_contracts[artifact_id]
            self.assertEqual(artifact_id, artifact["artifact_id"])
            self.assertEqual(contract["filename"], artifact["filename"])
            self.assertEqual(contract["media_type"], artifact["media_type"])
            self.assertEqual(contract["artifact_kind"], artifact["artifact_kind"])
            path = self.output / Path(artifact["storage_key"])
            self.assertTrue(path.is_file())
            self.assertEqual(path.stat().st_size, artifact["byte_count"])
            self.assertEqual(
                hashlib.sha256(path.read_bytes()).hexdigest(),
                artifact["sha256"],
            )
        self.assertEqual(
            api_artifacts._canonical_manifest_sha256(manifest),
            manifest["manifest_sha256"],
        )
        self.assertEqual("passed", manifest["tie_outs"]["status"])
        self.assertEqual([], manifest["tie_outs"]["failed_check_ids"])

    def test_v2_exports_applied_capacity_as_distinct_authority(self) -> None:
        request_payload = TechnoeconomicSubmissionRequest.model_validate(
            _applied_site_request_payload(
                source_id="annual-export-source",
                n=24,
            )
        ).model_dump(mode="json", exclude_none=False)
        request_payload.pop("standalone_commercial", None)
        snapshot = deepcopy(self.snapshot)
        snapshot["source_annual_job"] = {
            "request": {
                "curtailment_enabled": True,
                "curtailment_limit_kw": 125.0,
            }
        }
        snapshot_sha256 = technoeconomic_api.canonical_json_sha256(snapshot)
        request = technoeconomic_api.build_technoeconomic_kernel_request(
            request_payload,
            snapshot,
        )
        provenance = technoeconomic_api.build_technoeconomic_submission_provenance(
            request_payload,
            {
                "source_snapshot": snapshot,
                "source_snapshot_sha256": snapshot_sha256,
            },
            request,
        )
        calculation = kernel.run_technoeconomic(request)
        job_id = "tea_export_applied_v2"
        lease_token = "lease_export_applied_v2"
        sealed_artifact = run_technoeconomic._write_sealed_calculation_payload(
            job_id,
            lease_token,
            calculation,
            request_sha256=technoeconomic_api.canonical_json_sha256(
                request_payload
            ),
            source_snapshot_sha256=snapshot_sha256,
            submission_provenance_sha256=technoeconomic_api.canonical_json_sha256(
                provenance
            ),
            publish_check=lambda: None,
        )
        sealed_path = self.output / Path(sealed_artifact["storage_key"])
        routine_result = run_technoeconomic._routine_result(
            request,
            calculation,
            sealed_artifact,
            provenance,
        )

        self.assertEqual(2, routine_result["schema_version"])
        self.assertEqual(
            "frozen_annual_applied_capacity_w",
            routine_result["capacity_basis"],
        )
        self.assertEqual(
            125_000.0,
            routine_result["applied_capacities"]["solectria"][
                "applied_capacity_w"
            ],
        )
        self.assertEqual(
            139_180.8,
            routine_result["capacities"]["solectria"]["installed_wdc"],
        )

        manifest = reporting.generate_technoeconomic_exports(
            job_id=job_id,
            attempt_directory=sealed_path.parent,
            sealed_calculation_path=sealed_path,
            sealed_calculation_artifact=sealed_artifact,
            request_payload=request_payload,
            source_snapshot=snapshot,
            submission_provenance=provenance,
            routine_result=routine_result,
            cancellation_check=lambda: None,
            publish_check=lambda: None,
        )
        self.assertEqual(
            reporting.APPLIED_EXPORT_MANIFEST_SCHEMA_VERSION,
            manifest["schema_version"],
        )
        self.assertEqual(
            reporting.APPLIED_CSV_FORMAT_VERSION,
            manifest["csv_format_version"],
        )
        v2_versions = reporting.export_contract_versions(
            kernel.CALCULATION_CONTRACT_VERSION
        )
        self.assertEqual(reporting.PNG_SCHEMA_VERSION, v2_versions["png"])
        self.assertEqual(
            reporting.XLSX_LOGICAL_HASH_VERSION,
            v2_versions["xlsx_logical_hash"],
        )
        self.assertEqual("passed", manifest["tie_outs"]["status"])
        self.assertIn(
            "USD/applied W",
            reporting._human_metric(
                "lifecycle_cost_delta_se_minus_sol",
                applied_capacity_contract=True,
            ),
        )
        run_technoeconomic._verify_export_manifest(
            job_id,
            lease_token,
            manifest,
            request_sha256=technoeconomic_api.canonical_json_sha256(
                request_payload
            ),
            source_snapshot_sha256=snapshot_sha256,
            submission_provenance_sha256=technoeconomic_api.canonical_json_sha256(
                provenance
            ),
            calculation_contract_version=request.calculation_contract_version,
            sampling_version=request.sampling_version,
            sealed_calculation_sha256=sealed_artifact["sha256"],
        )

        archive_path = self.output / Path(
            manifest["artifacts"]["csv_bundle"]["storage_key"]
        )
        with zipfile.ZipFile(archive_path) as archive:
            self.assertIn("csv-bundle-manifest-v2.json", archive.namelist())
            realizations = list(
                csv.DictReader(
                    io.StringIO(
                        archive.read("realizations.csv").decode("utf-8")
                    )
                )
            )
            self.assertIn(kernel.APPLIED_FIELD_DELTA_COST, realizations[0])
            self.assertNotIn(kernel.FIELD_DELTA_COST, realizations[0])

            capacity_rows = list(
                csv.DictReader(
                    io.StringIO(
                        archive.read("capacity-and-basis.csv").decode("utf-8")
                    )
                )
            )
            self.assertTrue(
                all(row["installed_wdc"] == "139180.8" for row in capacity_rows)
            )
            self.assertTrue(
                all(
                    row["applied_capacity_w"] == "125000.0"
                    for row in capacity_rows
                )
            )
            self.assertTrue(
                all(
                    row["applied_capacity_rating_basis"]
                    == "ac_operating_limit"
                    for row in capacity_rows
                )
            )

            per_year_rows = list(
                csv.DictReader(
                    io.StringIO(
                        archive.read("per-year-summary.csv").decode("utf-8")
                    )
                )
            )
            self.assertEqual("125000.0", per_year_rows[0]["solectria_applied_w"])
            self.assertIn(
                "source_delta_specific_se_minus_sol_kwh_ac_per_applied_w_year",
                per_year_rows[0],
            )

            input_rows = list(
                csv.DictReader(
                    io.StringIO(
                        archive.read("input-specifications.csv").decode("utf-8")
                    )
                )
            )
            cost_row = next(
                row for row in input_rows if row["input_id"] == "cost.sol.capex"
            )
            self.assertEqual("usd_per_applied_w", cost_row["normalized_unit"])
            self.assertEqual(
                "125000.0", cost_row["solectria_applied_capacity_w"]
            )

    def test_v3_exports_dynamic_commercial_scaling_and_independent_tie_outs(self) -> None:
        payload = _commercial_scaling_request_payload(
            target_capacity=100.0,
            target_capacity_unit="mw",
            marginal_cost_value=-2_500_000.0,
        )
        payload["source_annual_job_id"] = "annual-export-source"
        payload["n"] = 8
        request_payload = TechnoeconomicSubmissionRequest.model_validate(
            payload
        ).model_dump(mode="json", exclude_none=False)
        request_payload.pop("standalone_commercial", None)
        snapshot = deepcopy(self.snapshot)
        snapshot["eligible_paired_energy_rows"] = [
            {
                "year": 2024,
                "period_start": "2024-01-01",
                "period_end": "2024-12-31",
                "sol_predicted_kwh": 172_263.0,
                "se_predicted_kwh": 174_227.0,
            }
        ]
        snapshot["source_annual_job"] = {
            "request": {
                "curtailment_enabled": True,
                "curtailment_limit_kw": 125.0,
            }
        }
        snapshot_sha256 = technoeconomic_api.canonical_json_sha256(snapshot)
        request = technoeconomic_api.build_technoeconomic_kernel_request(
            request_payload,
            snapshot,
        )
        provenance = technoeconomic_api.build_technoeconomic_submission_provenance(
            request_payload,
            {
                "source_snapshot": snapshot,
                "source_snapshot_sha256": snapshot_sha256,
            },
            request,
        )
        calculation = kernel.run_technoeconomic(request)
        # Keep the nested attempt path below legacy Windows MAX_PATH.
        job_id = "tea_export_v3"
        lease_token = "lease_v3"
        sealed_artifact = run_technoeconomic._write_sealed_calculation_payload(
            job_id,
            lease_token,
            calculation,
            request_sha256=technoeconomic_api.canonical_json_sha256(request_payload),
            source_snapshot_sha256=snapshot_sha256,
            submission_provenance_sha256=technoeconomic_api.canonical_json_sha256(
                provenance
            ),
            publish_check=lambda: None,
        )
        sealed_path = self.output / Path(sealed_artifact["storage_key"])
        routine_result = run_technoeconomic._routine_result(
            request,
            calculation,
            sealed_artifact,
            provenance,
        )

        self.assertEqual(3, routine_result["schema_version"])
        self.assertEqual(
            100_000_000.0,
            routine_result["commercial_scaling"]["target_capacity_w"],
        )
        self.assertEqual(
            "ac_operating_limit",
            routine_result["commercial_scaling"]["target_rating_basis"],
        )

        manifest = reporting.generate_technoeconomic_exports(
            job_id=job_id,
            attempt_directory=sealed_path.parent,
            sealed_calculation_path=sealed_path,
            sealed_calculation_artifact=sealed_artifact,
            request_payload=request_payload,
            source_snapshot=snapshot,
            submission_provenance=provenance,
            routine_result=routine_result,
            cancellation_check=lambda: None,
            publish_check=lambda: None,
        )
        self.assertEqual(
            reporting.COMMERCIAL_SCALING_EXPORT_MANIFEST_SCHEMA_VERSION,
            manifest["schema_version"],
        )
        self.assertEqual("passed", manifest["tie_outs"]["status"])
        archive_path = self.output / Path(
            manifest["artifacts"]["csv_bundle"]["storage_key"]
        )
        with zipfile.ZipFile(archive_path) as archive:
            self.assertIn("csv-bundle-manifest-v3.json", archive.namelist())
            realizations = list(
                csv.DictReader(
                    io.StringIO(
                        archive.read("realizations.csv").decode("utf-8")
                    )
                )
            )
            first = realizations[0]
            self.assertEqual(
                1_571_200.0,
                float(first[kernel.COMMERCIAL_FIELD_YEAR1_DELTA_ENERGY]),
            )
            self.assertEqual(
                100_000_000.0,
                float(first[kernel.COMMERCIAL_FIELD_TARGET_CAPACITY]),
            )
            lifecycle_ratio = (
                float(first[kernel.COMMERCIAL_FIELD_LIFECYCLE_MARGINAL_COST])
                / float(first[kernel.COMMERCIAL_FIELD_LIFECYCLE_DELTA_ENERGY])
            )
            self.assertAlmostEqual(
                lifecycle_ratio,
                float(first[kernel.COMMERCIAL_FIELD_MARGINAL_LCOO]),
            )
            input_rows = list(
                csv.DictReader(
                    io.StringIO(
                        archive.read("input-specifications.csv").decode("utf-8")
                    )
                )
            )
            marginal_row = next(
                row
                for row in input_rows
                if row["input_id"]
                == kernel.COMMERCIAL_MARGINAL_COST_DIFFERENCE_INPUT_ID
            )
            self.assertEqual("constant_usd", marginal_row["unit"])
            checks = {
                row["check_id"]: row
                for row in csv.DictReader(
                    io.StringIO(archive.read("checks.csv").decode("utf-8"))
                )
            }
            for check_id in (
                "commercial_target_capacity_receipt",
                "commercial_year1_energy_scaling",
                "commercial_lifecycle_energy_scaling",
                "commercial_marginal_cost_sampled_input_authority",
                "commercial_marginal_cost_crf_transform",
                "commercial_marginal_lcoo_lifecycle_ratio",
                "commercial_marginal_lcoo_equivalent_annual_ratio",
                "commercial_zero_energy_reason_matches_energy_class",
            ):
                self.assertEqual("OK", checks[check_id]["status_authority"])

    def test_v4_exports_standalone_lcoe_cdf_and_independent_tie_outs(self) -> None:
        payload = _standalone_commercial_request_payload(n=16)
        payload["source_annual_job_id"] = "annual-export-source"
        # Exercise the finite support boundary where an algebraically equivalent
        # ``(1 + r) ** -year`` implementation differs from the kernel's pinned
        # exp/log1p primitive by enough ULPs to produce a false export failure.
        payload["finance"]["real_discount_rate"]["distribution"] = {
            "family": "fixed",
            "value": -0.999999999999,
        }
        request_payload = TechnoeconomicSubmissionRequest.model_validate(
            payload
        ).model_dump(mode="json", exclude_none=False)
        snapshot = deepcopy(self.snapshot)
        snapshot["source_annual_job"] = {
            "request": {
                "curtailment_enabled": True,
                "curtailment_limit_kw": 125.0,
            }
        }
        snapshot_sha256 = technoeconomic_api.canonical_json_sha256(snapshot)
        request = technoeconomic_api.build_technoeconomic_kernel_request(
            request_payload,
            snapshot,
        )
        provenance = technoeconomic_api.build_technoeconomic_submission_provenance(
            request_payload,
            {
                "source_snapshot": snapshot,
                "source_snapshot_sha256": snapshot_sha256,
            },
            request,
        )
        calculation = kernel.run_technoeconomic(request)
        job_id = "tea_export_v4"
        sealed_artifact = run_technoeconomic._write_sealed_calculation_payload(
            job_id,
            "lease_v4",
            calculation,
            request_sha256=technoeconomic_api.canonical_json_sha256(request_payload),
            source_snapshot_sha256=snapshot_sha256,
            submission_provenance_sha256=technoeconomic_api.canonical_json_sha256(
                provenance
            ),
            publish_check=lambda: None,
        )
        sealed_path = self.output / Path(sealed_artifact["storage_key"])
        routine_result = run_technoeconomic._routine_result(
            request,
            calculation,
            sealed_artifact,
            provenance,
        )

        manifest = reporting.generate_technoeconomic_exports(
            job_id=job_id,
            attempt_directory=sealed_path.parent,
            sealed_calculation_path=sealed_path,
            sealed_calculation_artifact=sealed_artifact,
            request_payload=request_payload,
            source_snapshot=snapshot,
            submission_provenance=provenance,
            routine_result=routine_result,
            cancellation_check=lambda: None,
            publish_check=lambda: None,
        )

        self.assertEqual(
            reporting.STANDALONE_COMMERCIAL_EXPORT_MANIFEST_SCHEMA_VERSION,
            manifest["schema_version"],
        )
        self.assertEqual(
            reporting.STANDALONE_COMMERCIAL_CSV_FORMAT_VERSION,
            manifest["csv_format_version"],
        )
        self.assertEqual("passed", manifest["tie_outs"]["status"])
        self.assertEqual(
            reporting.STANDALONE_COMMERCIAL_CDF_CHART_CONTRACT_ID,
            manifest["artifacts"]["cdf_plot"]["chart_contract_id"],
        )
        self.assertEqual(
            set(reporting.STANDALONE_COMMERCIAL_CHART_CONTRACTS),
            set(manifest["chart_contracts"]),
        )

        archive_path = self.output / Path(
            manifest["artifacts"]["csv_bundle"]["storage_key"]
        )
        with zipfile.ZipFile(archive_path) as archive:
            self.assertIn("csv-bundle-manifest-v4.json", archive.namelist())
            self.assertIn("standalone-commercial-summary.csv", archive.namelist())
            realization = next(
                csv.DictReader(
                    io.StringIO(archive.read("realizations.csv").decode("utf-8"))
                )
            )
            self.assertEqual(
                100_000_000.0,
                float(
                    realization[
                        kernel.COMMERCIAL_STANDALONE_FIELD_TARGET_CAPACITY
                    ]
                ),
            )
            self.assertEqual(
                800.0,
                float(
                    realization[
                        kernel.COMMERCIAL_STANDALONE_FIELD_CAPACITY_SCALE_FACTOR
                    ]
                ),
            )
            self.assertGreater(
                float(realization[kernel.COMMERCIAL_STANDALONE_FIELD_LCOE]),
                0.0,
            )
            cdf_rows = list(
                csv.DictReader(
                    io.StringIO(archive.read("metric-cdfs.csv").decode("utf-8"))
                )
            )
            self.assertEqual(
                routine_result["standalone_commercial"]["cdf"]["source_point_count"],
                len(cdf_rows),
            )
            self.assertEqual(
                {kernel.COMMERCIAL_STANDALONE_FIELD_LCOE},
                {row["metric_id"] for row in cdf_rows},
            )
            self.assertEqual("1.0", cdf_rows[-1]["cumulative_probability"])
            summary_rows = list(
                csv.DictReader(
                    io.StringIO(
                        archive.read("standalone-commercial-summary.csv").decode(
                            "utf-8"
                        )
                    )
                )
            )
            self.assertEqual(4, len(summary_rows))
            self.assertEqual("headline", summary_rows[0]["record_type"])
            self.assertEqual("800.0", summary_rows[0]["capacity_scale_factor"])
            self.assertTrue(summary_rows[0]["p10"])
            self.assertTrue(summary_rows[0]["p50"])
            self.assertTrue(summary_rows[0]["p90"])
            cost_summary = next(
                row
                for row in summary_rows
                if row["input_id"] == "commercial.solaredge.capex"
            )
            self.assertEqual("full_initial_capex", cost_summary["cost_category"])
            self.assertEqual("2026", cost_summary["constant_dollar_cost_year"])
            self.assertEqual(
                ["commercial.solaredge.full-initial-system"],
                json.loads(cost_summary["coverage_ids_json"]),
            )
            input_rows = list(
                csv.DictReader(
                    io.StringIO(
                        archive.read("input-specifications.csv").decode("utf-8")
                    )
                )
            )
            capex_input = next(
                row
                for row in input_rows
                if row["input_id"] == "commercial.solaredge.capex"
            )
            self.assertEqual("full_initial_capex", capex_input["cost_type"])
            self.assertEqual("2026", capex_input["constant_dollar_cost_year"])
            self.assertEqual(
                ["commercial.solaredge.full-initial-system"],
                json.loads(
                    capex_input["coverage_include_ids_json_part_1"]
                    + capex_input["coverage_include_ids_json_part_2"]
                ),
            )
            per_year_rows = list(
                csv.DictReader(
                    io.StringIO(
                        archive.read("per-year-summary.csv").decode("utf-8")
                    )
                )
            )
            self.assertEqual(1, len(per_year_rows))
            per_year = per_year_rows[0]
            self.assertNotIn("source_sol_predicted_kwh_ac", per_year)
            self.assertNotIn("solectria_applied_w", per_year)
            self.assertEqual("125000.0", per_year["source_solaredge_applied_capacity_w"])
            self.assertEqual("ac_operating_limit", per_year["source_solaredge_rating_basis"])
            self.assertEqual("100000000.0", per_year["commercial_target_capacity_w"])
            self.assertEqual("ac_operating_limit", per_year["commercial_target_rating_basis"])
            self.assertEqual(
                "800.0",
                per_year["commercial_capacity_scale_factor_target_w_per_source_w"],
            )
            self.assertEqual("direct_capacity_scaling", per_year["commercial_transfer_method"])
            self.assertEqual(
                "160000000.0",
                per_year["commercial_scaled_target_year1_energy_kwh_ac"],
            )
            lcoe_prefix = kernel.COMMERCIAL_STANDALONE_FIELD_LCOE
            self.assertIn(f"{lcoe_prefix}::p10", per_year)
            self.assertIn(f"{lcoe_prefix}::p50", per_year)
            self.assertIn(f"{lcoe_prefix}::p90", per_year)
            self.assertNotIn(f"{lcoe_prefix}::p5", per_year)
            self.assertNotIn(f"{lcoe_prefix}::p95", per_year)
            checks = {
                row["check_id"]: row
                for row in csv.DictReader(
                    io.StringIO(archive.read("checks.csv").decode("utf-8"))
                )
            }
            for check_id in (
                "standalone_target_capacity_receipt",
                "standalone_source_capacity_receipt",
                "standalone_capacity_scale_factor",
                "standalone_capacity_scale_factor_realizations",
                "standalone_rating_basis_bridge",
                "standalone_full_system_cost_categories",
                "standalone_cost_line_coverage::commercial.solaredge.capex",
                "standalone_cost_line_constant_dollar_year::commercial.solaredge.capex",
                "standalone_year1_energy_scaling",
                "standalone_initial_cost_stack",
                "standalone_recurring_cost_stack",
                "standalone_scheduled_cost_stack",
                "standalone_lifecycle_cost_component_sum",
                "standalone_lcoe_lifecycle_ratio",
                "standalone_lcoe_equivalent_annual_ratio",
            ):
                self.assertEqual("OK", checks[check_id]["status_authority"])

        workbook_path = self.output / Path(
            manifest["artifacts"]["xlsx_workbook"]["storage_key"]
        )
        workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
        try:
            self.assertIn("Commercial LCOE", workbook.sheetnames)
            self.assertIn("Summary", workbook.sheetnames)
            per_year_sheet = workbook["Per-Year Summary"]
            per_year_headers = [cell.value for cell in next(per_year_sheet.iter_rows())]
            self.assertIn("source_solaredge_applied_capacity_w", per_year_headers)
            self.assertIn(
                f"{kernel.COMMERCIAL_STANDALONE_FIELD_LCOE}::p10",
                per_year_headers,
            )
            self.assertNotIn("source_sol_predicted_kwh_ac", per_year_headers)
        finally:
            workbook.close()

        sealed_calculation = reporting._load_sealed_calculation(
            attempt_directory=sealed_path.parent,
            sealed_calculation_path=sealed_path,
            sealed_calculation_artifact=sealed_artifact,
            request_payload=request_payload,
            source_snapshot=snapshot,
            submission_provenance=provenance,
        )
        mismatched_year_provenance = deepcopy(provenance)
        mismatched_year_provenance["standalone_commercial_receipt"][
            "cost_lines"
        ][0]["constant_dollar_cost_year"] = 2025
        year_checks = {
            row[0]: row
            for row in reporting._build_checks(
                sealed_calculation,
                snapshot,
                mismatched_year_provenance,
                routine_result,
            )
        }
        self.assertEqual(
            "FAIL",
            year_checks[
                "standalone_cost_line_constant_dollar_year::commercial.solaredge.capex"
            ][5],
        )

        tampered = deepcopy(routine_result)
        tampered["standalone_commercial"]["capacity_scale_factor"] = 799.0
        with self.assertRaisesRegex(
            reporting.TechnoeconomicExportError,
            "routine result differs",
        ):
            reporting.generate_technoeconomic_exports(
                job_id="tea_export_v4_tampered",
                attempt_directory=sealed_path.parent,
                sealed_calculation_path=sealed_path,
                sealed_calculation_artifact=sealed_artifact,
                request_payload=request_payload,
                source_snapshot=snapshot,
                submission_provenance=provenance,
                routine_result=tampered,
                cancellation_check=lambda: None,
                publish_check=lambda: None,
            )

    def test_v5_exports_paired_lcoe_cdfs_and_per_system_tie_outs(self) -> None:
        payload = _paired_commercial_request_payload(
            target_rating_basis="dc_installed_nameplate",
            n=16,
        )
        payload["source_annual_job_id"] = "annual-export-source"
        request_payload = TechnoeconomicSubmissionRequest.model_validate(
            payload
        ).model_dump(mode="json", exclude_none=False)
        snapshot = deepcopy(self.snapshot)
        snapshot["source_annual_job"] = {
            "request": {
                "curtailment_enabled": False,
                "curtailment_limit_kw": 125.0,
            }
        }
        snapshot_sha256 = technoeconomic_api.canonical_json_sha256(snapshot)
        request = technoeconomic_api.build_technoeconomic_kernel_request(
            request_payload,
            snapshot,
        )
        provenance = technoeconomic_api.build_technoeconomic_submission_provenance(
            request_payload,
            {
                "source_snapshot": snapshot,
                "source_snapshot_sha256": snapshot_sha256,
            },
            request,
        )
        calculation = kernel.run_technoeconomic(request)
        job_id = "tea_export_v5"
        sealed_artifact = run_technoeconomic._write_sealed_calculation_payload(
            job_id,
            "lease_v5",
            calculation,
            request_sha256=technoeconomic_api.canonical_json_sha256(request_payload),
            source_snapshot_sha256=snapshot_sha256,
            submission_provenance_sha256=technoeconomic_api.canonical_json_sha256(
                provenance
            ),
            publish_check=lambda: None,
        )
        sealed_path = self.output / Path(sealed_artifact["storage_key"])
        routine_result = run_technoeconomic._routine_result(
            request,
            calculation,
            sealed_artifact,
            provenance,
        )

        manifest = reporting.generate_technoeconomic_exports(
            job_id=job_id,
            attempt_directory=sealed_path.parent,
            sealed_calculation_path=sealed_path,
            sealed_calculation_artifact=sealed_artifact,
            request_payload=request_payload,
            source_snapshot=snapshot,
            submission_provenance=provenance,
            routine_result=routine_result,
            cancellation_check=lambda: None,
            publish_check=lambda: None,
        )

        self.assertEqual(
            reporting.PAIRED_COMMERCIAL_EXPORT_MANIFEST_SCHEMA_VERSION,
            manifest["schema_version"],
        )
        self.assertEqual(
            reporting.PAIRED_COMMERCIAL_CSV_FORMAT_VERSION,
            manifest["csv_format_version"],
        )
        self.assertEqual("passed", manifest["tie_outs"]["status"])
        self.assertEqual(
            reporting.PAIRED_COMMERCIAL_CDF_CHART_CONTRACT_ID,
            manifest["artifacts"]["cdf_plot"]["chart_contract_id"],
        )
        self.assertEqual(
            set(reporting.PAIRED_COMMERCIAL_CHART_CONTRACTS),
            set(manifest["chart_contracts"]),
        )

        paired = routine_result["paired_commercial"]
        for technology in ("solectria", "solaredge"):
            system = paired["systems"][technology]
            self.assertEqual(
                provenance["paired_commercial_receipt"]["systems"][technology][
                    "source_capacity"
                ]["applied_capacity_w"],
                system["source_applied_capacity_w"],
            )
            self.assertEqual(
                paired["target_capacity_w"] / system["source_applied_capacity_w"],
                system["capacity_scale_factor"],
            )
        archive_path = self.output / Path(
            manifest["artifacts"]["csv_bundle"]["storage_key"]
        )
        with zipfile.ZipFile(archive_path) as archive:
            self.assertIn("csv-bundle-manifest-v5.json", archive.namelist())
            self.assertIn("paired-commercial-summary.csv", archive.namelist())
            realization = next(
                csv.DictReader(
                    io.StringIO(archive.read("realizations.csv").decode("utf-8"))
                )
            )
            for field_name in (
                kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_LCOE,
                kernel.COMMERCIAL_STANDALONE_FIELD_LCOE,
                kernel.COMMERCIAL_PAIRED_FIELD_LCOE_DELTA,
            ):
                self.assertIn(field_name, realization)
            self.assertAlmostEqual(
                float(realization[kernel.COMMERCIAL_PAIRED_FIELD_LCOE_DELTA]),
                float(realization[kernel.COMMERCIAL_STANDALONE_FIELD_LCOE])
                - float(realization[kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_LCOE]),
            )

            cdf_rows = list(
                csv.DictReader(
                    io.StringIO(archive.read("metric-cdfs.csv").decode("utf-8"))
                )
            )
            self.assertEqual(
                {"solectria", "solaredge"},
                {row["technology"] for row in cdf_rows},
            )
            self.assertEqual(
                {
                    kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_LCOE,
                    kernel.COMMERCIAL_STANDALONE_FIELD_LCOE,
                },
                {row["metric_id"] for row in cdf_rows},
            )
            for technology in ("solectria", "solaredge"):
                technology_rows = [
                    row for row in cdf_rows if row["technology"] == technology
                ]
                self.assertEqual(
                    paired["systems"][technology]["cdf"]["source_point_count"],
                    len(technology_rows),
                )
                self.assertEqual(
                    "1.0",
                    technology_rows[-1]["cumulative_probability"],
                )

            summary_rows = list(
                csv.DictReader(
                    io.StringIO(
                        archive.read("paired-commercial-summary.csv").decode("utf-8")
                    )
                )
            )
            self.assertEqual(9, len(summary_rows))
            self.assertEqual(
                {"solectria", "solaredge"},
                {
                    row["technology"]
                    for row in summary_rows
                    if row["record_type"] == "headline"
                },
            )
            self.assertEqual(
                6,
                sum(row["record_type"] == "cost_line" for row in summary_rows),
            )
            self.assertEqual(
                1,
                sum(row["record_type"] == "diagnostic" for row in summary_rows),
            )

            input_rows = list(
                csv.DictReader(
                    io.StringIO(
                        archive.read("input-specifications.csv").decode("utf-8")
                    )
                )
            )
            paired_cost_rows = [
                row
                for row in input_rows
                if row["input_category"] == "paired_commercial_cost"
            ]
            self.assertEqual(6, len(paired_cost_rows))
            self.assertEqual(
                {"solectria_only", "solaredge_only"},
                {row["ownership"] for row in paired_cost_rows},
            )

            per_year_rows = list(
                csv.DictReader(
                    io.StringIO(
                        archive.read("per-year-summary.csv").decode("utf-8")
                    )
                )
            )
            self.assertEqual(1, len(per_year_rows))
            per_year = per_year_rows[0]
            self.assertTrue(per_year["solectria_source_applied_capacity_w"])
            self.assertTrue(per_year["solaredge_source_applied_capacity_w"])
            self.assertIn(
                f"{kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_LCOE}::p10",
                per_year,
            )
            self.assertIn(
                f"{kernel.COMMERCIAL_STANDALONE_FIELD_LCOE}::p90",
                per_year,
            )

            checks = {
                row["check_id"]: row
                for row in csv.DictReader(
                    io.StringIO(archive.read("checks.csv").decode("utf-8"))
                )
            }
            for check_id in (
                "paired_target_capacity_receipt",
                "paired_solectria_source_capacity_receipt",
                "paired_solaredge_source_capacity_receipt",
                "paired_solectria_year1_energy_scaling",
                "paired_solaredge_year1_energy_scaling",
                "paired_solectria_initial_cost_stack",
                "paired_solaredge_recurring_cost_stack",
                "paired_solectria_lcoe_lifecycle_ratio",
                "paired_solaredge_lcoe_equivalent_annual_ratio",
                "paired_lcoe_delta_se_minus_sol",
                "paired_per_year_realization_partition",
            ):
                self.assertEqual("OK", checks[check_id]["status_authority"])

        workbook_path = self.output / Path(
            manifest["artifacts"]["xlsx_workbook"]["storage_key"]
        )
        workbook = openpyxl.load_workbook(
            workbook_path,
            read_only=True,
            data_only=False,
        )
        try:
            self.assertIn("Commercial LCOE", workbook.sheetnames)
            self.assertIn("Summary", workbook.sheetnames)
            per_year_headers = [
                cell.value
                for cell in next(workbook["Per-Year Summary"].iter_rows())
            ]
            self.assertIn("solectria_source_applied_capacity_w", per_year_headers)
            self.assertIn("solaredge_source_applied_capacity_w", per_year_headers)
        finally:
            workbook.close()

        sealed_calculation = reporting._load_sealed_calculation(
            attempt_directory=sealed_path.parent,
            sealed_calculation_path=sealed_path,
            sealed_calculation_artifact=sealed_artifact,
            request_payload=request_payload,
            source_snapshot=snapshot,
            submission_provenance=provenance,
        )
        mismatched_provenance = deepcopy(provenance)
        mismatched_provenance["paired_commercial_receipt"]["systems"][
            "solectria"
        ]["cost_lines"][0]["constant_dollar_cost_year"] = 2025
        mismatched_checks = {
            row[0]: row
            for row in reporting._build_checks(
                sealed_calculation,
                snapshot,
                mismatched_provenance,
                routine_result,
            )
        }
        self.assertEqual(
            "FAIL",
            mismatched_checks[
                "paired_solectria_cost_line_identity::commercial.solectria.capex"
            ][5],
        )

    def test_v6_exports_sealed_audit_workbook_and_explicit_csv_manifest(self) -> None:
        payload = _v6_lifecycle_request_payload(n=8)
        payload["source_annual_job_id"] = "annual-export-source"
        parsed = TechnoeconomicSubmissionRequest.model_validate(payload)
        request_payload = technoeconomic_api.canonical_submission_request_payload(
            parsed
        )
        snapshot = deepcopy(self.snapshot)
        snapshot["source_annual_job"] = {
            "request": {
                "curtailment_enabled": True,
                "curtailment_limit_kw": 125.0,
            }
        }
        snapshot_sha256 = technoeconomic_api.canonical_json_sha256(snapshot)
        request = technoeconomic_api.build_technoeconomic_kernel_request(
            request_payload,
            snapshot,
        )
        provenance = technoeconomic_api.build_technoeconomic_submission_provenance(
            request_payload,
            {
                "source_snapshot": snapshot,
                "source_snapshot_sha256": snapshot_sha256,
            },
            request,
        )
        calculation = kernel.run_technoeconomic(request)
        job_id = "tea_export_v6"
        sealed_artifact = run_technoeconomic._write_sealed_calculation_payload(
            job_id,
            "lease_v6",
            calculation,
            request_sha256=technoeconomic_api.canonical_json_sha256(
                request_payload
            ),
            source_snapshot_sha256=snapshot_sha256,
            submission_provenance_sha256=technoeconomic_api.canonical_json_sha256(
                provenance
            ),
            publish_check=lambda: None,
        )
        sealed_path = self.output / Path(sealed_artifact["storage_key"])
        routine_result = run_technoeconomic._routine_result(
            request,
            calculation,
            sealed_artifact,
            provenance,
        )

        manifest = reporting.generate_technoeconomic_exports(
            job_id=job_id,
            attempt_directory=sealed_path.parent,
            sealed_calculation_path=sealed_path,
            sealed_calculation_artifact=sealed_artifact,
            request_payload=request_payload,
            source_snapshot=snapshot,
            submission_provenance=provenance,
            routine_result=routine_result,
            cancellation_check=lambda: None,
            publish_check=lambda: None,
        )

        self.assertEqual(
            reporting.LIFECYCLE_EXPORT_MANIFEST_SCHEMA_VERSION,
            manifest["schema_version"],
        )
        self.assertEqual(
            reporting.LIFECYCLE_CSV_FORMAT_VERSION,
            manifest["csv_format_version"],
        )
        self.assertEqual("passed", manifest["tie_outs"]["status"])
        self.assertEqual(
            kernel.formula_registry_hash(),
            manifest["formula_registry"]["sha256"],
        )
        self.assertEqual(9, len(manifest["workbook_audit"]["native_charts"]))
        self.assertEqual(3, len(manifest["workbook_audit"]["embedded_images"]))
        self.assertEqual(
            "passed", manifest["workbook_audit"]["formula_scan_status"]
        )
        self.assertTrue(
            all(
                sheet["logical_hash_version"]
                == reporting.LIFECYCLE_XLSX_LOGICAL_HASH_VERSION
                for sheet in manifest["artifacts"]["xlsx_workbook"]["sheets"]
            )
        )

        csv_path = self.output / Path(
            manifest["artifacts"]["csv_bundle"]["storage_key"]
        )
        with zipfile.ZipFile(csv_path) as archive:
            self.assertIn("csv-bundle-manifest-v6.json", archive.namelist())
            csv_manifest = json.loads(
                archive.read("csv-bundle-manifest-v6.json").decode("utf-8")
            )
            self.assertEqual(21, csv_manifest["table_count"])
            self.assertEqual(
                kernel.formula_registry_hash(),
                csv_manifest["provenance"]["formula_registry_sha256"],
            )
            self.assertIn("weather-summary.csv", archive.namelist())
            self.assertIn("representative-event-traces.csv", archive.namelist())
            convergence_rows = list(
                csv.DictReader(
                    io.StringIO(
                        archive.read("convergence.csv").decode("utf-8")
                    )
                )
            )
            decision_probability_rows = [
                row
                for row in convergence_rows
                if row["record_type"] == "decision_probability"
            ]
            self.assertTrue(decision_probability_rows)
            self.assertEqual(
                {"upgrade_npv", "delta_lcoe"},
                {row["metric_id"] for row in decision_probability_rows},
            )
            self.assertTrue(
                all(
                    row["decision_probability_change_threshold"]
                    for row in decision_probability_rows
                )
            )

        workbook_path = self.output / Path(
            manifest["artifacts"]["xlsx_workbook"]["storage_key"]
        )
        repeat_attempt = self.output / "v6-determinism-repeat"
        repeat_attempt.mkdir(parents=True)
        repeat_sealed_path = repeat_attempt / sealed_path.name
        shutil.copy2(sealed_path, repeat_sealed_path)
        repeat_manifest = reporting.generate_technoeconomic_exports(
            job_id=job_id,
            attempt_directory=repeat_attempt,
            sealed_calculation_path=repeat_sealed_path,
            sealed_calculation_artifact=sealed_artifact,
            request_payload=request_payload,
            source_snapshot=snapshot,
            submission_provenance=provenance,
            routine_result=routine_result,
            cancellation_check=lambda: None,
            publish_check=lambda: None,
        )
        repeat_workbook_path = self.output / Path(
            repeat_manifest["artifacts"]["xlsx_workbook"]["storage_key"]
        )
        self.assertEqual(workbook_path.read_bytes(), repeat_workbook_path.read_bytes())
        self.assertEqual(
            manifest["artifacts"]["xlsx_workbook"]["sha256"],
            repeat_manifest["artifacts"]["xlsx_workbook"]["sha256"],
        )
        self.assertEqual(
            manifest["artifacts"]["xlsx_workbook"]["sheets"],
            repeat_manifest["artifacts"]["xlsx_workbook"]["sheets"],
        )
        with zipfile.ZipFile(workbook_path) as first_archive, zipfile.ZipFile(
            repeat_workbook_path
        ) as second_archive:
            first_sheet_hashes = tuple(
                hashlib.sha256(
                    first_archive.read(f"xl/worksheets/sheet{index}.xml")
                ).hexdigest()
                for index in range(1, len(reporting.LIFECYCLE_WORKBOOK_SHEET_ORDER) + 1)
            )
            second_sheet_hashes = tuple(
                hashlib.sha256(
                    second_archive.read(f"xl/worksheets/sheet{index}.xml")
                ).hexdigest()
                for index in range(1, len(reporting.LIFECYCLE_WORKBOOK_SHEET_ORDER) + 1)
            )
        self.assertEqual(first_sheet_hashes, second_sheet_hashes)
        workbook = openpyxl.load_workbook(
            workbook_path,
            read_only=False,
            data_only=False,
        )
        try:
            self.assertEqual(
                list(reporting.LIFECYCLE_WORKBOOK_SHEET_ORDER),
                workbook.sheetnames,
            )
            self.assertEqual(
                len(kernel.formula_registry()),
                workbook["Formula Catalog"].max_row - 1,
            )
            labels = {
                row[1].value
                for row in workbook["Representative Event Traces"].iter_rows(
                    min_row=2
                )
                if row[1].value is not None
            }
            self.assertEqual({"NPV-P10", "NPV-P50", "NPV-P90"}, labels)
            audit = workbook["Calculation Audit"]
            self.assertGreater(audit.max_row, 1)
            self.assertEqual("f", audit["I2"].data_type)
            self.assertEqual("f", audit["J2"].data_type)
            self.assertEqual("f", audit["L2"].data_type)
            self.assertEqual("E7E6E6", audit["H2"].fill.fgColor.rgb[-6:])
            for coordinate in ("I2", "J2", "L2"):
                self.assertEqual("000000", audit[coordinate].font.color.rgb[-6:])
            summary_rows = {
                row[0].value: row
                for row in workbook["Summary"].iter_rows(min_row=5)
                if row[0].value is not None
            }
            weather_interpretation = summary_rows["Weather-path method"][3].value
            self.assertIn("balances weather years across realizations", weather_interpretation)
            self.assertIn("samples project years independently", weather_interpretation)
            self.assertIn("v5's one-weather-year-for-life", weather_interpretation)
            self.assertIn("not a causal performance claim", weather_interpretation)
            self.assertLess(workbook["Provenance"].max_row, 1_000)
            self.assertGreaterEqual(
                workbook["Decision Charts"].column_dimensions["A"].width,
                52.0,
            )
            self.assertGreaterEqual(len(workbook["Decision Charts"]._charts), 2)
            self.assertGreaterEqual(len(workbook["Lifecycle Charts"]._charts), 4)
            self.assertGreaterEqual(len(workbook["Reliability Charts"]._charts), 3)
            decision_probability_axis = workbook["Decision Charts"]._charts[1].y_axis
            self.assertEqual("0%", decision_probability_axis.numFmt.formatCode)
            self.assertEqual(0.0, decision_probability_axis.scaling.min)
            self.assertEqual(1.0, decision_probability_axis.scaling.max)
            self.assertEqual(
                "0.0%",
                workbook["Lifecycle Charts"]._charts[1].y_axis.numFmt.formatCode,
            )
            self.assertEqual(
                "0.000%",
                workbook["Reliability Charts"]._charts[1].y_axis.numFmt.formatCode,
            )
        finally:
            workbook.close()

    def test_v6_logical_formula_identity_ignores_only_a1_row_references(self) -> None:
        first = reporting._new_lifecycle_logical_sheet_hash()
        second = reporting._new_lifecycle_logical_sheet_hash()
        reporting._update_lifecycle_logical_sheet_hash(
            first,
            ("audit", "=I2-H2"),
            formula_identities={1: "audit_difference"},
        )
        reporting._update_lifecycle_logical_sheet_hash(
            second,
            ("audit", "=I901-H901"),
            formula_identities={1: "audit_difference"},
        )
        self.assertEqual(first.hexdigest(), second.hexdigest())

        registry = [dict(kernel.formula_registry()[0])]
        changed = deepcopy(registry)
        changed[0]["excel_template"] = "=SEMANTICALLY_DIFFERENT(InputA,InputB)"
        self.assertNotEqual(
            reporting._formula_template_sha256(registry),
            reporting._formula_template_sha256(changed),
        )

    def test_v6_routine_result_is_exactly_bound_to_sealed_authority(self) -> None:
        payload = _v6_lifecycle_request_payload(n=8)
        payload["source_annual_job_id"] = "annual-export-source"
        request_payload = technoeconomic_api.canonical_submission_request_payload(
            TechnoeconomicSubmissionRequest.model_validate(payload)
        )
        snapshot = deepcopy(self.snapshot)
        snapshot["source_annual_job"] = {
            "request": {
                "curtailment_enabled": True,
                "curtailment_limit_kw": 125.0,
            }
        }
        snapshot_sha256 = technoeconomic_api.canonical_json_sha256(snapshot)
        request = technoeconomic_api.build_technoeconomic_kernel_request(
            request_payload,
            snapshot,
        )
        provenance = technoeconomic_api.build_technoeconomic_submission_provenance(
            request_payload,
            {
                "source_snapshot": snapshot,
                "source_snapshot_sha256": snapshot_sha256,
            },
            request,
        )
        calculation = kernel.run_technoeconomic(request)
        artifact = run_technoeconomic._write_sealed_calculation_payload(
            "tea_export_v6_binding",
            "lease_v6_binding",
            calculation,
            request_sha256=technoeconomic_api.canonical_json_sha256(
                request_payload
            ),
            source_snapshot_sha256=snapshot_sha256,
            submission_provenance_sha256=technoeconomic_api.canonical_json_sha256(
                provenance
            ),
            publish_check=lambda: None,
        )
        sealed_path = self.output / Path(artifact["storage_key"])
        sealed = reporting._load_sealed_calculation(
            attempt_directory=sealed_path.parent,
            sealed_calculation_path=sealed_path,
            sealed_calculation_artifact=artifact,
            request_payload=request_payload,
            source_snapshot=snapshot,
            submission_provenance=provenance,
        )
        routine_result = run_technoeconomic._routine_result(
            request,
            calculation,
            artifact,
            provenance,
        )
        arguments = {
            "metadata": sealed.metadata,
            "request_payload": request_payload,
            "source_snapshot": snapshot,
            "submission_provenance": provenance,
            "sealed_calculation_artifact": artifact,
        }
        reporting._verify_lifecycle_routine_result(
            routine_result=routine_result,
            **arguments,
        )

        tampered_results = []
        for path, value in (
            (("schema_version",), 5),
            (("result_version",), None),
            (("paired_lifecycle", "headline_metric_id"), "delta_lcoe"),
            (("summaries", "headline_decision", "status"), "tampered"),
            (("sealed_calculation", "sha256"), "0" * 64),
        ):
            tampered = deepcopy(routine_result)
            cursor = tampered
            for key in path[:-1]:
                cursor = cursor[key]
            cursor[path[-1]] = value
            tampered_results.append((".".join(path), tampered))
        extra = deepcopy(routine_result)
        extra["unsealed_extra"] = True
        tampered_results.append(("extra field", extra))

        for label, tampered in tampered_results:
            with self.subTest(label=label):
                with self.assertRaisesRegex(
                    reporting.TechnoeconomicExportError,
                    "differs from frozen or sealed authority",
                ):
                    reporting._verify_lifecycle_routine_result(
                        routine_result=tampered,
                        **arguments,
                    )

    def test_v6_finance_palette_marks_warnings_failures_and_formula_roles(self) -> None:
        raw_path = self.output / "v6-finance-palette.xlsx"
        workbook = openpyxl.Workbook(write_only=True)
        summary_sheet = workbook.create_sheet("Summary")
        summary_metadata = {
            "summaries": {
                "headline_decision": {
                    "decision": "Decision suppressed",
                    "status": "suppressed",
                    "reason_codes": ["unstable_convergence"],
                },
                "probability_counts": {
                    "upgrade_npv": {
                        "positive": 0,
                        "negative": 0,
                        "tie": 1,
                        "denominator": 1,
                    },
                    "delta_lcoe": {
                        "positive": 0,
                        "negative": 0,
                        "tie": 1,
                        "denominator": 1,
                    },
                },
                "cost_energy_quadrants": {
                    "cost_neutral_zero_energy_change": {
                        "count": 1,
                        "probability": 1.0,
                    }
                },
                "reliability_summary": [],
                "warnings": [
                    {
                        "code": "accepted_provisional_inputs",
                        "inputs": ["component:solaredge:fixture"],
                    }
                ],
            },
            "convergence": {"status": "unstable"},
        }
        summary_calculation = reporting._SealedCalculation(
            metadata=summary_metadata,
            column_names=(
                "Realization",
                "UpgradeNPV_se_minus_sol_USD",
                "NPVTolerance_USD",
                "LifecycleLCOE_SOL_USD_per_kWh_AC",
                "LifecycleLCOE_SE_USD_per_kWh_AC",
                "IncrementalLCOO_se_minus_sol_USD_per_kWh_AC",
            ),
            columns=(
                np.asarray([0]),
                np.asarray([0.0]),
                np.asarray([1.0]),
                np.asarray([0.10]),
                np.asarray([0.11]),
                np.asarray([0.0]),
            ),
            row_count=1,
        )
        reporting._write_lifecycle_summary_sheet(
            summary_sheet,
            summary_calculation,
            {
                "paired_commercial": {
                    "target_capacity": 100.0,
                    "target_capacity_unit": "mw",
                    "target_rating_basis": "ac_operating_limit",
                    "lifecycle": {
                        "decision_probability_threshold": 0.75,
                        "decision_npv_tolerance_usd_per_target_w": 0.01,
                    },
                }
            },
            {
                "realization_count": 1,
                "calculation_contract_version": "tea-calculation-v6",
                "sampling_version": "tea-lhs-v2",
            },
            summary_metadata,
            (("fixture", 1, 1, 0, 0, "OK", "fixture"),),
        )
        input_sheet = workbook.create_sheet("Input Specifications")
        reporting._write_lifecycle_table_sheet(
            input_sheet,
            reporting._Table(
                "input-specifications.csv",
                "Input Specifications",
                ("field_path", "value_json"),
                lambda: iter((("systems.0.evidence.status", "provisional"),)),
            ),
            lambda: None,
        )
        audit_sheet = workbook.create_sheet("Calculation Audit")
        reporting._write_lifecycle_table_sheet(
            audit_sheet,
            reporting._Table(
                "calculation-audit.csv",
                "Calculation Audit",
                reporting.LIFECYCLE_AUDIT_COLUMNS,
                lambda: iter(
                    (
                        (
                            "audit-1",
                            "V6-F001",
                            "NPV-P50",
                            1,
                            "solaredge",
                            1,
                            "generic-inverter",
                            1.0,
                            "='Representative Event Traces'!B2",
                            0.25,
                            1e-12,
                            "FAIL",
                            100.0,
                            "fixture",
                        ),
                    )
                ),
            ),
            lambda: None,
        )
        checks_sheet = workbook.create_sheet("Checks")
        reporting._write_lifecycle_table_sheet(
            checks_sheet,
            reporting._Table(
                "checks.csv",
                "Checks",
                reporting.CHECK_COLUMNS,
                lambda: iter((("failed-check", 1.0, 2.0, -1.0, 0.0, "FAIL", "fixture"),)),
            ),
            lambda: None,
        )
        workbook.save(raw_path)

        loaded = openpyxl.load_workbook(raw_path, read_only=False, data_only=False)
        try:
            summary = loaded["Summary"]
            warning_labels = {
                "Headline decision",
                "Convergence condition",
                "Provisional-input condition",
                "Warnings",
            }
            seen_warning_labels = set()
            for row in summary.iter_rows(min_row=5):
                if row[0].value in warning_labels:
                    seen_warning_labels.add(row[0].value)
                    self.assertEqual("FFF2CC", row[1].fill.fgColor.rgb[-6:])
            self.assertEqual(warning_labels, seen_warning_labels)
            inputs = loaded["Input Specifications"]
            self.assertEqual("FFF2CC", inputs["B2"].fill.fgColor.rgb[-6:])
            self.assertEqual("008000", inputs["B2"].font.color.rgb[-6:])
            audit = loaded["Calculation Audit"]
            self.assertEqual("E7E6E6", audit["H2"].fill.fgColor.rgb[-6:])
            for coordinate in ("I2", "J2", "L2"):
                self.assertEqual("000000", audit[coordinate].font.color.rgb[-6:])
            self.assertEqual("F4CCCC", audit["L2"].fill.fgColor.rgb[-6:])
            checks = loaded["Checks"]
            self.assertEqual("F4CCCC", checks["F2"].fill.fgColor.rgb[-6:])
            self.assertEqual("000000", checks["H2"].font.color.rgb[-6:])
        finally:
            loaded.close()

    def test_v3_commercial_cost_timing_and_zero_reason_checks_are_independent(self) -> None:
        for timing in ("lifecycle_present_value", "equivalent_annual"):
            with self.subTest(timing=timing):
                payload = _commercial_scaling_request_payload(
                    marginal_cost_timing=timing,
                    marginal_cost_value=-1_250_000.0,
                )
                payload["source_annual_job_id"] = "annual-export-source"
                payload["n"] = 8
                request_payload = TechnoeconomicSubmissionRequest.model_validate(
                    payload
                ).model_dump(mode="json", exclude_none=False)
                request_payload.pop("standalone_commercial", None)
                snapshot = deepcopy(self.snapshot)
                snapshot["eligible_paired_energy_rows"] = [
                    {
                        "year": 2024,
                        "period_start": "2024-01-01",
                        "period_end": "2024-12-31",
                        "sol_predicted_kwh": 172_263.0,
                        "se_predicted_kwh": 174_227.0,
                    }
                ]
                snapshot["source_annual_job"] = {
                    "request": {
                        "curtailment_enabled": True,
                        "curtailment_limit_kw": 125.0,
                    }
                }
                snapshot_sha256 = technoeconomic_api.canonical_json_sha256(snapshot)
                request = technoeconomic_api.build_technoeconomic_kernel_request(
                    request_payload,
                    snapshot,
                )
                provenance = (
                    technoeconomic_api.build_technoeconomic_submission_provenance(
                        request_payload,
                        {
                            "source_snapshot": snapshot,
                            "source_snapshot_sha256": snapshot_sha256,
                        },
                        request,
                    )
                )
                calculation = kernel.run_technoeconomic(request)
                metadata = run_technoeconomic._sealed_metadata(
                    calculation,
                    request_sha256=technoeconomic_api.canonical_json_sha256(
                        request_payload
                    ),
                    source_snapshot_sha256=snapshot_sha256,
                    submission_provenance_sha256=(
                        technoeconomic_api.canonical_json_sha256(provenance)
                    ),
                )
                artifact = dict(self.sealed_artifact)
                artifact["row_count"] = request.n
                routine_result = run_technoeconomic._routine_result(
                    request,
                    calculation,
                    artifact,
                    provenance,
                )

                def sealed_with_columns(columns):
                    return reporting._SealedCalculation(
                        metadata=metadata,
                        column_names=list(calculation.realization_table),
                        columns=columns,
                        row_count=request.n,
                    )

                column_names = list(calculation.realization_table)
                base_columns = [
                    np.asarray(values).copy()
                    for values in calculation.realization_table.values()
                ]
                checks = {
                    row[0]: row
                    for row in reporting._build_checks(
                        sealed_with_columns(base_columns),
                        snapshot,
                        provenance,
                        routine_result,
                    )
                }
                self.assertEqual(
                    "OK",
                    checks[
                        "commercial_marginal_cost_sampled_input_authority"
                    ][5],
                )
                self.assertEqual(
                    "OK",
                    checks[
                        "commercial_zero_energy_reason_matches_energy_class"
                    ][5],
                )

                corrupted_sample_columns = [values.copy() for values in base_columns]
                sampled_name = (
                    "SampledInput::"
                    f"{kernel.COMMERCIAL_MARGINAL_COST_DIFFERENCE_INPUT_ID}"
                )
                corrupted_sample_columns[column_names.index(sampled_name)][0] += 1.0
                corrupted_sample_checks = {
                    row[0]: row
                    for row in reporting._build_checks(
                        sealed_with_columns(corrupted_sample_columns),
                        snapshot,
                        provenance,
                        routine_result,
                    )
                }
                self.assertEqual(
                    "FAIL",
                    corrupted_sample_checks[
                        "commercial_marginal_cost_sampled_input_authority"
                    ][5],
                )
                self.assertEqual(
                    "OK",
                    corrupted_sample_checks[
                        "commercial_marginal_cost_crf_transform"
                    ][5],
                )

                corrupted_class_columns = [values.copy() for values in base_columns]
                corrupted_class_columns[column_names.index("energy_class")][0] = (
                    "zero_lifecycle_gain"
                )
                corrupted_class_checks = {
                    row[0]: row
                    for row in reporting._build_checks(
                        sealed_with_columns(corrupted_class_columns),
                        snapshot,
                        provenance,
                        routine_result,
                    )
                }
                self.assertEqual(
                    "FAIL",
                    corrupted_class_checks[
                        "commercial_zero_energy_reason_matches_energy_class"
                    ][5],
                )
                self.assertEqual(
                    "OK",
                    corrupted_class_checks[
                        "commercial_zero_energy_lcoo_null_and_reason"
                    ][5],
                )

    def test_csv_bundle_is_complete_stable_and_retains_zero_negative_rows(self) -> None:
        manifest = self._generate()
        artifact = manifest["artifacts"]["csv_bundle"]
        path = self.output / Path(artifact["storage_key"])
        expected_names = {
            "csv-bundle-manifest-v1.json",
            "realizations.csv",
            "input-specifications.csv",
            "energy-snapshot.csv",
            "capacity-and-basis.csv",
            "common-cost-audit.csv",
            "commercial-transfer.csv",
            "metric-cdfs.csv",
            "per-year-summary.csv",
            "sensitivity.csv",
            "convergence.csv",
            "provenance.csv",
            "checks.csv",
        }
        with zipfile.ZipFile(path) as archive:
            self.assertEqual(expected_names, set(archive.namelist()))
            for metadata in artifact["tables"]:
                body = archive.read(metadata["filename"])
                self.assertNotIn(b"\r\n", body)
                self.assertEqual(hashlib.sha256(body).hexdigest(), metadata["sha256"])
            bundle_manifest = json.loads(
                archive.read("csv-bundle-manifest-v1.json").decode("utf-8")
            )
            self.assertEqual("UTF-8", bundle_manifest["encoding"])
            self.assertEqual("LF", bundle_manifest["line_terminator"])
            rows = list(
                csv.DictReader(
                    io.StringIO(archive.read("realizations.csv").decode("utf-8"))
                )
            )
            self.assertEqual(24, len(rows))
            self.assertEqual(
                list(self.calculation.realization_table),
                list(rows[0]),
            )
            self.assertTrue(
                all(float(row[kernel.FIELD_DELTA_ENERGY]) == 0.0 for row in rows)
            )
            self.assertTrue(
                any(float(row[kernel.FIELD_DELTA_COST]) < 0.0 for row in rows)
            )
            self.assertEqual(
                float(rows[0][kernel.FIELD_DELTA_COST]),
                self.calculation.realization_table[kernel.FIELD_DELTA_COST][0],
            )
            audit_rows = list(
                csv.DictReader(
                    io.StringIO(archive.read("common-cost-audit.csv").decode("utf-8"))
                )
            )
            self.assertEqual("common_cancelled", audit_rows[0]["comparison_treatment"])
            self.assertEqual(
                "true",
                audit_rows[0]["delta_contribution_se_minus_sol_exactly_zero"],
            )
            input_rows = list(
                csv.DictReader(
                    io.StringIO(
                        archive.read("input-specifications.csv").decode("utf-8")
                    )
                )
            )
            injected = next(row for row in input_rows if row["input_id"] == "cost.sol.capex")
            self.assertTrue(injected["label"].startswith("'="))
            self.assertEqual("solartac_site", injected["analysis_basis"])
            self.assertEqual("full_system", injected["cost_stack_completeness"])
            self.assertEqual("1.0", injected["solectria_quantity"])
            self.assertEqual("0.0", injected["solaredge_quantity"])
            self.assertEqual(
                "Divide the project total by frozen SOL Wdc.",
                injected["normalization_derivation"],
            )
            self.assertEqual(
                "frozen_annual_source_capacity_manifest",
                injected["wdc_denominator_method"],
            )
            self.assertEqual("true", injected["solectria_wdc_denominator_applied"])
            self.assertEqual("139180.8", injected["solectria_wdc_denominator"])
            self.assertEqual("engineering_judgment", injected["evidence_class"])
            self.assertEqual(
                "2026-01-01", injected["citation_publication_or_as_of_date"]
            )
            self.assertEqual("2026-08-13", injected["citation_accessed_date"])
            self.assertEqual(
                "Synthetic values used only by tests.",
                injected["citation_excerpt_or_derivation_note"],
            )
            self.assertEqual(
                "metadata_excerpt_only", injected["citation_preservation_mode"]
            )
            self.assertEqual(
                "b" * 64, injected["citation_user_supplied_content_sha256"]
            )
            self.assertEqual(
                "The synthetic test source has no separately preserved evidence bytes.",
                injected["citation_metadata_only_rationale"],
            )
            self.assertEqual("true", injected["evidence_explicit_acceptance"])
            self.assertEqual(
                "Accepted explicitly for export testing.",
                injected["evidence_acceptance_rationale"],
            )
            self.assertEqual("", injected["citation_stable_reference"])
            bounded = next(row for row in input_rows if row["input_id"] == "cost.se.capex")
            self.assertEqual("bounded_normal", bounded["distribution_family"])
            self.assertEqual("2000.0", bounded["standard_deviation"])
            energy_rows = list(
                csv.DictReader(
                    io.StringIO(archive.read("energy-snapshot.csv").decode("utf-8"))
                )
            )
            self.assertEqual(["eligible", "excluded"], [row["eligibility_status"] for row in energy_rows])
            self.assertEqual("2023", energy_rows[1]["year"])
            self.assertEqual("2023-01-01", energy_rows[1]["period_start"])
            self.assertEqual("0.0", energy_rows[1]["solectria_predicted_kwh_ac"])
            self.assertEqual("0.0", energy_rows[1]["solaredge_predicted_kwh_ac"])
            self.assertEqual(
                '["incomplete_annual_coverage"]', energy_rows[1]["exclusion_reason"]
            )
            self.assertEqual(
                "source_snapshot.excluded_annual_energy_rows[0]",
                energy_rows[1]["source_record_provenance_path"],
            )
            capacity_rows = list(
                csv.DictReader(
                    io.StringIO(
                        archive.read("capacity-and-basis.csv").decode("utf-8")
                    )
                )
            )
            self.assertTrue(all(row["rating_basis"] for row in capacity_rows))
            self.assertTrue(all(row["strings"] for row in capacity_rows))
            self.assertTrue(
                all(row["calibration_physics_version"] for row in capacity_rows)
            )
            provenance_text = archive.read("provenance.csv").decode("utf-8")
            self.assertNotIn("private/source.csv", provenance_text)
            self.assertNotIn("C:\\private\\annual-source.csv", provenance_text)
            self.assertNotIn("C:\\private\\evidence-reference.txt", provenance_text)
            provenance_rows = list(csv.DictReader(io.StringIO(provenance_text)))
            diagnostic_chunks = sorted(
                (
                    row["field_path"],
                    row["value_json"],
                )
                for row in provenance_rows
                if row["section"] == "source_snapshot"
                and row["field_path"].startswith(
                    "excluded_annual_energy_rows[0].row.diagnostic_blob.__json_chunk__"
                )
            )
            self.assertEqual(2, len(diagnostic_chunks))
            self.assertEqual("x" * 40_000, json.loads("".join(row[1] for row in diagnostic_chunks)))
            self.assertTrue(all(len(row[1]) <= 30_000 for row in diagnostic_chunks))
            check_rows = list(
                csv.DictReader(
                    io.StringIO(archive.read("checks.csv").decode("utf-8"))
                )
            )
            checks_by_id = {row["check_id"]: row for row in check_rows}
            for check_id in (
                "lcoe_lifecycle_ratio::solectria",
                "lcoe_lifecycle_ratio::solaredge",
                "crf_cost_transform::solectria",
                "crf_cost_transform::solaredge",
                "crf_energy_transform::solectria",
                "crf_energy_transform::solaredge",
                "zero_energy_lcoo_null_and_reason",
                "negative_energy_signed_lcoo_retention",
            ):
                self.assertEqual("OK", checks_by_id[check_id]["status_authority"])

        repeat_attempt = (
            self.output
            / ".technoeconomic_attempts"
            / "tea_export_repeat"
            / "lease_export_repeat"
        )
        repeat_attempt.mkdir(parents=True)
        repeat_sealed = repeat_attempt / self.sealed_path.name
        shutil.copyfile(self.sealed_path, repeat_sealed)
        repeat_manifest = self._generate(
            job_id="tea_export_repeat",
            attempt_directory=repeat_attempt,
            sealed_calculation_path=repeat_sealed,
        )
        repeat_artifact = repeat_manifest["artifacts"]["csv_bundle"]
        repeat_path = self.output / Path(repeat_artifact["storage_key"])
        self.assertEqual(path.read_bytes(), repeat_path.read_bytes())
        self.assertEqual(artifact["tables"], repeat_artifact["tables"])
        self.assertEqual(
            manifest["artifacts"]["xlsx_workbook"]["sheets"],
            repeat_manifest["artifacts"]["xlsx_workbook"]["sheets"],
        )
        self.assertTrue(
            all(
                len(sheet["logical_sha256"]) == 64
                for sheet in repeat_manifest["artifacts"]["xlsx_workbook"]["sheets"]
            )
        )

    def test_xlsx_streams_all_sheets_with_frozen_checks_and_display_formulas(self) -> None:
        real_workbook = reporting.openpyxl.Workbook
        with patch.object(
            reporting.openpyxl,
            "Workbook",
            wraps=real_workbook,
        ) as workbook_constructor:
            manifest = self._generate()
        workbook_constructor.assert_called_once_with(write_only=True)
        artifact = manifest["artifacts"]["xlsx_workbook"]
        path = self.output / Path(artifact["storage_key"])
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        self.assertEqual(
            [
                "Summary",
                "Realizations",
                "Input Specifications",
                "Energy Snapshot",
                "Capacity and Basis",
                "Common-Cost Audit",
                "Commercial Transfer",
                "Metric CDFs",
                "Per-Year Summary",
                "Sensitivity",
                "Convergence",
                "Provenance",
                "Checks",
            ],
            workbook.sheetnames,
        )
        self.assertEqual(
            25,
            sum(1 for _row in workbook["Realizations"].iter_rows(values_only=True)),
        )
        check_rows = list(workbook["Checks"].iter_rows(values_only=True))
        self.assertEqual("display_formula_status", check_rows[0][-1])
        self.assertTrue(str(check_rows[1][-1]).startswith("=IF("))
        self.assertTrue(all(row[5] == "OK" for row in check_rows[1:]))
        summary_values = list(workbook["Summary"].values)
        self.assertIn("Frozen numeric values are authoritative; formulas are display aids.", summary_values[1])
        self.assertTrue(
            any(
                isinstance(value, str) and value.startswith("=IF(COUNTIF")
                for row in summary_values
                for value in row
            )
        )
        self.assertTrue(artifact["write_only_streaming"])
        self.assertEqual(13, artifact["sheet_count"])
        input_sheet = workbook["Input Specifications"]
        injected_row = next(
            row
            for row in input_sheet.iter_rows(min_row=2)
            if row[0].value == "cost.sol.capex"
        )
        injected_cell = injected_row[2]
        self.assertEqual("s", injected_cell.data_type)
        self.assertTrue(str(injected_cell.value).startswith("="))

    def test_large_sheet_fast_path_preserves_values_formats_security_and_hashes(self) -> None:
        realization_columns = (
            "realization_index",
            "dangerous_text",
            "discount_rate",
            "metric_value",
        )
        realization_rows = (
            (1, "=2+2", 0.05, -3.25),
            (2, "+2", 0.06, 0.0),
            (3, "-2", 0.07, 1.25),
            (4, "@SUM(A1)", 0.08, 2.5),
        )
        cdf_columns = ("metric_id", "cumulative_probability", "value")
        cdf_rows = (
            ("-signed-metric", 0.5, -1.0),
            ("signed-metric", 1.0, 0.0),
        )
        tables = (
            reporting._Table(
                "realizations.csv",
                "Realizations",
                realization_columns,
                lambda: iter(realization_rows),
            ),
            reporting._Table(
                "metric-cdfs.csv",
                "Metric CDFs",
                cdf_columns,
                lambda: iter(cdf_rows),
            ),
        )
        path = self.attempt_directory / "fast-streaming-fixture.xlsx"
        sheets, _row_count = reporting._write_workbook(
            path,
            tables,
            {"realization_count": 4},
            {"summaries": {}},
            (("fixture", 1, 1, 0, 0, "OK", "fixture check"),),
            lambda: None,
        )
        workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
        self.addCleanup(workbook.close)
        realizations = workbook["Realizations"]
        metric_cdfs = workbook["Metric CDFs"]
        self.assertTrue(realizations.sheet_view.showGridLines)
        self.assertTrue(metric_cdfs.sheet_view.showGridLines)
        self.assertEqual("A2", realizations.freeze_panes)
        self.assertEqual("A1:D5", realizations.auto_filter.ref)
        self.assertNotEqual("00000000", realizations["A1"].fill.fgColor.rgb)
        for row_number, expected in enumerate(("=2+2", "+2", "-2", "@SUM(A1)"), start=2):
            cell = realizations.cell(row=row_number, column=2)
            self.assertEqual(expected, cell.value)
            self.assertEqual("s", cell.data_type)
        self.assertEqual("#,##0", realizations.column_dimensions["A"].number_format)
        self.assertEqual("0.0000%", realizations.column_dimensions["C"].number_format)
        self.assertEqual(
            "0.###############",
            realizations.column_dimensions["D"].number_format,
        )
        self.assertFalse(realizations["D2"].has_style)
        self.assertEqual("s", metric_cdfs["A2"].data_type)
        self.assertEqual(
            "0.0000%", metric_cdfs.column_dimensions["B"].number_format
        )
        with self.assertRaisesRegex(
            reporting.TechnoeconomicExportError,
            "lossless XLSX cell limit",
        ):
            reporting._fast_streaming_value(realizations, "x" * 32_768)

        sheet_records = {record["sheet_name"]: record for record in sheets}
        for sheet_name, columns, rows in (
            ("Realizations", realization_columns, realization_rows),
            ("Metric CDFs", cdf_columns, cdf_rows),
        ):
            digest = reporting._new_logical_sheet_hash()
            reporting._update_logical_sheet_hash(digest, columns)
            for row in rows:
                reporting._update_logical_sheet_hash(digest, row)
            self.assertEqual(
                digest.hexdigest(),
                sheet_records[sheet_name]["logical_sha256"],
            )
            self.assertEqual(
                reporting.XLSX_LOGICAL_HASH_VERSION,
                sheet_records[sheet_name]["logical_hash_version"],
            )

        with zipfile.ZipFile(path) as archive:
            worksheet_xml = archive.read("xl/worksheets/sheet2.xml").decode("utf-8")
        body_row = worksheet_xml.split('<row r="2">', 1)[1].split("</row>", 1)[0]
        self.assertNotIn(' s="', body_row)

    def test_maximum_commercial_design_round_trips_without_xlsx_truncation(self) -> None:
        payload = _commercial_request_payload(include_transfer=True)
        payload["source_annual_job_id"] = "annual-export-source"
        payload["n"] = 8
        maximum_text = "\u754c" * 4_000
        for system in ("solectria", "solaredge"):
            for field in (
                "inverter_topology",
                "transformer_topology",
                "bos_scope",
                "labor_productivity_and_rates",
                "commissioning_scope",
            ):
                payload["commercial_reference_design"][system][field] = maximum_text
        parsed = TechnoeconomicSubmissionRequest.model_validate(payload)
        request_payload = parsed.model_dump(mode="json", exclude_none=False)
        request_payload.pop("standalone_commercial", None)
        self.assertGreater(
            len(
                reporting._canonical_json_text(
                    request_payload["commercial_reference_design"]
                )
            ),
            32_767,
        )
        request = technoeconomic_api.build_technoeconomic_kernel_request(
            request_payload,
            self.snapshot,
        )
        provenance = technoeconomic_api.build_technoeconomic_submission_provenance(
            request_payload,
            {
                "source_snapshot": self.snapshot,
                "source_snapshot_sha256": self.source_snapshot_sha256,
            },
            request,
        )
        calculation = kernel.run_technoeconomic(request)
        job_id = "tea_export_commercial_max"
        lease_token = "lease_export_commercial_max"
        sealed_artifact = run_technoeconomic._write_sealed_calculation_payload(
            job_id,
            lease_token,
            calculation,
            request_sha256=technoeconomic_api.canonical_json_sha256(request_payload),
            source_snapshot_sha256=self.source_snapshot_sha256,
            submission_provenance_sha256=technoeconomic_api.canonical_json_sha256(
                provenance
            ),
            publish_check=lambda: None,
        )
        sealed_path = self.output / Path(sealed_artifact["storage_key"])
        routine_result = run_technoeconomic._routine_result(
            request,
            calculation,
            sealed_artifact,
            provenance,
        )
        manifest = reporting.generate_technoeconomic_exports(
            job_id=job_id,
            attempt_directory=sealed_path.parent,
            sealed_calculation_path=sealed_path,
            sealed_calculation_artifact=sealed_artifact,
            request_payload=request_payload,
            source_snapshot=self.snapshot,
            submission_provenance=provenance,
            routine_result=routine_result,
            cancellation_check=lambda: None,
            publish_check=lambda: None,
        )

        csv_path = self.output / Path(
            manifest["artifacts"]["csv_bundle"]["storage_key"]
        )
        with zipfile.ZipFile(csv_path) as archive:
            csv_rows = list(
                csv.DictReader(
                    io.StringIO(
                        archive.read("commercial-transfer.csv").decode("utf-8")
                    )
                )
            )
        csv_fields = {
            row["commercial_reference_design_field_path"]: row[
                "commercial_reference_design_value_json"
            ]
            for row in csv_rows
            if row["record_type"] == "commercial_reference_design_field"
        }
        self.assertTrue(
            any(row["mechanism_rationale"] for row in csv_rows if row["record_type"] == "mechanism")
        )
        self.assertTrue(
            any(
                row["mechanism_evidence_value_json"]
                for row in csv_rows
                if row["record_type"] == "mechanism_evidence_field"
            )
        )

        xlsx_path = self.output / Path(
            manifest["artifacts"]["xlsx_workbook"]["storage_key"]
        )
        workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=False)
        self.addCleanup(workbook.close)
        rows = workbook["Commercial Transfer"].iter_rows(values_only=True)
        headers = next(rows)
        xlsx_records = [dict(zip(headers, row)) for row in rows]
        xlsx_fields = {
            row["commercial_reference_design_field_path"]: row[
                "commercial_reference_design_value_json"
            ]
            for row in xlsx_records
            if row["record_type"] == "commercial_reference_design_field"
        }
        self.assertEqual(csv_fields, xlsx_fields)
        self.assertTrue(all(len(value) <= 32_767 for value in xlsx_fields.values()))

        rebuilt: dict[str, object] = {}
        for field_path, encoded in xlsx_fields.items():
            cursor = rebuilt
            parts = field_path.split(".")
            for part in parts[:-1]:
                cursor = cursor.setdefault(part, {})  # type: ignore[assignment]
            cursor[parts[-1]] = json.loads(encoded)
        self.assertEqual(request_payload["commercial_reference_design"], rebuilt)

    def test_paired_cdf_decision_view_uses_exact_full_width_chart(self) -> None:
        from matplotlib import pyplot as plt
        from matplotlib.collections import PathCollection

        solectria_metric = kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_LCOE
        solaredge_metric = kernel.COMMERCIAL_STANDALONE_FIELD_LCOE
        metadata = {
            "kernel_provenance": {
                "commercial_paired": {"constant_dollar_cost_year": 2022}
            },
            "summaries": {
                solectria_metric: {
                    "cdf": {
                        "values": [0.0, 10.0],
                        "cumulative_probability": [0.5, 1.0],
                        "population_count": 4,
                    },
                    "percentiles": {"p10": 0.0, "p50": 5.0, "p90": 10.0},
                },
                solaredge_metric: {
                    "cdf": {
                        "values": [1.0, 11.0],
                        "cumulative_probability": [0.5, 1.0],
                        "population_count": 4,
                    },
                    "percentiles": {"p10": 1.0, "p50": 6.0, "p90": 11.0},
                },
            }
        }
        captured: dict[str, object] = {}

        def capture_figure(figure, _path, *, layout_bottom=0.035):
            captured["figure"] = figure
            captured["layout_bottom"] = layout_bottom
            return 1600, 1000

        with patch.object(reporting, "_save_figure", side_effect=capture_figure):
            result = reporting._render_cdf_plot(
                metadata,
                self.output / "paired-cdf.png",
                paired_headlines_only=True,
            )

        figure = captured["figure"]
        try:
            self.assertEqual((1600, 1000, 4, 4), result)
            self.assertEqual(0.16, captured["layout_bottom"])
            self.assertEqual(1, len(figure.axes))
            axis = figure.axes[0]
            self.assertGreater(axis.get_position().width, 0.7)
            step_lines = [
                line
                for line in axis.lines
                if line.get_drawstyle() == "steps-post"
            ]
            self.assertEqual(2, len(step_lines))
            self.assertTrue(all(line.get_linewidth() == 3.0 for line in step_lines))
            self.assertEqual(
                [[0.0, 10_000.0], [1_000.0, 11_000.0]],
                sorted(line.get_xdata().tolist() for line in step_lines),
            )
            self.assertIsNone(axis.get_legend())
            marker_offsets = np.vstack(
                [
                    collection.get_offsets()
                    for collection in axis.collections
                    if isinstance(collection, PathCollection)
                ]
            )
            np.testing.assert_allclose(
                np.asarray(sorted(marker_offsets.tolist())),
                np.asarray([[5_000.0, 0.5], [6_000.0, 0.5]]),
            )
            self.assertEqual(
                "Lifecycle LCOE (real 2022 USD/MWh AC)",
                axis.get_xlabel(),
            )
            self.assertEqual("0%", axis.yaxis.get_major_formatter()(0.0))
            self.assertEqual("50%", axis.yaxis.get_major_formatter()(0.5))
            self.assertEqual("100%", axis.yaxis.get_major_formatter()(1.0))
            self.assertEqual(
                {"SolarEdge", "Solectria"},
                {text.get_text() for text in axis.texts},
            )
            self.assertIn(
                "4 runs per system • 2 distinct outcomes per system • "
                "Exact empirical CDF",
                {text.get_text() for text in figure.texts},
            )
            figure_copy = " ".join(text.get_text() for text in figure.texts)
            self.assertIn("SolarEdge P10  1000.00", figure_copy)
            self.assertIn("P50  6000.00", figure_copy)
            self.assertIn("P90  11000.00   USD/MWh AC", figure_copy)
            self.assertIn(
                "Lower median: Solectria by 1000.00 USD/MWh AC",
                figure_copy,
            )
            self.assertEqual(
                "paired_commercial_lcoe_cdf_v3",
                reporting.PAIRED_COMMERCIAL_CDF_CHART_CONTRACT_ID,
            )
            chart_contract = reporting.PAIRED_COMMERCIAL_CHART_CONTRACTS[
                reporting.PAIRED_COMMERCIAL_CDF_CHART_CONTRACT_ID
            ]
            self.assertEqual(
                "paired_right_continuous_empirical_cdf_decision_view",
                chart_contract["variant"],
            )
            self.assertIn("direct curve labels", chart_contract["non_color_cues"])
            self.assertIn(
                "P10/P50/P90 value rows with P50 chart markers",
                chart_contract["non_color_cues"],
            )
            self.assertEqual(
                "constant USD/kWh_AC",
                chart_contract["source_value_unit"],
            )
            self.assertEqual(
                "constant-dollar-year USD/MWh_AC",
                chart_contract["display_value_unit"],
            )
            self.assertEqual(
                [0.0, 10.0],
                metadata["summaries"][solectria_metric]["cdf"]["values"],
            )
        finally:
            plt.close(figure)

    def test_paired_cdf_subtitle_names_systems_when_counts_differ(self) -> None:
        available = [
            (
                kernel.COMMERCIAL_PAIRED_SOLECTRIA_FIELD_LCOE,
                {"cdf": {"values": [1.0, 2.0], "population_count": 16}},
            ),
            (
                kernel.COMMERCIAL_STANDALONE_FIELD_LCOE,
                {"cdf": {"values": [1.0, 2.0, 3.0], "population_count": 12}},
            ),
        ]
        self.assertEqual(
            "Solectria: 16 runs, 2 outcomes • "
            "SolarEdge: 12 runs, 3 outcomes • Exact empirical CDF",
            reporting._paired_cdf_subtitle(available),
        )

    def test_plots_are_nonempty_pngs_with_auditable_contracts(self) -> None:
        manifest = self._generate()
        self.assertEqual(set(reporting.CHART_CONTRACTS), set(manifest["chart_contracts"]))
        for artifact_id in ("cdf_plot", "sensitivity_plot", "convergence_plot"):
            artifact = manifest["artifacts"][artifact_id]
            self.assertEqual("image/png", artifact["media_type"])
            self.assertEqual(1600, artifact["width_px"])
            self.assertEqual(1000, artifact["height_px"])
            self.assertGreater(artifact["byte_count"], 10_000)
            self.assertIn(artifact["chart_contract_id"], manifest["chart_contracts"])
        cdf_artifact = manifest["artifacts"]["cdf_plot"]
        self.assertLessEqual(
            cdf_artifact["display_point_count"],
            cdf_artifact["source_point_count"],
        )
        dense_probability = np.arange(1, 10_001, dtype=np.float64) / 10_000
        selected = reporting._cdf_display_indices(dense_probability)
        self.assertLessEqual(len(selected), 1_200)
        self.assertEqual(0, selected[0])
        self.assertEqual(9_999, selected[-1])
        for quantile in (0.05, 0.5, 0.95):
            expected = int(np.searchsorted(dense_probability, quantile, side="left"))
            self.assertIn(expected, selected)

        sensitivity_path = self.attempt_directory / "qa-sensitivity.png"
        fake_steps = [
            {
                "entry_order": index + 1,
                "predictor_id": f"input.{index:02d}",
                "incremental_r_squared": (30 - index) / 1000,
                "standardized_beta": 0.1,
                "sign": "positive",
            }
            for index in range(30)
        ]
        _width, _height, source_count, display_count = (
            reporting._render_sensitivity_plot(
                {
                    "sensitivity": {
                        "synthetic_response": {
                            "status": "available",
                            "sample_count": 100,
                            "steps": fake_steps,
                        }
                    }
                },
                sensitivity_path,
            )
        )
        self.assertEqual(30, source_count)
        self.assertEqual(15, display_count)

    def test_sealed_payload_and_frozen_identity_tampering_fail_before_exports(self) -> None:
        original = self.sealed_path.read_bytes()
        self.sealed_path.write_bytes(original + b"tamper")
        with self.assertRaisesRegex(reporting.TechnoeconomicExportError, "byte count changed"):
            self._generate()
        self.assertFalse(any(path.suffix in {".zip", ".xlsx", ".png"} for path in self.attempt_directory.iterdir()))

        self.sealed_path.write_bytes(original)
        forged_snapshot = deepcopy(self.snapshot)
        forged_snapshot["eligible_paired_energy_rows"][0]["se_predicted_kwh"] = 999_999.0
        with self.assertRaisesRegex(reporting.TechnoeconomicExportError, "source_snapshot_sha256"):
            self._generate(source_snapshot=forged_snapshot)

    def test_complete_routine_result_projection_is_bound_before_writes(self) -> None:
        drifted_results = []
        for key, replacement in (
            ("summaries", {}),
            ("per_weather_year", {}),
            ("sensitivity", {}),
            ("convergence", {}),
            ("common_cost_audit", []),
            ("energy_available", False),
            ("seed", self.routine_result["seed"] + 1),
        ):
            drifted = deepcopy(self.routine_result)
            drifted[key] = replacement
            drifted_results.append((key, drifted))
        sealed_identity_drift = deepcopy(self.routine_result)
        sealed_identity_drift["sealed_calculation"]["sha256"] = "c" * 64
        drifted_results.append(("sealed_calculation.sha256", sealed_identity_drift))

        for label, drifted in drifted_results:
            with self.subTest(label=label):
                with self.assertRaisesRegex(
                    reporting.TechnoeconomicExportError,
                    "Durable routine result differs",
                ):
                    self._generate(routine_result=drifted)
        self.assertFalse(
            any(
                path.suffix in {".zip", ".xlsx", ".png"}
                for path in self.attempt_directory.iterdir()
            )
        )

    def test_distribution_cdf_and_scale_aware_numeric_contracts(self) -> None:
        self.assertEqual(
            ("fixed", 1.0, None, None, None, None, None),
            reporting._distribution_columns({"family": "fixed", "value": 1.0}),
        )
        self.assertEqual(
            ("uniform", None, 1.0, None, 2.0, None, None),
            reporting._distribution_columns(
                {"family": "uniform", "low": 1.0, "high": 2.0}
            ),
        )
        self.assertEqual(
            ("triangular", None, 1.0, 2.0, 3.0, None, None),
            reporting._distribution_columns(
                {"family": "triangular", "low": 1.0, "mode": 2.0, "high": 3.0}
            ),
        )
        self.assertEqual(
            ("bounded_normal", None, 1.0, None, 3.0, 2.0, 0.25),
            reporting._distribution_columns(
                {
                    "family": "bounded_normal",
                    "low": 1.0,
                    "high": 3.0,
                    "mean": 2.0,
                    "sd": 0.25,
                }
            ),
        )
        tolerance = reporting._binary64_tie_out_tolerance(
            np.asarray([3_850_000.0]),
            np.asarray([-3_850_000.0]),
        )
        self.assertGreater(tolerance, 1.21e-9)
        self.assertLess(tolerance, 1e-7)
        coverage_ids = [
            f"coverage.{index:03d}." + "a" * 147 for index in range(256)
        ]
        coverage_parts = reporting._coverage_columns(coverage_ids)
        self.assertGreater(len("".join(coverage_parts)), 32_767)
        self.assertTrue(all(len(part) <= 30_000 for part in coverage_parts))
        self.assertEqual(coverage_ids, json.loads("".join(coverage_parts)))

        calculation = reporting._load_sealed_calculation(
            attempt_directory=self.attempt_directory,
            sealed_calculation_path=self.sealed_path,
            sealed_calculation_artifact=self.sealed_artifact,
            request_payload=self.request_payload,
            source_snapshot=self.snapshot,
            submission_provenance=self.submission_provenance,
        )
        corrupted_metadata = deepcopy(calculation.metadata)
        metric_id = next(
            key
            for key, summary in corrupted_metadata["summaries"].items()
            if isinstance(summary, dict) and isinstance(summary.get("cdf"), dict)
        )
        corrupted_metadata["summaries"][metric_id]["cdf"]["cumulative_count"][0] += 1
        corrupted = reporting._SealedCalculation(
            metadata=corrupted_metadata,
            column_names=calculation.column_names,
            columns=calculation.columns,
            row_count=calculation.row_count,
        )
        checks = reporting._build_checks(
            corrupted,
            self.snapshot,
            self.submission_provenance,
            self.routine_result,
        )
        cdf_check = next(
            row for row in checks if row[0] == f"cdf_full_identity::{metric_id}"
        )
        self.assertEqual("FAIL", cdf_check[5])

    def test_negative_energy_rows_retain_signed_lcoo(self) -> None:
        snapshot = deepcopy(self.snapshot)
        snapshot["eligible_paired_energy_rows"][0]["se_predicted_kwh"] = 150_000.0
        snapshot_sha256 = technoeconomic_api.canonical_json_sha256(snapshot)
        request = technoeconomic_api.build_technoeconomic_kernel_request(
            self.request_payload,
            snapshot,
        )
        provenance = technoeconomic_api.build_technoeconomic_submission_provenance(
            self.request_payload,
            {
                "source_snapshot": snapshot,
                "source_snapshot_sha256": snapshot_sha256,
            },
            request,
        )
        calculation = kernel.run_technoeconomic(request)
        sealed_metadata = run_technoeconomic._sealed_metadata(
            calculation,
            request_sha256=technoeconomic_api.canonical_json_sha256(
                self.request_payload
            ),
            source_snapshot_sha256=snapshot_sha256,
            submission_provenance_sha256=technoeconomic_api.canonical_json_sha256(
                provenance
            ),
        )
        sealed = reporting._SealedCalculation(
            metadata=sealed_metadata,
            column_names=list(calculation.realization_table),
            columns=list(calculation.realization_table.values()),
            row_count=request.n,
        )
        routine_result = run_technoeconomic._routine_result(
            request,
            calculation,
            self.sealed_artifact,
            provenance,
        )
        checks = reporting._build_checks(
            sealed,
            snapshot,
            provenance,
            routine_result,
        )
        retained = next(
            row for row in checks if row[0] == "negative_energy_signed_lcoo_retention"
        )
        self.assertEqual("OK", retained[5])
        self.assertEqual(request.n, retained[1])
        self.assertTrue(
            np.all(np.isfinite(calculation.realization_table[kernel.FIELD_LCOO]))
        )
        self.assertTrue(
            np.all(
                calculation.realization_table["energy_class"]
                == "negative_lifecycle_gain"
            )
        )

    def test_cancellation_leaves_no_pending_or_staging_files(self) -> None:
        calls = 0

        def cancel_during_second_publish() -> None:
            nonlocal calls
            calls += 1
            if calls == 4:
                raise RuntimeError("cancelled")

        with self.assertRaisesRegex(RuntimeError, "cancelled"):
            self._generate(cancellation_check=cancel_during_second_publish)
        names = {path.name for path in self.attempt_directory.iterdir()}
        self.assertFalse(any(name.startswith(".pending") or name.startswith(".raw") for name in names))
        self.assertNotIn(".csv-staging", names)

    def test_xlsx_archive_normalization_checks_cancellation_during_copy(self) -> None:
        source = self.attempt_directory / "normalization-source.xlsx"
        target = self.attempt_directory / "normalization-target.xlsx"
        with zipfile.ZipFile(source, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("xl/worksheets/sheet1.xml", b"x" * (2 * 1024 * 1024))
        calls = 0

        def cancelled() -> None:
            nonlocal calls
            calls += 1
            raise RuntimeError("cancelled during normalization")

        with self.assertRaisesRegex(RuntimeError, "cancelled during normalization"):
            reporting._normalize_xlsx_archive(source, target, cancelled)
        self.assertEqual(1, calls)

    def test_v6_xlsx_normalization_canonicalizes_only_modified_timestamp(self) -> None:
        def write_source(path: Path, modified: str) -> None:
            core_properties = (
                '<cp:coreProperties xmlns:cp="urn:cp" xmlns:dcterms="urn:dcterms">'
                f'<dcterms:modified xsi:type="dcterms:W3CDTF" '
                f'xmlns:xsi="urn:xsi">{modified}</dcterms:modified>'
                "</cp:coreProperties>"
            ).encode("utf-8")
            with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
                archive.writestr("docProps/core.xml", core_properties)
                archive.writestr("xl/workbook.xml", b"same workbook payload")

        source_a = self.attempt_directory / "v6-core-a.xlsx"
        source_b = self.attempt_directory / "v6-core-b.xlsx"
        normalized_a = self.attempt_directory / "v6-normalized-a.xlsx"
        normalized_b = self.attempt_directory / "v6-normalized-b.xlsx"
        legacy_normalized = self.attempt_directory / "legacy-normalized.xlsx"
        write_source(source_a, "2026-09-03T01:02:03Z")
        write_source(source_b, "2026-09-03T04:05:06Z")

        reporting._normalize_xlsx_archive(
            source_a,
            normalized_a,
            lambda: None,
            canonicalize_core_properties=True,
        )
        reporting._normalize_xlsx_archive(
            source_b,
            normalized_b,
            lambda: None,
            canonicalize_core_properties=True,
        )
        self.assertEqual(normalized_a.read_bytes(), normalized_b.read_bytes())
        with zipfile.ZipFile(normalized_a) as archive:
            core = archive.read("docProps/core.xml")
        self.assertIn(b"1980-01-01T00:00:00Z", core)

        reporting._normalize_xlsx_archive(
            source_a,
            legacy_normalized,
            lambda: None,
        )
        with zipfile.ZipFile(legacy_normalized) as archive:
            legacy_core = archive.read("docProps/core.xml")
        self.assertIn(b"2026-09-03T01:02:03Z", legacy_core)


if __name__ == "__main__":
    unittest.main()
